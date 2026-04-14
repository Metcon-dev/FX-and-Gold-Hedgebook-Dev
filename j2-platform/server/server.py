"""
JÃ‚Â² FX & Gold Hedging Platform Ã¢â‚¬â€ Flask API Server
Wraps existing Python services (models/, services/) into REST endpoints.
"""
import sys
import os
import re
import io
import json
import math
import csv
import base64
import hashlib
import hmac
import smtplib
import sqlite3
import threading
import traceback
import time
from collections import Counter
from copy import copy
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from email.message import EmailMessage
from typing import Any, Dict, List, Optional, Tuple

# Add the original project to sys.path so we can import its modules
# server.py lives at JupyterProject2/j2-platform/server/server.py
# so the parent project (JupyterProject2) is 2 levels up
ORIGINAL_PROJECT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, ORIGINAL_PROJECT)
# Also set CWD so the db path resolves correctly
os.chdir(ORIGINAL_PROJECT)
# Shared DBs are hosted on the network T: drive for this deployment.
SHARED_DB_ROOT = r"T:\Trading Platform - db"


def _load_env_file_early(path: str) -> None:
    try:
        if not os.path.exists(path):
            return
        with open(path, "r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip()
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception:
        return


_SERVER_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.abspath(os.path.join(_SERVER_DIR, "..", ".."))
_load_env_file_early(os.path.join(_PROJECT_ROOT, ".env"))
_load_env_file_early(os.path.join(_SERVER_DIR, ".env"))


def _resolve_db_path(file_name: str, env_keys: Optional[List[str]] = None) -> str:
    if env_keys:
        for k in env_keys:
            raw = str(os.getenv(k, "") or "").strip()
            if not raw:
                continue
            expanded = os.path.abspath(os.path.expanduser(raw))
            try:
                if os.path.exists(expanded) and os.path.isfile(expanded):
                    return expanded
            except Exception:
                continue
    server_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.abspath(os.path.join(server_dir, "..", ".."))
    candidates = [
        os.path.join(SHARED_DB_ROOT, file_name),
        os.path.join(project_root, file_name),
        os.path.join(server_dir, file_name),
        os.path.join(server_dir, "..", file_name),
    ]
    for path in candidates:
        try:
            if os.path.exists(path) and os.path.isfile(path):
                return os.path.abspath(path)
        except Exception:
            continue
    return os.path.abspath(candidates[0])


FORCED_LEDGER_DB_PATH = _resolve_db_path("fx_trading_ledger.db", env_keys=["LEDGER_DB_PATH", "FX_LEDGER_DB_PATH"])
FORCED_PMX_DB_PATH = _resolve_db_path("pmx_database.db", env_keys=["PMX_DB_PATH", "PMX_DATABASE_PATH"])
os.environ["LEDGER_DB_PATH"] = FORCED_LEDGER_DB_PATH
os.environ["PMX_DB_PATH"] = FORCED_PMX_DB_PATH

from flask import Flask, request, jsonify, send_file, Response, g
from flask_cors import CORS
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
import pandas as pd
import requests
import ssl

from models.database import initialize_database, get_db_connection
from models.trade import (
    load_all_trades,
    add_new_trade,
    update_trade_order_id,
    backup_manual_trades_to_json,
    restore_manual_trades_from_json,
)
from services.trade_service import (
    format_ledger_like_example,
    calculate_breakdown_excel_format,
)
from services.trademc_service import (
    sync_trademc_trades,
    sync_trademc_companies,
    sync_trademc_weight_transactions,
    load_trademc_trades,
    get_trademc_stats,
    get_trademc_trades_for_trade_number,
    get_unique_weight_types,
    initialize_trademc_table,
    get_latest_trademc_market_prices,
    update_trademc_trade_ref_number,
    get_local_trademc_snapshot_stats,
    get_remote_trademc_snapshot_stats,
    fetch_trademc_trade_by_id,
)
from services.rest_service import (
    fetch_pmx_alldeal_filter_report,
    fetch_pmx_account_statement_report,
    fetch_pmx_load_account,
    fetch_pmx_fixinvoice_pdf,
    extract_pmx_report_rows,
    extract_pmx_statement_report_rows,
)
from services.clean_data_pipeline import (
    initialize_clean_pipeline_db,
    run_clean_data_pipeline,
)
from services.forecast_service import (
    get_forecast as _forecast_get,
    fetch_current_price as _forecast_current_price,
    invalidate_cache as _forecast_invalidate_cache,
)
from services.purchases_ml_service import (
    train_model as _purchases_train_model,
    get_ml_forecast as _purchases_get_forecast,
    get_model_status as _purchases_model_status,
)

def _load_env_file(path: str = ".env"):
    try:
        if not os.path.exists(path):
            return
        with open(path, "r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip()
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception:
        return

_load_env_file()
try:
    initialize_database()
except Exception as exc:
    print(f"[WARN] initialize_database failed: {exc}")
try:
    initialize_trademc_table()
except Exception as exc:
    print(f"[WARN] initialize_trademc_table failed: {exc}")
try:
    restore_manual_trades_from_json(only_blank=True)
except Exception as exc:
    print(f"[WARN] Startup restore of manual trade numbers failed: {exc}")
API_BUILD = "2026-03-17-t-drive-db-no-implicit-fiscal-cutoff"
PMX_DB_PATH = FORCED_PMX_DB_PATH
LEDGER_DB_PATH = FORCED_LEDGER_DB_PATH
PMX_SUPPORT_DOC_PATTERN = re.compile(r"(?:FNC|SWT|FCT)/[^\s,;]+", re.IGNORECASE)
# Hardcoded PMX session defaults per user request.
# Note: x-auth/sid can expire and may need to be replaced.
PMX_HARDCODED_X_AUTH = "3f4d9533-9e8e-43e1-b95c-ec7a1b9490a1"
PMX_HARDCODED_SID = "505775"
PMX_HARDCODED_USERNAME = "TL1671"
PMX_HARDCODED_PLATFORM = "Desktop"
PMX_HARDCODED_LOCATION = "LD"
PMX_HARDCODED_CACHE_CONTROL = "no-cache"
PMX_HARDCODED_CONTENT_TYPE = "application/json; charset=utf-8"
PMX_LOGIN_DEFAULT_PATH = "/restlogin"
PMX_EXPORT_TRADES_DIR = os.getenv("PMX_EXPORT_TRADES_DIR", r"T:\Platform Doc Testing")
SALES_ORDER_TEMPLATE_PATH = os.path.join(ORIGINAL_PROJECT, "Sales Order Format.xlsx")
GRAMS_PER_TROY_OUNCE = 31.1035
RECON_USD_SYMBOLS = {"USDZAR", "XAUUSD", "XPTUSD", "XPDUSD", "XAGUSD"}
RECON_USD_METAL_SYMBOLS = {"XAUUSD", "XPTUSD", "XPDUSD", "XAGUSD"}
RECON_JOURNAL_LIKE_ROW_TYPES = {"JRV", "MER", "JRC", "OTHER"}
_TRUTHY_VALUES = {"1", "true", "yes", "y", "on"}
FISCAL_TRADES_START_DATE = str(os.getenv("FISCAL_TRADES_START_DATE", "") or "").strip()
ENABLE_FISCAL_DATE_FILTER = str(os.getenv("ENABLE_FISCAL_DATE_FILTER", "false")).strip().lower() in _TRUTHY_VALUES
ENABLE_FISCAL_PURGE = str(os.getenv("ENABLE_FISCAL_PURGE", "false")).strip().lower() in _TRUTHY_VALUES
TRADE_DATA_START_DATE = str(os.getenv("TRADE_DATA_START_DATE", "2026-03-02") or "2026-03-02").strip() or "2026-03-02"
PMX_SESSION_CACHE: Dict[str, Any] = {}
PMX_SESSION_LOCK = threading.Lock()
try:
    TRADEMC_LIVE_PRICES_TTL_SECONDS = max(2, int(os.getenv("TRADEMC_LIVE_PRICES_TTL_SECONDS", "15") or 15))
except Exception:
    TRADEMC_LIVE_PRICES_TTL_SECONDS = 15
try:
    TRADEMC_LIVE_PRICES_SAMPLE_SIZE = max(10, int(os.getenv("TRADEMC_LIVE_PRICES_SAMPLE_SIZE", "40") or 40))
except Exception:
    TRADEMC_LIVE_PRICES_SAMPLE_SIZE = 40
TRADEMC_LIVE_PRICES_CACHE: Dict[str, Any] = {}
TRADEMC_LIVE_PRICES_LOCK = threading.Lock()
try:
    HEAVY_ROUTE_CACHE_TTL_SECONDS = max(5, int(os.getenv("HEAVY_ROUTE_CACHE_TTL_SECONDS", "20") or 20))
except Exception:
    HEAVY_ROUTE_CACHE_TTL_SECONDS = 20
HEAVY_ROUTE_CACHE: Dict[str, Any] = {}
HEAVY_ROUTE_CACHE_LOCK = threading.Lock()
AUTH_USER_TABLE = "app_users"
AUTH_COOKIE_NAME = "j2_auth_token"
AUTH_EXEMPT_PATHS = {
    "/api/health",
    "/api/auth/login",
    "/api/auth/logout",
    "/api/auth/me",
    "/api/reports/daily-trading/send-test",
}
try:
    AUTH_PASSWORD_ITERATIONS = max(120000, int(os.getenv("APP_AUTH_PBKDF2_ITERATIONS", "240000") or 240000))
except Exception:
    AUTH_PASSWORD_ITERATIONS = 240000
try:
    AUTH_SESSION_SECONDS = max(900, int(os.getenv("APP_AUTH_SESSION_SECONDS", "43200") or 43200))
except Exception:
    AUTH_SESSION_SECONDS = 43200
AUTH_COOKIE_SECURE = str(os.getenv("APP_AUTH_COOKIE_SECURE", "false")).strip().lower() in {"1", "true", "yes", "y", "on"}
_AUTH_SERIALIZER: Optional[URLSafeTimedSerializer] = None
CLEAN_PIPELINE_LOCK = threading.Lock()
_CLEAN_PIPELINE_STATE: Dict[str, Any] = {
    "running": False,
    "last_reason": "",
    "last_result": None,
    "last_started_at": None,
}


def _run_clean_pipeline(reason: str) -> Dict[str, Any]:
    with CLEAN_PIPELINE_LOCK:
        if bool(_CLEAN_PIPELINE_STATE.get("running")):
            return {
                "ok": False,
                "skipped": True,
                "error": "clean_pipeline_already_running",
                "reason": reason,
            }
        _CLEAN_PIPELINE_STATE["running"] = True
        _CLEAN_PIPELINE_STATE["last_reason"] = reason
        _CLEAN_PIPELINE_STATE["last_started_at"] = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

    try:
        result = run_clean_data_pipeline(
            ledger_db_path=LEDGER_DB_PATH,
            pmx_db_path=PMX_DB_PATH,
        )
        result["reason"] = reason
        return result
    except Exception as exc:
        return {"ok": False, "error": str(exc), "reason": reason}
    finally:
        with CLEAN_PIPELINE_LOCK:
            _CLEAN_PIPELINE_STATE["running"] = False
            _CLEAN_PIPELINE_STATE["last_result"] = result if "result" in locals() else {"ok": False, "reason": reason}


def _trigger_clean_pipeline(reason: str, wait: bool = False) -> Dict[str, Any]:
    if wait:
        return _run_clean_pipeline(reason)

    def _bg() -> None:
        res = _run_clean_pipeline(reason)
        if not bool(res.get("ok")) and not bool(res.get("skipped")):
            print(f"[CLEAN-PIPELINE][WARN] {reason}: {res.get('error')}")

    t = threading.Thread(target=_bg, daemon=True)
    t.start()
    return {"ok": True, "queued": True, "reason": reason}

# Daily PMX balances PDF email scheduler (runs in-process with Flask server).
DAILY_BALANCE_EMAIL_ENABLED = str(os.getenv("PMX_DAILY_BALANCE_EMAIL_ENABLED", "true")).strip().lower() in {
    "1",
    "true",
    "yes",
    "y",
    "on",
}
try:
    DAILY_BALANCE_EMAIL_HOUR = min(23, max(0, int(os.getenv("PMX_DAILY_BALANCE_EMAIL_HOUR", "22") or 22)))
except Exception:
    DAILY_BALANCE_EMAIL_HOUR = 22
try:
    DAILY_BALANCE_EMAIL_MINUTE = min(59, max(0, int(os.getenv("PMX_DAILY_BALANCE_EMAIL_MINUTE", "0") or 0)))
except Exception:
    DAILY_BALANCE_EMAIL_MINUTE = 0
try:
    DAILY_BALANCE_EMAIL_CHECK_INTERVAL_SECONDS = max(
        15, int(os.getenv("PMX_DAILY_BALANCE_EMAIL_CHECK_INTERVAL_SECONDS", "30") or 30)
    )
except Exception:
    DAILY_BALANCE_EMAIL_CHECK_INTERVAL_SECONDS = 30
try:
    DAILY_BALANCE_EMAIL_RETRY_SECONDS = max(
        60, int(os.getenv("PMX_DAILY_BALANCE_EMAIL_RETRY_SECONDS", "900") or 900)
    )
except Exception:
    DAILY_BALANCE_EMAIL_RETRY_SECONDS = 900
try:
    DAILY_BALANCE_EMAIL_REQUEST_TIMEOUT_SECONDS = max(
        15, int(os.getenv("PMX_DAILY_BALANCE_EMAIL_REQUEST_TIMEOUT_SECONDS", "60") or 60)
    )
except Exception:
    DAILY_BALANCE_EMAIL_REQUEST_TIMEOUT_SECONDS = 60
DAILY_BALANCE_EMAIL_PDF_PATH = (
    str(os.getenv("PMX_BALANCE_PDF_PATH", "/user/export_NOPMgrPos_pdf") or "/user/export_NOPMgrPos_pdf").strip()
    or "/user/export_NOPMgrPos_pdf"
)
DAILY_BALANCE_EMAIL_TRADE_NAME = (
    str(os.getenv("PMX_TRADE_NAME", "Metal Concentrators") or "Metal Concentrators").strip()
    or "Metal Concentrators"
)
DAILY_BALANCE_EMAIL_SUBJECT_PREFIX = (
    str(os.getenv("BALANCE_EMAIL_SUBJECT_PREFIX", "StoneX Account Balances") or "StoneX Account Balances").strip()
    or "StoneX Account Balances"
)
DAILY_BALANCE_EMAIL_JOB_NAME = "pmx_daily_balance_pdf_email"
_daily_balance_email_scheduler_lock = threading.Lock()
_daily_balance_email_scheduler_started = False
_daily_balance_email_last_attempt_epoch: Dict[str, float] = {}

# Daily Trading Report HTML email scheduler.
DAILY_TRADING_REPORT_EMAIL_ENABLED = str(
    os.getenv("DAILY_TRADING_REPORT_EMAIL_ENABLED", "true")
).strip().lower() in {"1", "true", "yes", "y", "on"}
try:
    DAILY_TRADING_REPORT_EMAIL_HOUR = min(
        23, max(0, int(os.getenv("DAILY_TRADING_REPORT_EMAIL_HOUR", "18") or 18))
    )
except Exception:
    DAILY_TRADING_REPORT_EMAIL_HOUR = 18
try:
    DAILY_TRADING_REPORT_EMAIL_MINUTE = min(
        59, max(0, int(os.getenv("DAILY_TRADING_REPORT_EMAIL_MINUTE", "0") or 0))
    )
except Exception:
    DAILY_TRADING_REPORT_EMAIL_MINUTE = 0
try:
    DAILY_TRADING_REPORT_EMAIL_CHECK_INTERVAL_SECONDS = max(
        15, int(os.getenv("DAILY_TRADING_REPORT_EMAIL_CHECK_INTERVAL_SECONDS", "30") or 30)
    )
except Exception:
    DAILY_TRADING_REPORT_EMAIL_CHECK_INTERVAL_SECONDS = 30
try:
    DAILY_TRADING_REPORT_EMAIL_RETRY_SECONDS = max(
        60, int(os.getenv("DAILY_TRADING_REPORT_EMAIL_RETRY_SECONDS", "900") or 900)
    )
except Exception:
    DAILY_TRADING_REPORT_EMAIL_RETRY_SECONDS = 900
DAILY_TRADING_REPORT_EMAIL_SUBJECT_PREFIX = str(
    os.getenv("DAILY_TRADING_REPORT_EMAIL_SUBJECT_PREFIX", "Daily Trading Report") or "Daily Trading Report"
).strip() or "Daily Trading Report"
DAILY_TRADING_REPORT_EMAIL_JOB_NAME = "daily_trading_summary_html_email"
_daily_trading_report_email_scheduler_lock = threading.Lock()
_daily_trading_report_email_scheduler_started = False
_daily_trading_report_email_last_attempt_epoch: Dict[str, float] = {}


def _pmx_non_empty(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text and text.lower() != "none":
            return text
    return ""


def _pmx_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "on"}


def _auth_serializer() -> URLSafeTimedSerializer:
    global _AUTH_SERIALIZER
    if _AUTH_SERIALIZER is None:
        secret = _pmx_non_empty(
            os.getenv("APP_AUTH_SECRET", ""),
            os.getenv("FLASK_SECRET_KEY", ""),
        )
        if not secret:
            seed = _pmx_non_empty(
                os.getenv("STONEX_SUBSCRIPTION_KEY", ""),
                os.getenv("PMX_LOGIN_USERNAME", ""),
                os.getenv("COMPUTERNAME", ""),
                ORIGINAL_PROJECT,
                "j2-platform",
            )
            secret = hashlib.sha256(f"auth::{seed}".encode("utf-8")).hexdigest()
        _AUTH_SERIALIZER = URLSafeTimedSerializer(secret_key=secret, salt="j2-platform-auth-v1")
    return _AUTH_SERIALIZER


def _auth_hash_password(password: str) -> str:
    text = str(password or "")
    salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", text.encode("utf-8"), salt, AUTH_PASSWORD_ITERATIONS)
    return f"pbkdf2_sha256${AUTH_PASSWORD_ITERATIONS}${salt.hex()}${digest.hex()}"


def _auth_verify_password(password: str, stored_hash: str) -> bool:
    raw = str(stored_hash or "").strip()
    if not raw:
        return False
    parts = raw.split("$")
    if len(parts) != 4 or parts[0] != "pbkdf2_sha256":
        return False
    try:
        iterations = int(parts[1])
        salt = bytes.fromhex(parts[2])
        expected = bytes.fromhex(parts[3])
    except Exception:
        return False
    actual = hashlib.pbkdf2_hmac("sha256", str(password or "").encode("utf-8"), salt, iterations)
    return hmac.compare_digest(actual, expected)


def _auth_user_from_row(row: Optional[Tuple[Any, ...]]) -> Optional[Dict[str, Any]]:
    if not row:
        return None
    return {
        "id": int(row[0]),
        "username": str(row[1] or ""),
        "display_name": str(row[2] or row[1] or ""),
        "password_hash": str(row[3] or ""),
        "role": str(row[4] or "admin"),
        "can_read": bool(row[5]),
        "can_write": bool(row[6]),
        "is_admin": bool(row[7]),
        "is_active": bool(row[8]),
        "created_at": str(row[9] or ""),
    }


def _auth_public_user(user: Dict[str, Any]) -> Dict[str, Any]:
    permissions: List[str] = []
    if bool(user.get("can_read")):
        permissions.append("read")
    if bool(user.get("can_write")):
        permissions.append("write")
    if bool(user.get("is_admin")):
        permissions.append("admin")
    return {
        "id": int(user.get("id", 0) or 0),
        "username": str(user.get("username", "") or ""),
        "display_name": str(user.get("display_name", "") or ""),
        "role": str(user.get("role", "admin") or "admin"),
        "permissions": permissions,
    }


def _auth_admin_user(user: Dict[str, Any]) -> Dict[str, Any]:
    """Extended user payload for admin user-management endpoints."""
    base = _auth_public_user(user)
    base.update(
        {
            "can_read": bool(user.get("can_read")),
            "can_write": bool(user.get("can_write")),
            "is_admin": bool(user.get("is_admin")),
            "is_active": bool(user.get("is_active")),
            "created_at": str(user.get("created_at", "") or ""),
        }
    )
    return base


def _auth_has_permission(user: Dict[str, Any], permission: str) -> bool:
    if not isinstance(user, dict):
        return False
    if bool(user.get("is_admin")):
        return True
    perm = str(permission or "").strip().lower()
    if perm == "read":
        return bool(user.get("can_read"))
    if perm == "write":
        return bool(user.get("can_write"))
    return False


def _auth_issue_token(user: Dict[str, Any]) -> str:
    payload = {
        "uid": int(user.get("id", 0) or 0),
        "username": str(user.get("username", "") or ""),
        "role": str(user.get("role", "admin") or "admin"),
    }
    return _auth_serializer().dumps(payload)


def _auth_parse_token(token: str) -> Optional[Dict[str, Any]]:
    raw = str(token or "").strip()
    if not raw:
        return None
    try:
        parsed = _auth_serializer().loads(raw, max_age=AUTH_SESSION_SECONDS)
        return parsed if isinstance(parsed, dict) else None
    except (BadSignature, SignatureExpired):
        return None
    except Exception:
        return None


def _auth_find_user_by_username(username: str) -> Optional[Dict[str, Any]]:
    lookup = str(username or "").strip()
    if not lookup:
        return None
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT id, username, display_name, password_hash, role, can_read, can_write, is_admin, is_active, created_at
            FROM {AUTH_USER_TABLE}
            WHERE lower(username) = lower(?)
            LIMIT 1
            """,
            (lookup,),
        )
        row = cur.fetchone()
        return _auth_user_from_row(row)
    except Exception as exc:
        print(f"[WARN] _auth_find_user_by_username failed: {exc}")
        return None
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass


def _auth_find_user_by_id(user_id: int) -> Optional[Dict[str, Any]]:
    if int(user_id or 0) <= 0:
        return None
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT id, username, display_name, password_hash, role, can_read, can_write, is_admin, is_active, created_at
            FROM {AUTH_USER_TABLE}
            WHERE id = ?
            LIMIT 1
            """,
            (int(user_id),),
        )
        row = cur.fetchone()
        return _auth_user_from_row(row)
    except Exception as exc:
        print(f"[WARN] _auth_find_user_by_id failed: {exc}")
        return None
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass


def _auth_list_users() -> List[Dict[str, Any]]:
    """List all app users ordered by username."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT id, username, display_name, password_hash, role, can_read, can_write, is_admin, is_active, created_at
        FROM {AUTH_USER_TABLE}
        ORDER BY lower(username), id
        """
    )
    rows = cur.fetchall()
    conn.close()
    users: List[Dict[str, Any]] = []
    for row in rows:
        user = _auth_user_from_row(row)
        if user:
            users.append(user)
    return users


def _auth_seed_default_users() -> None:
    admin_username = _pmx_non_empty(
        os.getenv("APP_AUTH_USERNAME", ""),
        os.getenv("STONEX_USERNAME", ""),
        os.getenv("PMX_LOGIN_USERNAME", ""),
    )
    admin_password = _pmx_non_empty(
        os.getenv("APP_AUTH_PASSWORD", ""),
        os.getenv("STONEX_PASSWORD", ""),
        os.getenv("PMX_LOGIN_PASSWORD", ""),
    )
    admin_display_name = _pmx_non_empty(os.getenv("APP_AUTH_DISPLAY_NAME", ""), admin_username)
    readonly_username = _pmx_non_empty(
        os.getenv("APP_READONLY_USERNAME", ""),
        os.getenv("APP_VIEWER_USERNAME", ""),
    )
    readonly_password = _pmx_non_empty(
        os.getenv("APP_READONLY_PASSWORD", ""),
        os.getenv("APP_VIEWER_PASSWORD", ""),
    )
    readonly_display_name = _pmx_non_empty(
        os.getenv("APP_READONLY_DISPLAY_NAME", ""),
        readonly_username,
    )
    force_reseed = _pmx_bool(os.getenv("APP_AUTH_FORCE_RESEED", "false"), default=False)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {AUTH_USER_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            display_name TEXT,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'admin',
            can_read INTEGER NOT NULL DEFAULT 1,
            can_write INTEGER NOT NULL DEFAULT 1,
            is_admin INTEGER NOT NULL DEFAULT 1,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        f"CREATE UNIQUE INDEX IF NOT EXISTS idx_{AUTH_USER_TABLE}_username ON {AUTH_USER_TABLE}(username)"
    )
    cur.execute(
        f"CREATE UNIQUE INDEX IF NOT EXISTS idx_{AUTH_USER_TABLE}_username_lower ON {AUTH_USER_TABLE}(lower(username))"
    )

    def _upsert_user(
        username: str,
        display_name: str,
        password: str,
        role: str,
        can_read: bool,
        can_write: bool,
        is_admin: bool,
    ) -> None:
        normalized_username = str(username or "").strip()
        if not normalized_username:
            return
        normalized_display_name = _pmx_non_empty(display_name, normalized_username)
        role_text = str(role or "viewer")
        can_read_i = int(bool(can_read))
        can_write_i = int(bool(can_write))
        is_admin_i = int(bool(is_admin))

        cur.execute(
            f"""
            SELECT id, display_name, password_hash, role, can_read, can_write, is_admin, is_active
            FROM {AUTH_USER_TABLE}
            WHERE lower(username) = lower(?)
            LIMIT 1
            """,
            (normalized_username,),
        )
        existing = cur.fetchone()
        if existing and not force_reseed:
            existing_id = int(existing[0])
            existing_display = str(existing[1] or "")
            existing_role = str(existing[3] or "viewer")
            existing_can_read = int(bool(existing[4]))
            existing_can_write = int(bool(existing[5]))
            existing_is_admin = int(bool(existing[6]))
            existing_active = int(bool(existing[7]))

            # Startup seed should be idempotent and fast: avoid expensive PBKDF2
            # on every boot if the user already exists.
            if (
                existing_display != normalized_display_name
                or existing_role != role_text
                or existing_can_read != can_read_i
                or existing_can_write != can_write_i
                or existing_is_admin != is_admin_i
                or existing_active != 1
            ):
                cur.execute(
                    f"""
                    UPDATE {AUTH_USER_TABLE}
                    SET
                        display_name = ?,
                        role = ?,
                        can_read = ?,
                        can_write = ?,
                        is_admin = ?,
                        is_active = 1
                    WHERE id = ?
                    """,
                    (
                        normalized_display_name,
                        role_text,
                        can_read_i,
                        can_write_i,
                        is_admin_i,
                        existing_id,
                    ),
                )
            return

        # Keep username lookups case-insensitive and avoid duplicate-lower collisions.
        cur.execute(
            f"DELETE FROM {AUTH_USER_TABLE} WHERE lower(username) = lower(?) AND username <> ?",
            (normalized_username, normalized_username),
        )
        cur.execute(
            f"""
            INSERT INTO {AUTH_USER_TABLE} (
                username, display_name, password_hash, role, can_read, can_write, is_admin, is_active
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 1)
            ON CONFLICT(username) DO UPDATE SET
                display_name=excluded.display_name,
                password_hash=excluded.password_hash,
                role=excluded.role,
                can_read=excluded.can_read,
                can_write=excluded.can_write,
                is_admin=excluded.is_admin,
                is_active=1
            """,
            (
                normalized_username,
                normalized_display_name,
                _auth_hash_password(password),
                role_text,
                can_read_i,
                can_write_i,
                is_admin_i,
            ),
        )

    seeded_users = 0

    if admin_username and admin_password:
        _upsert_user(
            username=admin_username,
            display_name=admin_display_name,
            password=admin_password,
            role="admin",
            can_read=True,
            can_write=True,
            is_admin=True,
        )
        seeded_users += 1
    else:
        print("[WARN] Admin auth user not seeded. Set APP_AUTH_USERNAME and APP_AUTH_PASSWORD in .env.")

    if readonly_username or readonly_password:
        if not readonly_username or not readonly_password:
            print("[WARN] Read-only user was skipped. Set both APP_READONLY_USERNAME and APP_READONLY_PASSWORD.")
        elif admin_username and readonly_username.strip().lower() == admin_username.strip().lower():
            print("[WARN] Read-only username matches admin username. Skipping read-only seed user.")
        else:
            _upsert_user(
                username=readonly_username,
                display_name=readonly_display_name,
                password=readonly_password,
                role="viewer",
                can_read=True,
                can_write=False,
                is_admin=False,
            )
            seeded_users += 1

    if seeded_users == 0:
        print(
            "[WARN] No auth users were seeded. Configure APP_AUTH_USERNAME/APP_AUTH_PASSWORD "
            "and optionally APP_READONLY_USERNAME/APP_READONLY_PASSWORD."
        )

    conn.commit()
    conn.close()


def _auth_cookie_token() -> str:
    token = str(request.cookies.get(AUTH_COOKIE_NAME, "") or "").strip()
    if token:
        return token
    auth_header = str(request.headers.get("Authorization", "") or "").strip()
    if auth_header.lower().startswith("bearer "):
        return auth_header.split(" ", 1)[1].strip()
    return ""


def _auth_request_user() -> Optional[Dict[str, Any]]:
    token = _auth_cookie_token()
    payload = _auth_parse_token(token)
    if not payload:
        return None
    user_id_raw = payload.get("uid")
    try:
        user_id = int(user_id_raw)
    except Exception:
        return None
    user = _auth_find_user_by_id(user_id)
    if not user or not bool(user.get("is_active")):
        return None
    return user


def _auth_set_cookie(res: Response, token: str) -> None:
    res.set_cookie(
        AUTH_COOKIE_NAME,
        value=token,
        max_age=AUTH_SESSION_SECONDS,
        httponly=True,
        secure=AUTH_COOKIE_SECURE,
        samesite="Lax",
        path="/",
    )


def _auth_clear_cookie(res: Response) -> None:
    res.set_cookie(
        AUTH_COOKIE_NAME,
        value="",
        expires=0,
        max_age=0,
        httponly=True,
        secure=AUTH_COOKIE_SECURE,
        samesite="Lax",
        path="/",
    )


def _pmx_get_cached_session() -> Dict[str, str]:
    with PMX_SESSION_LOCK:
        return {
            "x_auth": str(PMX_SESSION_CACHE.get("x_auth", "") or ""),
            "sid": str(PMX_SESSION_CACHE.get("sid", "") or ""),
            "username": str(PMX_SESSION_CACHE.get("username", "") or ""),
            "platform": str(PMX_SESSION_CACHE.get("platform", "") or ""),
            "location": str(PMX_SESSION_CACHE.get("location", "") or ""),
            "cache_control": str(PMX_SESSION_CACHE.get("cache_control", "") or ""),
            "content_type": str(PMX_SESSION_CACHE.get("content_type", "") or ""),
        }


def _pmx_store_session(session: Dict[str, Any]) -> None:
    if not isinstance(session, dict):
        return
    with PMX_SESSION_LOCK:
        for key in ("x_auth", "sid", "username", "platform", "location", "cache_control", "content_type"):
            value = str(session.get(key, "") or "").strip()
            if value:
                PMX_SESSION_CACHE[key] = value
        PMX_SESSION_CACHE["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _pmx_login_session(data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    data = data or {}
    login_username = _pmx_non_empty(data.get("pmx_login_username"), os.getenv("PMX_LOGIN_USERNAME", ""))
    login_password = _pmx_non_empty(data.get("pmx_login_password"), os.getenv("PMX_LOGIN_PASSWORD", ""))
    login_location = _pmx_non_empty(
        data.get("pmx_login_location"),
        data.get("location"),
        os.getenv("PMX_LOGIN_LOCATION", ""),
        PMX_HARDCODED_LOCATION,
        "LD",
    )
    platform = _pmx_non_empty(data.get("platform"), os.getenv("PMX_PLATFORM", ""), PMX_HARDCODED_PLATFORM, "Desktop")
    cache_control = _pmx_non_empty(
        data.get("cache_control"),
        os.getenv("PMX_CACHE_CONTROL", ""),
        PMX_HARDCODED_CACHE_CONTROL,
        "no-cache",
    )
    content_type = _pmx_non_empty(
        data.get("content_type"),
        os.getenv("PMX_CONTENT_TYPE", ""),
        PMX_HARDCODED_CONTENT_TYPE,
        "application/json; charset=utf-8",
    )
    forced_login = _pmx_bool(data.get("pmx_forced_login", os.getenv("PMX_LOGIN_FORCED", "true")), default=True)
    host = _pmx_non_empty(data.get("host"), os.getenv("PMX_API_HOST", ""), "pmxapi.stonex.com")
    path = _pmx_non_empty(data.get("pmx_login_path"), os.getenv("PMX_LOGIN_PATH", ""), PMX_LOGIN_DEFAULT_PATH)
    timeout = int(data.get("pmx_login_timeout", os.getenv("PMX_LOGIN_TIMEOUT", "30")) or 30)

    if not login_username or not login_password:
        return {
            "ok": False,
            "error": "PMX auto-login is not configured. Set PMX_LOGIN_USERNAME and PMX_LOGIN_PASSWORD.",
        }

    url = path if path.startswith("http") else f"https://{host}{path}"
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Content-Type": content_type,
        "Origin": "https://pmxecute.stonex.com",
        "Referer": "https://pmxecute.stonex.com/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
        "cache-control": cache_control,
        "pragma": "no-cache",
    }
    if platform:
        headers["platform"] = platform
    payload = {
        "username": login_username,
        "password": login_password,
        "location": login_location,
        "forcedLogin": bool(forced_login),
    }

    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=timeout)
        content_type_resp = resp.headers.get("Content-Type", "")
        parsed: Any = {}
        if "json" in content_type_resp.lower() or str(resp.text or "").lstrip().startswith(("{", "[")):
            try:
                parsed = resp.json()
            except Exception:
                parsed = {}
        status_text = str(parsed.get("status", "")).strip().lower() if isinstance(parsed, dict) else ""
        ok = bool(resp.ok) and status_text == "success"
        token = _pmx_non_empty(parsed.get("data", {}).get("authToken")) if isinstance(parsed, dict) else ""
        user_obj = parsed.get("userObj", {}) if isinstance(parsed, dict) and isinstance(parsed.get("userObj", {}), dict) else {}
        sid = _pmx_non_empty(user_obj.get("SID"))
        api_username = _pmx_non_empty(user_obj.get("UID"), data.get("username"), os.getenv("PMX_USERNAME", ""), PMX_HARDCODED_USERNAME)
        if not ok or not token:
            msg = _pmx_non_empty(
                parsed.get("message") if isinstance(parsed, dict) else "",
                resp.reason,
                "PMX login failed",
            )
            return {"ok": False, "status": resp.status_code, "error": msg}

        session = {
            "x_auth": token,
            "sid": sid,
            "username": api_username,
            "platform": platform,
            "location": login_location,
            "cache_control": cache_control,
            "content_type": content_type,
        }
        _pmx_store_session(session)

        os.environ["PMX_X_AUTH"] = token
        if sid:
            os.environ["PMX_SID"] = sid
        if api_username:
            os.environ["PMX_USERNAME"] = api_username
        if login_location:
            os.environ["PMX_LOCATION"] = login_location
        if platform:
            os.environ["PMX_PLATFORM"] = platform

        return {"ok": True, "status": resp.status_code, **session}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def _pmx_resolve_headers(data: Dict[str, Any], req_headers: Any, auto_login: bool = True) -> Dict[str, Any]:
    data = data or {}
    cached = _pmx_get_cached_session()

    def _header(name: str) -> str:
        if hasattr(req_headers, "get"):
            return str(req_headers.get(name, "") or "")
        return ""

    force_relogin = _pmx_bool(data.get("force_pmx_relogin", False), default=False)
    x_auth = _pmx_non_empty(
        data.get("x_auth"),
        _header("x-auth"),
        cached.get("x_auth"),
        os.getenv("PMX_X_AUTH", ""),
        PMX_HARDCODED_X_AUTH,
    )
    sid = _pmx_non_empty(
        data.get("sid"),
        _header("sid"),
        cached.get("sid"),
        os.getenv("PMX_SID", ""),
        PMX_HARDCODED_SID,
    )
    username = _pmx_non_empty(
        data.get("username"),
        _header("username"),
        cached.get("username"),
        os.getenv("PMX_USERNAME", ""),
        PMX_HARDCODED_USERNAME,
    )
    platform = _pmx_non_empty(
        data.get("platform"),
        _header("platform"),
        cached.get("platform"),
        os.getenv("PMX_PLATFORM", ""),
        PMX_HARDCODED_PLATFORM,
    )
    location = _pmx_non_empty(
        data.get("location"),
        _header("location"),
        cached.get("location"),
        os.getenv("PMX_LOCATION", ""),
        PMX_HARDCODED_LOCATION,
    )
    cache_control = _pmx_non_empty(
        data.get("cache_control"),
        _header("cache-control"),
        cached.get("cache_control"),
        os.getenv("PMX_CACHE_CONTROL", ""),
        PMX_HARDCODED_CACHE_CONTROL,
    )
    content_type = _pmx_non_empty(
        data.get("content_type"),
        _header("content-type"),
        cached.get("content_type"),
        os.getenv("PMX_CONTENT_TYPE", ""),
        PMX_HARDCODED_CONTENT_TYPE,
    )

    login_result: Dict[str, Any] = {}
    if auto_login and (force_relogin or not x_auth):
        login_result = _pmx_login_session(data)
        if login_result.get("ok"):
            x_auth = _pmx_non_empty(login_result.get("x_auth"), x_auth)
            sid = _pmx_non_empty(login_result.get("sid"), sid)
            username = _pmx_non_empty(login_result.get("username"), username)
            platform = _pmx_non_empty(login_result.get("platform"), platform)
            location = _pmx_non_empty(login_result.get("location"), location)
            cache_control = _pmx_non_empty(login_result.get("cache_control"), cache_control)
            content_type = _pmx_non_empty(login_result.get("content_type"), content_type)

    return {
        "x_auth": x_auth,
        "sid": sid,
        "username": username,
        "platform": platform,
        "location": location,
        "cache_control": cache_control,
        "content_type": content_type,
        "login_result": login_result,
    }


def _pmx_result_is_auth_failure(result: Dict[str, Any], payload: Any = None) -> bool:
    status = int(result.get("status") or 0) if str(result.get("status", "")).strip().isdigit() else 0
    reason = str(result.get("reason", "") or "")
    error = str(result.get("error", "") or "")
    message = ""
    if isinstance(payload, dict):
        message = str(payload.get("message", "") or "")
    combined = " ".join([reason, error, message]).lower()
    if status in {401, 403}:
        return True
    if any(
        token in combined
        for token in ("x-auth", "unauthor", "forbidden", "token", "session", "expired", "invalid login", "auth")
    ):
        return True
    if status >= 500 and "internal server error" in combined:
        return True
    return False


def _pmx_mark_failed_payload(result: Dict[str, Any], payload: Any = None) -> Dict[str, Any]:
    """Treat PMX JSON payloads with status=failed/error as request failures."""
    out = dict(result or {})
    if not bool(out.get("ok")):
        return out
    if not isinstance(payload, dict):
        return out
    pmx_status = str(payload.get("status", "") or "").strip().lower()
    if pmx_status not in {"failed", "error"}:
        return out
    out["ok"] = False
    if not out.get("status"):
        out["status"] = 502
    message = str(payload.get("message", "") or "").strip()
    if message and not out.get("error"):
        out["error"] = message
    elif not out.get("error"):
        out["error"] = "PMX account statement returned failed status"
    return out


def _pmx_human_error(result: Dict[str, Any], payload: Any = None, default: str = "") -> str:
    """Return a clean API-facing PMX error message (avoid leaking raw JSON blobs)."""
    raw_error = str((result or {}).get("error", "") or "").strip()
    payload_message = str(payload.get("message", "") or "").strip() if isinstance(payload, dict) else ""
    combined = " ".join([raw_error, payload_message]).lower()
    if "internal server error" in combined and _pmx_result_is_auth_failure(result or {}, payload):
        return (
            "PMX API returned Internal Server Error, which usually means the PMX session headers "
            "(x-auth/sid/username) are expired. Refresh PMX session headers or enable PMX auto-login."
        )
    return raw_error or payload_message or default


def _json_safe(value):
    """Recursively convert NaN/Inf/pandas nulls into JSON-safe values."""
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if isinstance(value, tuple):
        return tuple(_json_safe(v) for v in value)

    # numpy scalars (e.g. np.float64) -> python scalar
    if hasattr(value, "item") and callable(getattr(value, "item", None)):
        try:
            return _json_safe(value.item())
        except Exception:
            pass

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    return value


def _sanitize_filename_component(value: Any, fallback: str = "file") -> str:
    text = str(value or "").strip()
    if not text:
        return fallback
    text = re.sub(r'[<>:"/\\|?*\x00-\x1F]+', "_", text)
    text = text.strip(" .")
    return text or fallback


def _build_trade_filename_suffix(trade_nums: List[str], max_len: int = 120) -> str:
    safe_trade_tokens: List[str] = []
    seen_tokens = set()
    for raw in trade_nums:
        token = _sanitize_filename_component(normalize_trade_number(raw), "")
        if not token or token in seen_tokens:
            continue
        seen_tokens.add(token)
        safe_trade_tokens.append(token)

    if not safe_trade_tokens:
        return ""

    trade_suffix = ""
    for token in safe_trade_tokens:
        next_suffix = f"{trade_suffix}_{token}" if trade_suffix else token
        if len(next_suffix) > max_len:
            break
        trade_suffix = next_suffix

    included_count = trade_suffix.count("_") + 1 if trade_suffix else 0
    remaining_count = max(0, len(safe_trade_tokens) - included_count)
    if remaining_count > 0:
        trade_suffix = f"{trade_suffix}_plus{remaining_count}" if trade_suffix else f"plus{remaining_count}"

    return trade_suffix or str(len(safe_trade_tokens))


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        num = float(value)
        if math.isfinite(num):
            return num
        return default
    except Exception:
        return default


def _is_truthy(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on"}


def _get_cached_trademc_live_prices(force_refresh: bool = False) -> Dict[str, Any]:
    now_ts = time.time()
    stale_payload: Dict[str, Any] = {}
    stale_age: Optional[float] = None

    with TRADEMC_LIVE_PRICES_LOCK:
        cached = TRADEMC_LIVE_PRICES_CACHE.get("payload")
        cached_at = float(TRADEMC_LIVE_PRICES_CACHE.get("fetched_ts") or 0.0)
        if isinstance(cached, dict):
            age = max(0.0, now_ts - cached_at)
            if not force_refresh and age <= float(TRADEMC_LIVE_PRICES_TTL_SECONDS):
                out = dict(cached)
                out["cache_hit"] = True
                out["cache_age_seconds"] = round(age, 2)
                return out
            stale_payload = dict(cached)
            stale_age = age

    fresh = get_latest_trademc_market_prices(sample_size=TRADEMC_LIVE_PRICES_SAMPLE_SIZE)
    if not isinstance(fresh, dict):
        fresh = {"ok": False, "error": "Invalid TradeMC live-prices payload"}

    if bool(fresh.get("ok")):
        with TRADEMC_LIVE_PRICES_LOCK:
            TRADEMC_LIVE_PRICES_CACHE["payload"] = dict(fresh)
            TRADEMC_LIVE_PRICES_CACHE["fetched_ts"] = now_ts
        out = dict(fresh)
        out["cache_hit"] = False
        out["cache_age_seconds"] = 0.0
        return out

    if stale_payload:
        out = dict(stale_payload)
        out["stale"] = True
        out["stale_error"] = str(fresh.get("error") or "")
        out["cache_hit"] = True
        out["cache_age_seconds"] = round(float(stale_age or 0.0), 2)
        return out

    out = dict(fresh)
    out["cache_hit"] = False
    out["cache_age_seconds"] = 0.0
    return out


def _build_cache_key(prefix: str, args: Optional[Dict[str, Any]] = None) -> str:
    if not isinstance(args, dict) or not args:
        return prefix
    parts = []
    for k, v in sorted(args.items(), key=lambda item: str(item[0])):
        key = str(k).strip()
        if not key:
            continue
        parts.append(f"{key}={str(v)}")
    if not parts:
        return prefix
    return f"{prefix}?{'&'.join(parts)}"


def _get_cached_heavy_result(cache_key: str, builder: Any, ttl_seconds: Optional[int] = None) -> Any:
    ttl = int(ttl_seconds if ttl_seconds is not None else HEAVY_ROUTE_CACHE_TTL_SECONDS)
    ttl = max(1, ttl)
    now_ts = time.time()

    with HEAVY_ROUTE_CACHE_LOCK:
        entry = HEAVY_ROUTE_CACHE.get(cache_key)
        if isinstance(entry, dict):
            cached_ts = float(entry.get("ts") or 0.0)
            if (now_ts - cached_ts) <= float(ttl):
                return entry.get("value")

    value = builder()
    with HEAVY_ROUTE_CACHE_LOCK:
        HEAVY_ROUTE_CACHE[cache_key] = {"ts": now_ts, "value": value}
    return value


def _clear_heavy_route_cache(prefixes: Optional[List[str]] = None) -> None:
    with HEAVY_ROUTE_CACHE_LOCK:
        if not prefixes:
            HEAVY_ROUTE_CACHE.clear()
            return
        prefix_list = [str(p) for p in prefixes if str(p)]
        if not prefix_list:
            return
        keys = list(HEAVY_ROUTE_CACHE.keys())
        for key in keys:
            key_text = str(key)
            if any(key_text.startswith(prefix) for prefix in prefix_list):
                HEAVY_ROUTE_CACHE.pop(key, None)


def _set_row_value_by_header(ws: Any, header: str, value: Any, row_idx: int = 2) -> Optional[Any]:
    target = str(header or "").strip().lower()
    if not target:
        return None
    for col_idx in range(1, int(getattr(ws, "max_column", 0)) + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        if isinstance(header_val, str) and header_val.strip().lower() == target:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            return cell
    return None


def _compute_sales_order_metrics(trade_num: str) -> Dict[str, Any]:
    trade_num_norm = normalize_trade_number(trade_num)
    wa = build_weighted_average(trade_num_norm, source="pmx") or {}
    price_usd_oz = _safe_float(wa.get("xau_usd_wa_price"), 0.0)
    ex_rate = _safe_float(wa.get("usd_zar_wa_price"), 0.0)

    quantity_oz = 0.0
    df = load_all_pmx_trades()
    if isinstance(df, pd.DataFrame) and not df.empty:
        if "OrderID" in df.columns:
            trade_col = "OrderID"
        elif "Trade #" in df.columns:
            trade_col = "Trade #"
        else:
            trade_col = ""

        if trade_col:
            df = df.copy()
            df["_trade_num"] = df[trade_col].apply(normalize_trade_number)
            matched = df[df["_trade_num"] == trade_num_norm]
            for _, row in matched.iterrows():
                sym = str(row.get("Symbol", "")).upper()
                base, quote = split_symbol(sym)
                if base == "XAU" and quote == "USD":
                    quantity_oz += abs(_safe_float(row.get("Quantity", 0.0), 0.0))

    if quantity_oz <= 0:
        quantity_oz = abs(_safe_float(wa.get("xau_usd_total_qty"), 0.0))

    quantity_g = quantity_oz * GRAMS_PER_TROY_OUNCE
    price_excl = (price_usd_oz * ex_rate) / GRAMS_PER_TROY_OUNCE if price_usd_oz and ex_rate else 0.0

    return {
        "trade_num": trade_num_norm,
        "quantity_g": round(quantity_g, 3),
        "price_usd_oz": round(price_usd_oz, 5),
        "ex_rate": round(ex_rate, 5),
        "price_excl": round(price_excl, 5),
        "date_text": datetime.now().strftime("%Y/%m/%d"),
    }


def _build_sales_order_excel_for_trades(trade_nums: List[str], output_dir: str) -> Dict[str, Any]:
    normalized_trade_nums: List[str] = []
    seen_trade_nums = set()
    for raw in trade_nums:
        tn = normalize_trade_number(raw)
        if not tn or tn in seen_trade_nums:
            continue
        seen_trade_nums.add(tn)
        normalized_trade_nums.append(tn)

    if not normalized_trade_nums:
        return {"ok": False, "error": "No trade numbers provided"}
    if not os.path.exists(SALES_ORDER_TEMPLATE_PATH):
        return {"ok": False, "error": f"Missing template: {SALES_ORDER_TEMPLATE_PATH}"}

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        return {"ok": False, "error": f"openpyxl is required to generate Sales Order Excel: {exc}"}

    try:
        wb = load_workbook(SALES_ORDER_TEMPLATE_PATH)
        ws = wb.active
        template_row = 2
        max_col = int(getattr(ws, "max_column", 0))
        metrics_rows: List[Dict[str, Any]] = []

        for idx, tn in enumerate(normalized_trade_nums):
            row_idx = template_row + idx
            metrics = _compute_sales_order_metrics(tn)
            metrics_rows.append(metrics)

            if row_idx != template_row:
                for col_idx in range(1, max_col + 1):
                    src = ws.cell(row=template_row, column=col_idx)
                    dst = ws.cell(row=row_idx, column=col_idx)
                    dst.value = src.value
                    if src.has_style:
                        dst._style = copy(src._style)
                ws.row_dimensions[row_idx].height = ws.row_dimensions[template_row].height

            qty_cell = _set_row_value_by_header(ws, "Quantity", metrics["quantity_g"], row_idx=row_idx)
            px_cell = _set_row_value_by_header(ws, "Price($/oz)", metrics["price_usd_oz"], row_idx=row_idx)
            fx_cell = _set_row_value_by_header(ws, "ExRate", metrics["ex_rate"], row_idx=row_idx)
            excl_cell = _set_row_value_by_header(ws, "Price (excl)", metrics["price_excl"], row_idx=row_idx)
            _set_row_value_by_header(ws, "External order", tn, row_idx=row_idx)
            _set_row_value_by_header(ws, "Date", metrics["date_text"], row_idx=row_idx)

            if qty_cell is not None:
                qty_cell.number_format = "0.000"
            if px_cell is not None:
                px_cell.number_format = "0.00000"
            if fx_cell is not None:
                fx_cell.number_format = "0.00000"
            if excl_cell is not None:
                excl_cell.number_format = "0.00000"

        if len(normalized_trade_nums) == 1:
            file_name = f"{_sanitize_filename_component(normalized_trade_nums[0], 'trade')}_sales_order.xlsx"
        else:
            trade_suffix = _build_trade_filename_suffix(normalized_trade_nums)
            file_name = f"sales_orders_{trade_suffix}.xlsx"
        output_path = os.path.join(output_dir, file_name)
        wb.save(output_path)
        return {
            "ok": True,
            "path": output_path,
            "filename": file_name,
            "metrics_rows": metrics_rows,
            "trade_count": len(normalized_trade_nums),
        }
    except Exception as exc:
        return {"ok": False, "error": f"Failed to generate sales-order Excel: {exc}"}


def _build_pmx_fnc_pdf_result(
    cell: str,
    doc_type: str,
    data: Dict[str, Any],
    req_headers: Any,
) -> Dict[str, Any]:
    resolved_headers = _pmx_resolve_headers(data, req_headers, auto_login=True)
    x_auth = str(resolved_headers.get("x_auth", "") or "").strip()
    sid = str(resolved_headers.get("sid", "") or "").strip()
    username = str(resolved_headers.get("username", "") or "").strip()
    platform = str(resolved_headers.get("platform", "") or "").strip()
    location = str(resolved_headers.get("location", "") or "").strip()
    cache_control = str(resolved_headers.get("cache_control", "") or "").strip()
    content_type = str(resolved_headers.get("content_type", "") or "").strip()

    host = str(data.get("host", os.getenv("PMX_API_HOST", "pmxapi.stonex.com")) or "pmxapi.stonex.com")
    path = str(data.get("path", "/user/export_FixInvoice_pdf") or "/user/export_FixInvoice_pdf")
    authorization = str(data.get("authorization", "") or "")
    cookie = str(data.get("cookie", "") or "")
    extra_headers = data.get("headers", {}) if isinstance(data.get("headers", {}), dict) else {}
    origin = str(data.get("origin", "https://pmxecute.stonex.com") or "https://pmxecute.stonex.com")
    referer = str(data.get("referer", "https://pmxecute.stonex.com/") or "https://pmxecute.stonex.com/")
    timeout = int(data.get("timeout", 120) or 120)

    def _fetch_pdf(
        x_auth_value: str,
        username_value: str,
        sid_value: str,
        platform_value: str,
        location_value: str,
        cache_control_value: str,
        content_type_value: str,
    ) -> Dict[str, Any]:
        return fetch_pmx_fixinvoice_pdf(
            cell=cell,
            doc_type=doc_type,
            host=host,
            path=path,
            authorization=authorization,
            cookie=cookie,
            x_auth=x_auth_value,
            sid=sid_value,
            username=username_value,
            platform=platform_value,
            location=location_value,
            cache_control=cache_control_value,
            content_type=content_type_value,
            extra_headers=extra_headers,
            origin=origin,
            referer=referer,
            timeout=timeout,
        )

    result = _fetch_pdf(
        x_auth_value=x_auth,
        username_value=username,
        sid_value=sid,
        platform_value=platform,
        location_value=location,
        cache_control_value=cache_control,
        content_type_value=content_type,
    )
    if not result.get("ok") and any([sid, platform, location, cache_control, content_type]):
        result = _fetch_pdf(
            x_auth_value=x_auth,
            username_value=username,
            sid_value="",
            platform_value="",
            location_value="",
            cache_control_value="",
            content_type_value="",
        )

    session_refreshed = False
    relogin_error = ""
    if not result.get("ok") and _pmx_result_is_auth_failure(result):
        relogin = _pmx_login_session(data)
        if relogin.get("ok"):
            session_refreshed = True
            x_auth = _pmx_non_empty(relogin.get("x_auth"), x_auth)
            sid = _pmx_non_empty(relogin.get("sid"), sid)
            username = _pmx_non_empty(relogin.get("username"), username)
            platform = _pmx_non_empty(relogin.get("platform"), platform)
            location = _pmx_non_empty(relogin.get("location"), location)
            cache_control = _pmx_non_empty(relogin.get("cache_control"), cache_control)
            content_type = _pmx_non_empty(relogin.get("content_type"), content_type)

            result = _fetch_pdf(
                x_auth_value=x_auth,
                username_value=username,
                sid_value=sid,
                platform_value=platform,
                location_value=location,
                cache_control_value=cache_control,
                content_type_value=content_type,
            )
            if not result.get("ok") and any([sid, platform, location, cache_control, content_type]):
                result = _fetch_pdf(
                    x_auth_value=x_auth,
                    username_value=username,
                    sid_value="",
                    platform_value="",
                    location_value="",
                    cache_control_value="",
                    content_type_value="",
                )
        else:
            relogin_error = str(relogin.get("error", "") or "").strip()

    if not result.get("ok"):
        err_text = str(result.get("error", "") or "").strip()
        if relogin_error:
            err_text = (
                f"{err_text or 'PMX PDF request failed'}. "
                f"Auto-login attempt failed: {relogin_error}"
            )
        if _pmx_result_is_auth_failure(result) and "PMX auto-login is not configured" not in err_text:
            err_text = (
                "PMX PDF request failed due to expired session headers. "
                "Configure PMX_LOGIN_USERNAME and PMX_LOGIN_PASSWORD for automatic token refresh."
            )
        return {
            "ok": False,
            "error": err_text or result.get("error", "PMX PDF download failed"),
            "status": result.get("status"),
            "url": result.get("url", ""),
            "cell": cell,
            "DocType": doc_type,
            "session_refreshed": session_refreshed,
            "relogin_error": relogin_error,
        }

    pdf_bytes = result.get("body_bytes", b"")
    if not isinstance(pdf_bytes, (bytes, bytearray)):
        pdf_bytes = bytes(str(pdf_bytes), "utf-8")
    pdf_bytes = bytes(pdf_bytes)
    safe_name = "Fixing_Invoice_" + str(cell).replace("/", "_").replace("\\", "_") + ".pdf"
    return {
        "ok": True,
        "bytes": pdf_bytes,
        "content_type": result.get("content_type") or "application/pdf",
        "filename": safe_name,
        "cell": cell,
        "DocType": doc_type,
    }


def normalize_trade_number(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    return text.upper()


def _extract_trade_num_from_remarks(remarks: str) -> str:
    """Extract trade number from PMX remarks field.

    Patterns:
      - "ONG/2026/003475. SELL XAU/USD ... Limit, KAS-022."  → KAS-022
      - "Jos-070 USD/ZAR 10065.86  @ 16.74744"               → Jos-070
      - "ADJ-001 XAU/USD 41.484 OZ @ 5021.77"                → ADJ-001
    """
    import re
    if not remarks:
        return ""
    remarks = str(remarks).strip()
    # Pattern 1: trade num after last comma (e.g. "..., KAS-022.")
    m = re.search(r',\s*([A-Za-z]+-?\d+)\s*\.?\s*$', remarks)
    if m:
        return m.group(1).strip()
    # Pattern 2: trade num at start (e.g. "Jos-070 USD/ZAR ...")
    m = re.match(r'^([A-Za-z]+-?\d+)\s+(?:USD|XAU|ZAR|XAG|XPT)', remarks, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return ""


def split_symbol(sym):
    sym = str(sym or "").upper()
    if "/" in sym:
        base, quote = sym.split("/", 1)
        return base, quote
    if len(sym) == 6:
        return sym[:3], sym[3:]
    return sym, ""


def get_pmx_db_connection():
    """Create SQLite connection for PMX ledger database."""
    conn = sqlite3.connect(PMX_DB_PATH, timeout=30, check_same_thread=False)
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA busy_timeout=5000;")
    except Exception:
        pass
    return conn


def _pmx_swap_sql_predicate(alias: str = "") -> str:
    col = (lambda name: f"{alias}.{name}" if alias else name)
    return f"""
        (
            UPPER(TRIM(COALESCE({col("fnc_number")}, ''))) LIKE 'SWT/%'
            OR UPPER(TRIM(COALESCE({col("fnc_number")}, ''))) LIKE '%/SWT/%'
            OR UPPER(TRIM(COALESCE({col("doc_number")}, ''))) LIKE 'SWT/%'
            OR UPPER(TRIM(COALESCE({col("doc_number")}, ''))) LIKE '%/SWT/%'
            OR UPPER(TRIM(COALESCE({col("order_id")}, ''))) LIKE 'SWT/%'
            OR UPPER(TRIM(COALESCE({col("order_id")}, ''))) LIKE '%/SWT/%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%SWT/%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"DEAL_TYPE":"SWAP"%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"DEAL_TYPE"%SWAP%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"DEALTYPE":"SWAP"%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"DEALTYPE"%SWAP%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"TRADE_TYPE":"SWAP"%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"TRADE_TYPE"%SWAP%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"TRADETYPE":"SWAP"%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"TRADETYPE"%SWAP%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"TRD_OPT"%SWT%'
            OR UPPER(COALESCE({col("raw_payload")}, '')) LIKE '%"TRD_OPT"%SWAP%'
            OR UPPER(COALESCE({col("narration")}, '')) LIKE 'SWAP %'
            OR UPPER(COALESCE({col("narration")}, '')) LIKE '% SWAP %'
            OR UPPER(COALESCE({col("narration")}, '')) LIKE '%SWT/%'
        )
    """


def _pmx_delete_swap_rows(conn: sqlite3.Connection) -> int:
    cursor = conn.cursor()
    cursor.execute(f"DELETE FROM trades WHERE {_pmx_swap_sql_predicate()}")
    return max(int(cursor.rowcount or 0), 0)


def _purge_pre_fiscal_rows(
    cutoff_iso: Optional[str] = None,
    purge_fx: bool = True,
    purge_pmx: bool = True,
) -> Dict[str, Any]:
    cutoff = str(cutoff_iso or FISCAL_TRADES_START_DATE or "").strip()
    counts: Dict[str, Any] = {
        "cutoff": cutoff,
        "fx_trades": 0,
        "trademc_trades": 0,
        "trademc_weight_transactions": 0,
        "pmx_trades": 0,
    }
    if not cutoff:
        return counts

    parsed_cutoff = pd.to_datetime(cutoff, errors="coerce")
    if pd.isna(parsed_cutoff):
        return counts
    cutoff = parsed_cutoff.strftime("%Y-%m-%d")
    counts["cutoff"] = cutoff

    if purge_fx:
        conn_fx = sqlite3.connect(LEDGER_DB_PATH, timeout=30, check_same_thread=False)
        try:
            cur_fx = conn_fx.cursor()
            cur_fx.execute(
                """
                DELETE FROM trades
                WHERE COALESCE(NULLIF(substr(trade_date, 1, 10), ''), '0000-00-00') < ?
                """,
                (cutoff,),
            )
            counts["fx_trades"] = max(int(cur_fx.rowcount or 0), 0)

            cur_fx.execute(
                """
                DELETE FROM trademc_trades
                WHERE COALESCE(
                    NULLIF(substr(trade_timestamp, 1, 10), ''),
                    NULLIF(substr(date_created, 1, 10), ''),
                    '0000-00-00'
                ) < ?
                """,
                (cutoff,),
            )
            counts["trademc_trades"] = max(int(cur_fx.rowcount or 0), 0)

            cur_fx.execute(
                """
                DELETE FROM trademc_weight_transactions
                WHERE COALESCE(
                    NULLIF(substr(transaction_timestamp, 1, 10), ''),
                    NULLIF(substr(date_created, 1, 10), ''),
                    '0000-00-00'
                ) < ?
                """,
                (cutoff,),
            )
            counts["trademc_weight_transactions"] = max(int(cur_fx.rowcount or 0), 0)
            conn_fx.commit()
        finally:
            conn_fx.close()

    if purge_pmx:
        conn_pmx = get_pmx_db_connection()
        try:
            cur_pmx = conn_pmx.cursor()
            cur_pmx.execute(
                """
                DELETE FROM trades
                WHERE COALESCE(NULLIF(substr(trade_date, 1, 10), ''), '0000-00-00') < ?
                """,
                (cutoff,),
            )
            counts["pmx_trades"] = max(int(cur_pmx.rowcount or 0), 0)
            conn_pmx.commit()
        finally:
            conn_pmx.close()

    return counts


def initialize_pmx_database():
    """Initialize PMX trades database (separate from main ledger DB)."""
    conn = get_pmx_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS trades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trade_date TEXT NOT NULL,
            value_date TEXT NOT NULL,
            symbol TEXT NOT NULL,
            side TEXT NOT NULL,
            narration TEXT,
            quantity REAL NOT NULL DEFAULT 0,
            price REAL NOT NULL DEFAULT 0,
            settle_currency TEXT NOT NULL DEFAULT '',
            settle_amount REAL NOT NULL DEFAULT 0,
            doc_number TEXT UNIQUE,
            clord_id TEXT,
            order_id TEXT,
            fnc_number TEXT,
            debit_usd REAL DEFAULT 0,
            credit_usd REAL DEFAULT 0,
            debit_zar REAL DEFAULT 0,
            credit_zar REAL DEFAULT 0,
            debit_xau REAL DEFAULT 0,
            credit_xau REAL DEFAULT 0,
            balance_usd REAL DEFAULT 0,
            balance_zar REAL DEFAULT 0,
            balance_xau REAL DEFAULT 0,
            rest_trade_id TEXT,
            account TEXT,
            counter_currency TEXT,
            currency TEXT,
            currency_pair TEXT,
            last_px REAL,
            last_qty REAL,
            process_date TEXT,
            trade_currency TEXT,
            transact_time TEXT,
            source_system TEXT,
            trader_name TEXT,
            raw_payload TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pmx_symbol ON trades(symbol)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pmx_trade_date ON trades(trade_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pmx_order_id ON trades(order_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pmx_doc_number ON trades(doc_number)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_pmx_fnc_number ON trades(fnc_number)")
    removed_swaps = _pmx_delete_swap_rows(conn)
    conn.commit()
    if removed_swaps > 0:
        print(f"[PMX] Removed {removed_swaps} historical SWT/SWAP rows from PMX DB during startup.")
    conn.close()


def initialize_account_opening_balances_table():
    """Initialize the account_opening_balances table for reconciliation opening balances."""
    conn = get_pmx_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS account_opening_balances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month TEXT NOT NULL,
            currency TEXT NOT NULL,
            opening_balance REAL NOT NULL DEFAULT 0.0,
            updated_at TEXT NOT NULL,
            UNIQUE(month, currency)
        )
    """)
    conn.commit()
    conn.close()


initialize_pmx_database()
initialize_account_opening_balances_table()
try:
    initialize_clean_pipeline_db()
except Exception as exc:
    print(f"[WARN] Clean pipeline DB init failed: {exc}")
try:
    if ENABLE_FISCAL_PURGE and FISCAL_TRADES_START_DATE:
        _startup_purge = _purge_pre_fiscal_rows(cutoff_iso=FISCAL_TRADES_START_DATE)
        _removed = int(_startup_purge.get("fx_trades", 0)) + int(_startup_purge.get("trademc_trades", 0)) + int(
            _startup_purge.get("trademc_weight_transactions", 0)
        ) + int(_startup_purge.get("pmx_trades", 0))
        if _removed > 0:
            print(
                "[FISCAL] Removed pre-fiscal records on startup "
                f"(cutoff {FISCAL_TRADES_START_DATE}): {_startup_purge}"
            )
except Exception as exc:
    print(f"[WARN] Startup fiscal purge failed: {exc}")
try:
    _trigger_clean_pipeline("startup", wait=False)
except Exception as exc:
    print(f"[WARN] Startup clean pipeline trigger failed: {exc}")


# PMX helpers
def _pmx_first_non_empty(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        if isinstance(value, float) and pd.isna(value):
            continue
        text = str(value).strip()
        if text and text.lower() != "nan":
            return text
    return ""


def _pmx_to_float(value: Any) -> float:
    if value is None:
        return 0.0
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def _pmx_parse_date(value: Any, default_value: str = "") -> str:
    text = str(value or "").strip()
    if not text:
        return default_value

    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%Y%m%d", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass

    dt = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(dt):
        return dt.strftime("%Y-%m-%d")

    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) >= 8:
        try:
            return datetime.strptime(digits[:8], "%Y%m%d").strftime("%Y-%m-%d")
        except Exception:
            pass
    return default_value


def _pmx_parse_datetime(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    dt = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(dt):
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return text


def _pmx_normalize_symbol(value: Any) -> str:
    return str(value or "").upper().replace("/", "").replace("-", "").replace(" ", "").strip()


def _pmx_to_currency_pair(value: Any) -> str:
    text = str(value or "").upper().strip()
    if not text:
        return ""

    direct = text.replace(" ", "")
    if "/" in direct:
        left, right = direct.split("/", 1)
        if len(left) >= 3 and len(right) >= 3:
            return f"{left[:3]}/{right[:3]}"
    if "-" in direct:
        left, right = direct.split("-", 1)
        if len(left) >= 3 and len(right) >= 3:
            return f"{left[:3]}/{right[:3]}"

    letters = re.sub(r"[^A-Z]", "", text)
    if len(letters) >= 6:
        return f"{letters[:3]}/{letters[3:6]}"
    return ""


def _pmx_extract_support_doc(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        match = PMX_SUPPORT_DOC_PATTERN.search(text)
        if match:
            return match.group(0).strip()
    return ""


def _pmx_extract_quantity(row: Dict[str, Any]) -> float:
    for key in ("pcs_qty", "grs_qty", "grs", "qty", "last_qty", "LastQty", "Quantity"):
        qty = _pmx_to_float(row.get(key))
        if abs(qty) > 0:
            return qty
    return 0.0


def _pmx_extract_price(row: Dict[str, Any], narration_hint: str = "") -> float:
    for key in ("mtl_rate", "rate", "last_px", "LastPx", "Price", "price", "settlement_price"):
        px = _pmx_to_float(row.get(key))
        if px > 0:
            return px

    for text in (row.get("remarks"), row.get("remarks1"), row.get("comment"), narration_hint):
        match = re.search(r"@\s*([0-9][0-9,]*(?:\.[0-9]+)?)", str(text or ""))
        if match:
            px = _pmx_to_float(match.group(1))
            if px > 0:
                return px
    return 0.0


def _pmx_extract_side(row: Dict[str, Any], qty_hint: float = 0.0) -> str:
    raw = _pmx_first_non_empty(
        row.get("deal_type"),
        row.get("DealType"),
        row.get("side"),
        row.get("Side"),
        row.get("buy_sell"),
        row.get("BuySell"),
    ).upper()
    if raw in {"BUY", "B", "1"}:
        return "BUY"
    if raw in {"SELL", "S", "2"}:
        return "SELL"
    if "BUY" in raw:
        return "BUY"
    if "SELL" in raw:
        return "SELL"
    return "SELL" if qty_hint < 0 else "BUY"


def _pmx_build_doc_number(row: Dict[str, Any], fallback_index: int) -> str:
    raw_doc = _pmx_first_non_empty(
        row.get("docno"),
        row.get("DocNo"),
        row.get("doc_number"),
        row.get("DocNumber"),
        row.get("trd"),
        row.get("TradeId"),
        row.get("Id"),
        row.get("RecId"),
        row.get("NeoId"),
        row.get("TagNumber"),
    )
    normalized = normalize_trade_number(raw_doc)
    if normalized:
        return normalized

    stable_basis = "|".join(
        [
            _pmx_first_non_empty(row.get("docdate")),
            _pmx_first_non_empty(row.get("valdate")),
            _pmx_first_non_empty(row.get("inst_desc"), row.get("stk_type_name"), row.get("currency_pair")),
            _pmx_first_non_empty(row.get("deal_type"), row.get("side")),
            str(_pmx_extract_quantity(row)),
            str(_pmx_extract_price(row)),
            _pmx_first_non_empty(row.get("evt_ts"), row.get("event_ts")),
            _pmx_first_non_empty(row.get("remarks"), row.get("remarks1")),
        ]
    )
    digest = hashlib.md5(stable_basis.encode("utf-8")).hexdigest()[:16]
    return f"PMX-{digest}-{fallback_index}"


def _pmx_is_swap_trade(row: Dict[str, Any], support_doc: str = "", narration_hint: str = "") -> bool:
    if not isinstance(row, dict):
        return False

    support_doc_norm = str(support_doc or "").strip().upper()
    if support_doc_norm.startswith("SWT/"):
        return True

    deal_type = _pmx_first_non_empty(
        row.get("deal_type"),
        row.get("DealType"),
        row.get("trade_type"),
        row.get("TradeType"),
        row.get("trd_opt"),
        row.get("TrdOpt"),
    ).upper()
    if deal_type in {"SWT", "SWAP", "SWAPS"} or "SWAP" in deal_type:
        return True

    for value in (
        row.get("order_id"),
        row.get("OrderId"),
        row.get("trade_number"),
        row.get("trade_no"),
        row.get("ref_number"),
        row.get("docno"),
        row.get("DocNo"),
        row.get("NeoId"),
        row.get("TagNumber"),
        row.get("remarks"),
        row.get("remarks1"),
        row.get("comment"),
        row.get("notes"),
        row.get("description"),
        narration_hint,
    ):
        text = str(value or "").upper()
        if "SWT/" in text:
            return True
    return False


def _pmx_map_row_to_trade(row: Dict[str, Any], fallback_index: int) -> Optional[Dict[str, Any]]:
    if not isinstance(row, dict):
        return None

    inst_desc = _pmx_first_non_empty(row.get("inst_desc"), row.get("instrument"), row.get("stk_type_name"))
    currency_pair = _pmx_first_non_empty(
        _pmx_to_currency_pair(row.get("CurrencyPair")),
        _pmx_to_currency_pair(row.get("currency_pair")),
        _pmx_to_currency_pair(inst_desc),
        _pmx_to_currency_pair(row.get("cmdty")),
        _pmx_to_currency_pair(row.get("stk_type_name")),
    )

    symbol = _pmx_normalize_symbol(currency_pair)
    if not symbol:
        symbol = _pmx_normalize_symbol(
            _pmx_first_non_empty(row.get("stk_type_name"), row.get("cmdty"), row.get("inst_desc"), row.get("Symbol"))
        )
        if len(symbol) > 6:
            symbol = symbol[:6]
    if len(symbol) >= 6 and not currency_pair:
        currency_pair = f"{symbol[:3]}/{symbol[3:6]}"

    qty_raw = _pmx_extract_quantity(row)
    qty = abs(qty_raw)
    if not symbol or qty <= 0:
        return None

    narration = _pmx_first_non_empty(
        row.get("remarks"),
        row.get("remarks1"),
        row.get("comment"),
        row.get("notes"),
        row.get("description"),
        row.get("ContractDescription"),
        inst_desc,
    )
    px = _pmx_extract_price(row, narration_hint=narration)
    side = _pmx_extract_side(row, qty_hint=qty_raw)

    today_str = datetime.now().strftime("%Y-%m-%d")
    trade_date = _pmx_parse_date(_pmx_first_non_empty(row.get("docdate"), row.get("TradeDate")), default_value=today_str)
    value_date = _pmx_parse_date(
        _pmx_first_non_empty(row.get("valdate"), row.get("ValueDate"), row.get("settlement_date"), row.get("SettlementDate")),
        default_value=trade_date,
    )

    if not narration:
        if currency_pair:
            narration = f"{currency_pair} {qty:,.2f} @ {px:,.5f}" if px > 0 else f"{currency_pair} {qty:,.2f}"
        else:
            narration = f"{symbol} {qty:,.2f} @ {px:,.5f}" if px > 0 else f"{symbol} {qty:,.2f}"

    settle_currency = ""
    if currency_pair and "/" in currency_pair:
        _, settle_currency = currency_pair.split("/", 1)
    settle_currency = _pmx_first_non_empty(
        settle_currency,
        row.get("counter_currency"),
        row.get("CounterCurrency"),
        row.get("currency"),
        row.get("Currency"),
        row.get("trade_currency"),
        row.get("TradeCurrency"),
    ).upper()

    settle_amount = qty * px if px else 0.0
    rest_trade_id = normalize_trade_number(
        _pmx_first_non_empty(row.get("trd"), row.get("TradeId"), row.get("Id"), row.get("RecId"), row.get("deal_id"))
    )
    doc_number = _pmx_build_doc_number(row, fallback_index)
    order_id = normalize_trade_number(
        _pmx_first_non_empty(
            row.get("order_id"),
            row.get("OrderId"),
            row.get("trade_number"),
            row.get("trade_no"),
            row.get("ref_number"),
            _extract_trade_num_from_remarks(row.get("remarks", "")),
        )
    )
    order_id_upper = str(order_id or "").upper()
    if order_id_upper.startswith("SWT/") or "/SWT/" in order_id_upper:
        return None
    clord_id = _pmx_first_non_empty(row.get("clord_id"), row.get("ClOrdId"), row.get("TagNumber"))
    fnc_number = _pmx_extract_support_doc(
        row.get("docno"),
        row.get("remarks"),
        row.get("remarks1"),
        row.get("comment"),
        row.get("notes"),
        row.get("NeoId"),
        row.get("TagNumber"),
        order_id,
    )
    if _pmx_is_swap_trade(row, support_doc=fnc_number, narration_hint=narration):
        return None

    account = _pmx_first_non_empty(row.get("account"), row.get("accno"), row.get("account_no"), row.get("Acc_optKey"))
    trader_name = _pmx_first_non_empty(row.get("trader_name"), row.get("trader"), row.get("created_by"), row.get("username"))
    transact_time = _pmx_parse_datetime(_pmx_first_non_empty(row.get("evt_ts"), row.get("event_ts"), row.get("transact_time")))
    process_date = _pmx_parse_date(_pmx_first_non_empty(row.get("process_date"), row.get("docdate")), default_value=trade_date)

    try:
        raw_payload = json.dumps(row, ensure_ascii=False, default=str)
    except Exception:
        raw_payload = ""

    return {
        "trade_date": trade_date,
        "value_date": value_date,
        "symbol": symbol,
        "side": side,
        "narration": narration,
        "quantity": qty,
        "price": px,
        "settle_currency": settle_currency,
        "settle_amount": settle_amount,
        "doc_number": doc_number,
        "clord_id": clord_id,
        "order_id": order_id,
        "fnc_number": fnc_number,
        "debit_usd": 0.0,
        "credit_usd": 0.0,
        "debit_zar": 0.0,
        "credit_zar": 0.0,
        "debit_xau": 0.0,
        "credit_xau": 0.0,
        "balance_usd": 0.0,
        "balance_zar": 0.0,
        "balance_xau": 0.0,
        "rest_trade_id": rest_trade_id,
        "account": account,
        "counter_currency": _pmx_first_non_empty(row.get("counter_currency"), row.get("CounterCurrency")),
        "currency": _pmx_first_non_empty(row.get("currency"), row.get("Currency")),
        "currency_pair": currency_pair,
        "last_px": px,
        "last_qty": qty,
        "process_date": process_date,
        "trade_currency": _pmx_first_non_empty(row.get("trade_currency"), row.get("TradeCurrency")),
        "transact_time": transact_time,
        "source_system": _pmx_first_non_empty(row.get("source_system"), "PMX"),
        "trader_name": trader_name,
        "raw_payload": raw_payload,
    }


def _normalize_pmx_date_param(value: Any, default_dt: Optional[datetime] = None) -> str:
    if default_dt is None:
        default_dt = datetime.now()
    text = str(value or "").strip()
    if not text:
        return default_dt.strftime("%d-%m-%Y")
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%Y%m%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d-%m-%Y")
        except Exception:
            pass
    dt = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(dt):
        return dt.strftime("%d-%m-%Y")
    return text


def _pmx_yyyy_mm_dd_to_dd_mm_yyyy(value: Any, fallback: str = "") -> str:
    text = str(value or "").strip()
    if not text:
        return fallback
    dt = pd.to_datetime(text, errors="coerce")
    if pd.notna(dt):
        return dt.strftime("%d-%m-%Y")
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) >= 8:
        try:
            return datetime.strptime(digits[:8], "%Y%m%d").strftime("%d-%m-%Y")
        except Exception:
            pass
    return fallback


def _pmx_parse_dd_mm_yyyy(value: str) -> Optional[datetime]:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass
    dt = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(dt):
        return dt.to_pydatetime()
    return None


def _get_latest_pmx_trade_date() -> Optional[str]:
    conn = get_pmx_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            f"""
            SELECT MAX(trade_date)
            FROM trades
            WHERE trade_date IS NOT NULL
              AND TRIM(trade_date) != ''
              AND NOT {_pmx_swap_sql_predicate()}
            """
        )
        row = cursor.fetchone()
        return str(row[0]).strip() if row and row[0] else None
    except Exception:
        return None
    finally:
        conn.close()


def _pmx_filter_text(value: Any) -> str:
    return str(value or "").strip()


def _pmx_filter_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = _pmx_parse_date(text)
    return normalized or text[:10]


def _pmx_like_contains(value: str) -> str:
    escaped = value.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    return f"%{escaped}%"


def load_all_pmx_trades(filters: Optional[Dict[str, Any]] = None) -> pd.DataFrame:
    conn = get_pmx_db_connection()
    try:
        where_clauses = [f"NOT {_pmx_swap_sql_predicate()}"]
        params: List[Any] = []
        if ENABLE_FISCAL_DATE_FILTER:
            cutoff_date = _pmx_filter_date(FISCAL_TRADES_START_DATE)
            if cutoff_date:
                where_clauses.append("COALESCE(NULLIF(substr(trade_date, 1, 10), ''), '0000-00-00') >= ?")
                params.append(cutoff_date)

        if filters:
            symbol = _pmx_filter_text(filters.get("symbol"))
            if symbol and symbol != "All":
                symbol_norm = symbol.replace("/", "").replace("-", "").replace(" ", "").upper()
                if symbol_norm:
                    where_clauses.append(
                        "REPLACE(REPLACE(REPLACE(UPPER(COALESCE(symbol, '')), '/', ''), '-', ''), ' ', '') = ?"
                    )
                    params.append(symbol_norm)

            trade_num = _pmx_filter_text(filters.get("trade_num"))
            if trade_num:
                where_clauses.append("UPPER(COALESCE(order_id, '')) LIKE ? ESCAPE '\\'")
                params.append(_pmx_like_contains(trade_num.upper()))

            fnc_number = _pmx_filter_text(filters.get("fnc_number"))
            if fnc_number:
                where_clauses.append("UPPER(COALESCE(fnc_number, '')) LIKE ? ESCAPE '\\'")
                params.append(_pmx_like_contains(fnc_number.upper()))

            narration = _pmx_filter_text(filters.get("narration"))
            if narration:
                where_clauses.append("UPPER(COALESCE(narration, '')) LIKE ? ESCAPE '\\'")
                params.append(_pmx_like_contains(narration.upper()))

            start_date = _pmx_filter_date(filters.get("start_date"))
            if start_date:
                where_clauses.append("COALESCE(NULLIF(substr(trade_date, 1, 10), ''), '0000-00-00') >= ?")
                params.append(start_date)

            end_date = _pmx_filter_date(filters.get("end_date"))
            if end_date:
                where_clauses.append("COALESCE(NULLIF(substr(trade_date, 1, 10), ''), '9999-12-31') <= ?")
                params.append(end_date)

        where_sql = " AND ".join(where_clauses)
        query = """
            SELECT
                id,
                trade_date AS "Trade Date",
                value_date AS "Value Date",
                symbol AS "Symbol",
                side AS "Side",
                narration AS "Narration",
                quantity AS "Quantity",
                price AS "Price",
                settle_currency AS "Settle Currency",
                settle_amount AS "Settle Amount",
                doc_number AS "Doc #",
                clord_id AS "ClOrdID",
                order_id AS "OrderID",
                fnc_number AS "FNC #",
                debit_usd AS "Debit USD",
                credit_usd AS "Credit USD",
                debit_zar AS "Debit ZAR",
                credit_zar AS "Credit ZAR",
                debit_xau AS "Debit XAU",
                credit_xau AS "Credit XAU",
                balance_usd AS "Balance USD",
                balance_zar AS "Balance ZAR",
                balance_xau AS "Balance XAU",
                source_system AS "Source System",
                trader_name AS "Trader",
                created_at AS "Created At"
            FROM trades
            WHERE {where_sql}
            ORDER BY trade_date ASC, id ASC
        """.format(where_sql=where_sql)
        return pd.read_sql_query(query, conn, params=params)
    finally:
        conn.close()


def load_pmx_trades_for_trade_number(trade_num_input: Any) -> pd.DataFrame:
    """Load PMX ledger rows for one trade number only."""
    trade_num = normalize_trade_number(trade_num_input)
    if not trade_num:
        return pd.DataFrame()

    conn = get_pmx_db_connection()
    try:
        order_expr = "UPPER(TRIM(COALESCE(CAST(order_id AS TEXT), '')))"
        where_clauses = [
            f"({order_expr} = ? OR {order_expr} = ?)",
        ]
        params: List[Any] = [trade_num, f"{trade_num}.0"]

        if ENABLE_FISCAL_DATE_FILTER:
            cutoff_date = _pmx_filter_date(FISCAL_TRADES_START_DATE)
            if cutoff_date:
                where_clauses.append("COALESCE(NULLIF(substr(trade_date, 1, 10), ''), '0000-00-00') >= ?")
                params.append(cutoff_date)

        query = """
            SELECT
                id,
                trade_date AS "Trade Date",
                value_date AS "Value Date",
                symbol AS "Symbol",
                side AS "Side",
                narration AS "Narration",
                quantity AS "Quantity",
                price AS "Price",
                settle_currency AS "Settle Currency",
                settle_amount AS "Settle Amount",
                doc_number AS "Doc #",
                clord_id AS "ClOrdID",
                order_id AS "OrderID",
                fnc_number AS "FNC #",
                debit_usd AS "Debit USD",
                credit_usd AS "Credit USD",
                debit_zar AS "Debit ZAR",
                credit_zar AS "Credit ZAR",
                debit_xau AS "Debit XAU",
                credit_xau AS "Credit XAU",
                balance_usd AS "Balance USD",
                balance_zar AS "Balance ZAR",
                balance_xau AS "Balance XAU",
                source_system AS "Source System",
                trader_name AS "Trader",
                created_at AS "Created At"
            FROM trades
            WHERE {where_sql}
            ORDER BY trade_date ASC, id ASC
        """.format(where_sql=" AND ".join(where_clauses))
        df = pd.read_sql_query(query, conn, params=params)
    finally:
        conn.close()

    if df.empty:
        return df
    if "OrderID" in df.columns:
        df = df[df["OrderID"].apply(normalize_trade_number) == trade_num].copy()
    if df.empty:
        return df

    def _u(col: str) -> pd.Series:
        if col not in df.columns:
            return pd.Series([""] * len(df), index=df.index, dtype="object")
        return df[col].fillna("").astype(str).str.upper().str.strip()

    fnc = _u("FNC #")
    doc = _u("Doc #")
    oid = _u("OrderID")
    nar = _u("Narration")
    swap_mask = (
        fnc.str.startswith("SWT/") | fnc.str.contains("/SWT/", regex=False)
        | doc.str.startswith("SWT/") | doc.str.contains("/SWT/", regex=False)
        | oid.str.startswith("SWT/") | oid.str.contains("/SWT/", regex=False)
        | nar.str.startswith("SWAP ") | nar.str.contains(" SWAP ", regex=False) | nar.str.contains("SWT/", regex=False)
    )
    df = df[~swap_mask].copy()
    return df


def update_pmx_trade_order_id(trade_id: int, order_id: str) -> bool:
    conn = get_pmx_db_connection()
    cursor = conn.cursor()
    try:
        order_id_value = normalize_trade_number(order_id) if order_id and str(order_id).strip() else None
        cursor.execute(
            """
            UPDATE trades
            SET order_id = ?
            WHERE id = ?
            """,
            (order_id_value, trade_id),
        )
        if cursor.rowcount == 0:
            cursor.execute("SELECT 1 FROM trades WHERE id = ?", (trade_id,))
            if cursor.fetchone() is None:
                conn.rollback()
                return False
        conn.commit()
        return True
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _normalize_symbol_for_validation(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _fetch_trade_symbol_for_validation(trade_id: int, use_pmx: bool) -> Optional[str]:
    conn = None
    try:
        conn = get_pmx_db_connection() if use_pmx else get_db_connection()
        row = conn.execute("SELECT symbol FROM trades WHERE id = ?", (int(trade_id),)).fetchone()
        if not row:
            return None
        return str(row[0] or "")
    except Exception:
        return None
    finally:
        if conn is not None:
            conn.close()


def _validate_integer_trade_number_recent_trademc(
    trade_number: str,
    trade_symbol: str,
    days: int = 7,
) -> Tuple[bool, str]:
    """
    Validation rules:
    - Only validate pure-integer trade refs.
    - Skip validation for excluded symbols (XAGUSD, XPTUSD, XPDUSD).
    - For validated refs, require a confirmed TradeMC booking in the last N days.
    """
    normalized_trade = normalize_trade_number(trade_number)
    if not normalized_trade:
        return True, ""

    # Rule: only apply to pure-integer references.
    if not re.fullmatch(r"\d+", normalized_trade):
        return True, ""

    # Rule: skip excluded symbols.
    symbol_key = _normalize_symbol_for_validation(trade_symbol)
    if symbol_key in {"XAGUSD", "XPTUSD", "XPDUSD"}:
        return True, ""

    conn = None
    try:
        conn = get_db_connection()
        rows = pd.read_sql_query(
            """
            SELECT
                ref_number,
                trade_timestamp,
                date_created,
                date_updated,
                last_synced
            FROM trademc_trades
            WHERE LOWER(TRIM(COALESCE(status, ''))) = 'confirmed'
              AND UPPER(REPLACE(REPLACE(REPLACE(TRIM(COALESCE(ref_number, '')), ' ', ''), '-', ''), '/', '')) = ?
            """,
            conn,
            params=[normalized_trade],
        )
    except Exception as exc:
        return False, f"Invalid trade number: could not validate against TradeMC ({exc})."
    finally:
        if conn is not None:
            conn.close()

    if rows.empty:
        return (
            False,
            f"Invalid trade number '{normalized_trade}'. "
            f"No confirmed TradeMC booking exists for this reference in the last {int(days)} days.",
        )

    cutoff_utc = pd.Timestamp.utcnow() - pd.Timedelta(days=max(1, int(days)))
    latest_ts: Optional[pd.Timestamp] = None
    for _, row in rows.iterrows():
        for col in ("trade_timestamp", "date_created", "date_updated", "last_synced"):
            raw_val = row.get(col)
            if raw_val is None or str(raw_val).strip() == "":
                continue
            parsed = pd.to_datetime(raw_val, errors="coerce", utc=True)
            if pd.isna(parsed):
                continue
            if latest_ts is None or parsed > latest_ts:
                latest_ts = parsed

    if latest_ts is None or latest_ts < cutoff_utc:
        latest_label = latest_ts.strftime("%Y-%m-%d %H:%M:%S UTC") if latest_ts is not None else "unknown"
        return (
            False,
            f"Invalid trade number '{normalized_trade}'. "
            f"No confirmed TradeMC booking exists in the last {int(days)} days "
            f"(latest found: {latest_label}).",
        )

    return True, ""


def sync_pmx_trades_to_db(data: Dict[str, Any], req_headers: Any) -> Dict[str, Any]:
    data = data or {}

    replace = bool(data.get("replace", False))
    now = datetime.now()
    fiscal_cutoff = _pmx_filter_date(FISCAL_TRADES_START_DATE) if ENABLE_FISCAL_DATE_FILTER else ""
    fiscal_floor_dt: Optional[datetime] = None
    if fiscal_cutoff:
        try:
            fiscal_floor_dt = datetime.strptime(fiscal_cutoff, "%Y-%m-%d")
        except Exception:
            fiscal_floor_dt = None
    fallback_start_dt = fiscal_floor_dt or now

    start_date_raw = str(data.get("start_date", "") or "").strip()
    end_date_raw = str(data.get("end_date", "") or "").strip()

    history_start_default = os.getenv("PMX_HISTORY_START_DATE", fiscal_cutoff if fiscal_cutoff else "")
    default_start = _normalize_pmx_date_param(
        str(data.get("default_start_date", history_start_default) or ""),
        default_dt=fallback_start_dt,
    )
    default_end = _normalize_pmx_date_param("", default_dt=now)

    start_date = _normalize_pmx_date_param(start_date_raw, default_dt=fallback_start_dt)
    end_date = _normalize_pmx_date_param(end_date_raw, default_dt=now)

    if not start_date_raw:
        latest_local = _get_latest_pmx_trade_date() if not replace else None
        if latest_local:
            start_date = _pmx_yyyy_mm_dd_to_dd_mm_yyyy(latest_local, fallback=default_start)
        else:
            start_date = default_start
    if not end_date_raw:
        end_date = default_end

    start_dt = _pmx_parse_dd_mm_yyyy(start_date)
    end_dt = _pmx_parse_dd_mm_yyyy(end_date)
    if fiscal_floor_dt is not None:
        fiscal_floor_dd = _pmx_yyyy_mm_dd_to_dd_mm_yyyy(fiscal_cutoff, fallback=start_date)
        if start_dt is None or start_dt < fiscal_floor_dt:
            start_date = fiscal_floor_dd
            start_dt = fiscal_floor_dt
        if end_dt is None or end_dt < fiscal_floor_dt:
            end_date = fiscal_floor_dd
            end_dt = fiscal_floor_dt
    if start_dt and end_dt and start_dt > end_dt:
        start_date = end_date

    resolved_headers = _pmx_resolve_headers(data, req_headers, auto_login=True)
    x_auth = str(resolved_headers.get("x_auth", "") or "").strip()
    sid = str(resolved_headers.get("sid", "") or "").strip()
    username = str(resolved_headers.get("username", "") or "").strip()
    platform = str(resolved_headers.get("platform", "") or "").strip()
    location = str(resolved_headers.get("location", "") or "").strip()
    cache_control = str(resolved_headers.get("cache_control", "") or "").strip()
    content_type = str(resolved_headers.get("content_type", "") or "").strip()

    extra_headers = data.get("headers", {})
    if not isinstance(extra_headers, dict):
        extra_headers = {}

    cmdty = str(data.get("cmdty", "All") or "All")
    trd_opt = str(data.get("trd_opt", "All") or "All")
    created_by = str(data.get("created_by", os.getenv("PMX_CREATED_BY", "2")) or "2")
    acc_opt_key = str(data.get("acc_opt_key", os.getenv("PMX_ACC_OPT_KEY", "MT0601")) or "MT0601")
    trade_type = str(data.get("trade_type", "TD") or "TD")
    non_trd_cmdty = str(data.get("non_trd_cmdty", "") or "")
    host = str(data.get("host", os.getenv("PMX_API_HOST", "pmxapi.stonex.com")) or "pmxapi.stonex.com")
    path = str(data.get("path", "/user/alldealFilter_report") or "/user/alldealFilter_report")
    authorization = str(data.get("authorization", "") or "")
    cookie = str(data.get("cookie", "") or "")
    origin = str(data.get("origin", "https://pmxecute.stonex.com") or "https://pmxecute.stonex.com")
    referer = str(data.get("referer", "https://pmxecute.stonex.com/") or "https://pmxecute.stonex.com/")
    timeout = int(data.get("timeout", 180) or 180)

    def _decode_payload(fetch_result: Dict[str, Any]) -> Any:
        parsed_payload = fetch_result.get("json")
        if parsed_payload is None:
            body_text = fetch_result.get("body", "")
            if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
                try:
                    parsed_payload = json.loads(body_text)
                except Exception:
                    parsed_payload = {}
            else:
                parsed_payload = {}
        return parsed_payload

    def _fetch_report(
        x_auth_value: str,
        sid_value: str,
        username_value: str,
        platform_value: str,
        location_value: str,
        cache_control_value: str,
        content_type_value: str,
    ) -> Dict[str, Any]:
        return fetch_pmx_alldeal_filter_report(
            start_date=start_date,
            end_date=end_date,
            cmdty=cmdty,
            trd_opt=trd_opt,
            created_by=created_by,
            acc_opt_key=acc_opt_key,
            trade_type=trade_type,
            non_trd_cmdty=non_trd_cmdty,
            host=host,
            path=path,
            authorization=authorization,
            cookie=cookie,
            x_auth=x_auth_value,
            sid=sid_value,
            username=username_value,
            platform=platform_value,
            location=location_value,
            cache_control=cache_control_value,
            content_type=content_type_value,
            extra_headers=extra_headers,
            origin=origin,
            referer=referer,
            timeout=timeout,
        )

    result = _fetch_report(
        x_auth_value=x_auth,
        sid_value=sid,
        username_value=username,
        platform_value=platform,
        location_value=location,
        cache_control_value=cache_control,
        content_type_value=content_type,
    )
    payload = _decode_payload(result)
    session_refreshed = False
    relogin_error = ""

    if not result.get("ok") and _pmx_result_is_auth_failure(result, payload):
        relogin = _pmx_login_session(data)
        if relogin.get("ok"):
            session_refreshed = True
            x_auth = _pmx_non_empty(relogin.get("x_auth"), x_auth)
            sid = _pmx_non_empty(relogin.get("sid"), sid)
            username = _pmx_non_empty(relogin.get("username"), username)
            platform = _pmx_non_empty(relogin.get("platform"), platform)
            location = _pmx_non_empty(relogin.get("location"), location)
            cache_control = _pmx_non_empty(relogin.get("cache_control"), cache_control)
            content_type = _pmx_non_empty(relogin.get("content_type"), content_type)
            result = _fetch_report(
                x_auth_value=x_auth,
                sid_value=sid,
                username_value=username,
                platform_value=platform,
                location_value=location,
                cache_control_value=cache_control,
                content_type_value=content_type,
            )
            payload = _decode_payload(result)
        else:
            relogin_error = str(relogin.get("error", "") or "").strip()

    pmx_rows = extract_pmx_report_rows(payload)
    out: Dict[str, Any] = {
        "ok": bool(result.get("ok")),
        "status": result.get("status"),
        "error": result.get("error", ""),
        "message": payload.get("message", "") if isinstance(payload, dict) else "",
        "url": result.get("url", ""),
        "cmdty": cmdty,
        "trd_opt": trd_opt,
        "start_date": start_date,
        "end_date": end_date,
        "fetched_rows": len(pmx_rows),
        "inserted": 0,
        "updated": 0,
        "skipped": 0,
        "skipped_swaps": 0,
        "removed_swaps": 0,
        "replace": replace,
        "session_refreshed": session_refreshed,
    }

    if not result.get("ok"):
        missing_auth_headers = []
        if not x_auth:
            missing_auth_headers.append("x-auth")
        if not sid:
            missing_auth_headers.append("sid")
        if not username:
            missing_auth_headers.append("username")

        raw_error = str(out.get("error", "") or "").strip()
        if relogin_error:
            out["error"] = (
                f"{raw_error or 'PMX request failed'}. "
                f"Auto-login attempt failed: {relogin_error}"
            )
            if not out.get("message"):
                out["message"] = out["error"]
        elif missing_auth_headers:
            missing_txt = ", ".join(missing_auth_headers)
            out["error"] = (
                f"PMX request failed. Missing required PMX headers: {missing_txt}. "
                "Provide them in PMX Ledger or configure PMX_LOGIN_USERNAME/PMX_LOGIN_PASSWORD for auto-login."
            )
            if not out.get("message"):
                out["message"] = out["error"]
        elif raw_error.lower() in {"internal server error", "server error"}:
            out["error"] = (
                "PMX API returned Internal Server Error. "
                "This usually means your x-auth/sid session values are expired and must be refreshed."
            )
            if not out.get("message"):
                out["message"] = out["error"]

        if data.get("debug"):
            out["request_header_presence"] = {
                "x_auth": bool(x_auth),
                "sid": bool(sid),
                "username": bool(username),
                "platform": bool(platform),
                "location": bool(location),
                "session_refreshed": session_refreshed,
                "relogin_error": relogin_error,
            }
        return out

    insert_cols = [
        "trade_date",
        "value_date",
        "symbol",
        "side",
        "narration",
        "quantity",
        "price",
        "settle_currency",
        "settle_amount",
        "doc_number",
        "clord_id",
        "order_id",
        "fnc_number",
        "debit_usd",
        "credit_usd",
        "debit_zar",
        "credit_zar",
        "debit_xau",
        "credit_xau",
        "balance_usd",
        "balance_zar",
        "balance_xau",
        "rest_trade_id",
        "account",
        "counter_currency",
        "currency",
        "currency_pair",
        "last_px",
        "last_qty",
        "process_date",
        "trade_currency",
        "transact_time",
        "source_system",
        "trader_name",
        "raw_payload",
    ]
    placeholders = ", ".join(["?"] * len(insert_cols))
    insert_sql = f"""
        INSERT INTO trades ({", ".join(insert_cols)})
        VALUES ({placeholders})
        ON CONFLICT(doc_number) DO UPDATE SET
            trade_date = excluded.trade_date,
            value_date = excluded.value_date,
            symbol = excluded.symbol,
            side = excluded.side,
            narration = excluded.narration,
            quantity = excluded.quantity,
            price = excluded.price,
            settle_currency = excluded.settle_currency,
            settle_amount = excluded.settle_amount,
            clord_id = excluded.clord_id,
            order_id = CASE
                WHEN excluded.order_id IS NOT NULL AND TRIM(excluded.order_id) != '' THEN excluded.order_id
                ELSE trades.order_id
            END,
            fnc_number = CASE
                WHEN excluded.fnc_number IS NOT NULL AND TRIM(excluded.fnc_number) != '' THEN excluded.fnc_number
                ELSE trades.fnc_number
            END,
            debit_usd = excluded.debit_usd,
            credit_usd = excluded.credit_usd,
            debit_zar = excluded.debit_zar,
            credit_zar = excluded.credit_zar,
            debit_xau = excluded.debit_xau,
            credit_xau = excluded.credit_xau,
            balance_usd = excluded.balance_usd,
            balance_zar = excluded.balance_zar,
            balance_xau = excluded.balance_xau,
            rest_trade_id = excluded.rest_trade_id,
            account = excluded.account,
            counter_currency = excluded.counter_currency,
            currency = excluded.currency,
            currency_pair = excluded.currency_pair,
            last_px = excluded.last_px,
            last_qty = excluded.last_qty,
            process_date = excluded.process_date,
            trade_currency = excluded.trade_currency,
            transact_time = excluded.transact_time,
            source_system = excluded.source_system,
            trader_name = excluded.trader_name,
            raw_payload = excluded.raw_payload
    """

    conn = get_pmx_db_connection()
    cursor = conn.cursor()
    try:
        if replace:
            cursor.execute("DELETE FROM trades")

        inserted = 0
        updated = 0
        skipped = 0
        skipped_swaps = 0
        skipped_pre_start = 0
        removed_swaps = 0

        if not replace:
            removed_swaps = _pmx_delete_swap_rows(conn)

        for idx, row in enumerate(pmx_rows, start=1):
            if _pmx_is_swap_trade(row):
                skipped += 1
                skipped_swaps += 1
                continue

            mapped = _pmx_map_row_to_trade(row, idx)
            if not mapped:
                skipped += 1
                continue
            trade_date_value = str(mapped.get("trade_date", "") or "").strip()[:10]
            if trade_date_value and trade_date_value < TRADE_DATA_START_DATE:
                skipped += 1
                skipped_pre_start += 1
                continue

            cursor.execute("SELECT 1 FROM trades WHERE doc_number = ?", (mapped["doc_number"],))
            exists = cursor.fetchone() is not None
            values = tuple(mapped[col] for col in insert_cols)
            cursor.execute(insert_sql, values)
            if exists:
                updated += 1
            else:
                inserted += 1

        cursor.execute(
            """
            DELETE FROM trades
            WHERE COALESCE(NULLIF(substr(trade_date, 1, 10), ''), '0000-00-00') < ?
            """,
            (TRADE_DATA_START_DATE,),
        )
        removed_pre_start_existing = max(int(cursor.rowcount or 0), 0)
        conn.commit()
        out["inserted"] = inserted
        out["updated"] = updated
        out["skipped"] = skipped
        out["skipped_swaps"] = skipped_swaps
        out["skipped_pre_start"] = skipped_pre_start
        out["removed_pre_start_existing"] = removed_pre_start_existing
        out["trade_data_start_date"] = TRADE_DATA_START_DATE
        out["removed_swaps"] = removed_swaps
        out["fiscal_cutoff"] = fiscal_cutoff
        out["removed_pre_fiscal"] = 0
        if ENABLE_FISCAL_PURGE and fiscal_cutoff:
            fiscal_purge = _purge_pre_fiscal_rows(cutoff_iso=fiscal_cutoff, purge_fx=False, purge_pmx=True)
            out["removed_pre_fiscal"] = int(fiscal_purge.get("pmx_trades", 0))
    except Exception as exc:
        conn.rollback()
        out["ok"] = False
        out["error"] = str(exc)
    finally:
        conn.close()

    if data.get("debug"):
        out["preview"] = pmx_rows[:5]
        out["request_header_presence"] = {
            "x_auth": bool(x_auth),
            "sid": bool(sid),
            "username": bool(username),
            "platform": bool(platform),
            "location": bool(location),
        }
    return out


def _fetch_pmx_raw_report_rows(data: Dict[str, Any], req_headers: Any) -> Dict[str, Any]:
    """Fetch raw PMX all-deal report rows without writing to local DB."""
    data = data or {}
    now = datetime.now()
    fiscal_cutoff = _pmx_filter_date(FISCAL_TRADES_START_DATE) if ENABLE_FISCAL_DATE_FILTER else ""
    fiscal_floor_dt: Optional[datetime] = None
    if fiscal_cutoff:
        try:
            fiscal_floor_dt = datetime.strptime(fiscal_cutoff, "%Y-%m-%d")
        except Exception:
            fiscal_floor_dt = None

    start_date_raw = str(data.get("start_date", "") or "").strip()
    end_date_raw = str(data.get("end_date", "") or "").strip()

    history_start_default = os.getenv("PMX_HISTORY_START_DATE", fiscal_cutoff if fiscal_cutoff else "")
    default_start = _normalize_pmx_date_param(
        str(data.get("default_start_date", history_start_default) or ""),
        default_dt=fiscal_floor_dt or now,
    )
    default_end = _normalize_pmx_date_param("", default_dt=now)

    start_date = _normalize_pmx_date_param(start_date_raw, default_dt=fiscal_floor_dt or now)
    end_date = _normalize_pmx_date_param(end_date_raw, default_dt=now)
    if not start_date_raw:
        start_date = default_start
    if not end_date_raw:
        end_date = default_end

    start_dt = _pmx_parse_dd_mm_yyyy(start_date)
    end_dt = _pmx_parse_dd_mm_yyyy(end_date)
    if fiscal_floor_dt is not None:
        fiscal_floor_dd = _pmx_yyyy_mm_dd_to_dd_mm_yyyy(fiscal_cutoff, fallback=start_date)
        if start_dt is None or start_dt < fiscal_floor_dt:
            start_date = fiscal_floor_dd
            start_dt = fiscal_floor_dt
        if end_dt is None or end_dt < fiscal_floor_dt:
            end_date = fiscal_floor_dd
            end_dt = fiscal_floor_dt
    if start_dt and end_dt and start_dt > end_dt:
        start_date = end_date

    resolved_headers = _pmx_resolve_headers(data, req_headers, auto_login=True)
    x_auth = str(resolved_headers.get("x_auth", "") or "").strip()
    sid = str(resolved_headers.get("sid", "") or "").strip()
    username = str(resolved_headers.get("username", "") or "").strip()
    platform = str(resolved_headers.get("platform", "") or "").strip()
    location = str(resolved_headers.get("location", "") or "").strip()
    cache_control = str(resolved_headers.get("cache_control", "") or "").strip()
    content_type = str(resolved_headers.get("content_type", "") or "").strip()

    extra_headers = data.get("headers", {})
    if not isinstance(extra_headers, dict):
        extra_headers = {}

    cmdty = str(data.get("cmdty", "All") or "All")
    trd_opt = str(data.get("trd_opt", "All") or "All")
    created_by = str(data.get("created_by", os.getenv("PMX_CREATED_BY", "2")) or "2")
    acc_opt_key = str(data.get("acc_opt_key", os.getenv("PMX_ACC_OPT_KEY", "MT0601")) or "MT0601")
    trade_type = str(data.get("trade_type", "TD") or "TD")
    non_trd_cmdty = str(data.get("non_trd_cmdty", "") or "")
    host = str(data.get("host", os.getenv("PMX_API_HOST", "pmxapi.stonex.com")) or "pmxapi.stonex.com")
    path = str(data.get("path", "/user/alldealFilter_report") or "/user/alldealFilter_report")
    authorization = str(data.get("authorization", "") or "")
    cookie = str(data.get("cookie", "") or "")
    origin = str(data.get("origin", "https://pmxecute.stonex.com") or "https://pmxecute.stonex.com")
    referer = str(data.get("referer", "https://pmxecute.stonex.com/") or "https://pmxecute.stonex.com/")
    timeout = int(data.get("timeout", 180) or 180)

    def _decode_payload(fetch_result: Dict[str, Any]) -> Any:
        parsed_payload = fetch_result.get("json")
        if parsed_payload is None:
            body_text = fetch_result.get("body", "")
            if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
                try:
                    parsed_payload = json.loads(body_text)
                except Exception:
                    parsed_payload = {}
            else:
                parsed_payload = {}
        return parsed_payload

    def _fetch_report(
        x_auth_value: str,
        sid_value: str,
        username_value: str,
        platform_value: str,
        location_value: str,
        cache_control_value: str,
        content_type_value: str,
    ) -> Dict[str, Any]:
        return fetch_pmx_alldeal_filter_report(
            start_date=start_date,
            end_date=end_date,
            cmdty=cmdty,
            trd_opt=trd_opt,
            created_by=created_by,
            acc_opt_key=acc_opt_key,
            trade_type=trade_type,
            non_trd_cmdty=non_trd_cmdty,
            host=host,
            path=path,
            authorization=authorization,
            cookie=cookie,
            x_auth=x_auth_value,
            sid=sid_value,
            username=username_value,
            platform=platform_value,
            location=location_value,
            cache_control=cache_control_value,
            content_type=content_type_value,
            extra_headers=extra_headers,
            origin=origin,
            referer=referer,
            timeout=timeout,
        )

    result = _fetch_report(
        x_auth_value=x_auth,
        sid_value=sid,
        username_value=username,
        platform_value=platform,
        location_value=location,
        cache_control_value=cache_control,
        content_type_value=content_type,
    )
    payload = _decode_payload(result)
    session_refreshed = False

    if not result.get("ok") and _pmx_result_is_auth_failure(result, payload):
        relogin = _pmx_login_session(data)
        if relogin.get("ok"):
            session_refreshed = True
            x_auth = _pmx_non_empty(relogin.get("x_auth"), x_auth)
            sid = _pmx_non_empty(relogin.get("sid"), sid)
            username = _pmx_non_empty(relogin.get("username"), username)
            platform = _pmx_non_empty(relogin.get("platform"), platform)
            location = _pmx_non_empty(relogin.get("location"), location)
            cache_control = _pmx_non_empty(relogin.get("cache_control"), cache_control)
            content_type = _pmx_non_empty(relogin.get("content_type"), content_type)
            result = _fetch_report(
                x_auth_value=x_auth,
                sid_value=sid,
                username_value=username,
                platform_value=platform,
                location_value=location,
                cache_control_value=cache_control,
                content_type_value=content_type,
            )
            payload = _decode_payload(result)

    rows = extract_pmx_report_rows(payload)
    return {
        "ok": bool(result.get("ok")),
        "status": result.get("status"),
        "error": str(result.get("error", "") or ""),
        "rows": rows,
        "start_date": start_date,
        "end_date": end_date,
        "session_refreshed": session_refreshed,
    }


def get_all_companies_df():
    """Load companies table into a DataFrame."""
    import sqlite3
    db_path = LEDGER_DB_PATH
    conn = sqlite3.connect(db_path)
    try:
        df = pd.read_sql_query("SELECT * FROM trademc_companies ORDER BY company_name", conn)
    except Exception:
        df = pd.DataFrame()
    conn.close()
    return df


def load_trademc_trades_with_companies(**kwargs):
    """Load TradeMC trades joined with company names."""
    import sqlite3
    db_path = LEDGER_DB_PATH
    conn = sqlite3.connect(db_path)
    query = """
        SELECT t.*, c.company_name, c.refining_rate AS company_refining_rate
        FROM trademc_trades t
        LEFT JOIN trademc_companies c ON t.company_id = c.id
        WHERE 1=1
    """
    params = []
    status = kwargs.get("status")
    if status:
        query += " AND t.status = ?"
        params.append(status)
    ref_filter = kwargs.get("ref_filter")
    if ref_filter:
        query += " AND t.ref_number LIKE ?"
        params.append(f"%{ref_filter}%")
    company_id = kwargs.get("company_id")
    if company_id:
        query += " AND t.company_id = ?"
        params.append(company_id)
    start_date = kwargs.get("start_date")
    if start_date:
        query += " AND DATE(t.trade_timestamp) >= DATE(?)"
        params.append(start_date)
    end_date = kwargs.get("end_date")
    if end_date:
        query += " AND DATE(t.trade_timestamp) <= DATE(?)"
        params.append(end_date)
    query += " ORDER BY t.trade_timestamp DESC"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    for col in ["trade_timestamp", "date_created", "date_updated", "last_synced"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


# Load weight transactions with company join
def load_weight_transactions(**kwargs):
    import sqlite3
    db_path = LEDGER_DB_PATH
    conn = sqlite3.connect(db_path)
    query = """
        SELECT w.*, c.company_name
        FROM trademc_weight_transactions w
        LEFT JOIN trademc_companies c ON w.company_id = c.id
        WHERE 1=1
    """
    params = []
    company_id = kwargs.get("company_id")
    if company_id:
        query += " AND w.company_id = ?"
        params.append(company_id)
    txn_type = kwargs.get("type")
    if txn_type:
        query += " AND w.type = ?"
        params.append(txn_type)
    start_date = kwargs.get("start_date")
    if start_date:
        query += " AND w.transaction_timestamp >= ?"
        params.append(start_date)
    end_date = kwargs.get("end_date")
    if end_date:
        query += " AND w.transaction_timestamp <= ?"
        params.append(end_date)
    query += " ORDER BY w.transaction_timestamp DESC"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def build_ledger_view(source_df):
    """Build the trading ledger view from raw trades."""
    if source_df is None or source_df.empty:
        return pd.DataFrame()

    df = source_df.sort_values(by=["Trade Date", "Value Date", "id"]).copy()

    _safe_str = lambda col: (
        df[col].fillna("").astype(str).str.strip()
        if col in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )
    trade_nums = _safe_str("OrderID")
    docs = _safe_str("Doc #")
    fnc_nums = _safe_str("FNC #")
    symbols = _safe_str("Symbol").str.upper()
    sides = _safe_str("Side").str.upper()
    narrations = _safe_str("Narration")
    traders = _safe_str("Trader")
    qty_raw = pd.to_numeric(df.get("Quantity"), errors="coerce").fillna(0.0)
    price_raw = pd.to_numeric(df.get("Price"), errors="coerce").fillna(0.0)

    # Split symbols
    has_slash = symbols.str.contains("/", na=False)
    bases = pd.Series("", index=df.index)
    quotes = pd.Series("", index=df.index)
    splits = symbols[has_slash].str.split("/", n=1, expand=True)
    if splits is not None and not splits.empty:
        bases.loc[has_slash] = splits[0].values
        quotes.loc[has_slash] = splits[1].values if 1 in splits.columns else ""
    len6 = (~has_slash) & (symbols.str.len() == 6)
    bases.loc[len6] = symbols[len6].str[:3]
    quotes.loc[len6] = symbols[len6].str[3:]
    other = (~has_slash) & (~len6)
    bases.loc[other] = symbols[other]

    is_metal = bases.isin({"XAU", "XAG", "XPT", "XPD"}) & (quotes == "USD")
    is_fx = (bases == "USD") & (quotes == "ZAR")
    is_buy = sides == "BUY"

    debit_usd = pd.Series(0.0, index=df.index)
    credit_usd = pd.Series(0.0, index=df.index)
    debit_zar = pd.Series(0.0, index=df.index)
    credit_zar = pd.Series(0.0, index=df.index)
    credit_xau = pd.Series(0.0, index=df.index)
    debit_xau = pd.Series(0.0, index=df.index)

    m_buy = is_metal & is_buy
    debit_usd.loc[m_buy] = qty_raw[m_buy] * price_raw[m_buy]
    credit_xau.loc[m_buy] = qty_raw[m_buy]
    m_sell = is_metal & (~is_buy)
    credit_usd.loc[m_sell] = qty_raw[m_sell] * price_raw[m_sell]
    debit_xau.loc[m_sell] = qty_raw[m_sell]
    fx_sell = is_fx & (sides == "SELL")
    debit_usd.loc[fx_sell] = qty_raw[fx_sell]
    credit_zar.loc[fx_sell] = qty_raw[fx_sell] * price_raw[fx_sell]
    fx_buy = is_fx & is_buy
    credit_usd.loc[fx_buy] = qty_raw[fx_buy]
    debit_zar.loc[fx_buy] = qty_raw[fx_buy] * price_raw[fx_buy]

    trade_keys = trade_nums.where(trade_nums != "", docs)
    trade_keys = trade_keys.where(trade_keys != "", df.index.astype(str))

    net_usd = credit_usd - debit_usd
    net_zar = credit_zar - debit_zar
    net_xau = credit_xau - debit_xau

    bal_usd = net_usd.groupby(trade_keys).cumsum()
    bal_zar = net_zar.groupby(trade_keys).cumsum()
    bal_xau = net_xau.groupby(trade_keys).cumsum()

    # Narration auto-fill
    narr_result = narrations.copy()
    empty_narr = narr_result == ""
    metal_no_narr = empty_narr & is_metal
    fx_no_narr = empty_narr & is_fx

    if metal_no_narr.any():
        pair = bases[metal_no_narr] + "/" + quotes[metal_no_narr]
        narr_result.loc[metal_no_narr] = (
            pair + " " + qty_raw[metal_no_narr].apply(lambda x: f"{x:.3f}") +
            " OZ @ " + price_raw[metal_no_narr].apply(lambda x: f"{x:.2f}")
        )
    if fx_no_narr.any():
        narr_result.loc[fx_no_narr] = (
            "USD/ZAR " + qty_raw[fx_no_narr].apply(lambda x: f"{x:,.2f}") +
            " @ " + price_raw[fx_no_narr].apply(lambda x: f"{x:.5f}")
        )

    trade_dates = pd.to_datetime(df.get("Trade Date"), errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
    value_dates = pd.to_datetime(df.get("Value Date"), errors="coerce").dt.strftime("%Y-%m-%d").fillna("")

    ledger_df = pd.DataFrame({
        "id": df["id"].values,
        "Trade #": trade_nums.values,
        "FNC #": fnc_nums.values,
        "Doc #": docs.values,
        "Trade Date": trade_dates.values,
        "Value Date": value_dates.values,
        "Symbol": symbols.values,
        "Side": sides.values,
        # Keep raw PMX trade fields in ledger payload so hedge quick-details can
        # summarize even when ticket rows are unavailable.
        "Quantity": qty_raw.values,
        "Price": price_raw.values,
        "Narration": narr_result.values,
        "Debit USD": debit_usd.values,
        "Credit USD": credit_usd.values,
        "Balance USD": bal_usd.values,
        "Net XAU oz": bal_xau.values,
        "Net XAU g": bal_xau.values * 31.1035,
        "Debit ZAR": debit_zar.values,
        "Credit ZAR": credit_zar.values,
        "Balance ZAR": bal_zar.values,
        "Trader": traders.values,
    })

    # Status
    if not ledger_df.empty:
        status_key = ledger_df["Trade #"].astype(str).str.strip()
        blank = status_key == ""
        status_key.loc[blank] = ledger_df["Doc #"].astype(str).str.strip()
        blank = status_key == ""
        status_key.loc[blank] = ledger_df.index[blank].astype(str)

        last_rows = ledger_df.assign(_sk=status_key).groupby("_sk", dropna=False).tail(1)
        open_keys = set(
            last_rows[
                (last_rows["Balance USD"].abs() > 1e-6) |
                (last_rows["Balance ZAR"].abs() > 1e-6)
            ]["_sk"].tolist()
        )
        ledger_df["Status"] = status_key.apply(lambda k: "Open" if k in open_keys else "Closed")

    return ledger_df


def build_open_positions(source_df):
    """Get open positions summary and detail."""
    if source_df is None or source_df.empty:
        return [], {"open_trades": 0, "open_usd": 0, "open_zar": 0}

    ledger = build_ledger_view(source_df)
    if ledger.empty:
        return [], {"open_trades": 0, "open_usd": 0, "open_zar": 0}

    open_rows = ledger[ledger.get("Status") == "Open"]
    if open_rows.empty:
        return [], {"open_trades": 0, "open_usd": 0, "open_zar": 0}

    # Get unique trade keys and their last row
    trade_key = open_rows["Trade #"].astype(str).str.strip()
    blank = trade_key == ""
    trade_key.loc[blank] = open_rows["Doc #"].astype(str).str.strip()

    unique_keys = trade_key.unique().tolist()
    last_bal_usd = 0
    last_bal_zar = 0
    for k in unique_keys:
        mask = trade_key == k
        rows = open_rows[mask]
        if not rows.empty:
            last_bal_usd += float(rows["Balance USD"].iloc[-1])
            last_bal_zar += float(rows["Balance ZAR"].iloc[-1])

    summary = {
        "open_trades": len(unique_keys),
        "open_usd": round(last_bal_usd, 2),
        "open_zar": round(last_bal_zar, 2),
    }

    # Convert to records for JSON
    records = open_rows.to_dict(orient="records")
    return records, summary



def _parse_loose_number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        num = float(value)
        return num if math.isfinite(num) else None
    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace(",", "").replace(" ", "")
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1]
    try:
        num = float(cleaned)
        return num if math.isfinite(num) else None
    except Exception:
        # Fallback for values like "R1,234.56", "257.21 oz", etc.
        match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?(?:[eE][-+]?\d+)?", cleaned)
        if not match:
            return None
        token = match.group(0).replace(",", "")
        try:
            num = float(token)
            return num if math.isfinite(num) else None
        except Exception:
            return None


def _normalize_ccy_label(value: Any) -> str:
    raw = re.sub(r"[^A-Z]", "", str(value or "").strip().upper())
    if raw in {"USD", "DOLLAR", "USDDOLLAR"}:
        return "USD"
    if raw in {"ZAR", "RAND", "SOUTHAFRICANRAND"}:
        return "ZAR"
    if raw in {"XAU", "GOLD"}:
        return "XAU"
    if raw in {"XAG", "SILVER"}:
        return "XAG"
    return ""


def _collect_nested_dicts(payload: Any) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    stack: List[Any] = [payload]
    while stack:
        node = stack.pop()
        if isinstance(node, dict):
            out.append(node)
            for value in node.values():
                if isinstance(value, (dict, list, tuple)):
                    stack.append(value)
        elif isinstance(node, (list, tuple)):
            for item in node:
                if isinstance(item, (dict, list, tuple)):
                    stack.append(item)
        elif isinstance(node, str):
            text = node.strip()
            if text.startswith("{") or text.startswith("["):
                try:
                    parsed = json.loads(text)
                except Exception:
                    parsed = None
                if isinstance(parsed, (dict, list, tuple)):
                    stack.append(parsed)
    return out


def _extract_pmx_account_balances(payload: Any) -> Dict[str, Any]:
    best: Dict[str, Dict[str, Any]] = {
        "USD": {"score": -1.0, "value": None},
        "ZAR": {"score": -1.0, "value": None},
        "XAU": {"score": -1.0, "value": None},
        "XAG": {"score": -1.0, "value": None},
    }
    summary: Dict[str, Optional[float]] = {
        "total_funds_usd": None,
        "position_valuation_usd": None,
        "net_equity_usd": None,
        "loss_limit": None,
        "margin_requirement_usd": None,
        "free_equity_usd": None,
        "margin_percent": None,
    }
    as_of_date = ""
    dict_nodes = _collect_nested_dicts(payload)

    def _take(ccy: str, value: Any, score: float) -> None:
        num = _parse_loose_number(value)
        if num is None or ccy not in best:
            return
        if score > float(best[ccy]["score"]):
            best[ccy] = {"score": float(score), "value": float(num)}

    # Direct PMX loadAccount payload mapping:
    # {
    #   "status":"success",
    #   "data":{"STKA":[{"D":"XAU","Q":...},{"D":"ZAR","Q":...}], "FA":[{"C":"USD","Q":...}]}
    # }
    if isinstance(payload, dict):
        data_obj = payload.get("data")
        if isinstance(data_obj, dict):
            stka_rows = data_obj.get("STKA")
            if isinstance(stka_rows, list):
                for row in stka_rows:
                    if not isinstance(row, dict):
                        continue
                    ccy = _normalize_ccy_label(
                        row.get("D")
                        or row.get("C")
                        or row.get("CCY")
                        or row.get("currency")
                        or row.get("INSTID")
                    )
                    if ccy in {"XAU", "XAG", "ZAR"}:
                        for amount_key, score in (("Q", 140.0), ("V", 138.0), ("TQ", 132.0), ("UQ", 130.0), ("OQ", 126.0)):
                            if amount_key in row:
                                _take(ccy, row.get(amount_key), score)
            fa_rows = data_obj.get("FA")
            if isinstance(fa_rows, list):
                for row in fa_rows:
                    if not isinstance(row, dict):
                        continue
                    ccy = _normalize_ccy_label(
                        row.get("C")
                        or row.get("D")
                        or row.get("CCY")
                        or row.get("currency")
                    )
                    if ccy == "USD":
                        for amount_key, score in (("Q", 140.0), ("V", 138.0), ("TQ", 132.0), ("UQ", 130.0)):
                            if amount_key in row:
                                _take(ccy, row.get(amount_key), score)

                        # Account summary values from FA(USD) row:
                        # Q = Total Funds, V = Position Valuation, FA/FE = Free Equity.
                        if summary["total_funds_usd"] is None:
                            summary["total_funds_usd"] = _parse_loose_number(row.get("Q"))
                        if summary["position_valuation_usd"] is None:
                            summary["position_valuation_usd"] = _parse_loose_number(
                                row.get("V", row.get("PV"))
                            )
                        if summary["free_equity_usd"] is None:
                            summary["free_equity_usd"] = _parse_loose_number(
                                row.get("FA", row.get("FE"))
                            )
                        if summary["margin_requirement_usd"] is None:
                            summary["margin_requirement_usd"] = _parse_loose_number(
                                row.get("MR", row.get("MARGIN_REQ"))
                            )

            # Header-level summary fields (when PMX provides them).
            if summary["loss_limit"] is None:
                summary["loss_limit"] = _parse_loose_number(data_obj.get("LL", data_obj.get("LOSS_LIMIT")))
            if summary["margin_percent"] is None:
                summary["margin_percent"] = _parse_loose_number(data_obj.get("MP", data_obj.get("MARGIN_PCT")))
            if summary["margin_requirement_usd"] is None:
                summary["margin_requirement_usd"] = _parse_loose_number(
                    data_obj.get("MR", data_obj.get("MARGIN_REQUIREMENT"))
                )
            if summary["free_equity_usd"] is None:
                summary["free_equity_usd"] = _parse_loose_number(data_obj.get("FE", data_obj.get("FREE_EQUITY")))
            if summary["net_equity_usd"] is None:
                summary["net_equity_usd"] = _parse_loose_number(data_obj.get("NE", data_obj.get("NET_EQUITY")))

    for node in dict_nodes:
        if not as_of_date:
            for key in (
                "AsOfDate",
                "asOfDate",
                "as_of_date",
                "as_of",
                "valueDate",
                "value_date",
                "updatedAt",
                "updated_at",
                "lastUpdated",
                "last_update",
                "timestamp",
                "date",
            ):
                val = node.get(key)
                if val is not None and str(val).strip():
                    as_of_date = str(val).strip()
                    break

        # Record-style pattern (currency field + amount field)
        row_ccy = ""
        for node_key, node_value in node.items():
            key_norm = re.sub(r"[^a-z0-9]", "", str(node_key).lower())
            if key_norm in {
                "currency",
                "currencycode",
                "ccy",
                "ccycode",
                "cmdty",
                "cmdtycode",
                "commodity",
                "metal",
                "metalcode",
                "symbol",
                "asset",
                "instrument",
                "name",
            }:
                row_ccy = _normalize_ccy_label(node_value)
                if row_ccy:
                    break
        if row_ccy:
            for key, raw_val in node.items():
                key_norm = re.sub(r"[^a-z0-9]", "", str(key).lower())
                if key_norm in {"currency", "ccy", "cmdty", "commodity", "metal", "symbol", "asset", "instrument", "name"}:
                    continue
                if any(token in key_norm for token in ("rate", "price", "spot", "wa")):
                    continue
                score = 72.0
                if "balance" in key_norm or key_norm.endswith("bal"):
                    score = 98.0
                elif "net" in key_norm or "position" in key_norm:
                    score = 92.0
                elif "amount" in key_norm or "value" in key_norm or "cash" in key_norm or "equity" in key_norm:
                    score = 84.0
                elif "qty" in key_norm or "quantity" in key_norm or "oz" in key_norm or "gram" in key_norm:
                    score = 78.0
                if isinstance(raw_val, (dict, list, tuple)):
                    for nested in _collect_nested_dicts(raw_val):
                        for nested_val in nested.values():
                            _take(row_ccy, nested_val, score - 4.0)
                else:
                    _take(row_ccy, raw_val, score)

        # Key-style pattern (e.g. xauBalance/usd_bal/zar)
        for key, raw_val in node.items():
            if isinstance(raw_val, (dict, list, tuple)):
                continue
            key_norm = re.sub(r"[^a-z0-9]", "", str(key).lower())
            if not key_norm:
                continue
            if any(token in key_norm for token in ("rate", "price", "spot")) and not any(
                token in key_norm for token in ("bal", "balance", "net", "position")
            ):
                continue
            for ccy, tokens in {
                "USD": ("usd", "dollar"),
                "ZAR": ("zar", "rand"),
                "XAU": ("xau", "gold"),
                "XAG": ("xag", "silver"),
            }.items():
                if not any(token in key_norm for token in tokens):
                    continue
                score = 52.0
                if key_norm in set(tokens) | {ccy.lower()}:
                    score = 100.0
                elif "balance" in key_norm or key_norm.endswith("bal"):
                    score = 96.0
                elif "net" in key_norm or "position" in key_norm:
                    score = 90.0
                _take(ccy, raw_val, score)

    xau = best["XAU"]["value"]
    xag = best["XAG"]["value"]
    usd = best["USD"]["value"]
    zar = best["ZAR"]["value"]

    # Derive missing summary values when PMX omits explicit fields.
    total_funds = summary["total_funds_usd"]
    position_valuation = summary["position_valuation_usd"]
    net_equity = summary["net_equity_usd"]
    margin_requirement = summary["margin_requirement_usd"]
    free_equity = summary["free_equity_usd"]

    if net_equity is None and total_funds is not None and position_valuation is not None:
        net_equity = float(total_funds) + float(position_valuation)
    if margin_requirement is None and free_equity is not None and net_equity is not None:
        margin_requirement = float(free_equity) - float(net_equity)
    if free_equity is None and margin_requirement is not None and net_equity is not None:
        free_equity = float(net_equity) + float(margin_requirement)

    return {
        "xau": round(float(xau), 4) if xau is not None else None,
        "xag": round(float(xag), 4) if xag is not None else None,
        "usd": round(float(usd), 2) if usd is not None else None,
        "zar": round(float(zar), 2) if zar is not None else None,
        "total_funds_usd": round(float(total_funds), 2) if total_funds is not None else None,
        "position_valuation_usd": round(float(position_valuation), 2) if position_valuation is not None else None,
        "net_equity_usd": round(float(net_equity), 2) if net_equity is not None else None,
        "loss_limit": round(float(summary["loss_limit"]), 2) if summary["loss_limit"] is not None else None,
        "margin_requirement_usd": round(float(margin_requirement), 2) if margin_requirement is not None else None,
        "free_equity_usd": round(float(free_equity), 2) if free_equity is not None else None,
        "margin_percent": round(float(summary["margin_percent"]), 2) if summary["margin_percent"] is not None else None,
        "as_of_date": as_of_date,
    }


def _empty_account_balances_payload(account_code: str) -> Dict[str, Any]:
    return {
        "account_code": account_code,
        "xau": None,
        "xag": None,
        "usd": None,
        "zar": None,
        "total_funds_usd": None,
        "position_valuation_usd": None,
        "net_equity_usd": None,
        "loss_limit": None,
        "margin_requirement_usd": None,
        "free_equity_usd": None,
        "margin_percent": None,
        "as_of_date": "",
        "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ok": False,
        "status": None,
        "error": "",
    }


def _fetch_open_positions_account_balances(data: Optional[Dict[str, Any]] = None, req_headers: Any = None) -> Dict[str, Any]:
    data = data or {}
    account_code = _pmx_non_empty(
        data.get("acccode"),
        data.get("account_code"),
        data.get("acc_opt_key"),
        os.getenv("PMX_ACC_OPT_KEY", ""),
        "MT0601",
    )
    out = _empty_account_balances_payload(account_code)

    resolved_headers = _pmx_resolve_headers(data, req_headers, auto_login=True)
    x_auth = str(resolved_headers.get("x_auth", "") or "").strip()
    sid = str(resolved_headers.get("sid", "") or "").strip()
    username = str(resolved_headers.get("username", "") or "").strip()
    platform = str(resolved_headers.get("platform", "") or "").strip()
    location = str(resolved_headers.get("location", "") or "").strip()
    cache_control = str(resolved_headers.get("cache_control", "") or "").strip()
    content_type = str(resolved_headers.get("content_type", "") or "").strip()

    host = str(data.get("host", os.getenv("PMX_API_HOST", "pmxapi.stonex.com")) or "pmxapi.stonex.com")
    path = str(data.get("load_account_path", data.get("account_path", "/user/loadAccount")) or "/user/loadAccount")
    authorization = str(data.get("authorization", "") or "")
    cookie = str(data.get("cookie", "") or "")
    extra_headers = data.get("headers", {}) if isinstance(data.get("headers", {}), dict) else {}
    origin = str(data.get("origin", "https://pmxecute.stonex.com") or "https://pmxecute.stonex.com")
    referer = str(data.get("referer", "https://pmxecute.stonex.com/") or "https://pmxecute.stonex.com/")
    timeout = int(data.get("timeout", 60) or 60)

    def _decode_payload(fetch_result: Dict[str, Any]) -> Any:
        parsed_payload = fetch_result.get("json")
        if parsed_payload is None:
            body_text = fetch_result.get("body", "")
            if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
                try:
                    parsed_payload = json.loads(body_text)
                except Exception:
                    parsed_payload = {}
            else:
                parsed_payload = {}
        return parsed_payload

    def _fetch(
        x_auth_value: str,
        sid_value: str,
        username_value: str,
        platform_value: str,
        location_value: str,
        cache_control_value: str,
        content_type_value: str,
    ) -> Dict[str, Any]:
        return fetch_pmx_load_account(
            acccode=account_code,
            host=host,
            path=path,
            authorization=authorization,
            cookie=cookie,
            x_auth=x_auth_value,
            sid=sid_value,
            username=username_value,
            platform=platform_value,
            location=location_value,
            cache_control=cache_control_value,
            content_type=content_type_value,
            extra_headers=extra_headers,
            origin=origin,
            referer=referer,
            timeout=timeout,
        )

    result = _fetch(
        x_auth_value=x_auth,
        sid_value=sid,
        username_value=username,
        platform_value=platform,
        location_value=location,
        cache_control_value=cache_control,
        content_type_value=content_type,
    )
    payload = _decode_payload(result)
    session_refreshed = False
    relogin_error = ""

    if not result.get("ok") and _pmx_result_is_auth_failure(result, payload):
        relogin = _pmx_login_session(data)
        if relogin.get("ok"):
            session_refreshed = True
            x_auth = _pmx_non_empty(relogin.get("x_auth"), x_auth)
            sid = _pmx_non_empty(relogin.get("sid"), sid)
            username = _pmx_non_empty(relogin.get("username"), username)
            platform = _pmx_non_empty(relogin.get("platform"), platform)
            location = _pmx_non_empty(relogin.get("location"), location)
            cache_control = _pmx_non_empty(relogin.get("cache_control"), cache_control)
            content_type = _pmx_non_empty(relogin.get("content_type"), content_type)
            result = _fetch(
                x_auth_value=x_auth,
                sid_value=sid,
                username_value=username,
                platform_value=platform,
                location_value=location,
                cache_control_value=cache_control,
                content_type_value=content_type,
            )
            payload = _decode_payload(result)
        else:
            relogin_error = str(relogin.get("error", "") or "").strip()

    parsed_balances = _extract_pmx_account_balances(payload)
    out.update(parsed_balances)
    out["status"] = result.get("status")
    out["session_refreshed"] = session_refreshed
    out["fetched_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    parsed_any = any(out.get(ccy.lower()) is not None for ccy in ("XAU", "XAG", "ZAR", "USD"))
    out["ok"] = bool(result.get("ok")) and parsed_any

    err_text = str(result.get("error", "") or "").strip()
    if relogin_error:
        err_text = f"{err_text or 'PMX loadAccount request failed'}. Auto-login attempt failed: {relogin_error}"
    if not out["ok"]:
        if err_text:
            out["error"] = err_text
        elif not parsed_any:
            out["error"] = "PMX loadAccount returned no parseable XAU/XAG/ZAR balances."

    return out


def _pmx_recon_build_base_fetch_args(args_dict: Dict[str, Any], req_headers: Any, start_date: str, end_date: str) -> Dict[str, Any]:
    """Build the base fetch args dict for PMX account statement fetches."""
    resolved_headers = _pmx_resolve_headers(args_dict, req_headers, auto_login=True)
    x_auth = str(resolved_headers.get("x_auth", "") or "").strip()
    sid = str(resolved_headers.get("sid", "") or "").strip()
    username = str(resolved_headers.get("username", "") or "").strip()
    platform = str(resolved_headers.get("platform", "") or "").strip()
    location = str(resolved_headers.get("location", "") or "").strip()
    cache_control = str(resolved_headers.get("cache_control", "") or "").strip()
    content_type = str(resolved_headers.get("content_type", "") or "").strip()
    return {
        "start_date": start_date,
        "end_date": end_date,
        "acc_code": str(os.getenv("PMX_ACC_OPT_KEY", "MT0601") or "MT0601"),
        "report_type": "docDate",
        "col1": "LC",
        "col2": "GLD",
        "unit_code1": "",
        "unit_code2": "OZ",
        "option": "1",
        "host": str(os.getenv("PMX_API_HOST", "pmxapi.stonex.com") or "pmxapi.stonex.com"),
        "path": "/user/account_statementReport",
        "authorization": str(args_dict.get("authorization", "") or ""),
        "cookie": str(args_dict.get("cookie", "") or ""),
        "x_auth": x_auth,
        "sid": sid,
        "username": username,
        "platform": platform,
        "location": location,
        "cache_control": cache_control,
        "content_type": content_type,
        "origin": "https://pmxecute.stonex.com",
        "referer": "https://pmxecute.stonex.com/",
        "timeout": int(args_dict.get("timeout", "180") or "180"),
    }


def _pmx_recon_fetch_view(
    base_fetch_args: Dict[str, Any],
    col1: str,
    col2: str,
    unit1: str,
    unit2: str,
    args_dict: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], bool, str]:
    """Fetch one PMX statement view with auth retry.
    Returns (raw_rows, ok, error_msg).
    raw_rows are the dicts from extract_pmx_statement_report_rows().
    """
    fetch_args = dict(base_fetch_args)
    fetch_args.update({"col1": col1, "col2": col2, "unit_code1": unit1, "unit_code2": unit2})

    def _decode(r: Dict[str, Any]) -> Any:
        p = r.get("json")
        if p is None:
            body = r.get("body", "")
            if isinstance(body, str) and body.strip().startswith(("{", "[")):
                try:
                    p = json.loads(body)
                except Exception:
                    p = {}
            else:
                p = {}
        return p

    result = fetch_pmx_account_statement_report(**fetch_args)
    payload = _decode(result)
    result = _pmx_mark_failed_payload(result, payload)
    if not result.get("ok") and _pmx_result_is_auth_failure(result, payload):
        relogin = _pmx_login_session(args_dict)
        if relogin.get("ok"):
            for k in ("x_auth", "sid", "username", "platform", "location", "cache_control", "content_type"):
                fetch_args[k] = _pmx_non_empty(relogin.get(k), fetch_args.get(k, ""))
            result = fetch_pmx_account_statement_report(**fetch_args)
            payload = _decode(result)
            result = _pmx_mark_failed_payload(result, payload)
    if not result.get("ok"):
        return [], False, str(result.get("error", "") or "Statement fetch failed")
    # Recon must include all statement rows; filtering by FNC/JRV tokens can drop valid cash legs.
    rows = [r for r in extract_pmx_statement_report_rows(payload, require_doc_token=False) if isinstance(r, dict)]
    return rows, True, ""


def build_account_recon(args_dict: Dict[str, Any], req_headers: Any) -> Dict[str, Any]:
    """Build the account balance reconciliation payload using deterministic column extraction."""
    now = datetime.now()
    first_of_month_dt = now.replace(day=1)
    start_date = _normalize_pmx_date_param(
        str(args_dict.get("start_date", "") or "").strip(),
        default_dt=first_of_month_dt,
    )
    end_date = _normalize_pmx_date_param(
        str(args_dict.get("end_date", "") or "").strip(),
        default_dt=now,
    )
    start_dt = _pmx_parse_dd_mm_yyyy(start_date) or first_of_month_dt
    month = start_dt.strftime("%Y-%m")

    error_msg = ""
    actual_balances_ok = False

    # Compute ISO date strings for DB query
    end_dt = _pmx_parse_dd_mm_yyyy(end_date) or now
    db_start_iso = start_dt.strftime("%Y-%m-%d")
    db_end_iso = end_dt.strftime("%Y-%m-%d")

    tx_xau_list: List[float] = []
    tx_usd_list: List[float] = []
    tx_zar_list: List[float] = []
    canonical_rows: List[Dict[str, Any]] = []
    transactions_ok = False

    try:
        base_fetch_args = _pmx_recon_build_base_fetch_args(args_dict, req_headers, start_date, end_date)
        lc_rows, lc_ok, lc_err = _pmx_recon_fetch_view(
            base_fetch_args=base_fetch_args,
            col1="LC",
            col2="GLD",
            unit1="",
            unit2="OZ",
            args_dict=args_dict,
        )
        zar_rows, zar_ok, zar_err = _pmx_recon_fetch_view(
            base_fetch_args=base_fetch_args,
            col1="ZAR",
            col2="GLD",
            unit1="",
            unit2="OZ",
            args_dict=args_dict,
        )
        xau_rows, xau_ok, xau_err = _pmx_recon_fetch_view(
            base_fetch_args=base_fetch_args,
            col1="XAU",
            col2="GLD",
            unit1="OZ",
            unit2="OZ",
            args_dict=args_dict,
        )

        def _ci_get(raw_row: Dict[str, Any], target_key: str) -> Any:
            target = str(target_key or "").strip().upper()
            for key, value in raw_row.items():
                if str(key or "").strip().upper() == target:
                    return value
            return None

        def _row_type_from_doc(doc_number: str) -> str:
            doc_u = str(doc_number or "").upper()
            if doc_u.startswith("FNC/"):
                return "FNC"
            if doc_u.startswith("JRV/"):
                return "JRV"
            if doc_u.startswith("MER/"):
                return "MER"
            if doc_u.startswith("SWT/"):
                return "SWT"
            if "/" in doc_u:
                return doc_u.split("/")[0]
            return "OTHER"

        def _append_col_net_rows(
            raw_rows: List[Dict[str, Any]],
            ccy: str,
            source_view: str,
            credit_col: str,
            debit_col: str,
        ) -> None:
            iter_rows = raw_rows[:-1] if len(raw_rows) > 0 else raw_rows
            for raw in iter_rows:
                if not isinstance(raw, dict):
                    continue
                doc_number = str(_r_first_non_empty(raw, ["docno", "DocNo", "doc_number", "DocNumber", "document_no", "Doc #"]) or "").strip()
                trade_date_v = _r_extract_date(raw, ["docdate", "TradeDate", "trade_date", "DocDate", "date", "Trade Date", "Doc Date"])
                value_date_v = _r_extract_date(raw, ["valdate", "ValueDate", "value_date", "settlement_date", "Value Date", "Settlement Date"])
                symbol_v = _pmx_to_currency_pair(
                    _r_first_non_empty(raw, ["CurrencyPair", "currency_pair", "cmdty", "stk_type_name", "inst_desc", "Symbol"])
                ) or ""
                side_v = str(_r_first_non_empty(raw, ["side", "Side", "deal_type", "trd_opt"]) or "").upper()
                narration_v = str(_r_first_non_empty(raw, ["remarks", "remarks1", "comment", "notes", "description", "ContractDescription", "Narration"]) or "")
                if "BALANCE B/F" in narration_v.upper():
                    continue
                credit_val = _parse_loose_number(_ci_get(raw, credit_col))
                debit_val = _parse_loose_number(_ci_get(raw, debit_col))
                if credit_val is None and debit_val is None:
                    continue
                net_val = float(credit_val or 0.0) - float(debit_val or 0.0)
                if abs(net_val) <= 1e-12:
                    continue
                row_type = _row_type_from_doc(doc_number)
                if ccy == "USD":
                    tx_usd_list.append(net_val)
                    canonical_rows.append({
                        "doc_number": doc_number,
                        "trade_date": trade_date_v,
                        "value_date": value_date_v,
                        "row_type": row_type,
                        "symbol": symbol_v,
                        "side": side_v,
                        "narration": narration_v,
                        "movement_xau": None,
                        "movement_usd": net_val,
                        "movement_zar": None,
                        "source_view": source_view,
                        "included_xau": False,
                        "included_usd": True,
                        "included_zar": False,
                    })
                elif ccy == "ZAR":
                    tx_zar_list.append(net_val)
                    canonical_rows.append({
                        "doc_number": doc_number,
                        "trade_date": trade_date_v,
                        "value_date": value_date_v,
                        "row_type": row_type,
                        "symbol": symbol_v,
                        "side": side_v,
                        "narration": narration_v,
                        "movement_xau": None,
                        "movement_usd": None,
                        "movement_zar": net_val,
                        "source_view": source_view,
                        "included_xau": False,
                        "included_usd": False,
                        "included_zar": True,
                    })
                elif ccy == "XAU":
                    tx_xau_list.append(net_val)
                    canonical_rows.append({
                        "doc_number": doc_number,
                        "trade_date": trade_date_v,
                        "value_date": value_date_v,
                        "row_type": row_type,
                        "symbol": symbol_v,
                        "side": side_v,
                        "narration": narration_v,
                        "movement_xau": net_val,
                        "movement_usd": None,
                        "movement_zar": None,
                        "source_view": source_view,
                        "included_xau": True,
                        "included_usd": False,
                        "included_zar": False,
                    })

        if lc_ok and isinstance(lc_rows, list):
            _append_col_net_rows(lc_rows, "USD", "PMX statement LC col1", "COL1_CREDIT", "COL1_DEBIT")
        if zar_ok and isinstance(zar_rows, list):
            _append_col_net_rows(zar_rows, "ZAR", "PMX statement ZAR col1", "COL1_CREDIT", "COL1_DEBIT")
        if xau_ok and isinstance(xau_rows, list):
            _append_col_net_rows(xau_rows, "XAU", "PMX statement XAU col1", "COL1_CREDIT", "COL1_DEBIT")
        # XAU safety: some PMX views carry XAU movement in COL2 rather than COL1.
        if len(tx_xau_list) == 0:
            if xau_ok and isinstance(xau_rows, list):
                _append_col_net_rows(xau_rows, "XAU", "PMX statement XAU col2 fallback", "COL2_CREDIT", "COL2_DEBIT")
        if len(tx_xau_list) == 0:
            if zar_ok and isinstance(zar_rows, list):
                _append_col_net_rows(zar_rows, "XAU", "PMX statement ZAR col2 fallback", "COL2_CREDIT", "COL2_DEBIT")
        if len(tx_xau_list) == 0:
            if lc_ok and isinstance(lc_rows, list):
                _append_col_net_rows(lc_rows, "XAU", "PMX statement LC col2 fallback", "COL2_CREDIT", "COL2_DEBIT")

        transactions_ok = bool(lc_ok and zar_ok and xau_ok)
        if not transactions_ok:
            err_bits = [msg for msg in [lc_err, zar_err, xau_err] if str(msg or "").strip()]
            error_msg = "; ".join(err_bits) if err_bits else "Statement transaction fetch failed"
    except Exception as _exc:
        error_msg = f"Statement transaction fetch failed: {_exc}"
        transactions_ok = False

    tx_xau: Optional[float] = sum(tx_xau_list) if transactions_ok else None
    tx_usd: Optional[float] = sum(tx_usd_list) if transactions_ok else None
    tx_zar: Optional[float] = sum(tx_zar_list) if transactions_ok else None

    # Fetch actual live balances
    try:
        actual = _fetch_open_positions_account_balances(args_dict, req_headers)
        actual_balances_ok = bool(actual.get("ok"))
    except Exception as exc:
        actual = {}
        error_msg = (error_msg + f"; Balance fetch failed: {exc}").lstrip("; ")

    # Load opening balances from DB
    opening: Dict[str, Optional[float]] = {}
    try:
        conn = get_pmx_db_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT currency, opening_balance FROM account_opening_balances WHERE month = ?",
            (month,),
        )
        for db_row in cur.fetchall():
            opening[str(db_row[0]).upper()] = float(db_row[1])
        conn.close()
    except Exception:
        pass

    for _override_key, _ccy_key in (("opening_xau", "XAU"), ("opening_usd", "USD"), ("opening_zar", "ZAR")):
        _override_raw = str(args_dict.get(_override_key, "") or "").strip()
        if _override_raw:
            _override_val = _parse_loose_number(_override_raw)
            if _override_val is not None:
                opening[_ccy_key] = float(_override_val)

    currencies: Dict[str, Dict[str, Any]] = {}
    for ccy, tx_total in (("XAU", tx_xau), ("USD", tx_usd), ("ZAR", tx_zar)):
        opening_bal = opening.get(ccy)
        actual_key = ccy.lower()
        actual_v = _parse_loose_number(actual.get(actual_key)) if isinstance(actual, dict) else None
        expected = (float(opening_bal) + tx_total) if (opening_bal is not None and tx_total is not None) else None
        delta = (actual_v - expected) if (actual_v is not None and expected is not None) else None
        currencies[ccy] = {
            "opening_balance": opening_bal,
            "transaction_total": tx_total,
            "expected_balance": expected,
            "actual_balance": actual_v,
            "delta": delta,
        }

    # Flag likely setup issue: recon run with zero openings and no included transactions.
    try:
        opening_vals = [currencies.get(ccy, {}).get("opening_balance") for ccy in ("XAU", "USD", "ZAR")]
        tx_vals = [currencies.get(ccy, {}).get("transaction_total") for ccy in ("XAU", "USD", "ZAR")]
        actual_vals = [currencies.get(ccy, {}).get("actual_balance") for ccy in ("XAU", "USD", "ZAR")]
        openings_all_zero = all((v is not None) and (abs(float(v)) <= 1e-12) for v in opening_vals)
        tx_all_zero = all((v is not None) and (abs(float(v)) <= 1e-12) for v in tx_vals)
        actual_has_nonzero = any((v is not None) and (abs(float(v)) > 1e-12) for v in actual_vals)
        if openings_all_zero and tx_all_zero and actual_has_nonzero:
            setup_msg = (
                "Recon warning: opening balances are all zero and no eligible statement movements were included. "
                "Set monthly opening balances (XAU/USD/ZAR) and verify statement filters."
            )
            error_msg = (error_msg + f"; {setup_msg}").lstrip("; ")
    except Exception:
        pass

    return {
        "start_date": start_date,
        "end_date": end_date,
        "month": month,
        "currencies": currencies,
        "actual_balances_ok": actual_balances_ok,
        "transactions_ok": transactions_ok,
        "error": error_msg,
        "rows": canonical_rows,
        "diagnostics": {
            "row_count_total": len(canonical_rows),
            "row_count_included_xau": sum(1 for r in canonical_rows if r.get("included_xau")),
            "row_count_included_usd": sum(1 for r in canonical_rows if r.get("included_usd")),
            "row_count_included_zar": sum(1 for r in canonical_rows if r.get("included_zar")),
            "row_count_view_xau_raw": sum(1 for r in canonical_rows if r.get("included_xau")),
            "row_count_view_usd_raw": sum(1 for r in canonical_rows if r.get("included_usd")),
            "row_count_view_zar_raw": sum(1 for r in canonical_rows if r.get("included_zar")),
            "view_a_ok": transactions_ok,
            "view_b_ok": transactions_ok,
            "view_c_ok": transactions_ok,
            "delta_formula": "actual - expected",
            "source": "PMX statement",
        },
    }


def _daily_balance_email_get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(LEDGER_DB_PATH, timeout=30, check_same_thread=False)
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA busy_timeout=5000;")
    except Exception:
        pass
    return conn


def _daily_balance_email_ensure_log_table() -> None:
    conn = _daily_balance_email_get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS scheduled_job_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_name TEXT NOT NULL,
                run_date TEXT NOT NULL,
                status TEXT NOT NULL,
                message TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            )
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_scheduled_job_runs_lookup
            ON scheduled_job_runs (job_name, run_date, status, created_at)
            """
        )
        conn.commit()
    finally:
        conn.close()


def _daily_balance_email_log_run(job_name: str, run_date: str, status: str, message: str = "") -> None:
    conn = _daily_balance_email_get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO scheduled_job_runs (job_name, run_date, status, message, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                str(job_name or "").strip(),
                str(run_date or "").strip(),
                str(status or "").strip(),
                str(message or "").strip()[:2000],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()
    finally:
        conn.close()


def _daily_balance_email_has_success(job_name: str, run_date: str) -> bool:
    conn = _daily_balance_email_get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT 1
            FROM scheduled_job_runs
            WHERE job_name = ?
              AND run_date = ?
              AND status = 'success'
            LIMIT 1
            """,
            (str(job_name or "").strip(), str(run_date or "").strip()),
        )
        return cur.fetchone() is not None
    finally:
        conn.close()


def _daily_balance_email_iso_to_mm_dd_yyyy(value: str) -> str:
    text = str(value or "").strip()
    try:
        return datetime.strptime(text, "%Y-%m-%d").strftime("%m/%d/%Y")
    except Exception:
        return datetime.now().strftime("%m/%d/%Y")


def _daily_balance_email_fetch_pdf(run_date_iso: str, force_relogin: bool = False) -> Dict[str, Any]:
    account_code = _pmx_non_empty(os.getenv("PMX_ACC_OPT_KEY", ""), "MT0601")
    start_end = _daily_balance_email_iso_to_mm_dd_yyyy(run_date_iso)
    host = _pmx_non_empty(os.getenv("PMX_API_HOST", ""), "pmxapi.stonex.com")
    pdf_path = DAILY_BALANCE_EMAIL_PDF_PATH
    url = pdf_path if str(pdf_path).startswith("http") else f"https://{host}{pdf_path}"

    request_data = {"force_pmx_relogin": True} if force_relogin else {}
    resolved_headers = _pmx_resolve_headers(request_data, None, auto_login=True)
    x_auth = str(resolved_headers.get("x_auth", "") or "").strip()
    sid = str(resolved_headers.get("sid", "") or "").strip()
    username = str(resolved_headers.get("username", "") or "").strip()
    platform = str(resolved_headers.get("platform", "") or "").strip()
    location = str(resolved_headers.get("location", "") or "").strip()
    cache_control = str(resolved_headers.get("cache_control", "") or "").strip()
    content_type = str(resolved_headers.get("content_type", "") or "").strip()

    params = {
        "startDate": start_end,
        "endDate": start_end,
        "trd": DAILY_BALANCE_EMAIL_TRADE_NAME,
        "trd_key": account_code,
    }
    headers: Dict[str, str] = {
        "Accept": "application/pdf,application/octet-stream,*/*",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
        "Origin": "https://pmxecute.stonex.com",
        "Referer": "https://pmxecute.stonex.com/",
    }
    if x_auth:
        headers["x-auth"] = x_auth
    if sid:
        headers["sid"] = sid
    if username:
        headers["username"] = username
        headers["usercode"] = username
    if platform:
        headers["platform"] = platform
    if location:
        headers["location"] = location
    if cache_control:
        headers["cache-control"] = cache_control
    if content_type:
        headers["content-type"] = content_type

    try:
        resp = requests.get(url, headers=headers, params=params, timeout=DAILY_BALANCE_EMAIL_REQUEST_TIMEOUT_SECONDS)
        body_bytes = resp.content if isinstance(resp.content, (bytes, bytearray)) else b""
        content_type_resp = str(resp.headers.get("Content-Type", "") or "")
        content_disposition = str(resp.headers.get("Content-Disposition", "") or "")
        is_pdf_content_type = "application/pdf" in content_type_resp.lower()
        is_pdf_disposition = "pdf" in content_disposition.lower()
        is_pdf_signature = bytes(body_bytes).startswith(b"%PDF")
        ok = bool(resp.ok) and (is_pdf_content_type or is_pdf_disposition or is_pdf_signature)

        body_text = ""
        if not ok:
            try:
                body_text = resp.text if isinstance(resp.text, str) else ""
            except Exception:
                body_text = ""

        return {
            "ok": ok,
            "status": int(resp.status_code),
            "reason": str(resp.reason or ""),
            "url": str(resp.url),
            "content_type": content_type_resp,
            "content_disposition": content_disposition,
            "body_bytes": bytes(body_bytes),
            "body_text": body_text,
            "run_date_mmddyyyy": start_end,
            "account_code": account_code,
            "error": (
                "" if ok else (body_text[:300].strip() if body_text.strip() else f"{resp.reason} (HTTP {resp.status_code})")
            ),
        }
    except Exception as exc:
        return {
            "ok": False,
            "status": 0,
            "reason": "",
            "error": str(exc),
            "body_bytes": b"",
            "run_date_mmddyyyy": start_end,
            "account_code": account_code,
        }


def _daily_balance_email_send_with_pdf(run_date_iso: str, pdf_bytes: bytes, account_code: str) -> Dict[str, Any]:
    recipient = str(os.getenv("BALANCE_EMAIL_TO", "") or "").strip()
    smtp_host = str(os.getenv("SMTP_HOST", "") or "").strip()
    smtp_user = str(os.getenv("SMTP_USER", "") or "").strip()
    smtp_password = str(os.getenv("SMTP_PASSWORD", "") or "").strip()
    smtp_from = _pmx_non_empty(os.getenv("SMTP_FROM", ""), smtp_user)
    try:
        smtp_port = int(os.getenv("SMTP_PORT", "587") or 587)
    except Exception:
        smtp_port = 587
    use_ssl = str(os.getenv("SMTP_SSL", "") or "").strip().lower() in {"1", "true", "yes", "y", "on"}
    use_starttls = str(os.getenv("SMTP_STARTTLS", "true") or "true").strip().lower() in {"1", "true", "yes", "y", "on"}

    if not recipient:
        return {"ok": False, "error": "BALANCE_EMAIL_TO is not configured."}
    if not smtp_host:
        return {"ok": False, "error": "SMTP_HOST is not configured."}
    if not smtp_from:
        return {"ok": False, "error": "SMTP_FROM/SMTP_USER is not configured."}
    if smtp_user and not smtp_password:
        return {"ok": False, "error": "SMTP_PASSWORD is not configured."}
    if not isinstance(pdf_bytes, (bytes, bytearray)) or len(pdf_bytes) == 0:
        return {"ok": False, "error": "Missing PDF bytes for attachment."}

    run_date = _daily_balance_email_iso_to_mm_dd_yyyy(run_date_iso).replace("/", "-")
    subject = f"{DAILY_BALANCE_EMAIL_SUBJECT_PREFIX} | {account_code} | {run_date}"
    body = (
        f"PMX account balances report attached.\n"
        f"Account: {account_code}\n"
        f"Trade Name: {DAILY_BALANCE_EMAIL_TRADE_NAME}\n"
        f"Date: {run_date}\n"
        f"Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
    )
    file_name = f"PMX_Account_Balances_{account_code}_{run_date}.pdf"

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_from
    msg["To"] = recipient
    msg.set_content(body)
    msg.add_attachment(bytes(pdf_bytes), maintype="application", subtype="pdf", filename=file_name)

    try:
        if use_ssl:
            with smtplib.SMTP_SSL(host=smtp_host, port=smtp_port, context=ssl.create_default_context(), timeout=60) as client:
                if smtp_user:
                    client.login(smtp_user, smtp_password)
                client.send_message(msg)
        else:
            with smtplib.SMTP(host=smtp_host, port=smtp_port, timeout=60) as client:
                client.ehlo()
                if use_starttls:
                    client.starttls(context=ssl.create_default_context())
                    client.ehlo()
                if smtp_user:
                    client.login(smtp_user, smtp_password)
                client.send_message(msg)
        return {"ok": True, "message": f"Email sent to {recipient} ({file_name})"}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def _daily_balance_email_run_once_for_date(run_date_iso: str) -> Dict[str, Any]:
    # First attempt with cached session; retry once with forced relogin if needed.
    first = _daily_balance_email_fetch_pdf(run_date_iso, force_relogin=False)
    pdf_result = first
    if not first.get("ok"):
        status = int(first.get("status") or 0)
        body_text = str(first.get("body_text", "") or "")
        payload = None
        if body_text.strip().startswith(("{", "[")):
            try:
                payload = json.loads(body_text)
            except Exception:
                payload = None
        pmx_failure = _pmx_result_is_auth_failure(
            {"status": status, "reason": first.get("reason"), "error": first.get("error")},
            payload,
        )
        if pmx_failure or status in {401, 403, 500}:
            pdf_result = _daily_balance_email_fetch_pdf(run_date_iso, force_relogin=True)

    if not pdf_result.get("ok"):
        return {"ok": False, "error": f"PDF download failed: {pdf_result.get('error', 'unknown error')}"}

    send_result = _daily_balance_email_send_with_pdf(
        run_date_iso=run_date_iso,
        pdf_bytes=bytes(pdf_result.get("body_bytes", b"")),
        account_code=str(pdf_result.get("account_code", "MT0601") or "MT0601"),
    )
    if not send_result.get("ok"):
        return {"ok": False, "error": f"Email send failed: {send_result.get('error', 'unknown error')}"}
    return {"ok": True, "message": str(send_result.get("message", "Daily PMX balance email sent."))}


def _daily_balance_email_scheduler_loop() -> None:
    print(
        f"[SCHED] Daily PMX balances email scheduler active at "
        f"{DAILY_BALANCE_EMAIL_HOUR:02d}:{DAILY_BALANCE_EMAIL_MINUTE:02d}."
    )
    _daily_balance_email_ensure_log_table()

    while True:
        try:
            now = datetime.now()
            run_date_iso = now.strftime("%Y-%m-%d")
            due_now = (now.hour > DAILY_BALANCE_EMAIL_HOUR) or (
                now.hour == DAILY_BALANCE_EMAIL_HOUR and now.minute >= DAILY_BALANCE_EMAIL_MINUTE
            )
            if due_now and not _daily_balance_email_has_success(DAILY_BALANCE_EMAIL_JOB_NAME, run_date_iso):
                now_epoch = time.time()
                last_attempt = float(_daily_balance_email_last_attempt_epoch.get(run_date_iso, 0.0))
                if (now_epoch - last_attempt) >= float(DAILY_BALANCE_EMAIL_RETRY_SECONDS):
                    _daily_balance_email_last_attempt_epoch[run_date_iso] = now_epoch
                    result = _daily_balance_email_run_once_for_date(run_date_iso)
                    status = "success" if result.get("ok") else "failed"
                    message = str(result.get("message") or result.get("error") or "").strip()
                    _daily_balance_email_log_run(DAILY_BALANCE_EMAIL_JOB_NAME, run_date_iso, status, message)
                    print(f"[SCHED] {DAILY_BALANCE_EMAIL_JOB_NAME} {status} for {run_date_iso}: {message}")

            # Keep in-memory attempt cache small.
            for key in list(_daily_balance_email_last_attempt_epoch.keys()):
                if key != run_date_iso:
                    _daily_balance_email_last_attempt_epoch.pop(key, None)
        except Exception as exc:
            print(f"[SCHED] Daily PMX balances email scheduler error: {exc}")
            try:
                traceback.print_exc()
            except Exception:
                pass

        time.sleep(DAILY_BALANCE_EMAIL_CHECK_INTERVAL_SECONDS)


def _start_daily_balance_email_scheduler() -> None:
    global _daily_balance_email_scheduler_started
    if not DAILY_BALANCE_EMAIL_ENABLED:
        print("[SCHED] Daily PMX balances email scheduler disabled (PMX_DAILY_BALANCE_EMAIL_ENABLED=false).")
        return
    with _daily_balance_email_scheduler_lock:
        if _daily_balance_email_scheduler_started:
            return
        t = threading.Thread(target=_daily_balance_email_scheduler_loop, daemon=True)
        t.start()
        _daily_balance_email_scheduler_started = True


def _daily_trading_report_parse_recipients(raw: Any) -> List[str]:
    if raw is None:
        return []
    if isinstance(raw, (list, tuple, set)):
        values = [str(v or "").strip() for v in raw]
    else:
        values = [v.strip() for v in str(raw).replace(";", ",").split(",")]
    out: List[str] = []
    for item in values:
        if item and "@" in item:
            out.append(item)
    return out


def _daily_trading_report_html_escape(value: Any) -> str:
    text = str(value or "")
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def _daily_trading_report_money(value: Any, prefix: str = "R", decimals: int = 2) -> str:
    try:
        num = float(value)
        if not math.isfinite(num):
            return "--"
        return f"{prefix}{num:,.{decimals}f}"
    except Exception:
        return "--"


def _daily_trading_report_num(value: Any, decimals: int = 2) -> str:
    try:
        num = float(value)
        if not math.isfinite(num):
            return "--"
        return f"{num:,.{decimals}f}"
    except Exception:
        return "--"


def _daily_trading_report_pct(value: Any, decimals: int = 2) -> str:
    try:
        num = float(value)
        if not math.isfinite(num):
            return "--"
        return f"{num:.{decimals}f}%"
    except Exception:
        return "--"


def _build_dashboard_summary_payload(run_date_iso: Optional[str] = None) -> Dict[str, Any]:
    run_date = str(run_date_iso or datetime.now().strftime("%Y-%m-%d")).strip()
    dashboard_profit_min_date = "2026-03-01"

    profit = build_profit_monthly_report() or {}
    months = profit.get("months", []) if isinstance(profit, dict) else []
    summary = profit.get("summary", {}) if isinstance(profit, dict) else {}
    if not isinstance(months, list):
        months = []
    if not isinstance(summary, dict):
        summary = {}

    all_trades: List[Dict[str, Any]] = []
    for month in months:
        if not isinstance(month, dict):
            continue
        trades = month.get("trades", [])
        if isinstance(trades, list):
            for tr in trades:
                if isinstance(tr, dict):
                    all_trades.append(tr)

    day_profit_map: Dict[str, float] = {}
    day_trade_count_map: Dict[str, int] = {}
    normalized_profit_rows: List[Tuple[str, float, Optional[float]]] = []
    for tr in all_trades:
        d = str(tr.get("trade_date") or "").strip()[:10]
        if not d or d < dashboard_profit_min_date:
            continue
        try:
            pv = float(tr.get("total_profit_zar") or 0.0)
            if not math.isfinite(pv):
                continue
        except Exception:
            continue
        pct_val: Optional[float] = None
        try:
            p = tr.get("profit_pct")
            if p is not None and str(p).strip() != "":
                pct = float(str(p).replace("%", "").strip())
                if math.isfinite(pct):
                    pct_val = pct
        except Exception:
            pct_val = None
        normalized_profit_rows.append((d, pv, pct_val))
        day_profit_map[d] = float(day_profit_map.get(d, 0.0)) + pv
        day_trade_count_map[d] = int(day_trade_count_map.get(d, 0)) + 1

    # Strict Profit-tab style aggregation:
    # - Daily = sum for run_date only
    # - MTD = sum from month start up to run_date
    selected_day = run_date
    daily_profit = float(day_profit_map.get(run_date, 0.0))
    daily_trades = int(day_trade_count_map.get(run_date, 0))

    month_key = run_date[:7]
    mtd_profit = 0.0
    mtd_trades = 0
    for d, pv, _pct in normalized_profit_rows:
        if d.startswith(month_key) and d <= run_date:
            mtd_profit += float(pv)
            mtd_trades += 1

    hedging_rows = build_hedging_comparison(source="pmx") or []
    if not isinstance(hedging_rows, list):
        hedging_rows = []

    unhedged_tol_g = GRAMS_PER_TROY_OUNCE
    total_tm_g = 0.0
    under_hedged_total_g = 0.0
    under_hedged_count = 0
    hedged_rows = 0
    for row in hedging_rows:
        if not isinstance(row, dict):
            continue
        try:
            tm_g_signed = float(row.get("tm_weight_g") or 0.0)
        except Exception:
            tm_g_signed = 0.0
        total_tm_g += abs(tm_g_signed)

        try:
            need_g_signed = float(row.get("hedge_need_g") or 0.0)
        except Exception:
            need_g_signed = 0.0

        is_under_hedged = (
            (tm_g_signed > 0 and need_g_signed > unhedged_tol_g)
            or (tm_g_signed < 0 and need_g_signed < -unhedged_tol_g)
        )
        if is_under_hedged:
            under_hedged_count += 1
            under_hedged_total_g += abs(need_g_signed)
        else:
            hedged_rows += 1

    total_to_hedge_g = float(under_hedged_total_g)
    covered_g = max(0.0, float(total_tm_g) - total_to_hedge_g)

    if normalized_profit_rows:
        total_profit_all = float(sum(r[1] for r in normalized_profit_rows))
        total_trades_all = int(len(normalized_profit_rows))
        pct_values = [r[2] for r in normalized_profit_rows if r[2] is not None]
        overall_profit_pct = float(sum(pct_values) / len(pct_values)) if pct_values else float(summary.get("average_profit_margin_pct") or 0.0)
    else:
        total_profit_all = float(summary.get("total_profit_zar") or 0.0)
        total_trades_all = int(summary.get("trades") or 0)
        overall_profit_pct = float(summary.get("average_profit_margin_pct") or 0.0)

    month_plot_rows: List[Dict[str, Any]] = []
    for m in months[-6:]:
        if not isinstance(m, dict):
            continue
        month_plot_rows.append(
            {
                "label": str(m.get("month_label") or m.get("month_key") or ""),
                "value": float(m.get("total_profit_zar") or 0.0),
            }
        )
    sorted_days = sorted(day_profit_map.keys())
    daily_plot_rows = [{"label": d, "value": float(day_profit_map[d])} for d in sorted_days[-12:]]

    return {
        "run_date": run_date,
        "selected_day": selected_day,
        "daily_profit": daily_profit,
        "daily_trades": daily_trades,
        "month_to_date_profit": mtd_profit,
        "month_to_date_trades": mtd_trades,
        "total_profit_all": total_profit_all,
        "total_trades_all": total_trades_all,
        "overall_profit_pct": overall_profit_pct,
        "total_grams_to_hedge": total_to_hedge_g,
        "covered_g": covered_g,
        "hedged_rows": hedged_rows,
        "unhedged_rows": under_hedged_count,
        "month_plot_rows": month_plot_rows,
        "daily_plot_rows": daily_plot_rows,
    }


def _daily_trading_report_build_payload(run_date_iso: Optional[str] = None) -> Dict[str, Any]:
    dash = _build_dashboard_summary_payload(run_date_iso)
    run_date = str(dash.get("run_date") or (run_date_iso or datetime.now().strftime("%Y-%m-%d"))).strip()

    reval = build_open_positions_reval({}, None) or {}
    reval_summary = reval.get("summary", {}) if isinstance(reval, dict) else {}
    if not isinstance(reval_summary, dict):
        reval_summary = {}
    open_trades = int(reval_summary.get("open_trades") or 0)
    open_positions_pnl = float(reval_summary.get("total_pnl_zar") or 0.0)

    return {
        "date": run_date,
        "subtitle": f"{run_date} | Executive overview + desk detail",
        "source": "local:rule-based",
        "daily_profit": float(dash.get("daily_profit") or 0.0),
        "daily_trades": int(dash.get("daily_trades") or 0),
        "month_to_date_profit": float(dash.get("month_to_date_profit") or 0.0),
        "month_to_date_trades": int(dash.get("month_to_date_trades") or 0),
        "open_positions_pnl": open_positions_pnl,
        "open_positions_trades": open_trades,
        "total_grams_to_hedge": float(dash.get("total_grams_to_hedge") or 0.0),
        "hedged_rows": int(dash.get("hedged_rows") or 0),
        "unhedged_rows": int(dash.get("unhedged_rows") or 0),
        "covered_g": float(dash.get("covered_g") or 0.0),
        "overall_profit_pct": float(dash.get("overall_profit_pct") or 0.0),
        "total_profit_all": float(dash.get("total_profit_all") or 0.0),
        "total_trades_all": int(dash.get("total_trades_all") or 0),
        "month_plot_rows": list(dash.get("month_plot_rows") or []),
        "daily_plot_rows": list(dash.get("daily_plot_rows") or []),
    }


def _daily_trading_report_render_html(payload: Dict[str, Any]) -> str:
    def _bar_rows(rows: List[Dict[str, Any]], currency_prefix: str = "R", units: str = "") -> str:
        if not rows:
            return "<tr><td colspan='3' style='padding:8px 0;color:#738095;'>No data</td></tr>"
        max_abs = max(max(abs(float(r.get("value") or 0.0)) for r in rows), 1.0)
        out = []
        for r in rows:
            label = _daily_trading_report_html_escape(r.get("label", ""))
            value = float(r.get("value") or 0.0)
            width = max(2.0, min(100.0, (abs(value) / max_abs) * 100.0))
            color = "#1f8b4c" if value >= 0 else "#d83b2d"
            val_txt = f"{currency_prefix}{value:,.2f}{units}"
            out.append(
                "<tr>"
                f"<td style='padding:4px 0; width:86px; color:#34445d; font-size:12px;'>{label}</td>"
                "<td style='padding:4px 8px;'>"
                "<div style='height:10px;background:#dce3ee;border-radius:999px;overflow:hidden;'>"
                f"<div style='width:{width:.2f}%;height:100%;background:{color};'></div>"
                "</div>"
                "</td>"
                f"<td style='padding:4px 0; width:120px; text-align:right; font-weight:700; color:#1f2a3d; font-size:12px;'>{_daily_trading_report_html_escape(val_txt)}</td>"
                "</tr>"
            )
        return "".join(out)

    date_txt = _daily_trading_report_html_escape(payload.get("date", ""))
    subtitle = _daily_trading_report_html_escape(payload.get("subtitle", ""))
    source = _daily_trading_report_html_escape(payload.get("source", ""))
    daily_profit = float(payload.get("daily_profit") or 0.0)
    daily_is_loss = daily_profit < 0
    daily_card_bg = "#f7e6e8" if daily_is_loss else "#e9f6ec"
    daily_card_bd = "#d95a66" if daily_is_loss else "#9fd2ad"
    daily_card_tx = "#8b1220" if daily_is_loss else "#1f7a46"

    month_rows_html = _bar_rows(payload.get("month_plot_rows", []), "R")
    daily_rows_html = _bar_rows(payload.get("daily_plot_rows", []), "R")
    hedge_rows_html = _bar_rows(
        [
            {"label": "Covered g", "value": float(payload.get("covered_g") or 0.0)},
            {"label": "To Hedge g", "value": float(payload.get("total_grams_to_hedge") or 0.0)},
        ],
        "",
        "g",
    )

    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<style>
  @media only screen and (max-width: 640px) {{
    .kpi-title {{ font-size:16px !important; }}
    .kpi-value {{ font-size:38px !important; }}
    .kpi-sub {{ font-size:16px !important; }}
    .main-title {{ font-size:20px !important; }}
    .main-subtitle {{ font-size:14px !important; }}
    .main-meta {{ font-size:12px !important; }}
  }}
</style>
</head>
<body style="margin:0;padding:0;background:#f4f6fa;font-family:Segoe UI,Arial,sans-serif;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f4f6fa;"><tr><td align="center" style="padding:10px;">
  <table role="presentation" width="760" cellspacing="0" cellpadding="0" style="max-width:760px;width:100%;background:#ffffff;border:1px solid #d6dce6;">
    <tr><td style="padding:14px 14px 10px;background:#cdd5e2;border-bottom:1px solid #bcc6d5;">
      <div class="main-title" style="font-size:38px;line-height:1;color:#1b273b;font-weight:800;">Daily Trading Report</div>
      <div class="main-subtitle" style="margin-top:6px;color:#1f2e46;font-size:18px;">{subtitle}</div>
      <div class="main-meta" style="margin-top:4px;color:#3f4f67;font-size:14px;">Date: {date_txt} | Source: {source}</div>
    </td></tr>
    <tr><td style="padding:10px;">
      <table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>
        <td width="50%" style="padding:4px;"><div style="border:1px solid {daily_card_bd};background:{daily_card_bg};border-radius:10px;padding:10px;">
          <div class="kpi-title" style="font-size:20px;color:{daily_card_tx};font-weight:700;">Daily Profit{ " (LOSS)" if daily_is_loss else ""}</div>
          <div class="kpi-value" style="font-size:50px;line-height:1.05;color:{daily_card_tx};font-weight:800;">{_daily_trading_report_money(payload.get("daily_profit"), "R", 2)}</div>
          <div class="kpi-sub" style="font-size:22px;color:{daily_card_tx};">Trades: {int(payload.get("daily_trades") or 0)} | Profit %: {_daily_trading_report_pct(payload.get("overall_profit_pct"), 2)}</div>
        </div></td>
        <td width="50%" style="padding:4px;"><div style="border:1px solid #9fd2ad;background:#e9f6ec;border-radius:10px;padding:10px;">
          <div class="kpi-title" style="font-size:20px;color:#1f7a46;font-weight:700;">Month-to-date Profit</div>
          <div class="kpi-value" style="font-size:50px;line-height:1.05;color:#1f7a46;font-weight:800;">{_daily_trading_report_money(payload.get("month_to_date_profit"), "R", 2)}</div>
          <div class="kpi-sub" style="font-size:22px;color:#1f7a46;">Trades: {int(payload.get("month_to_date_trades") or 0)} | Profit %: {_daily_trading_report_pct(payload.get("overall_profit_pct"), 2)}</div>
        </div></td>
      </tr><tr>
        <td width="50%" style="padding:4px;"><div style="border:1px solid #9db2e5;background:#edf2fc;border-radius:10px;padding:10px;">
          <div class="kpi-title" style="font-size:20px;color:#2753bc;font-weight:700;">Open Positions Reval</div>
          <div class="kpi-value" style="font-size:50px;line-height:1.05;color:#2753bc;font-weight:800;">{_daily_trading_report_money(payload.get("open_positions_pnl"), "R", 2)}</div>
          <div class="kpi-sub" style="font-size:22px;color:#2753bc;">Open trades: {int(payload.get("open_positions_trades") or 0)} | Overall Profit %: {_daily_trading_report_pct(payload.get("overall_profit_pct"), 2)}</div>
        </div></td>
        <td width="50%" style="padding:4px;"><div style="border:1px solid #e7bea6;background:#fcf3eb;border-radius:10px;padding:10px;">
          <div class="kpi-title" style="font-size:20px;color:#a74917;font-weight:700;">Total Grams To Hedge</div>
          <div class="kpi-value" style="font-size:50px;line-height:1.05;color:#c01e1e;font-weight:800;">{_daily_trading_report_num(payload.get("total_grams_to_hedge"), 2)}g</div>
          <div class="kpi-sub" style="font-size:22px;color:#a74917;">Hedged rows: {int(payload.get("hedged_rows") or 0)} | Unhedged rows: {int(payload.get("unhedged_rows") or 0)}</div>
        </div></td>
      </tr></table>

      <div style="margin-top:8px;border-left:4px solid #2d5ad0;background:#eaf0ff;padding:8px 10px;font-size:15px;font-weight:700;color:#2c4ea3;">Executive Summary</div>
      <ul style="margin:8px 0 10px 20px;padding:0;color:#1f2b3c;font-size:14px;line-height:1.6;">
        <li>Daily closed P&amp;L was {_daily_trading_report_money(payload.get("daily_profit"), "R", 2)} across {int(payload.get("daily_trades") or 0)} trades, with month-to-date at {_daily_trading_report_money(payload.get("month_to_date_profit"), "R", 2)}.</li>
        <li>Total cumulative profit stands at {_daily_trading_report_money(payload.get("total_profit_all"), "R", 2)} across {int(payload.get("total_trades_all") or 0)} trades.</li>
        <li>Hedge coverage tracks {int(payload.get("hedged_rows") or 0)} hedged rows with {_daily_trading_report_num(payload.get("total_grams_to_hedge"), 2)}g remaining to hedge.</li>
      </ul>

      <div style="margin-top:8px;border-left:4px solid #5f8f28;background:#eef6df;padding:8px 10px;font-size:15px;font-weight:700;color:#4a6b1f;">Monthly Profit Plot (ZAR)</div>
      <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin-top:6px;">{month_rows_html}</table>

      <div style="margin-top:10px;border-left:4px solid #2f68b2;background:#e8f2ff;padding:8px 10px;font-size:15px;font-weight:700;color:#21508d;">Daily Profit Plot (ZAR)</div>
      <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin-top:6px;">{daily_rows_html}</table>

      <div style="margin-top:10px;border-left:4px solid #bd7a2f;background:#fbf3e8;padding:8px 10px;font-size:15px;font-weight:700;color:#9a5d1d;">Hedging Plot (grams)</div>
      <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin-top:6px;">{hedge_rows_html}</table>
      <div style="color:#364860;font-size:12px;margin-top:6px;">Rows: Hedged {int(payload.get("hedged_rows") or 0)} | Unhedged {int(payload.get("unhedged_rows") or 0)} | To hedge {_daily_trading_report_num(payload.get("total_grams_to_hedge"), 2)}g</div>

      <div style="margin-top:10px;border-left:4px solid #cc4f86;background:#fdeff6;padding:8px 10px;font-size:15px;font-weight:700;color:#a42f64;">Open Positions Revaluation Detail</div>
      <div style="padding:8px 2px;color:#1f2b3c;font-size:14px;">Open Trades: {int(payload.get("open_positions_trades") or 0)} | Total P&amp;L: <strong>{_daily_trading_report_money(payload.get("open_positions_pnl"), "R", 2)}</strong></div>
      <div style="padding-top:8px;border-top:1px solid #d4dce8;color:#6d7d95;font-size:11px;">Auto-generated for mobile and desktop viewing.</div>
    </td></tr>
  </table>
</td></tr></table></body></html>"""
    return html


def _daily_trading_report_send_html(
    run_date_iso: str,
    html_body: str,
    recipients_override: Optional[List[str]] = None,
) -> Dict[str, Any]:
    recipients = recipients_override or _daily_trading_report_parse_recipients(
        os.getenv("DAILY_TRADING_REPORT_EMAIL_TO", "")
    )
    if not recipients:
        recipients = _daily_trading_report_parse_recipients(os.getenv("DASHBOARD_REPORT_EMAIL_TO", ""))
    if not recipients:
        return {"ok": False, "error": "DAILY_TRADING_REPORT_EMAIL_TO is not configured."}

    smtp_host = str(os.getenv("SMTP_HOST", "") or "").strip()
    smtp_user = str(os.getenv("SMTP_USER", "") or "").strip()
    smtp_password = str(os.getenv("SMTP_PASSWORD", "") or "").strip()
    smtp_from = _pmx_non_empty(os.getenv("SMTP_FROM", ""), smtp_user)
    try:
        smtp_port = int(os.getenv("SMTP_PORT", "587") or 587)
    except Exception:
        smtp_port = 587
    use_ssl = str(os.getenv("SMTP_SSL", "") or "").strip().lower() in {"1", "true", "yes", "y", "on"}
    use_starttls = str(os.getenv("SMTP_STARTTLS", "true") or "true").strip().lower() in {"1", "true", "yes", "y", "on"}

    if not smtp_host:
        return {"ok": False, "error": "SMTP_HOST is not configured."}
    if not smtp_from:
        return {"ok": False, "error": "SMTP_FROM/SMTP_USER is not configured."}
    if not smtp_password:
        return {"ok": False, "error": "SMTP_PASSWORD is not configured."}

    run_date = str(run_date_iso or datetime.now().strftime("%Y-%m-%d")).strip()
    subject = f"{DAILY_TRADING_REPORT_EMAIL_SUBJECT_PREFIX} | {run_date}"
    plain_body = (
        f"Daily Trading Report for {run_date}\n"
        f"Generated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
    )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_from
    msg["To"] = ", ".join(recipients)
    msg.set_content(plain_body)
    msg.add_alternative(html_body, subtype="html")

    try:
        if use_ssl:
            with smtplib.SMTP_SSL(host=smtp_host, port=smtp_port, context=ssl.create_default_context(), timeout=60) as client:
                if smtp_user:
                    client.login(smtp_user, smtp_password)
                client.send_message(msg)
        else:
            with smtplib.SMTP(host=smtp_host, port=smtp_port, timeout=60) as client:
                client.ehlo()
                if use_starttls:
                    client.starttls(context=ssl.create_default_context())
                    client.ehlo()
                if smtp_user:
                    client.login(smtp_user, smtp_password)
                client.send_message(msg)
        return {"ok": True, "message": f"Daily trading report email sent to {', '.join(recipients)}"}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def _daily_trading_report_run_once_for_date(
    run_date_iso: str,
    recipients_override: Optional[List[str]] = None,
) -> Dict[str, Any]:
    try:
        # Use current dashboard-ready data/cache; avoid forced TradeMC sync here.
        try:
            _clear_heavy_route_cache(["profit_monthly", "hedging:", "pmx_open_positions_reval"])
        except Exception:
            pass
        payload = _daily_trading_report_build_payload(run_date_iso)
        html_body = _daily_trading_report_render_html(payload)
    except Exception as exc:
        return {"ok": False, "error": f"Failed to build report payload: {exc}"}
    return _daily_trading_report_send_html(run_date_iso, html_body, recipients_override=recipients_override)


def _daily_trading_report_scheduler_loop() -> None:
    print(
        f"[SCHED] Daily Trading Report email scheduler active at "
        f"{DAILY_TRADING_REPORT_EMAIL_HOUR:02d}:{DAILY_TRADING_REPORT_EMAIL_MINUTE:02d}."
    )
    _daily_balance_email_ensure_log_table()
    while True:
        try:
            now = datetime.now()
            run_date_iso = now.strftime("%Y-%m-%d")
            due_now = (
                now.hour == DAILY_TRADING_REPORT_EMAIL_HOUR
                and now.minute >= DAILY_TRADING_REPORT_EMAIL_MINUTE
            )
            if due_now and not _daily_balance_email_has_success(DAILY_TRADING_REPORT_EMAIL_JOB_NAME, run_date_iso):
                now_epoch = time.time()
                last_attempt = float(_daily_trading_report_email_last_attempt_epoch.get(run_date_iso, 0.0))
                if (now_epoch - last_attempt) >= float(DAILY_TRADING_REPORT_EMAIL_RETRY_SECONDS):
                    _daily_trading_report_email_last_attempt_epoch[run_date_iso] = now_epoch
                    result = _daily_trading_report_run_once_for_date(run_date_iso)
                    status = "success" if result.get("ok") else "failed"
                    message = str(result.get("message") or result.get("error") or "").strip()
                    _daily_balance_email_log_run(DAILY_TRADING_REPORT_EMAIL_JOB_NAME, run_date_iso, status, message)
                    print(f"[SCHED] {DAILY_TRADING_REPORT_EMAIL_JOB_NAME} {status} for {run_date_iso}: {message}")

            for key in list(_daily_trading_report_email_last_attempt_epoch.keys()):
                if key != run_date_iso:
                    _daily_trading_report_email_last_attempt_epoch.pop(key, None)
        except Exception as exc:
            print(f"[SCHED] Daily Trading Report email scheduler error: {exc}")
            try:
                traceback.print_exc()
            except Exception:
                pass
        time.sleep(DAILY_TRADING_REPORT_EMAIL_CHECK_INTERVAL_SECONDS)


def _start_daily_trading_report_email_scheduler() -> None:
    global _daily_trading_report_email_scheduler_started
    if not DAILY_TRADING_REPORT_EMAIL_ENABLED:
        print("[SCHED] Daily Trading Report email scheduler disabled (DAILY_TRADING_REPORT_EMAIL_ENABLED=false).")
        return
    with _daily_trading_report_email_scheduler_lock:
        if _daily_trading_report_email_scheduler_started:
            return
        t = threading.Thread(target=_daily_trading_report_scheduler_loop, daemon=True)
        t.start()
        _daily_trading_report_email_scheduler_started = True


def _empty_open_positions_reval_payload(market: Any = None) -> Dict[str, Any]:
    market = market if isinstance(market, dict) else {}
    xau_val = _safe_float(market.get("xau_usd"), default=float("nan"))
    fx_val = _safe_float(market.get("usd_zar"), default=float("nan"))
    return {
        "rows": [],
        "summary": {
            "open_trades": 0,
            "total_fx_qty_usd": 0.0,
            "total_gold_qty_oz": 0.0,
            "total_gold_qty_g": 0.0,
            "total_fx_pnl_zar": 0.0,
            "total_gold_pnl_usd": 0.0,
            "total_pnl_zar": 0.0,
        },
        "market": {
            "xau_usd": xau_val if math.isfinite(xau_val) else None,
            "usd_zar": fx_val if math.isfinite(fx_val) else None,
            "timestamp": market.get("timestamp"),
            "fetched_at": market.get("fetched_at"),
            "xau_usd_source": market.get("xau_usd_source"),
            "usd_zar_source": market.get("usd_zar_source"),
        },
    }


def build_open_positions_reval(data: Optional[Dict[str, Any]] = None, req_headers: Any = None) -> Dict[str, Any]:
    """
    Revalue open PMX positions at current market rates.

    Workbook logic (Open Positions Reval.xlsx):
      Gold PnL (USD) = (CurrentGold - GoldWA) * NetGoldOz
      FX PnL (ZAR)   = (CurrentFx - FxWA) * NetUSD
      Total PnL (ZAR)= FX PnL + Gold PnL(USD) * CurrentFx
    """
    data = data or {}
    market = _get_cached_trademc_live_prices(force_refresh=False)
    if not isinstance(market, dict):
        market = {}

    market_xau = _safe_float(market.get("xau_usd"), default=float("nan"))
    market_fx = _safe_float(market.get("usd_zar"), default=float("nan"))
    has_xau = math.isfinite(market_xau)
    has_fx = math.isfinite(market_fx)
    metal_tol_g = 32.0
    usd_tol = 1.0

    df = load_all_pmx_trades()
    if df is None or df.empty:
        return _empty_open_positions_reval_payload(market)

    trade_col = "OrderID" if "OrderID" in df.columns else ("Trade #" if "Trade #" in df.columns else "")
    if not trade_col:
        return _empty_open_positions_reval_payload(market)

    work = df.copy()
    work["trade_num"] = work[trade_col].apply(normalize_trade_number)
    # Open Positions Reval is strictly based on unallocated trades.
    work = work[work["trade_num"] == ""]
    if work.empty:
        return _empty_open_positions_reval_payload(market)

    work["Symbol"] = work.get("Symbol", "").astype(str).str.upper().str.replace("-", "", regex=False).str.strip()
    split_pairs = work["Symbol"].apply(split_symbol)
    work["base"] = split_pairs.apply(lambda t: t[0] if isinstance(t, tuple) and len(t) > 0 else "")
    work["quote"] = split_pairs.apply(lambda t: t[1] if isinstance(t, tuple) and len(t) > 1 else "")
    work["Side"] = work.get("Side", "").astype(str).str.upper().str.strip()
    work["qty_abs"] = pd.to_numeric(work.get("Quantity"), errors="coerce").fillna(0.0).abs()
    work["price"] = pd.to_numeric(work.get("Price"), errors="coerce").fillna(0.0)
    side_sign = work["Side"].map({"BUY": 1.0, "SELL": -1.0}).fillna(0.0)
    work["signed_qty"] = work["qty_abs"] * side_sign

    fx_rows = work[(work["base"] == "USD") & (work["quote"] == "ZAR")].copy()
    gold_rows = work[(work["base"] == "XAU") & (work["quote"] == "USD")].copy()

    rows_out: List[Dict[str, Any]] = []

    if not fx_rows.empty:
        fx_rows["fx_zar_flow"] = fx_rows["signed_qty"] * fx_rows["price"] * -1.0
        fx_qty_usd = float(fx_rows["signed_qty"].sum())
        if abs(fx_qty_usd) > (usd_tol + 1e-9):
            fx_zar_flow = float(fx_rows["fx_zar_flow"].sum())
            fx_wa_rate = abs(fx_zar_flow / fx_qty_usd) if abs(fx_qty_usd) > 1e-12 else 0.0
            rows_out.append(
                {
                    "trade_num": "USD/ZAR",
                    "pair": "USD/ZAR",
                    "pair_symbol": "USD/ZAR",
                    "fx_qty_usd": fx_qty_usd,
                    "fx_zar_flow": fx_zar_flow,
                    "fx_wa_rate": fx_wa_rate,
                    "gold_qty_oz": 0.0,
                    "gold_usd_flow": 0.0,
                    "gold_wa_price": 0.0,
                }
            )

    if not gold_rows.empty:
        gold_rows["gold_usd_flow"] = gold_rows["signed_qty"] * gold_rows["price"] * -1.0
        gold_qty_oz = float(gold_rows["signed_qty"].sum())
        if abs(gold_qty_oz * GRAMS_PER_TROY_OUNCE) > (metal_tol_g + 1e-9):
            gold_usd_flow = float(gold_rows["gold_usd_flow"].sum())
            gold_wa_price = abs(gold_usd_flow / gold_qty_oz) if abs(gold_qty_oz) > 1e-12 else 0.0
            rows_out.append(
                {
                    "trade_num": "XAU/USD",
                    "pair": "XAU/USD",
                    "pair_symbol": "XAU/USD",
                    "fx_qty_usd": 0.0,
                    "fx_zar_flow": 0.0,
                    "fx_wa_rate": 0.0,
                    "gold_qty_oz": gold_qty_oz,
                    "gold_usd_flow": gold_usd_flow,
                    "gold_wa_price": gold_wa_price,
                }
            )

    if not rows_out:
        return _empty_open_positions_reval_payload(market)

    result = pd.DataFrame(rows_out)

    result["gold_qty_g"] = result["gold_qty_oz"] * GRAMS_PER_TROY_OUNCE
    result["fx_side"] = result["fx_qty_usd"].apply(lambda x: "BUY" if x > 0 else ("SELL" if x < 0 else ""))
    result["gold_side"] = result["gold_qty_oz"].apply(lambda x: "BUY" if x > 0 else ("SELL" if x < 0 else ""))
    result["market_xau_usd"] = market_xau if has_xau else float("nan")
    result["market_usd_zar"] = market_fx if has_fx else float("nan")

    if has_xau:
        result["gold_pnl_usd"] = (market_xau - result["gold_wa_price"]) * result["gold_qty_oz"]
    else:
        result["gold_pnl_usd"] = float("nan")

    if has_fx:
        result["fx_pnl_zar"] = (market_fx - result["fx_wa_rate"]) * result["fx_qty_usd"]
    else:
        result["fx_pnl_zar"] = float("nan")

    if has_xau and has_fx:
        result["gold_pnl_zar"] = result["gold_pnl_usd"] * market_fx
        result["total_pnl_zar"] = result["fx_pnl_zar"] + result["gold_pnl_zar"]
    else:
        result["gold_pnl_zar"] = float("nan")
        result["total_pnl_zar"] = float("nan")

    result = result.sort_values("trade_num")

    for col, decimals in [
        ("fx_qty_usd", 2),
        ("fx_wa_rate", 5),
        ("gold_qty_oz", 4),
        ("gold_qty_g", 2),
        ("gold_wa_price", 4),
        ("gold_pnl_usd", 2),
        ("gold_pnl_zar", 2),
        ("fx_pnl_zar", 2),
        ("total_pnl_zar", 2),
        ("market_xau_usd", 4),
        ("market_usd_zar", 5),
    ]:
        if col in result.columns:
            result[col] = pd.to_numeric(result[col], errors="coerce").round(decimals)

    summary = {
        "open_trades": int(len(result)),
        "total_fx_qty_usd": round(float(result["fx_qty_usd"].sum()), 2),
        "total_gold_qty_oz": round(float(result["gold_qty_oz"].sum()), 4),
        "total_gold_qty_g": round(float(result["gold_qty_g"].sum()), 2),
        "total_fx_pnl_zar": round(float(result["fx_pnl_zar"].sum()), 2),
        "total_gold_pnl_usd": round(float(result["gold_pnl_usd"].sum()), 2),
        "total_pnl_zar": round(float(result["total_pnl_zar"].sum()), 2),
    }

    market_out = {
        "xau_usd": round(market_xau, 4) if has_xau else None,
        "usd_zar": round(market_fx, 5) if has_fx else None,
        "timestamp": market.get("timestamp"),
        "fetched_at": market.get("fetched_at"),
        "xau_usd_source": market.get("xau_usd_source"),
        "usd_zar_source": market.get("usd_zar_source"),
    }

    return {
        "rows": result.to_dict(orient="records"),
        "summary": summary,
        "market": market_out,
    }

def build_forward_exposure(data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Build forward-only PMX exposure payload for the Forward Exposure tab."""
    data = data or {}

    def _safe_series(df: pd.DataFrame, col: str, default: Any) -> pd.Series:
        if col in df.columns:
            return df[col]
        return pd.Series([default] * len(df), index=df.index)

    def _normalize_symbol(value: Any) -> str:
        return str(value or "").upper().replace("/", "").replace("-", "").strip()

    def _business_days_between(start_day: pd.Timestamp, end_day: pd.Timestamp) -> int:
        """Business-day difference excluding start day, including end day if business day."""
        if pd.isna(start_day) or pd.isna(end_day):
            return 0
        start = pd.Timestamp(start_day).normalize()
        end = pd.Timestamp(end_day).normalize()
        if end >= start:
            return max(0, len(pd.bdate_range(start=start, end=end)) - 1)
        return -max(0, len(pd.bdate_range(start=end, end=start)) - 1)

    def _empty_payload() -> Dict[str, Any]:
        return {
            "rows": [],
            "calendar": [],
            "summary": {
                "rows": 0,
                "trade_numbers": 0,
                "usd_net": 0.0,
                "gold_net_oz": 0.0,
                "zar_flow": 0.0,
            },
            "tenors": [],
        }

    source = load_all_pmx_trades()
    if source.empty:
        return _empty_payload()

    work = pd.DataFrame(
        {
            "id": pd.to_numeric(_safe_series(source, "id", 0), errors="coerce").fillna(0).astype(int),
            "trade_num": _safe_series(source, "OrderID", "").apply(normalize_trade_number),
            "doc_number": _safe_series(source, "Doc #", "").fillna("").astype(str).str.strip(),
            "trade_date": pd.to_datetime(_safe_series(source, "Trade Date", ""), errors="coerce"),
            "value_date": pd.to_datetime(_safe_series(source, "Value Date", ""), errors="coerce"),
            "symbol": _safe_series(source, "Symbol", "").apply(_normalize_symbol),
            "side": _safe_series(source, "Side", "").fillna("").astype(str).str.upper().str.strip(),
            "quantity_raw": pd.to_numeric(_safe_series(source, "Quantity", 0.0), errors="coerce").fillna(0.0),
            "price": pd.to_numeric(_safe_series(source, "Price", 0.0), errors="coerce").fillna(0.0),
        }
    )

    work = work[work["symbol"].isin(["XAUUSD", "USDZAR"])].copy()
    work = work[work["side"].isin(["BUY", "SELL"])].copy()
    work = work[work["trade_date"].notna() & work["value_date"].notna()].copy()
    if work.empty:
        return _empty_payload()

    work["trade_date_day"] = work["trade_date"].dt.normalize()
    work["value_date_day"] = work["value_date"].dt.normalize()
    work["spot_date_day"] = work["trade_date_day"] + pd.offsets.BDay(2)
    work["days_from_spot"] = work.apply(
        lambda r: _business_days_between(r["spot_date_day"], r["value_date_day"]),
        axis=1,
    )
    # Forward-only: keep rows strictly after spot date.
    work = work[work["days_from_spot"] > 0].copy()
    today_day = pd.Timestamp(datetime.now().date())
    work = work[work["value_date_day"] > today_day].copy()
    if work.empty:
        return _empty_payload()

    symbol_filter = _normalize_symbol(data.get("symbol", ""))
    if symbol_filter and symbol_filter != "ALL":
        work = work[work["symbol"] == symbol_filter].copy()

    start_dt = pd.to_datetime(str(data.get("start_date", "")).strip(), errors="coerce")
    if pd.notna(start_dt):
        work = work[work["value_date_day"] >= start_dt.normalize()].copy()
    end_dt = pd.to_datetime(str(data.get("end_date", "")).strip(), errors="coerce")
    if pd.notna(end_dt):
        work = work[work["value_date_day"] <= end_dt.normalize()].copy()

    if work.empty:
        return _empty_payload()

    qty_abs = work["quantity_raw"].abs()
    is_gold = work["symbol"] == "XAUUSD"
    is_fx = work["symbol"] == "USDZAR"
    is_buy = work["side"] == "BUY"
    is_sell = work["side"] == "SELL"

    work["quantity"] = qty_abs
    work["usd_net"] = 0.0
    work["gold_net_oz"] = 0.0
    work["zar_flow"] = 0.0

    work.loc[is_gold & is_buy, "usd_net"] = -(qty_abs[is_gold & is_buy] * work.loc[is_gold & is_buy, "price"])
    work.loc[is_gold & is_sell, "usd_net"] = qty_abs[is_gold & is_sell] * work.loc[is_gold & is_sell, "price"]
    work.loc[is_fx & is_buy, "usd_net"] = qty_abs[is_fx & is_buy]
    work.loc[is_fx & is_sell, "usd_net"] = -qty_abs[is_fx & is_sell]

    work.loc[is_gold & is_buy, "gold_net_oz"] = qty_abs[is_gold & is_buy]
    work.loc[is_gold & is_sell, "gold_net_oz"] = -qty_abs[is_gold & is_sell]

    work.loc[is_fx & is_buy, "zar_flow"] = -(qty_abs[is_fx & is_buy] * work.loc[is_fx & is_buy, "price"])
    work.loc[is_fx & is_sell, "zar_flow"] = qty_abs[is_fx & is_sell] * work.loc[is_fx & is_sell, "price"]

    work["trade_date"] = work["trade_date_day"].dt.strftime("%Y-%m-%d")
    work["value_date"] = work["value_date_day"].dt.strftime("%Y-%m-%d")
    work["days_from_spot"] = pd.to_numeric(work["days_from_spot"], errors="coerce").fillna(0).astype(int)
    work["trade_key"] = work["trade_num"]
    missing_trade_key = work["trade_key"].astype(str).str.strip() == ""
    work.loc[missing_trade_key, "trade_key"] = work.loc[missing_trade_key, "doc_number"]
    missing_trade_key = work["trade_key"].astype(str).str.strip() == ""
    work.loc[missing_trade_key, "trade_key"] = work.loc[missing_trade_key, "id"].apply(lambda v: f"ID:{int(v)}")

    detail = work.sort_values(
        by=["value_date_day", "trade_date_day", "id"],
        ascending=[True, False, False],
    )

    detail_cols = [
        "id",
        "trade_num",
        "trade_date",
        "value_date",
        "symbol",
        "side",
        "quantity",
        "price",
        "usd_net",
        "gold_net_oz",
        "zar_flow",
        "doc_number",
        "days_from_spot",
        "trade_key",
    ]
    detail = detail[detail_cols]
    detail_out = detail.drop(columns=["trade_key"]).to_dict(orient="records")

    calendar = (
        detail.groupby(["value_date", "days_from_spot"], dropna=False)
        .agg(
            trade_count=("id", "count"),
            trade_numbers=("trade_key", lambda s: int(len(set(str(v).strip() for v in s if str(v).strip())))),
            usd_net=("usd_net", "sum"),
            gold_net_oz=("gold_net_oz", "sum"),
            zar_flow=("zar_flow", "sum"),
        )
        .reset_index()
        .sort_values(by=["value_date"], ascending=[True])
    )
    calendar_out = calendar.to_dict(orient="records")

    summary = {
        "rows": int(len(detail)),
        "trade_numbers": int(len(set(str(v).strip() for v in detail["trade_key"].tolist() if str(v).strip()))),
        "usd_net": float(detail["usd_net"].sum()),
        "gold_net_oz": float(detail["gold_net_oz"].sum()),
        "zar_flow": float(detail["zar_flow"].sum()),
    }

    return {
        "rows": detail_out,
        "calendar": calendar_out,
        "summary": summary,
        "tenors": [],
    }


def build_hedging_comparison(source: str = "pmx"):
    """Compare TradeMC metal flow vs PMX metal/USD hedge flows."""
    source_key = str(source or "pmx").strip().lower()
    use_pmx = source_key == "pmx"
    fiscal_cutoff = _pmx_filter_date(FISCAL_TRADES_START_DATE) if ENABLE_FISCAL_DATE_FILTER else ""

    def _with_zero_hedge(tm_grouped_df: pd.DataFrame):
        compare_df = tm_grouped_df.copy()
        compare_df["stonex_buy_oz"] = 0.0
        compare_df["stonex_sell_oz"] = 0.0
        compare_df["stonex_net_oz"] = 0.0
        compare_df["stonex_hedge_g"] = 0.0
        compare_df["pmx_net_oz"] = 0.0
        compare_df["pmx_hedge_g"] = 0.0
        compare_df["pmx_net_usd"] = 0.0
        compare_df["hedge_need_g"] = compare_df["tm_weight_g"] - compare_df["stonex_hedge_g"]
        compare_df["metal_need_oz"] = compare_df["tm_weight_oz"] - compare_df["stonex_net_oz"]
        compare_df["usd_to_cut"] = compare_df["pmx_net_usd"].abs()
        compare_df["usd_need"] = compare_df["usd_to_cut"]
        compare_df["metal_hedged"] = compare_df["hedge_need_g"].abs() <= 31.1035  # 1 oz tolerance
        compare_df["usd_hedged"] = compare_df["usd_to_cut"] <= 1.0
        compare_df["hedged"] = compare_df["metal_hedged"] & compare_df["usd_hedged"]
        return compare_df.to_dict(orient="records")

    # Fast path: load only columns required for TradeMC and PMX hedging.
    tm_conn = None
    ledger_conn = None
    try:
        tm_conn = get_db_connection()
        tm_date_sql = ""
        tm_params: List[Any] = []
        if fiscal_cutoff:
            tm_date_sql = "AND COALESCE(NULLIF(substr(trade_timestamp, 1, 10), ''), '0000-00-00') >= ?"
            tm_params.append(fiscal_cutoff)
        tm = pd.read_sql_query(
            f"""
            SELECT
                ref_number,
                weight,
                trade_timestamp
            FROM trademc_trades
            WHERE status = 'confirmed'
              AND ref_number IS NOT NULL
              AND TRIM(ref_number) <> ''
              {tm_date_sql}
            """,
            tm_conn,
            params=tm_params,
        )
        ledger_conn = get_pmx_db_connection() if use_pmx else get_db_connection()
        ledger_date_sql = ""
        ledger_params: List[Any] = []
        if fiscal_cutoff:
            ledger_date_sql = "AND COALESCE(NULLIF(substr(trade_date, 1, 10), ''), '0000-00-00') >= ?"
            ledger_params.append(fiscal_cutoff)
        ledger = pd.read_sql_query(
            f"""
            SELECT
                order_id AS OrderID,
                trade_date AS TradeDate,
                symbol AS Symbol,
                side AS Side,
                quantity AS Quantity,
                price AS Price
            FROM trades
            WHERE order_id IS NOT NULL
              AND TRIM(order_id) <> ''
              AND UPPER(COALESCE(side, '')) IN ('BUY', 'SELL')
              {ledger_date_sql}
            """,
            ledger_conn,
            params=ledger_params,
        )
    except Exception:
        # Fallback to existing loaders if direct SQL path fails.
        trademc_df = load_trademc_trades_with_companies(status="confirmed")
        ledger_df = load_all_pmx_trades() if use_pmx else load_all_trades()
        tm = trademc_df[[c for c in ["ref_number", "weight", "trade_timestamp"] if c in trademc_df.columns]].copy()
        if "trade_timestamp" in tm.columns and fiscal_cutoff:
            tm["_trade_date"] = pd.to_datetime(tm["trade_timestamp"], errors="coerce").dt.strftime("%Y-%m-%d")
            tm = tm[tm["_trade_date"].fillna("") >= fiscal_cutoff].copy()
            tm.drop(columns=["_trade_date"], inplace=True, errors="ignore")
        order_col = "OrderID" if "OrderID" in ledger_df.columns else ("Trade #" if "Trade #" in ledger_df.columns else "")
        if order_col:
            ledger = pd.DataFrame(
                {
                    "OrderID": ledger_df[order_col],
                    "TradeDate": ledger_df.get("Trade Date", ""),
                    "Symbol": ledger_df.get("Symbol", ""),
                    "Side": ledger_df.get("Side", ""),
                    "Quantity": ledger_df.get("Quantity", 0.0),
                    "Price": ledger_df.get("Price", 0.0),
                }
            )
        else:
            ledger = pd.DataFrame(columns=["OrderID", "TradeDate", "Symbol", "Side", "Quantity", "Price"])
    finally:
        if tm_conn is not None:
            tm_conn.close()
        if ledger_conn is not None:
            ledger_conn.close()

    tm_grouped = pd.DataFrame(columns=["trade_num", "tm_weight_g", "tm_weight_oz", "trade_date"])
    if not tm.empty:
        tm["trade_num"] = tm["ref_number"].apply(normalize_trade_number) if "ref_number" in tm.columns else ""
        tm = tm[tm["trade_num"] != ""]
        if not tm.empty:
            tm["weight"] = pd.to_numeric(tm.get("weight"), errors="coerce").fillna(0.0)
            tm["tm_weight_oz"] = tm["weight"] / 31.1035
            # Extract trade date (YYYY-MM-DD) from trade_timestamp for daily grouping
            if "trade_timestamp" in tm.columns:
                tm["trade_date"] = pd.to_datetime(tm["trade_timestamp"], errors="coerce").dt.strftime("%Y-%m-%d")
            else:
                tm["trade_date"] = None
            tm_grouped = (
                tm.groupby("trade_num", dropna=False)
                .agg(
                    tm_weight_g=("weight", "sum"),
                    tm_weight_oz=("tm_weight_oz", "sum"),
                    trade_date=("trade_date", "first"),
                )
                .reset_index()
            )

    if ledger.empty:
        if tm_grouped.empty:
            return []
        return _with_zero_hedge(tm_grouped)

    ledger["trade_num"] = ledger["OrderID"].apply(normalize_trade_number) if "OrderID" in ledger.columns else ""
    ledger = ledger[ledger["trade_num"] != ""]
    if ledger.empty:
        if tm_grouped.empty:
            return []
        return _with_zero_hedge(tm_grouped)

    ledger["Symbol"] = (
        ledger.get("Symbol", "")
        .astype(str)
        .str.upper()
        .str.replace("/", "", regex=False)
        .str.replace("-", "", regex=False)
        .str.strip()
    )
    ledger["Side"] = ledger.get("Side", "").astype(str).str.upper()
    ledger["Quantity"] = pd.to_numeric(ledger.get("Quantity"), errors="coerce").fillna(0.0)
    ledger["Price"] = pd.to_numeric(ledger.get("Price"), errors="coerce").fillna(0.0)
    if "TradeDate" in ledger.columns:
        ledger["trade_date"] = pd.to_datetime(ledger.get("TradeDate"), errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
    else:
        ledger["trade_date"] = ""
    ledger = ledger[ledger["Symbol"].isin(["XAUUSD", "USDZAR"]) & ledger["Side"].isin(["BUY", "SELL"])].copy()
    if ledger.empty:
        if tm_grouped.empty:
            return []
        return _with_zero_hedge(tm_grouped)

    ledger["_qty_abs"] = ledger["Quantity"].abs()
    is_buy = ledger["Side"] == "BUY"
    is_metal = ledger["Symbol"] == "XAUUSD"
    is_fx = ledger["Symbol"] == "USDZAR"

    metal = ledger[is_metal].copy()

    side_totals = (
        metal.groupby(["trade_num", "Side"], dropna=False)["_qty_abs"]
        .sum()
        .unstack(fill_value=0.0)
        .reset_index()
    )
    if "BUY" not in side_totals.columns:
        side_totals["BUY"] = 0.0
    if "SELL" not in side_totals.columns:
        side_totals["SELL"] = 0.0

    side_totals["stonex_buy_oz"] = side_totals["BUY"]
    side_totals["stonex_sell_oz"] = side_totals["SELL"]
    # Keep directional sign so PMX BUY offsets negative TradeMC sales and vice versa.
    side_totals["stonex_net_oz"] = side_totals["stonex_sell_oz"] - side_totals["stonex_buy_oz"]
    side_totals["stonex_hedge_g"] = side_totals["stonex_net_oz"] * 31.1035

    # Match ledger Balance USD logic:
    # net_usd = credit_usd - debit_usd, where
    # - XAU BUY debits USD, XAU SELL credits USD
    # - USDZAR SELL debits USD, USDZAR BUY credits USD
    debit_usd = pd.Series(0.0, index=ledger.index)
    credit_usd = pd.Series(0.0, index=ledger.index)
    debit_usd.loc[is_metal & is_buy] = ledger.loc[is_metal & is_buy, "_qty_abs"] * ledger.loc[is_metal & is_buy, "Price"]
    credit_usd.loc[is_metal & (~is_buy)] = ledger.loc[is_metal & (~is_buy), "_qty_abs"] * ledger.loc[is_metal & (~is_buy), "Price"]
    debit_usd.loc[is_fx & (ledger["Side"] == "SELL")] = ledger.loc[is_fx & (ledger["Side"] == "SELL"), "_qty_abs"]
    credit_usd.loc[is_fx & is_buy] = ledger.loc[is_fx & is_buy, "_qty_abs"]
    ledger["_net_usd"] = credit_usd - debit_usd

    usd_totals = (
        ledger.groupby("trade_num", dropna=False)["_net_usd"]
        .sum()
        .reset_index()
        .rename(columns={"_net_usd": "pmx_net_usd"})
    )
    ledger_dates = (
        ledger.groupby("trade_num", dropna=False)["trade_date"]
        .max()
        .reset_index()
    )

    if tm_grouped.empty:
        compare = ledger_dates.copy()
    else:
        compare = tm_grouped.merge(
            ledger_dates.rename(columns={"trade_date": "ledger_trade_date"}),
            on="trade_num",
            how="outer",
        )
        compare["trade_date"] = compare.get("trade_date", "").fillna(compare.get("ledger_trade_date", ""))
        compare.drop(columns=["ledger_trade_date"], inplace=True, errors="ignore")

    compare = compare.merge(
        side_totals[["trade_num", "stonex_buy_oz", "stonex_sell_oz", "stonex_net_oz", "stonex_hedge_g"]],
        on="trade_num",
        how="left",
    )
    compare = compare.merge(usd_totals, on="trade_num", how="left")
    compare = compare[compare["trade_num"].astype(str).str.strip() != ""].copy()

    for col in [
        "tm_weight_g",
        "tm_weight_oz",
        "stonex_buy_oz",
        "stonex_sell_oz",
        "stonex_net_oz",
        "stonex_hedge_g",
        "pmx_net_usd",
    ]:
        if col not in compare.columns:
            compare[col] = 0.0
        compare[col] = pd.to_numeric(compare[col], errors="coerce").fillna(0.0)

    # Hedging ledger should be TradeMC-driven. Drop rows with effectively zero
    # TradeMC net weight so PMX-only cash rows do not appear as hedge trades.
    tm_non_zero_eps_g = 0.005
    compare = compare[compare["tm_weight_g"].abs() > tm_non_zero_eps_g].copy()
    if compare.empty:
        return []

    compare["hedge_need_g"] = compare["tm_weight_g"] - compare["stonex_hedge_g"]
    compare["metal_need_oz"] = compare["tm_weight_oz"] - compare["stonex_net_oz"]
    compare["usd_to_cut"] = compare["pmx_net_usd"].abs()
    compare["usd_need"] = compare["usd_to_cut"]

    compare["pmx_net_oz"] = compare["stonex_net_oz"]
    compare["pmx_hedge_g"] = compare["stonex_hedge_g"]

    compare["metal_hedged"] = compare["hedge_need_g"].abs() <= 31.1035  # 1 oz tolerance
    compare["usd_hedged"] = compare["usd_to_cut"] <= 1.0
    compare["hedged"] = compare["metal_hedged"] & compare["usd_hedged"]
    compare = compare.sort_values("trade_num")

    # Ensure strict JSON-safe output (no NaN/Infinity values).
    compare_safe = compare.astype(object).where(pd.notna(compare), None)
    return compare_safe.to_dict(orient="records")


def _build_weighted_average_from_rows(matched: pd.DataFrame, trade_num: str) -> Optional[Dict[str, Any]]:
    if matched is None or matched.empty:
        return None
    result = {"trade_num": trade_num, "xau_usd": [], "usd_zar": []}

    def _to_float(value: Any) -> float:
        try:
            out = float(value)
            if math.isfinite(out):
                return out
            return 0.0
        except Exception:
            return 0.0

    for _, row in matched.iterrows():
        sym = str(row.get("Symbol", "")).upper()
        base, quote = split_symbol(sym)
        raw_qty = _to_float(row.get("Quantity", 0))
        price = _to_float(row.get("Price", 0))
        side = str(row.get("Side", "")).upper()
        td = str(row.get("Trade Date", row.get("trade_date", "")))

        # Use signed flow so weighted average reflects net position:
        # BUY = positive quantity, SELL = negative quantity.
        qty_abs = abs(raw_qty)
        if side == "BUY":
            signed_qty = qty_abs
        elif side == "SELL":
            signed_qty = -qty_abs
        else:
            signed_qty = raw_qty

        entry = {
            "date": td,
            "side": side,
            "quantity": signed_qty,
            "price": price,
            "symbol": sym,
        }

        if base in {"XAU", "XAG", "XPT", "XPD"} and quote == "USD":
            entry["value"] = signed_qty * price
            result["xau_usd"].append(entry)
        elif base == "USD" and quote == "ZAR":
            entry["value"] = signed_qty * price
            result["usd_zar"].append(entry)

    # Compute weighted averages
    for key in ["xau_usd", "usd_zar"]:
        entries = result[key]
        if entries:
            total_qty = sum(e["quantity"] for e in entries)
            total_val = sum(e["value"] for e in entries)
            result[f"{key}_wa_price"] = round(total_val / total_qty, 5) if abs(total_qty) > 1e-12 else 0
            result[f"{key}_total_qty"] = round(total_qty, 4)
            result[f"{key}_total_val"] = round(total_val, 2)

    return result


def build_weighted_average(trade_num_input, source: str = "pmx"):
    """Calculate weighted average for a trade number from the selected ledger source."""
    trade_num = normalize_trade_number(trade_num_input)
    if not trade_num:
        return None

    source_key = str(source or "pmx").strip().lower()
    use_pmx = source_key == "pmx"

    if use_pmx:
        matched = load_pmx_trades_for_trade_number(trade_num)
        return _build_weighted_average_from_rows(matched, trade_num)

    df = load_all_trades()
    if df.empty:
        return None

    if "OrderID" in df.columns:
        trade_col = "OrderID"
    elif "Trade #" in df.columns:
        trade_col = "Trade #"
    else:
        return None

    df["trade_num"] = df[trade_col].apply(normalize_trade_number)
    matched = df[df["trade_num"] == trade_num].copy()
    return _build_weighted_average_from_rows(matched, trade_num)


def _weighted_avg(values, weights):
    if values is None or weights is None:
        return None
    mask = values.notna() & weights.notna()
    if not mask.any():
        return None
    w = weights[mask]
    v = values[mask]
    total_w = w.sum()
    if total_w == 0:
        return None
    return float((v * w).sum() / total_w)


def _df_to_records(df: pd.DataFrame):
    if df is None or df.empty:
        return []
    return df.where(pd.notna(df), None).to_dict(orient="records")


def build_trading_ticket_frames(trade_num_input, source: str = "pmx"):
    """Build trading ticket dataframes for a trade number."""
    trade_num = normalize_trade_number(trade_num_input)
    if not trade_num:
        return None
    source_key = str(source or "pmx").strip().lower()
    use_pmx = source_key == "pmx"

    tm_detail = pd.DataFrame()
    stonex_rows = pd.DataFrame()
    summary_df = pd.DataFrame()

    tm_rows = pd.DataFrame()
    tm_usd_value = None
    tm_zar_value = None
    tm_total_weight_g = None
    hedge_need_g = None

    tm = load_trademc_trades_with_companies(status="confirmed")
    if not tm.empty and "ref_number" in tm.columns:
        tm["_tn"] = tm["ref_number"].apply(normalize_trade_number)
        tm_rows = tm[tm["_tn"] == trade_num].copy()

        if not tm_rows.empty:
            tm_rows["weight"] = pd.to_numeric(tm_rows.get("weight"), errors="coerce")
            tm_rows["weight_oz"] = tm_rows["weight"] / 31.1035
            tm_total_weight_g = tm_rows["weight"].sum() if "weight" in tm_rows.columns else None
            tm_total_weight_g = float(tm_total_weight_g) if pd.notna(tm_total_weight_g) else None

            fx_rate_col = "zar_to_usd_confirmed" if "zar_to_usd_confirmed" in tm_rows.columns else "zar_to_usd"
            usd_rate_col = "usd_per_troy_ounce_confirmed" if "usd_per_troy_ounce_confirmed" in tm_rows.columns else ""
            zar_rate_col = "zar_per_troy_ounce_confirmed" if "zar_per_troy_ounce_confirmed" in tm_rows.columns else "zar_per_troy_ounce"

            fx_rate = pd.to_numeric(tm_rows.get(fx_rate_col), errors="coerce") if fx_rate_col else pd.Series([pd.NA] * len(tm_rows))
            usd_rate = pd.to_numeric(tm_rows.get(usd_rate_col), errors="coerce") if usd_rate_col in tm_rows.columns else pd.Series([pd.NA] * len(tm_rows))

            if usd_rate.isna().all() and zar_rate_col in tm_rows.columns and fx_rate_col:
                zar_rate = pd.to_numeric(tm_rows.get(zar_rate_col), errors="coerce")
                usd_rate = zar_rate / fx_rate

            tm_rows["usd_rate"] = usd_rate
            tm_rows["fx_rate"] = fx_rate
            tm_rows["usd_value"] = tm_rows["weight_oz"] * tm_rows["usd_rate"]
            tm_rows["zar_value"] = tm_rows["usd_value"] * tm_rows["fx_rate"]
            tm_rows["zar_spot_per_g"] = (tm_rows["usd_rate"] * tm_rows["fx_rate"]) / 31.1035
            tm_rows["zar_spot_value"] = tm_rows["zar_spot_per_g"] * tm_rows["weight"]
            if "company_refining_rate" in tm_rows.columns:
                refining_rate = pd.to_numeric(tm_rows["company_refining_rate"], errors="coerce").fillna(0.0)
            else:
                refining_rate = pd.Series([0.0] * len(tm_rows), index=tm_rows.index)
            tm_rows["company_refining_rate"] = refining_rate
            tm_rows["zar_value_less_refining"] = tm_rows["zar_value"] * (1.0 - (refining_rate / 100.0))

            detail_cols = [
                "company_name",
                "weight",
                "weight_oz",
                "usd_rate",
                "fx_rate",
                "usd_value",
                "zar_value",
                "company_refining_rate",
                "zar_value_less_refining",
            ]
            detail_cols = [c for c in detail_cols if c in tm_rows.columns]
            tm_detail = tm_rows[detail_cols].copy()
            tm_detail = tm_detail.rename(columns={
                "company_name": "Company",
                "weight": "Weight (g)",
                "weight_oz": "Weight (oz)",
                "usd_rate": "$/oz Booked",
                "fx_rate": "FX Rate",
                "usd_value": "USD Value",
                "zar_value": "ZAR Value",
                "company_refining_rate": "company_refining_rate",
                "zar_value_less_refining": "zar_value_less_refining",
            })

            tm_usd_value = tm_rows["usd_value"].sum() if "usd_value" in tm_rows.columns else None
            tm_usd_value = float(tm_usd_value) if pd.notna(tm_usd_value) else None
            tm_zar_value = tm_rows["zar_spot_value"].sum() if "zar_spot_value" in tm_rows.columns else None
            tm_zar_value = float(tm_zar_value) if pd.notna(tm_zar_value) else None

    # Control account base comes from hedging unhedged grams.
    try:
        hedge_rows = build_hedging_comparison(source=source_key)
        for row in hedge_rows or []:
            if normalize_trade_number(row.get("trade_num")) != trade_num:
                continue
            hv = pd.to_numeric(pd.Series([row.get("hedge_need_g")]), errors="coerce").iloc[0]
            if pd.notna(hv):
                hedge_need_g = float(hv)
            break
    except Exception:
        hedge_need_g = None

    ledger_rows = pd.DataFrame()
    if use_pmx:
        ledger_rows = load_pmx_trades_for_trade_number(trade_num)
    else:
        df = load_all_trades()
        trade_col = "OrderID" if "OrderID" in df.columns else ("Trade #" if "Trade #" in df.columns else "")
        if not df.empty and trade_col:
            df["_tn"] = df[trade_col].apply(normalize_trade_number)
            ledger_rows = df[df["_tn"] == trade_num].copy()

    if not ledger_rows.empty:
        ledger_rows["Symbol"] = ledger_rows.get("Symbol", "").astype(str).str.upper()
        ledger_rows["Side"] = ledger_rows.get("Side", "").astype(str).str.upper()
        ledger_rows["Quantity"] = pd.to_numeric(ledger_rows.get("Quantity"), errors="coerce")
        ledger_rows["Price"] = pd.to_numeric(ledger_rows.get("Price"), errors="coerce")

        # PMX trades in ticket: gold first, then USDZAR; newest trades at the top.
        symbol_norm = (
            ledger_rows["Symbol"]
            .astype(str)
            .str.upper()
            .str.replace("/", "", regex=False)
            .str.replace("-", "", regex=False)
            .str.strip()
        )
        ledger_rows["_symbol_rank"] = symbol_norm.map({"XAUUSD": 0, "USDZAR": 1}).fillna(2).astype(int)
        ledger_rows["_trade_dt"] = pd.to_datetime(ledger_rows.get("Trade Date"), errors="coerce")
        ledger_rows["_value_dt"] = pd.to_datetime(ledger_rows.get("Value Date"), errors="coerce")
        sort_cols = ["_symbol_rank", "_trade_dt", "_value_dt"]
        sort_asc = [True, False, False]
        if "Doc #" in ledger_rows.columns:
            sort_cols.append("Doc #")
            sort_asc.append(False)
        ledger_rows = ledger_rows.sort_values(sort_cols, ascending=sort_asc, na_position="last")

        show_cols = ["Doc #", "FNC #", "Trade Date", "Value Date", "Symbol", "Side", "Quantity", "Price", "Narration", "Settle Currency", "Settle Amount"]
        stonex_rows = ledger_rows[[c for c in show_cols if c in ledger_rows.columns]].copy()
        if "Trade Date" in stonex_rows.columns:
            stonex_rows["Trade Date"] = pd.to_datetime(stonex_rows["Trade Date"], errors="coerce").dt.strftime("%Y-%m-%d")
        if "Value Date" in stonex_rows.columns:
            stonex_rows["Value Date"] = pd.to_datetime(stonex_rows["Value Date"], errors="coerce").dt.strftime("%Y-%m-%d")

        # Keep ticket WA values identical to Weighted Average logic without reloading the full PMX table.
        wa_result = _build_weighted_average_from_rows(ledger_rows, trade_num)
        gold_avg = wa_result.get("xau_usd_wa_price") if wa_result else None
        fx_avg = wa_result.get("usd_zar_wa_price") if wa_result else None

        # For control account metal exposure:
        # BUY adds to held metal, SELL reduces held metal.
        position_sign = ledger_rows["Side"].map({"BUY": 1.0, "SELL": -1.0}).fillna(0.0)

        gold_rows = ledger_rows[ledger_rows["Symbol"] == "XAUUSD"].copy()
        fx_rows = ledger_rows[ledger_rows["Symbol"] == "USDZAR"].copy()

        # Signed flow convention for ticket PnL:
        # BUY = cash outflow (negative), SELL = cash inflow (positive).
        cash_flow_sign = ledger_rows["Side"].map({"BUY": -1.0, "SELL": 1.0}).fillna(0.0)

        gold_notional = None
        if not gold_rows.empty:
            gold_notional = float(
                (gold_rows["Quantity"].abs() * gold_rows["Price"] * cash_flow_sign.loc[gold_rows.index]).sum()
            )

        fx_net_zar = None
        if not fx_rows.empty:
            fx_net_zar = float(
                (fx_rows["Quantity"].abs() * fx_rows["Price"] * cash_flow_sign.loc[fx_rows.index]).sum()
            )

        gold_abs_wa_price = None
        if not gold_rows.empty:
            gold_abs_qty = gold_rows["Quantity"].abs()
            gold_abs_total = float(gold_abs_qty.sum())
            if gold_abs_total > 1e-12:
                gold_abs_wa_price = float((gold_rows["Price"] * gold_abs_qty).sum() / gold_abs_total)

        fx_abs_wa_price = None
        if not fx_rows.empty:
            fx_abs_qty = fx_rows["Quantity"].abs()
            fx_abs_total = float(fx_abs_qty.sum())
            if fx_abs_total > 1e-12:
                fx_abs_wa_price = float((fx_rows["Price"] * fx_abs_qty).sum() / fx_abs_total)

        gold_price_for_control = float(gold_avg) if gold_avg is not None else None
        if (gold_price_for_control is None or abs(gold_price_for_control) <= 1e-12) and gold_abs_wa_price is not None:
            gold_price_for_control = gold_abs_wa_price

        fx_rate_for_control = float(fx_avg) if fx_avg is not None else None
        if (fx_rate_for_control is None or abs(fx_rate_for_control) <= 1e-12) and fx_abs_wa_price is not None:
            fx_rate_for_control = fx_abs_wa_price

        spot_rate_zar_g = None
        if gold_avg is not None and fx_avg is not None:
            try:
                spot_rate_zar_g = (float(gold_avg) * float(fx_avg)) / 31.1035
            except Exception:
                spot_rate_zar_g = None
        if (
            (spot_rate_zar_g is None or abs(spot_rate_zar_g) <= 1e-12)
            and gold_price_for_control is not None
            and fx_rate_for_control is not None
        ):
            spot_rate_zar_g = (gold_price_for_control * fx_rate_for_control) / 31.1035

        pmx_signed_gold_oz = None
        if not gold_rows.empty:
            signed_gold_oz = (gold_rows["Quantity"].abs() * position_sign.loc[gold_rows.index]).sum()
            pmx_signed_gold_oz = float(signed_gold_oz)

        # StoneX ZAR flow should reflect traded gold grams at spot ZAR/g.
        stonex_zar_flow = None
        if pmx_signed_gold_oz is not None and spot_rate_zar_g is not None:
            traded_gold_g = abs(float(pmx_signed_gold_oz)) * 31.1035
            # Keep StoneX ZAR Flow consistent with the displayed ticket figures:
            # Total Traded (g) is shown to 2 dp and Spot ZAR/g to 4 dp.
            traded_gold_g_display = round(traded_gold_g, 2)
            spot_rate_zar_g_display = round(float(spot_rate_zar_g), 4)
            stonex_zar_flow = traded_gold_g_display * spot_rate_zar_g_display

        is_sell_hedge = tm_total_weight_g is not None and tm_total_weight_g < 0

        control_account_g = None
        control_account_oz = None
        control_account_usd_value = None
        control_account_spot_zar_g = None
        control_account_zar_value = None
        if hedge_need_g is not None:
            control_account_g = hedge_need_g
        elif tm_total_weight_g is not None and pmx_signed_gold_oz is not None:
            # Fallback if hedging row is unavailable.
            control_account_oz = pmx_signed_gold_oz + (tm_total_weight_g / 31.1035)
            control_account_g = control_account_oz * 31.1035

        # For sell hedges (negative TradeMC grams), show/control the remaining
        # PMX hedge requirement as a positive adjusting quantity in the ticket.
        if (
            control_account_g is not None
            and is_sell_hedge
            and control_account_g < 0
        ):
            control_account_g = abs(float(control_account_g))

        if control_account_g is not None:
            control_account_oz = float(control_account_g) / 31.1035
            control_account_spot_zar_g = spot_rate_zar_g
            if control_account_spot_zar_g is not None:
                control_account_zar_value = control_account_spot_zar_g * float(control_account_g)
            if (
                control_account_zar_value is not None
                and fx_rate_for_control is not None
                and abs(float(fx_rate_for_control)) > 1e-12
            ):
                control_account_usd_value = control_account_zar_value / float(fx_rate_for_control)
            elif gold_price_for_control is not None:
                control_account_usd_value = control_account_oz * float(gold_price_for_control)

        profit_usd = None
        profit_zar = None
        profit_pct = None
        stonex_sell_usd = gold_notional
        if stonex_sell_usd is None and fx_net_zar is not None and fx_rate_for_control not in (None, 0):
            stonex_sell_usd = fx_net_zar / float(fx_rate_for_control)

        stonex_sell_zar = stonex_zar_flow if stonex_zar_flow is not None else fx_net_zar
        if stonex_sell_zar is None and gold_notional is not None and fx_rate_for_control not in (None, 0):
            stonex_sell_zar = gold_notional * float(fx_rate_for_control)

        tm_side_usd = abs(float(tm_usd_value)) if tm_usd_value is not None else None
        # TradeMC total spot-value in ZAR.
        tm_side_zar = abs(float(tm_zar_value)) if tm_zar_value is not None else None

        stonex_leg_usd = None
        if stonex_sell_usd is not None or control_account_usd_value is not None:
            stonex_sell_usd_component = float(stonex_sell_usd) if stonex_sell_usd is not None else 0.0
            control_account_usd_component = (
                float(control_account_usd_value) if control_account_usd_value is not None else 0.0
            )
            if is_sell_hedge:
                # For sell hedges, both executed StoneX hedge and control-account
                # adjustment contribute to the sell-side notional.
                stonex_sell_usd_component = abs(stonex_sell_usd_component)
                control_account_usd_component = abs(control_account_usd_component)
            stonex_leg_usd = stonex_sell_usd_component + control_account_usd_component

        stonex_leg_zar = None
        if stonex_sell_zar is not None or control_account_zar_value is not None:
            stonex_leg_zar = (float(stonex_sell_zar) if stonex_sell_zar is not None else 0.0) + (
                float(control_account_zar_value) if control_account_zar_value is not None else 0.0
            )

        # Profit side logic:
        # - Client negative weight (short): TradeMC is sell side, StoneX(+control) is buy side.
        # - Client positive weight (long): StoneX(+control) is sell side, TradeMC is buy side.
        sell_side_usd = None
        buy_side_usd = None
        sell_side_zar = None
        buy_side_zar = None
        if is_sell_hedge:
            sell_side_usd = tm_side_usd
            buy_side_usd = abs(float(stonex_leg_usd)) if stonex_leg_usd is not None else None
            sell_side_zar = tm_side_zar
            buy_side_zar = abs(float(stonex_leg_zar)) if stonex_leg_zar is not None else None
        else:
            sell_side_usd = abs(float(stonex_leg_usd)) if stonex_leg_usd is not None else None
            buy_side_usd = tm_side_usd
            sell_side_zar = abs(float(stonex_leg_zar)) if stonex_leg_zar is not None else None
            buy_side_zar = tm_side_zar

        if sell_side_usd is not None and buy_side_usd is not None:
            profit_usd = sell_side_usd - buy_side_usd

        # Simple ticket rule: profit is total sell minus total buy (same currency only).
        if sell_side_zar is not None and buy_side_zar is not None:
            profit_zar = sell_side_zar - buy_side_zar

        if buy_side_zar not in (None, 0) and profit_zar is not None:
            profit_pct = (float(profit_zar) / float(buy_side_zar)) * 100.0

        if (spot_rate_zar_g is None or abs(spot_rate_zar_g) <= 1e-12) and control_account_spot_zar_g is not None:
            spot_rate_zar_g = control_account_spot_zar_g

        summary_df = pd.DataFrame([{
            "Gold WA $/oz": gold_avg,
            "FX WA USD/ZAR": fx_avg,
            "Sell Side (USD)": sell_side_usd,
            "Buy Side (USD)": buy_side_usd,
            "Sell Side (ZAR)": sell_side_zar,
            "Buy Side (ZAR)": buy_side_zar,
            "Total Traded (oz)": abs(float(pmx_signed_gold_oz)) if pmx_signed_gold_oz is not None else None,
            "Total Traded (g)": (abs(float(pmx_signed_gold_oz)) * 31.1035) if pmx_signed_gold_oz is not None else None,
            "Control Account (g)": control_account_g,
            "Control Account (oz)": control_account_oz,
            "Control Account (ZAR)": control_account_zar_value,
            "Spot ZAR/g": spot_rate_zar_g,
            "StoneX ZAR Flow": stonex_zar_flow if stonex_zar_flow is not None else fx_net_zar,
            "Profit (USD)": profit_usd,
            "Profit (ZAR)": profit_zar,
            "Profit % (ZAR Spot Cost)": profit_pct,
        }])

    return {
        "trade_num": trade_num,
        "tm_detail": tm_detail,
        "stonex_rows": stonex_rows,
        "summary": summary_df,
    }


def build_trading_ticket(trade_num_input):
    """Build trading ticket JSON payload for a trade number."""
    frames = build_trading_ticket_frames(trade_num_input)
    if frames is None:
        return None
    return {
        "trade_num": frames["trade_num"],
        "trademc": _df_to_records(frames["tm_detail"]),
        "stonex": _df_to_records(frames["stonex_rows"]),
        "summary": _df_to_records(frames["summary"]),
    }


def build_profit_monthly_report() -> Dict[str, Any]:
    """Build month-wise profit report with trade and transaction drill-down."""

    def _to_num(value: Any) -> Optional[float]:
        try:
            out = float(value)
            if math.isfinite(out):
                return out
            return None
        except Exception:
            return None

    def _to_dt(value: Any) -> Optional[datetime]:
        dt = pd.to_datetime(value, errors="coerce")
        if pd.isna(dt):
            return None
        if hasattr(dt, "to_pydatetime"):
            return dt.to_pydatetime()
        return None

    def _sym_norm(value: Any) -> str:
        return str(value or "").upper().replace("/", "").replace("-", "").replace(" ", "").strip()

    def _month_key(dt: Optional[datetime]) -> str:
        if dt is None:
            return "Unknown"
        return dt.strftime("%Y-%m")

    def _month_label(key: str) -> str:
        try:
            return datetime.strptime(key, "%Y-%m").strftime("%b %Y")
        except Exception:
            return key

    tm_df = load_trademc_trades_with_companies(status="confirmed")
    pmx_df = load_all_pmx_trades()
    if tm_df is None:
        tm_df = pd.DataFrame()
    if pmx_df is None:
        pmx_df = pd.DataFrame()

    hedge_need_map: Dict[str, float] = {}
    usd_to_cut_map: Dict[str, float] = {}
    try:
        hedging_rows = build_hedging_comparison(source="pmx")
        for raw in hedging_rows or []:
            if not isinstance(raw, dict):
                continue
            tn = normalize_trade_number(raw.get("trade_num"))
            if not tn:
                continue
            hv = _to_num(raw.get("hedge_need_g"))
            if hv is not None:
                hedge_need_map[tn] = float(hv)
            uv = _to_num(raw.get("usd_to_cut"))
            if uv is not None:
                usd_to_cut_map[tn] = abs(float(uv))
    except Exception:
        hedge_need_map = {}
        usd_to_cut_map = {}

    tm_by_trade: Dict[str, Dict[str, Any]] = {}
    if not tm_df.empty and "ref_number" in tm_df.columns:
        for rec in tm_df.to_dict(orient="records"):
            tn = normalize_trade_number(rec.get("ref_number"))
            if not tn:
                continue

            state = tm_by_trade.setdefault(
                tn,
                {
                    "tm_weight_g": 0.0,
                    "tm_usd_value": 0.0,
                    "tm_zar_value": 0.0,
                    "has_tm_usd": False,
                    "has_tm_zar": False,
                    "latest_dt": None,
                    "transactions": [],
                },
            )

            weight_g = _to_num(rec.get("weight"))
            weight_g = float(weight_g) if weight_g is not None else 0.0

            fx_rate = _to_num(rec.get("zar_to_usd_confirmed"))
            if fx_rate is None:
                fx_rate = _to_num(rec.get("zar_to_usd"))

            usd_rate = _to_num(rec.get("usd_per_troy_ounce_confirmed"))
            if usd_rate is None:
                zar_rate = _to_num(rec.get("zar_per_troy_ounce_confirmed"))
                if zar_rate is None:
                    zar_rate = _to_num(rec.get("zar_per_troy_ounce"))
                if zar_rate is not None and fx_rate is not None and abs(float(fx_rate)) > 1e-12:
                    usd_rate = float(zar_rate) / float(fx_rate)

            weight_oz = weight_g / GRAMS_PER_TROY_OUNCE
            usd_value = (weight_oz * float(usd_rate)) if usd_rate is not None else None
            zar_spot_per_g = None
            if usd_rate is not None and fx_rate is not None:
                zar_spot_per_g = (float(usd_rate) * float(fx_rate)) / GRAMS_PER_TROY_OUNCE
            zar_value = (zar_spot_per_g * weight_g) if zar_spot_per_g is not None else None

            state["tm_weight_g"] += weight_g
            if usd_value is not None:
                state["tm_usd_value"] += float(usd_value)
                state["has_tm_usd"] = True
            if zar_value is not None:
                state["tm_zar_value"] += float(zar_value)
                state["has_tm_zar"] = True

            tdt = _to_dt(rec.get("trade_timestamp")) or _to_dt(rec.get("date_created"))
            if tdt is not None and (state["latest_dt"] is None or tdt > state["latest_dt"]):
                state["latest_dt"] = tdt

            state["transactions"].append(
                {
                    "Source": "TradeMC",
                    "Date": (tdt.strftime("%Y-%m-%d") if tdt else ""),
                    "Company": str(rec.get("company_name") or ""),
                    "Weight (g)": weight_g if abs(weight_g) > 1e-12 else None,
                    "USD/oz": usd_rate,
                    "FX Rate": fx_rate,
                    "ZAR Value": zar_value,
                    "Reference": str(rec.get("ref_number") or ""),
                    "ID": rec.get("id"),
                }
            )

    pmx_by_trade: Dict[str, Dict[str, Any]] = {}
    if not pmx_df.empty:
        trade_col = "OrderID" if "OrderID" in pmx_df.columns else ("Trade #" if "Trade #" in pmx_df.columns else "")
        if trade_col:
            for rec in pmx_df.to_dict(orient="records"):
                tn = normalize_trade_number(rec.get(trade_col))
                if not tn:
                    continue

                state = pmx_by_trade.setdefault(
                    tn,
                    {
                        "gold_notional_usd": 0.0,
                        "fx_notional_zar": 0.0,
                        "gold_abs_qty": 0.0,
                        "gold_abs_price_qty": 0.0,
                        "fx_abs_qty": 0.0,
                        "fx_abs_price_qty": 0.0,
                        "gold_signed_oz": 0.0,
                        "xau_total_qty": 0.0,
                        "xau_total_val": 0.0,
                        "fx_total_qty": 0.0,
                        "fx_total_val": 0.0,
                        "latest_dt": None,
                        "transactions": [],
                    },
                )

                symbol = _sym_norm(rec.get("Symbol"))
                side = str(rec.get("Side") or "").upper().strip()
                side_sign = 1.0 if side == "BUY" else (-1.0 if side == "SELL" else 0.0)
                cash_flow_sign = -1.0 if side == "BUY" else (1.0 if side == "SELL" else 0.0)
                qty = abs(float(_to_num(rec.get("Quantity")) or 0.0))
                price = float(_to_num(rec.get("Price")) or 0.0)

                pdt = _to_dt(rec.get("Trade Date"))
                vdt = _to_dt(rec.get("Value Date"))
                if pdt is not None and (state["latest_dt"] is None or pdt > state["latest_dt"]):
                    state["latest_dt"] = pdt

                state["transactions"].append(
                    {
                        "Source": "PMX",
                        "Trade Date": pdt.strftime("%Y-%m-%d") if pdt else "",
                        "Value Date": vdt.strftime("%Y-%m-%d") if vdt else "",
                        "Symbol": symbol,
                        "Side": side,
                        "Quantity": qty if qty > 0 else None,
                        "Price": price if price > 0 else None,
                        "FNC #": str(rec.get("FNC #") or ""),
                        "Doc #": str(rec.get("Doc #") or ""),
                        "Narration": str(rec.get("Narration") or ""),
                    }
                )

                if symbol == "XAUUSD":
                    notional = qty * price
                    state["gold_notional_usd"] += (notional * cash_flow_sign)
                    state["gold_abs_qty"] += qty
                    state["gold_abs_price_qty"] += notional
                    signed_qty = qty * side_sign
                    state["gold_signed_oz"] += signed_qty
                    state["xau_total_qty"] += signed_qty
                    state["xau_total_val"] += signed_qty * price
                elif symbol == "USDZAR":
                    notional = qty * price
                    state["fx_notional_zar"] += (notional * cash_flow_sign)
                    state["fx_abs_qty"] += qty
                    state["fx_abs_price_qty"] += notional
                    signed_qty = qty * side_sign
                    state["fx_total_qty"] += signed_qty
                    state["fx_total_val"] += signed_qty * price

    trade_rows: List[Dict[str, Any]] = []
    for tn, tm_state in tm_by_trade.items():
        pmx_state = pmx_by_trade.get(tn, {})

        tm_weight_g = float(tm_state.get("tm_weight_g", 0.0))
        tm_weight_oz = float(tm_weight_g) / GRAMS_PER_TROY_OUNCE if abs(float(tm_weight_g)) > 1e-12 else 0.0
        is_short = tm_weight_g < 0

        tm_side_usd = abs(float(tm_state["tm_usd_value"])) if bool(tm_state.get("has_tm_usd")) else None
        tm_side_zar = abs(float(tm_state["tm_zar_value"])) if bool(tm_state.get("has_tm_zar")) else None

        gold_notional = _to_num(pmx_state.get("gold_notional_usd"))
        fx_net_zar = _to_num(pmx_state.get("fx_notional_zar"))
        pmx_signed_oz = _to_num(pmx_state.get("gold_signed_oz"))

        xau_total_qty = _to_num(pmx_state.get("xau_total_qty"))
        xau_total_val = _to_num(pmx_state.get("xau_total_val"))
        fx_total_qty = _to_num(pmx_state.get("fx_total_qty"))
        fx_total_val = _to_num(pmx_state.get("fx_total_val"))
        gold_avg = None
        if xau_total_qty is not None and xau_total_val is not None and abs(float(xau_total_qty)) > 1e-12:
            gold_avg = float(xau_total_val) / float(xau_total_qty)
        fx_avg = None
        if fx_total_qty is not None and fx_total_val is not None and abs(float(fx_total_qty)) > 1e-12:
            fx_avg = float(fx_total_val) / float(fx_total_qty)

        gold_abs_qty = _to_num(pmx_state.get("gold_abs_qty"))
        gold_abs_price_qty = _to_num(pmx_state.get("gold_abs_price_qty"))
        gold_abs_wa = None
        if gold_abs_qty is not None and gold_abs_price_qty is not None and abs(float(gold_abs_qty)) > 1e-12:
            gold_abs_wa = float(gold_abs_price_qty) / float(gold_abs_qty)

        fx_abs_qty = _to_num(pmx_state.get("fx_abs_qty"))
        fx_abs_price_qty = _to_num(pmx_state.get("fx_abs_price_qty"))
        fx_abs_wa = None
        if fx_abs_qty is not None and fx_abs_price_qty is not None and abs(float(fx_abs_qty)) > 1e-12:
            fx_abs_wa = float(fx_abs_price_qty) / float(fx_abs_qty)

        gold_price_for_control = gold_avg if gold_avg is not None and abs(float(gold_avg)) > 1e-12 else gold_abs_wa
        fx_rate_for_control = fx_avg if fx_avg is not None and abs(float(fx_avg)) > 1e-12 else fx_abs_wa

        spot_rate_zar_g = None
        if gold_avg is not None and fx_avg is not None:
            spot_rate_zar_g = (float(gold_avg) * float(fx_avg)) / GRAMS_PER_TROY_OUNCE
        elif gold_price_for_control is not None and fx_rate_for_control is not None:
            spot_rate_zar_g = (float(gold_price_for_control) * float(fx_rate_for_control)) / GRAMS_PER_TROY_OUNCE

        control_account_g = hedge_need_map.get(tn)
        if control_account_g is None:
            if pmx_signed_oz is not None:
                control_account_g = tm_weight_g + (float(pmx_signed_oz) * GRAMS_PER_TROY_OUNCE)
        if control_account_g is not None and is_short and control_account_g < 0:
            control_account_g = abs(float(control_account_g))

        control_account_oz = None
        control_account_zar = None
        control_account_usd = None
        if control_account_g is not None:
            control_account_oz = float(control_account_g) / GRAMS_PER_TROY_OUNCE
            if spot_rate_zar_g is not None:
                control_account_zar = float(spot_rate_zar_g) * float(control_account_g)
            if control_account_zar is not None and fx_rate_for_control is not None and abs(float(fx_rate_for_control)) > 1e-12:
                control_account_usd = float(control_account_zar) / float(fx_rate_for_control)
            elif control_account_oz is not None and gold_price_for_control is not None:
                control_account_usd = float(control_account_oz) * float(gold_price_for_control)

        stonex_sell_usd = gold_notional
        if stonex_sell_usd is None and fx_net_zar is not None and fx_rate_for_control is not None and abs(float(fx_rate_for_control)) > 1e-12:
            stonex_sell_usd = float(fx_net_zar) / float(fx_rate_for_control)

        stonex_sell_zar = fx_net_zar
        if stonex_sell_zar is None and gold_notional is not None and fx_rate_for_control is not None:
            stonex_sell_zar = float(gold_notional) * float(fx_rate_for_control)

        # Weighted-average proportional logic:
        # Compare TradeMC WA vs StoneX WA on the quantity actually traded on PMX,
        # capped by available booked TradeMC quantity.
        stonex_total_oz_traded = abs(float(pmx_signed_oz)) if pmx_signed_oz is not None else 0.0
        stonex_total_g_traded = stonex_total_oz_traded * GRAMS_PER_TROY_OUNCE
        tm_booked_oz = abs(float(tm_weight_oz))
        tm_booked_g = abs(float(tm_weight_g))
        metal_gap_g = abs(float(hedge_need_map.get(tn, 0.0) or 0.0))
        usd_to_cut = abs(float(usd_to_cut_map.get(tn, 0.0) or 0.0))
        metal_hedged = metal_gap_g <= 31.1035  # 1 oz tolerance
        usd_hedged = usd_to_cut <= 1.0
        hedged = bool(metal_hedged and usd_hedged)
        hedge_status = "Hedged" if hedged else "Unhedged"

        # Use net/signed WA for PMX first so offsetting intraday churn does not
        # distort profit on trades where buys and sells exist under one trade #.
        wa_gold_price = _to_num(gold_price_for_control if gold_price_for_control is not None else gold_abs_wa)
        wa_usdzar = _to_num(fx_rate_for_control if fx_rate_for_control is not None else fx_abs_wa)
        tm_wa_usd_per_oz = (abs(float(tm_side_usd)) / tm_booked_oz) if (tm_side_usd is not None and tm_booked_oz > 1e-12) else None
        tm_wa_fx = (abs(float(tm_side_zar)) / abs(float(tm_side_usd))) if (tm_side_zar is not None and tm_side_usd is not None and abs(float(tm_side_usd)) > 1e-12) else None

        # Running profit: use all PMX grams sold (not capped by TradeMC booked).
        pmx_sold_oz = stonex_total_oz_traded
        pmx_sold_g = pmx_sold_oz * GRAMS_PER_TROY_OUNCE
        matched_oz = pmx_sold_oz
        matched_g = pmx_sold_g

        # Compute each leg's value pro-rated to PMX sold quantity.
        # PMX leg (valued at PMX weighted averages):
        pmx_leg_usd = None
        if wa_gold_price is not None and pmx_sold_oz > 1e-12:
            pmx_leg_usd = float(pmx_sold_oz) * float(wa_gold_price)
        pmx_leg_zar = None
        if pmx_leg_usd is not None and wa_usdzar is not None:
            pmx_leg_zar = float(pmx_leg_usd) * float(wa_usdzar)

        # TradeMC leg (valued at TradeMC weighted averages, pro-rated to PMX sold):
        tm_leg_usd = None
        if tm_wa_usd_per_oz is not None and pmx_sold_oz > 1e-12:
            tm_leg_usd = float(pmx_sold_oz) * float(tm_wa_usd_per_oz)
        tm_leg_zar = None
        if tm_leg_usd is not None and tm_wa_fx is not None:
            tm_leg_zar = float(tm_leg_usd) * float(tm_wa_fx)

        # Assign sell/buy sides based on position direction:
        # Long (TradeMC grams > 0): Buy in TradeMC, Sell in PMX
        # Short (TradeMC grams < 0): Sell in TradeMC, Buy in PMX
        if is_short:
            sell_side_usd = tm_leg_usd
            buy_side_usd = pmx_leg_usd
            sell_side_zar = tm_leg_zar
            buy_side_zar = pmx_leg_zar
        else:
            sell_side_usd = pmx_leg_usd
            buy_side_usd = tm_leg_usd
            sell_side_zar = pmx_leg_zar
            buy_side_zar = tm_leg_zar

        profit_usd = (float(sell_side_usd) - float(buy_side_usd)) if (sell_side_usd is not None and buy_side_usd is not None) else None
        profit_zar = (float(sell_side_zar) - float(buy_side_zar)) if (sell_side_zar is not None and buy_side_zar is not None) else None
        profit_pct = None
        if buy_side_zar is not None and abs(float(buy_side_zar)) > 1e-12 and profit_zar is not None:
            profit_pct = (float(profit_zar) / abs(float(buy_side_zar))) * 100.0

        stonex_traded_g_row = stonex_total_g_traded
        stonex_zar_flow_row = pmx_leg_zar if pmx_leg_zar is not None else fx_net_zar

        # Split profit into metal-driven and exchange-driven components.
        # Metal profit: USD price spread converted to ZAR at the TradeMC WA FX rate
        # (cost-basis rate), so exchange_profit captures only the PMX FX rate differential.
        metal_profit_zar = None
        if profit_usd is not None and tm_wa_fx is not None:
            metal_profit_zar = float(profit_usd) * float(tm_wa_fx)
        elif profit_usd is not None and wa_usdzar is not None:
            metal_profit_zar = float(profit_usd) * float(wa_usdzar)
        exchange_profit_zar = None
        if profit_zar is not None and metal_profit_zar is not None:
            exchange_profit_zar = float(profit_zar) - float(metal_profit_zar)
        elif profit_zar is not None:
            exchange_profit_zar = float(profit_zar)

        latest_dt = tm_state.get("latest_dt") or pmx_state.get("latest_dt")
        month_key = _month_key(latest_dt)
        trade_date = latest_dt.strftime("%Y-%m-%d") if latest_dt else ""

        trade_rows.append(
            {
                "trade_num": tn,
                "trade_date": trade_date,
                "month_key": month_key,
                "month_label": _month_label(month_key),
                "client_weight_g": tm_weight_g,
                "exchange_profit_zar": exchange_profit_zar,
                "metal_profit_zar": metal_profit_zar,
                "total_profit_zar": profit_zar,
                "profit_pct": profit_pct,
                "sell_side_zar": sell_side_zar,
                "buy_side_zar": buy_side_zar,
                "stonex_zar_flow": stonex_zar_flow_row,
                "stonex_traded_g": stonex_traded_g_row,
                "pmx_wa_gold_usd_oz": wa_gold_price,
                "pmx_wa_usdzar": wa_usdzar,
                "trademc_wa_gold_usd_oz": tm_wa_usd_per_oz,
                "trademc_wa_usdzar": tm_wa_fx,
                "matched_oz": matched_oz,
                "unmatched_oz": max(0.0, abs(float(tm_weight_oz)) - float(matched_oz)),
                "metal_gap_g": metal_gap_g,
                "usd_to_cut": usd_to_cut,
                "metal_hedged": metal_hedged,
                "usd_hedged": usd_hedged,
                "hedged": hedged,
                "hedge_status": hedge_status,
                "control_account_g": control_account_g,
                "control_account_zar": control_account_zar,
                "trademc_transactions": tm_state.get("transactions", []),
                "pmx_transactions": pmx_state.get("transactions", []),
            }
        )

    months_map: Dict[str, Dict[str, Any]] = {}
    for row in trade_rows:
        mk = str(row.get("month_key") or "Unknown")
        month_state = months_map.setdefault(
            mk,
            {
                "month_key": mk,
                "month_label": _month_label(mk),
                "exchange_profit_zar": 0.0,
                "metal_profit_zar": 0.0,
                "total_profit_zar": 0.0,
                "trade_count": 0,
                "trades": [],
            },
        )
        month_state["trade_count"] += 1
        month_state["exchange_profit_zar"] += float(_to_num(row.get("exchange_profit_zar")) or 0.0)
        month_state["metal_profit_zar"] += float(_to_num(row.get("metal_profit_zar")) or 0.0)
        month_state["total_profit_zar"] += float(_to_num(row.get("total_profit_zar")) or 0.0)
        month_state["trades"].append(row)

    def _month_sort_key(item: Dict[str, Any]):
        mk = str(item.get("month_key") or "")
        if mk == "Unknown":
            return (0, datetime.min)
        try:
            return (1, datetime.strptime(mk, "%Y-%m"))
        except Exception:
            return (0, datetime.min)

    months = sorted(months_map.values(), key=_month_sort_key, reverse=True)
    for month in months:
        month["trades"] = sorted(
            month.get("trades", []),
            key=lambda r: (str(r.get("trade_date") or ""), str(r.get("trade_num") or "")),
            reverse=True,
        )

    profit_pct_values: List[float] = []
    for row in trade_rows:
        pct = _to_num(row.get("profit_pct"))
        if pct is not None:
            profit_pct_values.append(float(pct))

    average_profit_margin_pct = None
    if profit_pct_values:
        average_profit_margin_pct = float(sum(profit_pct_values) / len(profit_pct_values))

    summary = {
        "months": len(months),
        "trades": int(sum(int(m.get("trade_count", 0)) for m in months)),
        "exchange_profit_zar": float(sum(float(m.get("exchange_profit_zar", 0.0)) for m in months)),
        "metal_profit_zar": float(sum(float(m.get("metal_profit_zar", 0.0)) for m in months)),
        "total_profit_zar": float(sum(float(m.get("total_profit_zar", 0.0)) for m in months)),
        "average_profit_margin_pct": average_profit_margin_pct,
    }

    return {"months": months, "summary": summary}


def build_trading_ticket_pdf(trade_num_value: str,
                             trademc_rows: pd.DataFrame,
                             stonex_rows: pd.DataFrame,
                             summary_rows: pd.DataFrame):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, Image as RLImage, PageBreak, KeepTogether,
        )
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        import html as _html
        import os as _os
    except Exception:
        return _build_trading_ticket_pdf_fpdf(
            trade_num_value,
            trademc_rows,
            stonex_rows,
            summary_rows,
        )

    def _esc(v: str) -> str:
        return _html.escape(str(v), quote=False)

    # ── Font (Calibri → Helvetica fallback) ──────────────────────────
    FONT_REG  = "Helvetica"
    FONT_BOLD = "Helvetica-Bold"
    FONT_IT   = "Helvetica-Oblique"
    try:
        _rp = r"C:\Windows\Fonts\calibri.ttf"
        _bp = r"C:\Windows\Fonts\calibrib.ttf"
        _ip = r"C:\Windows\Fonts\calibrii.ttf"
        if _os.path.isfile(_rp) and _os.path.isfile(_bp):
            pdfmetrics.registerFont(TTFont("Calibri",      _rp))
            pdfmetrics.registerFont(TTFont("Calibri-Bold", _bp))
            FONT_REG  = "Calibri"
            FONT_BOLD = "Calibri-Bold"
            if _os.path.isfile(_ip):
                pdfmetrics.registerFont(TTFont("Calibri-Italic", _ip))
                FONT_IT = "Calibri-Italic"
    except Exception:
        pass

    # ── Brand colours ────────────────────────────────────────────────
    CHARCOAL   = colors.HexColor("#1C1C1C")
    COPPER     = colors.HexColor("#B07840")
    COPPER_LT  = colors.HexColor("#FBF5EC")
    COPPER_MD  = colors.HexColor("#E8D5B7")
    COPPER_DK  = colors.HexColor("#8A5C2E")
    NAVY       = colors.HexColor("#1E2A38")
    CREAM_TXT  = colors.HexColor("#F5EDE3")
    TBL_HDR    = NAVY
    ROW_ALT    = colors.HexColor("#FAFAFA")
    BORDER     = colors.HexColor("#E2E2E2")
    GREY_RULE  = colors.HexColor("#D0D0D0")
    MID_GREY   = colors.HexColor("#888888")
    DARK_GREY  = colors.HexColor("#555555")
    LIGHT_GREY = colors.HexColor("#F4F4F4")
    GREEN      = colors.HexColor("#2D7A4F")
    GREEN_LT   = colors.HexColor("#EBF5F0")
    RED_C      = colors.HexColor("#C0392B")
    RED_LT     = colors.HexColor("#FDECEB")
    WHITE      = colors.white

    # ── Numeric helpers ──────────────────────────────────────────────
    def _s(v) -> str:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return str(v).strip()

    def _n(v) -> Optional[float]:
        try:
            f = float(v)
            return None if pd.isna(f) else f
        except Exception:
            return None

    def _f(v, dp=2, prefix="") -> str:
        n = _n(v)
        return "\u2014" if n is None else f"{prefix}{n:,.{dp}f}"

    def _money(v, prefix="$", dp=2) -> str:
        n = _n(v)
        return "\u2014" if n is None else f"{prefix}{n:,.{dp}f}"

    def _grams(v) -> str:
        n = _n(v)
        return "\u2014" if n is None else f"{n:,.2f} g"

    def _ounces(v, dp=4) -> str:
        n = _n(v)
        return "\u2014" if n is None else f"{n:,.{dp}f} oz"

    def _pct(v, dp=2) -> str:
        n = _n(v)
        return "\u2014" if n is None else f"{n:,.{dp}f}%"

    def _tr(t: str, mx: int) -> str:
        return t if len(t) <= mx else t[:mx - 1] + "\u2026"

    # ── Document ─────────────────────────────────────────────────────
    buf  = BytesIO()
    PAGE = landscape(A4)
    LM = RM = 20 * mm
    TM = 18 * mm
    BM = 16 * mm
    doc  = SimpleDocTemplate(buf, pagesize=PAGE,
                             leftMargin=LM, rightMargin=RM,
                             topMargin=TM, bottomMargin=BM)
    PW = PAGE[0] - LM - RM    # usable width

    # ── Style factory ────────────────────────────────────────────────
    _style_count = [0]
    def _ps(name, **kw) -> ParagraphStyle:
        _style_count[0] += 1
        uname = f"{name}_{_style_count[0]}"
        return ParagraphStyle(
            uname,
            fontName=kw.pop("fontName", FONT_REG),
            fontSize=kw.pop("fontSize", 8),
            leading=kw.pop("leading", 11),
            textColor=kw.pop("textColor", CHARCOAL),
            **kw,
        )

    generated_at = datetime.now().strftime("%d %B %Y at %H:%M")
    story = []

    # ── Logo ─────────────────────────────────────────────────────────
    _logo_path = _os.path.join(
        _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
        "MetCon Logo.png",
    )
    _logo_cell: object = Spacer(60 * mm, 16 * mm)
    try:
        if _os.path.isfile(_logo_path):
            from PIL import Image as _PILImage
            with _PILImage.open(_logo_path) as _pimg:
                _pimg.load()
            _logo_cell = RLImage(_logo_path, width=60 * mm, height=16 * mm)
    except Exception:
        pass

    # ════════════════════════════════════════════════════════════════
    # HEADER — logo left | title block right
    # ════════════════════════════════════════════════════════════════
    title_para = Paragraph(
        f"<font name='{FONT_BOLD}' size='8' color='#B07840'>METAL CONCENTRATORS SA</font><br/>"
        f"<font name='{FONT_BOLD}' size='22' color='#1C1C1C'>Trade Breakdown Report</font><br/>"
        f"<font name='{FONT_BOLD}' size='12' color='#B07840'>Trade #{_esc(_s(trade_num_value))}</font>",
        _ps("_hdr", leading=26, alignment=TA_RIGHT),
    )
    date_para = Paragraph(
        f"<font size='7.5' color='#888888'>Report generated on {_esc(generated_at)}</font>",
        _ps("_hdr_date", fontSize=7.5, textColor=MID_GREY, alignment=TA_RIGHT),
    )
    hdr = Table([[_logo_cell, title_para]], colWidths=[PW * 0.38, PW * 0.62])
    hdr.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "BOTTOM"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(hdr)
    story.append(date_para)
    story.append(Spacer(1, 3 * mm))
    # Double-line header separator
    story.append(HRFlowable(width=PW, thickness=3, color=COPPER,
                           spaceBefore=0, spaceAfter=0.8 * mm))
    story.append(HRFlowable(width=PW, thickness=0.75, color=COPPER_MD,
                           spaceBefore=0, spaceAfter=6 * mm))

    # ════════════════════════════════════════════════════════════════
    # REUSABLE HELPERS
    # ════════════════════════════════════════════════════════════════
    def _sec_lbl(title: str, description: str = ""):
        """Section header with copper left bar, title, and optional description."""
        inner_rows = [[Paragraph(
            f"<font name='{FONT_BOLD}' size='10' color='#1C1C1C'>{_esc(title.upper())}</font>",
            _ps("_sl", fontName=FONT_BOLD, fontSize=10, textColor=CHARCOAL, leading=14),
        )]]
        if description:
            inner_rows.append([Paragraph(
                f"<font size='7.5' color='#777777'>{_esc(description)}</font>",
                _ps("_sld", fontSize=7.5, textColor=MID_GREY, leading=10),
            )])
        inner = Table(inner_rows, colWidths=[PW - 8])
        inner.setStyle(TableStyle([
            ("LEFTPADDING",   (0, 0), (-1, -1), 0),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
            ("TOPPADDING",    (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]))
        lbl = Table([[Spacer(1, 1), inner]], colWidths=[5, PW - 5])
        lbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (0, 0), COPPER),
            ("BACKGROUND",    (1, 0), (1, 0), COPPER_LT),
            ("LEFTPADDING",   (0, 0), (0, 0), 0),
            ("RIGHTPADDING",  (0, 0), (0, 0), 0),
            ("LEFTPADDING",   (1, 0), (-1, -1), 12),
            ("RIGHTPADDING",  (1, 0), (-1, -1), 12),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(lbl)
        story.append(Spacer(1, 4 * mm))

    def _audit_step_header(step_num: str, title: str, subtitle: str = ""):
        """Numbered audit step header with copper badge and description."""
        badge = Paragraph(
            f"<font name='{FONT_BOLD}' size='10' color='#FFFFFF'>\u00a0{_esc(step_num)}\u00a0</font>",
            _ps("_badge", fontName=FONT_BOLD, fontSize=10, textColor=WHITE, alignment=TA_CENTER),
        )
        title_p = Paragraph(
            f"<font name='{FONT_BOLD}' size='11' color='#1C1C1C'>{_esc(title)}</font>",
            _ps("_ash", fontName=FONT_BOLD, fontSize=11, leading=15),
        )
        row_tbl = Table([[badge, title_p]], colWidths=[26 * mm, PW - 26 * mm])
        row_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (0, 0), COPPER),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",   (0, 0), (0, 0), 6),
            ("RIGHTPADDING",  (0, 0), (0, 0), 6),
            ("LEFTPADDING",   (1, 0), (1, 0), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(row_tbl)
        if subtitle:
            story.append(Spacer(1, 2 * mm))
            story.append(Paragraph(
                f"<font name='{FONT_IT}' size='7.5' color='#666666'>{_esc(subtitle)}</font>",
                _ps("_ass", fontName=FONT_IT, fontSize=7.5, textColor=DARK_GREY, leftIndent=8, rightIndent=8),
            ))
        story.append(Spacer(1, 3 * mm))
        story.append(HRFlowable(width=PW, thickness=0.5, color=COPPER_MD,
                                spaceBefore=0, spaceAfter=3.5 * mm))

    def _audit_row(label: str, value: str, bold_value: bool = False,
                   value_color=None, indent: int = 0):
        """Key-value audit row with clean separator."""
        vc = value_color or CHARCOAL
        lpad = 12 + indent * 16
        lbl_cell = Paragraph(
            f"<font size='8' color='#444444'>{_esc(label)}</font>",
            _ps("_arl", fontSize=8, textColor=DARK_GREY, leading=11),
        )
        val_font = FONT_BOLD if bold_value else FONT_REG
        val_cell = Paragraph(
            f"<font name='{val_font}' size='8.5'>{_esc(value)}</font>",
            _ps("_arv", fontName=val_font, fontSize=8.5, textColor=vc,
                alignment=TA_RIGHT, leading=12),
        )
        tbl = Table([[lbl_cell, val_cell]], colWidths=[PW * 0.55, PW * 0.45])
        tbl.setStyle(TableStyle([
            ("LEFTPADDING",   (0, 0), (0, 0), lpad),
            ("RIGHTPADDING",  (-1, 0), (-1, 0), 12),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.25, BORDER),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(tbl)

    def _audit_row_formula(label: str, formula: str, result: str,
                           value_color=None, indent: int = 0):
        """Three-column audit row: label | formula \u2192 result."""
        vc = value_color or CHARCOAL
        lpad = 12 + indent * 16
        lbl_cell = Paragraph(
            f"<font size='7.5' color='#444444'>{_esc(label)}</font>",
            _ps("_arf_l", fontSize=7.5, textColor=DARK_GREY, leading=10),
        )
        formula_cell = Paragraph(
            f"<font name='{FONT_IT}' size='7' color='#999999'>{_esc(formula)}  \u2192</font>",
            _ps("_arf_f", fontName=FONT_IT, fontSize=7, textColor=MID_GREY, alignment=TA_RIGHT, leading=10),
        )
        val_cell = Paragraph(
            f"<font name='{FONT_REG}' size='8.5'>{_esc(result)}</font>",
            _ps("_arf_v", fontName=FONT_REG, fontSize=8.5, textColor=vc,
                alignment=TA_RIGHT, leading=12),
        )
        tbl = Table([[lbl_cell, formula_cell, val_cell]],
                    colWidths=[PW * 0.28, PW * 0.42, PW * 0.30])
        tbl.setStyle(TableStyle([
            ("LEFTPADDING",   (0, 0), (0, 0), lpad),
            ("RIGHTPADDING",  (-1, 0), (-1, 0), 12),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.25, BORDER),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(tbl)

    def _audit_sub_title(title: str):
        """Sub-section title within an audit step with left accent."""
        tbl = Table([[Paragraph(
            f"<font name='{FONT_BOLD}' size='8' color='#2A2A2A'>{_esc(title)}</font>",
            _ps("_ast", fontName=FONT_BOLD, fontSize=8, textColor=CHARCOAL),
        )]], colWidths=[PW])
        tbl.setStyle(TableStyle([
            ("LEFTPADDING",   (0, 0), (-1, -1), 12),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND",    (0, 0), (-1, -1), LIGHT_GREY),
            ("LINEBEFORE",    (0, 0), (0, 0), 2.5, COPPER),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.3, BORDER),
        ]))
        story.append(tbl)

    _pos_st = _ps("_pos", fontName=FONT_BOLD, fontSize=8.5, textColor=GREEN,  alignment=TA_RIGHT)
    _neg_st = _ps("_neg", fontName=FONT_BOLD, fontSize=8.5, textColor=RED_C,  alignment=TA_RIGHT)

    def _data_tbl(col_names, col_pcts, rows, right_cols=None, profit_col=None, totals=None):
        """Professional striped data table with clear hierarchy and optional totals row."""
        right_cols = right_cols or []
        if not rows:
            story.append(Paragraph(
                "<font color='#999999'>\u00a0\u00a0No data available for this section.</font>",
                _ps("_nd", textColor=MID_GREY, fontSize=8),
            ))
            story.append(Spacer(1, 5 * mm))
            return

        cw = [PW * p for p in col_pcts]

        hdr_row = [
            Paragraph(
                _esc(col),
                _ps(f"_th", fontName=FONT_BOLD, fontSize=7.5, leading=10,
                    textColor=CREAM_TXT, alignment=TA_RIGHT if ci in right_cols else TA_CENTER),
            )
            for ci, col in enumerate(col_names)
        ]
        tdata = [hdr_row]

        for row in rows:
            cells = []
            for ci, col in enumerate(col_names):
                raw = row.get(col, "")
                txt = _esc(_s(raw)) if raw != "" else "\u2014"
                if col == profit_col:
                    nv = _n(raw)
                    cells.append(Paragraph(
                        txt if txt != "\u2014" else "",
                        _pos_st if (nv is not None and nv >= 0) else _neg_st,
                    ))
                elif ci in right_cols:
                    cells.append(Paragraph(txt, _ps(f"_tdr", fontSize=8, alignment=TA_RIGHT)))
                else:
                    cells.append(Paragraph(_esc(_tr(_s(raw), 42)), _ps(f"_td", fontSize=8)))
            tdata.append(cells)

        if totals:
            tot_cells = []
            for ci, col in enumerate(col_names):
                raw = totals.get(col, "")
                txt = _esc(_s(raw)) if raw != "" else ""
                tot_cells.append(Paragraph(txt, _ps(
                    f"_tot", fontName=FONT_BOLD, fontSize=8,
                    alignment=TA_RIGHT if ci in right_cols else TA_LEFT,
                )))
            tdata.append(tot_cells)

        tbl = Table(tdata, colWidths=cw, repeatRows=1)
        ts = TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), TBL_HDR),
            ("LINEBELOW",     (0, 0), (-1, 0), 1.6, COPPER),
            ("TOPPADDING",    (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING",    (0, 1), (-1, -1), 4.5),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4.5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5.5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5.5),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW",     (0, 1), (-1, -1), 0.3, BORDER),
        ])
        for ri in range(1, len(tdata)):
            if totals and ri == len(tdata) - 1:
                ts.add("BACKGROUND", (0, ri), (-1, ri), COPPER_LT)
                ts.add("LINEABOVE",  (0, ri), (-1, ri), 1.5, COPPER_MD)
            else:
                ts.add("BACKGROUND", (0, ri), (-1, ri), ROW_ALT if ri % 2 == 0 else WHITE)
        tbl.setStyle(ts)
        story.append(tbl)
        story.append(Spacer(1, 4 * mm))

    def _append_source_dump(block_title: str, df: pd.DataFrame, preferred_cols: Optional[List[str]] = None):
        """Append every available source field/value for full reconciliation traceability."""
        _audit_sub_title(block_title)
        if df is None or df.empty:
            _audit_row("Rows", "0")
            return

        cols = list(df.columns)
        ordered_cols = cols
        if preferred_cols:
            ordered_cols = [c for c in preferred_cols if c in cols] + [c for c in cols if c not in preferred_cols]

        _audit_row("Rows", str(len(df)))
        for ridx, (_, row) in enumerate(df.iterrows(), start=1):
            _audit_sub_title(f"{block_title} — Row {ridx}")
            for col in ordered_cols:
                raw_val = row.get(col)
                if raw_val is None:
                    text_val = "—"
                else:
                    text_val = _s(raw_val)
                    if text_val == "":
                        text_val = "—"
                text_val = _tr(text_val, 180)
                _audit_row(str(col), text_val, indent=1)
            story.append(Spacer(1, 1.5 * mm))

    # ════════════════════════════════════════════════════════════════
    # PRE-COMPUTE ALL AUDIT TRAIL DATA
    # ════════════════════════════════════════════════════════════════
    tm_pdf = trademc_rows.copy() if trademc_rows is not None else pd.DataFrame()
    stx_pdf = stonex_rows.copy() if stonex_rows is not None else pd.DataFrame()
    sum_pdf = summary_rows.copy() if summary_rows is not None else pd.DataFrame()

    # -- TradeMC bookings with refining --
    bookings = []
    tm_total_g = tm_total_oz = tm_total_usd = tm_total_zar_gross = tm_total_zar_net = 0.0
    if not tm_pdf.empty:
        for _, r in tm_pdf.iterrows():
            weight_g = _n(r.get("Weight (g)")) or 0.0
            weight_oz = _n(r.get("Weight (oz)")) or (weight_g / 31.1035)
            booked_price = _n(r.get("$/oz Booked"))
            fx_rate = _n(r.get("FX Rate"))
            usd_val_raw = _n(r.get("USD Value"))
            usd_val = usd_val_raw if usd_val_raw is not None else (
                weight_oz * booked_price if booked_price is not None else None)
            zar_gross_raw = _n(r.get("ZAR Value"))
            zar_gross = zar_gross_raw if zar_gross_raw is not None else (
                usd_val * fx_rate if usd_val is not None and fx_rate is not None else None)
            refining_rate = _n(r.get("company_refining_rate")) or 0.0
            refining_deduction = zar_gross * (refining_rate / 100.0) if zar_gross is not None else None
            zar_net_raw = _n(r.get("zar_value_less_refining"))
            zar_net = zar_net_raw if zar_net_raw is not None else (
                zar_gross * (1.0 - refining_rate / 100.0) if zar_gross is not None else None)
            company = _s(r.get("Company")) or _s(r.get("company_name")) or "Unknown"
            bookings.append({
                "company": company, "weight_g": weight_g, "weight_oz": weight_oz,
                "booked_price": booked_price, "fx_rate": fx_rate,
                "usd_value": usd_val, "zar_gross": zar_gross,
                "refining_rate": refining_rate, "refining_deduction": refining_deduction,
                "zar_net": zar_net,
            })
            tm_total_g += weight_g
            tm_total_oz += weight_oz
            tm_total_usd += usd_val or 0.0
            tm_total_zar_gross += zar_gross or 0.0
            tm_total_zar_net += zar_net or 0.0

    # -- StoneX trade classification --
    xau_trades = []
    fx_trades = []
    all_pmx = []
    if not stx_pdf.empty:
        for _, r in stx_pdf.iterrows():
            sym = str(r.get("Symbol", "")).upper().replace("/", "").replace("-", "").replace(" ", "")
            side = str(r.get("Side", "")).upper().strip()
            qty = abs(_n(r.get("Quantity")) or 0.0)
            price = _n(r.get("Price"))
            notional = qty * price if price is not None and qty > 0 else None
            trade_info = {"symbol": sym, "side": side, "qty": qty, "price": price,
                          "notional": notional,
                          "trade_date": _s(r.get("Trade Date")),
                          "value_date": _s(r.get("Value Date")),
                          "fnc": _s(r.get("FNC #")), "narration": _s(r.get("Narration"))}
            if qty > 0:
                all_pmx.append(trade_info)
                if sym == "XAUUSD":
                    xau_trades.append(trade_info)
                elif sym == "USDZAR":
                    fx_trades.append(trade_info)

    # -- Weighted averages --
    gold_wa_total_notional = sum(t["notional"] or 0 for t in xau_trades)
    gold_wa_total_qty = sum(t["qty"] for t in xau_trades)
    gold_wa = gold_wa_total_notional / gold_wa_total_qty if gold_wa_total_qty > 1e-9 else None

    fx_wa_total_notional = sum(t["notional"] or 0 for t in fx_trades)
    fx_wa_total_qty = sum(t["qty"] for t in fx_trades)
    fx_wa = fx_wa_total_notional / fx_wa_total_qty if fx_wa_total_qty > 1e-9 else None

    # -- Spot ZAR/g --
    spot_zar_per_g = (gold_wa * fx_wa) / 31.1035 if gold_wa and fx_wa else None

    # -- Cash flows --
    xau_cash_flows = []
    net_stonex_usd = 0.0
    for t in xau_trades:
        signed = None
        if t["notional"] is not None:
            signed = t["notional"] if t["side"] == "SELL" else -t["notional"]
        xau_cash_flows.append({**t, "signed": signed})
        net_stonex_usd += signed or 0.0

    fx_cash_flows = []
    net_fx_zar = 0.0
    for t in fx_trades:
        signed = None
        if t["notional"] is not None:
            signed = t["notional"] if t["side"] == "SELL" else -t["notional"]
        fx_cash_flows.append({**t, "signed": signed})
        net_fx_zar += signed or 0.0

    # -- Summary data --
    base = sum_pdf.iloc[0].to_dict() if not sum_pdf.empty else {}
    sell_usd = _n(base.get("Sell Side (USD)"))
    buy_usd = _n(base.get("Buy Side (USD)"))
    sell_zar = _n(base.get("Sell Side (ZAR)"))
    buy_zar = _n(base.get("Buy Side (ZAR)"))
    profit_usd = _n(base.get("Profit (USD)"))
    profit_zar = _n(base.get("Profit (ZAR)"))
    profit_pct = _n(base.get("Profit % (ZAR Spot Cost)"))
    ctrl_g = _n(base.get("Control Account (g)"))
    ctrl_oz = _n(base.get("Control Account (oz)"))
    ctrl_zar = _n(base.get("Control Account (ZAR)"))
    total_traded_g = _n(base.get("Total Traded (g)"))
    total_traded_oz = _n(base.get("Total Traded (oz)"))
    stonex_zar_flow = _n(base.get("StoneX ZAR Flow"))
    sum_gold_wa = _n(base.get("Gold WA $/oz"))
    sum_fx_wa = _n(base.get("FX WA USD/ZAR"))
    sum_spot_zar_g = _n(base.get("Spot ZAR/g"))

    # Use summary values if available, else computed
    gold_wa_display = sum_gold_wa if sum_gold_wa is not None else gold_wa
    fx_wa_display = sum_fx_wa if sum_fx_wa is not None else fx_wa
    spot_display = sum_spot_zar_g if sum_spot_zar_g is not None else spot_zar_per_g
    gold_wa_total_notional_display = gold_wa_total_notional
    if gold_wa_display is not None and gold_wa_total_qty > 1e-9:
        gold_wa_total_notional_display = float(gold_wa_display) * float(gold_wa_total_qty)
    fx_wa_total_notional_display = fx_wa_total_notional
    if fx_wa_display is not None and fx_wa_total_qty > 1e-9:
        fx_wa_total_notional_display = float(fx_wa_display) * float(fx_wa_total_qty)

    if profit_zar is None and sell_zar is not None and buy_zar is not None:
        profit_zar = sell_zar - buy_zar
    if profit_usd is None and sell_usd is not None and buy_usd is not None:
        profit_usd = sell_usd - buy_usd
    if profit_pct is None and profit_zar is not None and buy_zar and abs(buy_zar) > 1e-12:
        profit_pct = (profit_zar / buy_zar) * 100.0
    if ctrl_oz is None and ctrl_g is not None:
        ctrl_oz = ctrl_g / 31.1035
    if total_traded_oz is None and total_traded_g is not None:
        total_traded_oz = total_traded_g / 31.1035

    stonex_g = total_traded_g or (total_traded_oz * 31.1035 if total_traded_oz else None)

    # ════════════════════════════════════════════════════════════════
    # SECTION 1 — EXECUTIVE SUMMARY (KPI dashboard)
    # ════════════════════════════════════════════════════════════════
    _sec_lbl("Executive Summary",
             "Key performance indicators for this trade at a glance.")

    # -- Primary KPI row: profit-focused --
    def _make_kpi_card(label: str, value: str, bg_color=WHITE,
                       value_color=CHARCOAL, border_color=BORDER,
                       accent_color=COPPER):
        lbl_p = Paragraph(
            f"<font name='{FONT_BOLD}' size='6.5' color='#777777'>{_esc(label.upper())}</font>",
            _ps("_kl", fontName=FONT_BOLD, fontSize=6.5, textColor=MID_GREY,
                alignment=TA_CENTER, leading=9),
        )
        val_p = Paragraph(
            f"<font name='{FONT_BOLD}' size='13'>{_esc(value)}</font>",
            _ps("_kv", fontName=FONT_BOLD, fontSize=13, leading=17,
                textColor=value_color, alignment=TA_CENTER),
        )
        card = Table([[lbl_p], [val_p]], colWidths=[PW / 4 - 5 * mm])
        card.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), bg_color),
            ("BOX",           (0, 0), (-1, -1), 0.5, border_color),
            ("LINEABOVE",     (0, 0), (-1, 0), 3, accent_color),
            ("TOPPADDING",    (0, 0), (-1, 0),  8),
            ("BOTTOMPADDING", (0, -1), (-1, -1), 10),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return card

    # Profit color logic
    profit_color = GREEN if (profit_zar is not None and profit_zar >= 0) else RED_C
    profit_bg = GREEN_LT if (profit_zar is not None and profit_zar >= 0) else RED_LT

    kpi_row = [
        _make_kpi_card("Total Weight Traded",
                       _grams(stonex_g)),
        _make_kpi_card("Gold WA Price",
                       _money(gold_wa_display, "$", 3) + "/oz" if gold_wa_display else "\u2014"),
        _make_kpi_card("Weighted Avg FX Rate",
                       f"R\u00a0{fx_wa_display:,.4f}" if fx_wa_display else "\u2014"),
        _make_kpi_card("Net Profit (ZAR)",
                       _money(profit_zar, "R\u00a0"),
                       bg_color=profit_bg, value_color=profit_color,
                       border_color=profit_color, accent_color=profit_color),
    ]
    kpi_table = Table([kpi_row], colWidths=[PW / 4] * 4)
    kpi_table.setStyle(TableStyle([
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 2),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 3 * mm))

    # -- Secondary KPI row --
    kpi_row2_items = []
    if spot_display is not None:
        kpi_row2_items.append(("Implied Spot ZAR/g", f"R\u00a0{spot_display:,.4f}"))
    if stonex_zar_flow is not None:
        kpi_row2_items.append(("StoneX ZAR Inflow", _money(stonex_zar_flow, "R\u00a0")))
    if sell_usd is not None:
        kpi_row2_items.append(("Sell Side (USD)", _money(sell_usd, "$")))
    if buy_usd is not None:
        kpi_row2_items.append(("Buy Side (USD)", _money(buy_usd, "$")))
    if profit_pct is not None:
        kpi_row2_items.append(("Profit Margin", _pct(profit_pct, 3)))
    if ctrl_g is not None:
        kpi_row2_items.append(("Control Account", _grams(ctrl_g)))

    if kpi_row2_items:
        n_items = len(kpi_row2_items)
        cw_each = PW / max(n_items, 1)
        lbl_cells = []
        val_cells = []
        for lbl, val in kpi_row2_items:
            lbl_cells.append(Paragraph(
                f"<font name='{FONT_BOLD}' size='6' color='#999999'>{_esc(lbl.upper())}</font>",
                _ps("_k2l", fontName=FONT_BOLD, fontSize=6, textColor=MID_GREY,
                    alignment=TA_CENTER, leading=8),
            ))
            val_cells.append(Paragraph(
                f"<font name='{FONT_BOLD}' size='9.5'>{_esc(val)}</font>",
                _ps("_k2v", fontName=FONT_BOLD, fontSize=9.5, leading=13,
                    textColor=CHARCOAL, alignment=TA_CENTER),
            ))
        k2_tbl = Table([lbl_cells, val_cells], colWidths=[cw_each] * n_items)
        k2_tbl.setStyle(TableStyle([
            ("TOPPADDING",    (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 3),
            ("TOPPADDING",    (0, 1), (-1, 1), 3),
            ("BOTTOMPADDING", (0, 1), (-1, 1), 7),
            ("LEFTPADDING",   (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("BOX",           (0, 0), (-1, -1), 0.3, BORDER),
            ("LINEBELOW",     (0, 0), (-1, 0),  0.3, BORDER),
            ("LINEBEFORE",    (1, 0), (-1, -1), 0.3, BORDER),
            ("BACKGROUND",    (0, 0), (-1, -1), LIGHT_GREY),
        ]))
        story.append(k2_tbl)
    story.append(Spacer(1, 7 * mm))

    # ════════════════════════════════════════════════════════════════
    # SECTION 2 — TradeMC Client Bookings (with refining columns)
    # ════════════════════════════════════════════════════════════════
    _sec_lbl("Client Bookings \u2014 Buy Side",
             "Gold purchased from TradeMC clients, valued in USD, converted to ZAR, and adjusted for refining deductions.")

    TM_COLS = ["Company", "Weight (g)", "Weight (oz)", "$/oz Booked", "FX Rate",
               "USD Value", "ZAR Gross", "Refining %", "ZAR Net"]
    _TM_W = {"Company": 0.19, "Weight (g)": 0.09, "Weight (oz)": 0.09,
             "$/oz Booked": 0.10, "FX Rate": 0.09, "USD Value": 0.11,
             "ZAR Gross": 0.11, "Refining %": 0.09, "ZAR Net": 0.13}

    tm_table_rows = []
    for b in bookings:
        tm_table_rows.append({
            "Company": b["company"],
            "Weight (g)": _f(b["weight_g"], 2),
            "Weight (oz)": _f(b["weight_oz"], 4),
            "$/oz Booked": _f(b["booked_price"], 2),
            "FX Rate": _f(b["fx_rate"], 4),
            "USD Value": _f(b["usd_value"], 2),
            "ZAR Gross": _f(b["zar_gross"], 2),
            "Refining %": _pct(b["refining_rate"]),
            "ZAR Net": _f(b["zar_net"], 2),
        })
    tm_right = list(range(1, len(TM_COLS)))  # all columns except Company
    tm_wids = [_TM_W[c] for c in TM_COLS]

    tm_totals_row = {
        "Company": "TOTAL",
        "Weight (g)": f"{tm_total_g:,.2f}",
        "Weight (oz)": f"{tm_total_oz:,.4f}",
        "USD Value": f"{tm_total_usd:,.2f}",
        "ZAR Gross": f"{tm_total_zar_gross:,.2f}",
        "ZAR Net": f"{tm_total_zar_net:,.2f}",
    } if bookings else None

    _data_tbl(TM_COLS, tm_wids, tm_table_rows,
              right_cols=tm_right, totals=tm_totals_row)

    # ════════════════════════════════════════════════════════════════
    # SECTION 3 — PMX / StoneX Trades (Sell Side)
    # ════════════════════════════════════════════════════════════════
    _sec_lbl("StoneX / PMX Trades \u2014 Sell Side",
             "Executed PMX trades used to derive weighted average gold and FX rates.")

    STX_COLS = ["Doc #", "FNC #", "Trade Date", "Value Date", "Symbol", "Side", "Narration", "Quantity", "Price", "Settle Currency", "Settle Amount"]
    stx_avail = [c for c in STX_COLS if c in stx_pdf.columns]
    _STX_W = {"Doc #": 0.08, "FNC #": 0.10, "Trade Date": 0.08, "Value Date": 0.08, "Symbol": 0.06,
              "Side": 0.05, "Narration": 0.18, "Quantity": 0.11, "Price": 0.10,
              "Settle Currency": 0.06, "Settle Amount": 0.10}
    _sw = sum(_STX_W.get(c, 0.10) for c in stx_avail) or 1
    stx_wids = [_STX_W.get(c, 0.10) / _sw for c in stx_avail]
    stx_right = [i for i, c in enumerate(stx_avail) if c in ("Quantity", "Price", "Settle Amount")]

    stx_rows_data: list = []
    for _, r in stx_pdf.iterrows():
        rd: Dict[str, str] = {}
        for c in stx_avail:
            if c == "Quantity":       rd[c] = _f(r.get(c), 2)
            elif c == "Price":        rd[c] = _f(r.get(c), 4)
            elif c == "Settle Amount": rd[c] = _f(r.get(c), 2)
            else:                     rd[c] = _s(r.get(c))
        stx_rows_data.append(rd)

    _data_tbl(stx_avail, stx_wids, stx_rows_data, right_cols=stx_right)

    # ════════════════════════════════════════════════════════════════
    # PAGE BREAK — AUDIT TRAIL
    # ════════════════════════════════════════════════════════════════
    story.append(PageBreak())

    # Audit trail header
    audit_hdr = Paragraph(
        f"<font name='{FONT_BOLD}' size='16' color='#1C1C1C'>TRADE BREAKDOWN AUDIT TRAIL</font>",
        _ps("_audit_hdr", fontName=FONT_BOLD, fontSize=16, leading=20),
    )
    audit_sub = Paragraph(
        f"<font name='{FONT_BOLD}' size='9' color='#B07840'>Trade #{_esc(_s(trade_num_value))}</font>"
        f"<font size='8' color='#888888'>  \u2014  Complete calculation breakdown for reconciliation and compliance review</font>",
        _ps("_audit_sub", fontSize=9, leading=13),
    )
    story.append(audit_hdr)
    story.append(Spacer(1, 1.5 * mm))
    story.append(audit_sub)
    story.append(Spacer(1, 2 * mm))
    story.append(HRFlowable(width=PW, thickness=2.5, color=COPPER,
                            spaceBefore=0, spaceAfter=0.5 * mm))
    story.append(HRFlowable(width=PW, thickness=0.5, color=COPPER_MD,
                            spaceBefore=0, spaceAfter=5 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 0 — Outcome Snapshot
    # ──────────────────────────────────────────────────────────────
    _audit_step_header(
        "Step 0.",
        "Outcome Snapshot",
        "Top-level results for this trade as shown on the Trade Breakdown page.",
    )
    _audit_row("Sell Side (USD)", _money(sell_usd, "$"))
    _audit_row("Buy Side (USD)", _money(buy_usd, "$"))
    _audit_row("Profit (USD)", _money(profit_usd, "$"),
               bold_value=True, value_color=GREEN if (profit_usd or 0) >= 0 else RED_C)
    _audit_row("Sell Side (ZAR)", _money(sell_zar, "R "))
    _audit_row("Buy Side (ZAR)", _money(buy_zar, "R "))
    _audit_row("Profit (ZAR)", _money(profit_zar, "R "),
               bold_value=True, value_color=GREEN if (profit_zar or 0) >= 0 else RED_C)
    _audit_row("Control Account (g)", _grams(ctrl_g))
    _audit_row("Control Account (oz)", _ounces(ctrl_oz, 4))
    _audit_row("Control Account (ZAR)", _money(ctrl_zar, "R "))
    _audit_row("Total Traded (g)", _grams(total_traded_g))
    _audit_row("Total Traded (oz)", _ounces(total_traded_oz, 4))
    _audit_row("StoneX ZAR Flow", _money(stonex_zar_flow, "R "))
    _audit_row("Profit Margin", _pct(profit_pct, 3), bold_value=True, value_color=COPPER)
    story.append(Spacer(1, 4 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 1 — Input Data Summary
    # ──────────────────────────────────────────────────────────────
    _audit_step_header("Step 1.", "Input Data Summary",
                       "Source data counts loaded from TradeMC (buy side) and StoneX/PMX (sell side) for this ticket.")
    _audit_row("TradeMC bookings (buy side)", str(len(bookings)))
    _audit_row("StoneX/PMX trades (sell side)", str(len(all_pmx)))
    _audit_row("XAU/USD (gold) trades", str(len(xau_trades)))
    _audit_row("USD/ZAR (FX) trades", str(len(fx_trades)))
    _audit_row("Conversion constant", "31.1035 g/troy oz")
    story.append(Spacer(1, 4 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 2 — TradeMC Valuation (Buy Side)
    # ──────────────────────────────────────────────────────────────
    _audit_step_header("Step 2.", "TradeMC Valuation (Buy Side)",
                       "Each client booking is converted: weight in grams \u2192 troy ounces \u2192 USD value (at booked $/oz) \u2192 ZAR value (at FX rate), then reduced by the refining charge.")

    for idx, b in enumerate(bookings):
        _audit_sub_title(f"Booking {idx + 1}: {b['company']}")
        _audit_row("Weight", _grams(b["weight_g"]), indent=1)
        _audit_row_formula("Troy ounces",
                           f"{_f(b['weight_g'], 2)} / 31.1035",
                           _ounces(b["weight_oz"], 6), indent=1)
        _audit_row("Booked gold price", _money(b["booked_price"], "$") + " /oz", indent=1)
        _audit_row_formula("USD value",
                           f"{_ounces(b['weight_oz'], 6)} x {_money(b['booked_price'], '$')}",
                           _money(b["usd_value"], "$"), indent=1)
        _audit_row("FX rate (ZAR/USD)", _f(b["fx_rate"], 4), indent=1)
        _audit_row_formula("ZAR value (gross)",
                           f"{_money(b['usd_value'], '$')} x {_f(b['fx_rate'], 4)}",
                           _money(b["zar_gross"], "R\u00a0"), indent=1)
        _audit_row("Refining rate", _pct(b["refining_rate"]), indent=1)
        _audit_row_formula("Refining deduction",
                           f"{_money(b['zar_gross'], 'R ')} x {_pct(b['refining_rate'])}",
                           _money(b["refining_deduction"], "R\u00a0"), indent=1)
        _audit_row("ZAR value (net of refining)", _money(b["zar_net"], "R\u00a0"),
                   indent=1)

    _audit_sub_title("Buy Side Totals")
    _audit_row("Total weight", f"{_grams(tm_total_g)} ({_ounces(tm_total_oz, 6)})")
    _audit_row("Total USD value", _money(tm_total_usd, "$"))
    _audit_row("Total ZAR (gross)", _money(tm_total_zar_gross, "R\u00a0"))
    _audit_row("Total ZAR (net of refining)", _money(tm_total_zar_net, "R\u00a0"))
    story.append(Spacer(1, 4 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 3 — StoneX Weighted Averages
    # ──────────────────────────────────────────────────────────────
    _audit_step_header("Step 3.", "StoneX Weighted Averages",
                       "The weighted average price is calculated as the total notional value divided by total quantity, ensuring larger trades carry proportional weight.")

    _audit_sub_title("Gold Weighted Average ($/oz)")
    if xau_trades:
        for idx, t in enumerate(xau_trades):
            lbl_parts = [f"{t['side']} {_ounces(t['qty'], 4)} @ {_money(t['price'], '$', 4)}"]
            if t.get("fnc"):
                lbl_parts.append(f"FNC: {t['fnc']}")
            if t.get("trade_date"):
                lbl_parts.append(t["trade_date"])
            if t.get("narration"):
                lbl_parts.append(_tr(t["narration"], 50))
            _audit_row_formula(
                " \u2014 ".join(lbl_parts),
                f"{_ounces(t['qty'], 4)} x {_money(t['price'], '$', 4)}",
                _money(t["notional"], "$"),
                indent=1,
            )
        _audit_row("Sum of notional values", _money(gold_wa_total_notional_display, "$"))
        _audit_row("Sum of quantities", _ounces(gold_wa_total_qty, 4))
        _audit_row_formula("Gold Weighted Average",
                           f"{_money(gold_wa_total_notional_display, '$')} / {_ounces(gold_wa_total_qty, 4)}",
                           _money(gold_wa_display, "$", 4),
                           value_color=COPPER)
    else:
        _audit_row("XAU/USD trades", "\u2014")

    story.append(Spacer(1, 2 * mm))

    _audit_sub_title("FX Weighted Average (ZAR/USD)")
    rand_nbsp = "R\u00a0"
    if fx_trades:
        for idx, t in enumerate(fx_trades):
            lbl_parts = [f"{t['side']} {_money(t['qty'], '$', 2)} @ {_money(t['price'], rand_nbsp, 4)}"]
            if t.get("fnc"):
                lbl_parts.append(f"FNC: {t['fnc']}")
            if t.get("trade_date"):
                lbl_parts.append(t["trade_date"])
            if t.get("narration"):
                lbl_parts.append(_tr(t["narration"], 50))
            _audit_row_formula(
                " \u2014 ".join(lbl_parts),
                f"{_money(t['qty'], '$', 2)} x {_money(t['price'], rand_nbsp, 4)}",
                _money(t["notional"], "R\u00a0"),
                indent=1,
            )
        _audit_row("Sum of notional values", _money(fx_wa_total_notional_display, "R\u00a0"))
        _audit_row("Sum of quantities", _money(fx_wa_total_qty, "$", 2))
        _audit_row_formula("FX Weighted Average",
                           f"{_money(fx_wa_total_notional_display, 'R ')} / {_money(fx_wa_total_qty, '$', 2)}",
                           _money(fx_wa_display, "R\u00a0", 4),
                           value_color=COPPER)
    else:
        _audit_row("USD/ZAR trades", "\u2014")
    story.append(Spacer(1, 4 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 4 — Spot Rate Derivation
    # ──────────────────────────────────────────────────────────────
    _audit_step_header("Step 4.", "Spot Rate Derivation",
                       "Derives the effective ZAR cost per gram by combining the gold WA ($/oz) with the FX WA (ZAR/USD) and converting from troy ounces to grams.")
    _audit_row_formula("Spot ZAR per gram",
                       f"({_money(gold_wa_display, '$', 4)} x {_money(fx_wa_display, 'R ', 4)}) / 31.1035",
                       _money(spot_display, "R\u00a0", 4),
                       value_color=COPPER)
    story.append(Spacer(1, 4 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 5 — StoneX USD Cash Flow (Sell Side)
    # ──────────────────────────────────────────────────────────────
    _audit_step_header("Step 5.", "StoneX USD Cash Flow (Sell Side)",
                       "Net USD proceeds from StoneX gold trades. Sells generate positive inflows (+), buys are negative outflows (\u2212).")
    for cf in xau_cash_flows:
        sign_str = ""
        if cf["signed"] is not None:
            sign_str = f"+{_money(cf['signed'], '$')}" if cf["signed"] >= 0 else f"-{_money(abs(cf['signed']), '$')}"
        else:
            sign_str = "\u2014"
        lbl_parts = [f"{cf['side']} {_ounces(cf['qty'], 4)} @ {_money(cf['price'], '$', 4)}"]
        if cf.get("fnc"):
            lbl_parts.append(f"FNC: {cf['fnc']}")
        if cf.get("trade_date"):
            lbl_parts.append(cf["trade_date"])
        if cf.get("narration"):
            lbl_parts.append(_tr(cf["narration"], 50))
        _audit_row(
            " \u2014 ".join(lbl_parts),
            sign_str,
            value_color=GREEN if (cf["signed"] or 0) >= 0 else RED_C,
        )
    _audit_row("Net StoneX USD cash flow", _money(net_stonex_usd, "$"),
               value_color=GREEN if net_stonex_usd >= 0 else RED_C)
    story.append(Spacer(1, 4 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 5b — StoneX FX Cash Flow (ZAR Leg)
    # ──────────────────────────────────────────────────────────────
    _audit_step_header("Step 5b.", "StoneX FX Cash Flow (ZAR Leg)",
                       "Net ZAR proceeds from StoneX FX (USD/ZAR) trades. Sells generate positive inflows (+), buys are negative outflows (\u2212).")
    for cf in fx_cash_flows:
        sign_str = ""
        if cf["signed"] is not None:
            sign_str = f"+{_money(cf['signed'], rand_nbsp)}" if cf["signed"] >= 0 else f"-{_money(abs(cf['signed']), rand_nbsp)}"
        else:
            sign_str = "\u2014"
        lbl_parts = [f"{cf['side']} {_money(cf['qty'], '$', 2)} @ {_money(cf['price'], rand_nbsp, 4)}"]
        if cf.get("fnc"):
            lbl_parts.append(f"FNC: {cf['fnc']}")
        if cf.get("trade_date"):
            lbl_parts.append(cf["trade_date"])
        if cf.get("narration"):
            lbl_parts.append(cf["narration"][:60])
        _audit_row(
            " \u2014 ".join(lbl_parts),
            sign_str,
            value_color=GREEN if (cf["signed"] or 0) >= 0 else RED_C,
        )
    _audit_row("Net StoneX FX ZAR cash flow", _money(net_fx_zar, "R\u00a0"),
               value_color=GREEN if net_fx_zar >= 0 else RED_C)
    story.append(Spacer(1, 4 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 6 — Control Account (Metal Exposure)
    # ──────────────────────────────────────────────────────────────
    _audit_step_header("Step 6.", "Control Account (Metal Exposure)",
                       "Remaining unhedged metal position. A zero balance means the trade is fully hedged; any residual represents open exposure.")
    _audit_row("Control Account (grams)", _grams(ctrl_g))
    _audit_row("Control Account (troy ounces)", _ounces(ctrl_oz, 4))
    _audit_row("Control Account (ZAR)", _money(ctrl_zar, "R\u00a0"))
    story.append(Spacer(1, 4 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 7 — ZAR Position (Sell Side Valuation)
    # ──────────────────────────────────────────────────────────────
    _audit_step_header("Step 7.", "ZAR Position (Sell Side Valuation)",
                       "StoneX USD flows converted to ZAR at the traded FX rate, establishing the sell side valuation against TradeMC buy side.")
    _audit_row("StoneX ZAR Flow", _money(stonex_zar_flow, "R\u00a0"))
    _audit_row("Sell Side ZAR", _money(sell_zar, "R\u00a0"))
    _audit_row("Buy Side ZAR (TradeMC net)", _money(buy_zar, "R\u00a0"))
    story.append(Spacer(1, 4 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 8 — Profit / Loss Calculation
    # ──────────────────────────────────────────────────────────────
    _audit_step_header("Step 8.", "Profit / Loss Calculation",
                       "Final profit is the difference between sell side (StoneX proceeds) and buy side (TradeMC cost, net of refining).")

    _audit_row_formula("Profit (USD)",
                       f"{_money(sell_usd, '$')} \u2212 {_money(buy_usd, '$')}",
                       _money(profit_usd, "$"),
                       value_color=GREEN if (profit_usd or 0) >= 0 else RED_C)
    _audit_row_formula("Profit (ZAR)",
                       f"{_money(sell_zar, 'R ')} \u2212 {_money(buy_zar, 'R ')}",
                       _money(profit_zar, "R\u00a0"),
                       value_color=GREEN if (profit_zar or 0) >= 0 else RED_C)
    _audit_row_formula("Profit Margin",
                       f"({_money(profit_zar, 'R ')} / {_money(buy_zar, 'R ')}) x 100",
                       _pct(profit_pct, 3),
                       value_color=COPPER)
    story.append(Spacer(1, 3 * mm))

    # Profit highlight box
    story.append(Spacer(1, 2 * mm))
    if profit_zar is not None:
        p_bg = GREEN_LT if profit_zar >= 0 else RED_LT
        p_border = GREEN if profit_zar >= 0 else RED_C
        p_color = GREEN if profit_zar >= 0 else RED_C
        p_label = "NET PROFIT" if profit_zar >= 0 else "NET LOSS"
        profit_box_data = [[
            Paragraph(
                f"<font name='{FONT_BOLD}' size='9' color='#555555'>{p_label} (ZAR)</font>",
                _ps("_pb_l", fontName=FONT_BOLD, fontSize=9, textColor=DARK_GREY,
                    leading=12),
            ),
            Paragraph(
                f"<font name='{FONT_BOLD}' size='16'>{_esc(_money(profit_zar, 'R '))}</font>",
                _ps("_pb_v", fontName=FONT_BOLD, fontSize=16, leading=20,
                    textColor=p_color, alignment=TA_RIGHT),
            ),
            Paragraph(
                f"<font name='{FONT_BOLD}' size='10'>({_esc(_pct(profit_pct, 3))} margin)</font>",
                _ps("_pb_p", fontName=FONT_BOLD, fontSize=10,
                    textColor=p_color, alignment=TA_RIGHT, leading=13),
            ),
        ]]
        pbox = Table(profit_box_data, colWidths=[PW * 0.28, PW * 0.42, PW * 0.30])
        pbox.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), p_bg),
            ("BOX",           (0, 0), (-1, -1), 1.5, p_border),
            ("LINEABOVE",     (0, 0), (-1, 0), 3, p_border),
            ("LEFTPADDING",   (0, 0), (-1, -1), 16),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 16),
            ("TOPPADDING",    (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(pbox)
    story.append(Spacer(1, 8 * mm))

    # ──────────────────────────────────────────────────────────────
    # STEP 9 — Complete Source Data Appendix
    # ──────────────────────────────────────────────────────────────
    story.append(PageBreak())
    _audit_step_header(
        "Step 9.",
        "Complete Source Data Appendix",
        "All available source fields loaded for this trade. Included for full traceability of the trade breakdown calculations.",
    )
    _append_source_dump(
        "TradeMC Source Rows",
        tm_pdf,
        preferred_cols=[
            "Company",
            "Weight (g)",
            "Weight (oz)",
            "$/oz Booked",
            "FX Rate",
            "USD Value",
            "ZAR Value",
            "company_refining_rate",
            "zar_value_less_refining",
        ],
    )
    story.append(Spacer(1, 2 * mm))
    _append_source_dump(
        "StoneX / PMX Source Rows",
        stx_pdf,
        preferred_cols=[
            "Trade Date",
            "Value Date",
            "Symbol",
            "Side",
            "Quantity",
            "Price",
            "FNC #",
            "OrderID",
            "Doc #",
            "Narration",
            "Settle Currency",
            "Settle Amount",
        ],
    )
    story.append(Spacer(1, 2 * mm))
    _append_source_dump("Summary Source Rows", sum_pdf)
    story.append(Spacer(1, 6 * mm))

    # ── Page numbering and footer via canvas callbacks ──────────
    def _footer_on_page(canvas_obj, doc_obj):
        canvas_obj.saveState()
        # Copper accent line above footer
        canvas_obj.setStrokeColor(COPPER_MD)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(LM, 11 * mm, PAGE[0] - RM, 11 * mm)
        # Footer text
        canvas_obj.setFont(FONT_REG, 6.5)
        canvas_obj.setFillColor(MID_GREY)
        footer_left = (
            f"Metal Concentrators SA  \u2022  Confidential  \u2022  "
            f"Trade Breakdown #{_s(trade_num_value)}"
        )
        footer_right = f"Generated {generated_at}  \u2022  Page {canvas_obj.getPageNumber()}"
        canvas_obj.drawString(LM, 7 * mm, footer_left)
        canvas_obj.drawRightString(PAGE[0] - RM, 7 * mm, footer_right)
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=_footer_on_page, onLaterPages=_footer_on_page)
    return buf.getvalue()


def _build_trading_ticket_pdf_fpdf(
    trade_num_value: str,
    trademc_rows: pd.DataFrame,
    stonex_rows: pd.DataFrame,
    summary_rows: pd.DataFrame,
):
    try:
        from fpdf import FPDF
    except Exception:
        return None
    import os as _os

    def _to_num(value: Any) -> Optional[float]:
        try:
            n = float(value)
            if math.isnan(n) or math.isinf(n):
                return None
            return n
        except Exception:
            return None

    def _txt(value: Any) -> str:
        text = str(value if value is not None else "--")
        try:
            text.encode("latin-1")
            return text
        except Exception:
            return text.encode("latin-1", errors="replace").decode("latin-1")

    def _fmt_num(value: Any, dp: int = 2) -> str:
        n = _to_num(value)
        if n is None:
            return "--"
        return f"{n:,.{dp}f}"

    def _fmt_money(value: Any, prefix: str = "", dp: int = 2) -> str:
        n = _to_num(value)
        if n is None:
            return "--"
        return f"{prefix}{n:,.{dp}f}"

    tm_df = trademc_rows.copy() if isinstance(trademc_rows, pd.DataFrame) else pd.DataFrame()
    st_df = stonex_rows.copy() if isinstance(stonex_rows, pd.DataFrame) else pd.DataFrame()
    sum_df = summary_rows.copy() if isinstance(summary_rows, pd.DataFrame) else pd.DataFrame()

    pdf = FPDF("L", "mm", "A4")
    pdf.set_auto_page_break(True, margin=10)
    pdf.add_page()

    c_charcoal = (28, 28, 28)
    c_copper = (176, 120, 64)
    c_copper_lt = (251, 245, 236)
    c_hdr = (42, 42, 42)
    c_stripe = (246, 246, 246)

    page_w = 297.0
    lm = 10.0
    usable_w = page_w - (lm * 2)
    generated_at = datetime.now().strftime("%d %b %Y  %H:%M")

    logo_path = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "MetCon Logo.png")
    if _os.path.isfile(logo_path):
        try:
            pdf.image(logo_path, x=lm, y=8, w=52, h=13)
        except Exception:
            pass

    pdf.set_xy(130, 8)
    pdf.set_text_color(*c_copper)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(157, 5, _txt("METAL CONCENTRATORS SA"), 0, 2, "R")
    pdf.set_text_color(*c_charcoal)
    pdf.set_font("Arial", "B", 18)
    pdf.cell(157, 8, _txt("TRADING TICKET"), 0, 2, "R")
    pdf.set_text_color(120, 120, 120)
    pdf.set_font("Arial", "", 8)
    pdf.cell(157, 5, _txt(f"Trade {trade_num_value}    Generated {generated_at}"), 0, 1, "R")
    pdf.set_draw_color(*c_copper)
    pdf.set_line_width(0.8)
    pdf.line(lm, 24, lm + usable_w, 24)
    pdf.ln(5)

    def _section_title(title: str):
        y = pdf.get_y()
        pdf.set_fill_color(*c_copper)
        pdf.rect(lm, y, 2, 6, style="F")
        pdf.set_fill_color(*c_copper_lt)
        pdf.rect(lm + 2, y, usable_w - 2, 6, style="F")
        pdf.set_xy(lm + 4, y + 1.2)
        pdf.set_text_color(*c_copper)
        pdf.set_font("Arial", "B", 9)
        pdf.cell(usable_w - 6, 3.6, _txt(title.upper()), 0, 1, "L")
        pdf.ln(1.5)

    def _trim_to_width(text: str, width: float) -> str:
        text = _txt(text)
        if text == "--":
            return text
        max_chars = max(6, int(width / 1.95))
        if len(text) <= max_chars:
            return text
        return text[: max_chars - 3] + "..."

    def _table(
        headers: List[str],
        rows_data: List[List[str]],
        col_widths: List[float],
        right_cols: Optional[List[int]] = None,
    ):
        right_cols = right_cols or []
        widths = [usable_w * w for w in col_widths]
        line_h = 6.4

        def _draw_header():
            pdf.set_fill_color(*c_hdr)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", "B", 8)
            for i, h in enumerate(headers):
                align = "R" if i in right_cols else "C"
                pdf.cell(widths[i], 7, _txt(h), 0, 0, align, True)
            pdf.ln()

        _draw_header()
        alt = False
        for row in rows_data:
            if pdf.get_y() > 190:
                pdf.add_page()
                _draw_header()
                alt = False
            pdf.set_fill_color(*(c_stripe if alt else (255, 255, 255)))
            pdf.set_text_color(*c_charcoal)
            pdf.set_font("Arial", "", 8)
            for i, raw in enumerate(row):
                align = "R" if i in right_cols else "L"
                txt = _trim_to_width(str(raw if raw is not None and str(raw).strip() != "" else "--"), widths[i])
                pdf.cell(widths[i], line_h, txt, 1, 0, align, True)
            pdf.ln()
            alt = not alt
        pdf.ln(2.5)

    _section_title("TradeMC - Client Bookings")
    tm_cols = ["Company", "Weight (g)", "Weight (oz)", "$/oz Booked", "FX Rate", "USD Value", "ZAR Value"]
    tm_avail = [c for c in tm_cols if c in tm_df.columns]
    tm_rows: List[List[str]] = []
    for _, row in tm_df.iterrows():
        tm_rows.append([
            _txt(row.get("Company", "--")),
            _fmt_num(row.get("Weight (g)"), 2) if "Weight (g)" in tm_avail else "--",
            _fmt_num(row.get("Weight (oz)"), 6) if "Weight (oz)" in tm_avail else "--",
            _fmt_num(row.get("$/oz Booked"), 2) if "$/oz Booked" in tm_avail else "--",
            _fmt_num(row.get("FX Rate"), 4) if "FX Rate" in tm_avail else "--",
            _fmt_money(row.get("USD Value"), "$", 2) if "USD Value" in tm_avail else "--",
            _fmt_money(row.get("ZAR Value"), "R ", 2) if "ZAR Value" in tm_avail else "--",
        ])
    if not tm_rows:
        tm_rows = [["--"] * 7]
    _table(
        ["Company", "Weight (g)", "Weight (oz)", "$/oz Booked", "FX Rate", "USD Value", "ZAR Value"],
        tm_rows,
        [0.28, 0.11, 0.13, 0.12, 0.11, 0.12, 0.13],
        right_cols=[1, 2, 3, 4, 5, 6],
    )

    _section_title("PMX / StoneX Trades")
    st_cols = ["Doc #", "FNC #", "Trade Date", "Value Date", "Symbol", "Side", "Narration", "Quantity", "Price", "Settle Currency", "Settle Amount"]
    st_avail = [c for c in st_cols if c in st_df.columns]
    st_widths_map = {"Doc #": 0.07, "FNC #": 0.09, "Trade Date": 0.08, "Value Date": 0.08, "Symbol": 0.06,
                     "Side": 0.05, "Narration": 0.16, "Quantity": 0.10, "Price": 0.10,
                     "Settle Currency": 0.06, "Settle Amount": 0.09}
    _sw_fpdf = sum(st_widths_map.get(c, 0.08) for c in st_avail) or 1
    st_widths = [st_widths_map.get(c, 0.08) / _sw_fpdf for c in st_avail]
    st_right = [i for i, c in enumerate(st_avail) if c in ("Quantity", "Price", "Settle Amount")]
    st_rows: List[List[str]] = []
    for _, row in st_df.iterrows():
        r_cells = []
        for c in st_avail:
            if c == "Quantity":
                r_cells.append(_fmt_num(row.get(c), 4))
            elif c == "Price":
                r_cells.append(_fmt_num(row.get(c), 4))
            elif c == "Settle Amount":
                r_cells.append(_fmt_num(row.get(c), 2))
            else:
                r_cells.append(_txt(row.get(c, "--")))
        st_rows.append(r_cells)
    if not st_rows:
        st_rows = [["--"] * len(st_avail)]
    _table(
        st_avail,
        st_rows,
        st_widths,
        right_cols=st_right,
    )

    _section_title("Ticket Summary")
    summary_row = sum_df.iloc[0].to_dict() if not sum_df.empty else {}
    summary_items = [
        ("Gold WA $/oz", _fmt_money(summary_row.get("Gold WA $/oz"), "$", 4)),
        ("FX WA USD/ZAR", _fmt_money(summary_row.get("FX WA USD/ZAR"), "R ", 4)),
        ("Spot ZAR/g", _fmt_money(summary_row.get("Spot ZAR/g"), "R ", 4)),
        ("Sell Side (USD)", _fmt_money(summary_row.get("Sell Side (USD)"), "$", 2)),
        ("Buy Side (USD)", _fmt_money(summary_row.get("Buy Side (USD)"), "$", 2)),
        ("Sell Side (ZAR)", _fmt_money(summary_row.get("Sell Side (ZAR)"), "R ", 2)),
        ("Buy Side (ZAR)", _fmt_money(summary_row.get("Buy Side (ZAR)"), "R ", 2)),
        ("Total Traded (g)", _fmt_num(summary_row.get("Total Traded (g)"), 2)),
        ("Total Traded (oz)", _fmt_num(summary_row.get("Total Traded (oz)"), 4)),
        ("Control Account (g)", _fmt_num(summary_row.get("Control Account (g)"), 2)),
        ("Control Account (oz)", _fmt_num(summary_row.get("Control Account (oz)"), 4)),
        ("Control Account (ZAR)", _fmt_money(summary_row.get("Control Account (ZAR)"), "R ", 2)),
        ("StoneX ZAR Flow", _fmt_money(summary_row.get("StoneX ZAR Flow"), "R ", 2)),
        ("Profit (USD)", _fmt_money(summary_row.get("Profit (USD)"), "$", 2)),
        ("Profit (ZAR)", _fmt_money(summary_row.get("Profit (ZAR)"), "R ", 2)),
        ("Profit % (ZAR Spot Cost)", _fmt_num(summary_row.get("Profit % (ZAR Spot Cost)"), 3) + "%"),
    ]
    _table(
        ["Metric", "Value"],
        [[k, v] for (k, v) in summary_items],
        [0.56, 0.44],
        right_cols=[1],
    )

    # -----------------------------------------------------------------
    # Full Calculation Audit Trail (mirrors Trade Breakdown page)
    # -----------------------------------------------------------------
    def _as_pair(value: Any) -> str:
        text = str(value or "").upper().replace(" ", "").replace("-", "")
        if not text:
            return ""
        if "/" in text:
            return text
        if len(text) == 6:
            return f"{text[:3]}/{text[3:]}"
        return text

    def _fmt_pct(value: Any, dp: int = 2) -> str:
        n = _to_num(value)
        if n is None:
            return "--"
        return f"{n:,.{dp}f}%"

    def _fmt_grams(value: Any, dp: int = 2) -> str:
        n = _to_num(value)
        if n is None:
            return "--"
        return f"{n:,.{dp}f} g"

    def _fmt_oz(value: Any, dp: int = 4) -> str:
        n = _to_num(value)
        if n is None:
            return "--"
        return f"{n:,.{dp}f} oz"

    def _none_if_zero(value: Optional[float], eps: float = 1e-12) -> Optional[float]:
        if value is None:
            return None
        try:
            fv = float(value)
            return None if abs(fv) <= eps else fv
        except Exception:
            return None

    tm_bookings: List[Dict[str, Any]] = []
    tm_totals = {
        "weight_g": 0.0,
        "weight_oz": 0.0,
        "usd_value": 0.0,
        "zar_gross": 0.0,
        "zar_net": 0.0,
    }
    for _, row in tm_df.iterrows():
        weight_g = _to_num(row.get("Weight (g)")) or 0.0
        weight_oz = _to_num(row.get("Weight (oz)"))
        if weight_oz is None:
            weight_oz = weight_g / GRAMS_PER_TROY_OUNCE
        booked_price = _to_num(row.get("$/oz Booked"))
        fx_rate = _to_num(row.get("FX Rate"))
        usd_value = _to_num(row.get("USD Value"))
        if usd_value is None and booked_price is not None:
            usd_value = float(weight_oz) * float(booked_price)
        zar_gross = _to_num(row.get("ZAR Value"))
        if zar_gross is None and usd_value is not None and fx_rate is not None:
            zar_gross = float(usd_value) * float(fx_rate)
        refining_rate = _to_num(row.get("company_refining_rate")) or 0.0
        refining_deduction = (float(zar_gross) * (float(refining_rate) / 100.0)) if zar_gross is not None else None
        zar_net = _to_num(row.get("zar_value_less_refining"))
        if zar_net is None and zar_gross is not None:
            zar_net = float(zar_gross) * (1.0 - (float(refining_rate) / 100.0))

        tm_bookings.append(
            {
                "company": str(row.get("Company") or row.get("company_name") or "Unknown Company"),
                "weight_g": float(weight_g),
                "weight_oz": float(weight_oz),
                "booked_price": booked_price,
                "fx_rate": fx_rate,
                "usd_value": usd_value,
                "zar_gross": zar_gross,
                "refining_rate": float(refining_rate),
                "refining_deduction": refining_deduction,
                "zar_net": zar_net,
            }
        )

        tm_totals["weight_g"] += float(weight_g)
        tm_totals["weight_oz"] += float(weight_oz)
        tm_totals["usd_value"] += float(usd_value or 0.0)
        tm_totals["zar_gross"] += float(zar_gross or 0.0)
        tm_totals["zar_net"] += float(zar_net or 0.0)

    pmx_trades: List[Dict[str, Any]] = []
    for _, row in st_df.iterrows():
        symbol = _as_pair(row.get("Symbol"))
        side = str(row.get("Side") or "").upper().strip()
        qty = abs(float(_to_num(row.get("Quantity")) or 0.0))
        price = _to_num(row.get("Price"))
        if qty <= 0:
            continue
        notional = (qty * float(price)) if price is not None else None
        pmx_trades.append(
            {
                "trade_date": str(row.get("Trade Date") or ""),
                "value_date": str(row.get("Value Date") or ""),
                "symbol": symbol,
                "side": side,
                "qty": qty,
                "price": price,
                "notional": notional,
                "fnc": str(row.get("FNC #") or ""),
                "doc": str(row.get("Doc #") or ""),
                "narration": str(row.get("Narration") or ""),
            }
        )

    xau_trades = [t for t in pmx_trades if t.get("symbol") == "XAU/USD" and t.get("price") is not None]
    fx_trades = [t for t in pmx_trades if t.get("symbol") == "USD/ZAR" and t.get("price") is not None]

    gold_total_qty = sum(float(t.get("qty") or 0.0) for t in xau_trades)
    gold_total_notional = sum(float(t.get("notional") or 0.0) for t in xau_trades)
    gold_wa = (gold_total_notional / gold_total_qty) if gold_total_qty > 1e-12 else None

    fx_total_qty = sum(float(t.get("qty") or 0.0) for t in fx_trades)
    fx_total_notional = sum(float(t.get("notional") or 0.0) for t in fx_trades)
    fx_wa = (fx_total_notional / fx_total_qty) if fx_total_qty > 1e-12 else None

    spot_zar_per_g = None
    if gold_wa is not None and fx_wa is not None:
        spot_zar_per_g = (float(gold_wa) * float(fx_wa)) / GRAMS_PER_TROY_OUNCE

    xau_cash_flow_rows: List[Dict[str, Any]] = []
    net_stonex_usd_flow = 0.0
    for t in xau_trades:
        signed = None
        if t.get("notional") is not None:
            signed = float(t["notional"]) if str(t.get("side")) == "SELL" else -float(t["notional"])
            net_stonex_usd_flow += signed
        xau_cash_flow_rows.append({**t, "signed": signed})

    summary_row = sum_df.iloc[0].to_dict() if not sum_df.empty else {}
    gold_wa_display = _to_num(summary_row.get("Gold WA $/oz"))
    if gold_wa_display is None:
        gold_wa_display = gold_wa
    fx_wa_display = _to_num(summary_row.get("FX WA USD/ZAR"))
    if fx_wa_display is None:
        fx_wa_display = fx_wa
    spot_zar_per_g_display = _to_num(summary_row.get("Spot ZAR/g"))
    if spot_zar_per_g_display is None and gold_wa_display is not None and fx_wa_display is not None:
        spot_zar_per_g_display = (float(gold_wa_display) * float(fx_wa_display)) / GRAMS_PER_TROY_OUNCE
    gold_total_notional_display = gold_total_notional
    if gold_wa_display is not None and gold_total_qty > 1e-12:
        gold_total_notional_display = float(gold_wa_display) * float(gold_total_qty)
    fx_total_notional_display = fx_total_notional
    if fx_wa_display is not None and fx_total_qty > 1e-12:
        fx_total_notional_display = float(fx_wa_display) * float(fx_total_qty)

    sell_side_usd = _to_num(summary_row.get("Sell Side (USD)"))
    if sell_side_usd is None:
        sell_side_usd = abs(net_stonex_usd_flow)
    sell_side_usd = _none_if_zero(sell_side_usd)

    buy_side_usd = _to_num(summary_row.get("Buy Side (USD)"))
    if buy_side_usd is None:
        buy_side_usd = abs(tm_totals["usd_value"])
    buy_side_usd = _none_if_zero(buy_side_usd)

    profit_usd_calc = _to_num(summary_row.get("Profit (USD)"))
    if profit_usd_calc is None and sell_side_usd is not None and buy_side_usd is not None:
        profit_usd_calc = float(sell_side_usd) - float(buy_side_usd)

    sell_side_zar = _to_num(summary_row.get("Sell Side (ZAR)"))
    if sell_side_zar is None:
        sell_side_zar = _to_num(summary_row.get("StoneX ZAR Flow"))
    sell_side_zar = _none_if_zero(sell_side_zar)

    buy_side_zar = _to_num(summary_row.get("Buy Side (ZAR)"))
    if buy_side_zar is None:
        buy_side_zar = abs(tm_totals["zar_net"])
    buy_side_zar = _none_if_zero(buy_side_zar)

    profit_zar_calc = _to_num(summary_row.get("Profit (ZAR)"))
    if profit_zar_calc is None and sell_side_zar is not None and buy_side_zar is not None:
        profit_zar_calc = float(sell_side_zar) - float(buy_side_zar)

    profit_margin_pct = _to_num(summary_row.get("Profit % (ZAR Spot Cost)"))
    if (
        profit_margin_pct is None
        and profit_zar_calc is not None
        and buy_side_zar is not None
        and abs(float(buy_side_zar)) > 1e-12
    ):
        profit_margin_pct = (float(profit_zar_calc) / abs(float(buy_side_zar))) * 100.0

    control_g = _to_num(summary_row.get("Control Account (g)"))
    control_oz = _to_num(summary_row.get("Control Account (oz)"))
    if control_oz is None and control_g is not None:
        control_oz = float(control_g) / GRAMS_PER_TROY_OUNCE
    control_zar = _to_num(summary_row.get("Control Account (ZAR)"))

    total_traded_g = _to_num(summary_row.get("Total Traded (g)"))
    if total_traded_g is None and abs(tm_totals["weight_g"]) > 1e-12:
        total_traded_g = abs(tm_totals["weight_g"])
    total_traded_oz = _to_num(summary_row.get("Total Traded (oz)"))
    if total_traded_oz is None and total_traded_g is not None:
        total_traded_oz = float(total_traded_g) / GRAMS_PER_TROY_OUNCE

    stonex_zar_flow = _to_num(summary_row.get("StoneX ZAR Flow"))
    if stonex_zar_flow is None and sell_side_zar is not None:
        stonex_zar_flow = sell_side_zar

    def _audit_step(step_no: str, title: str, subtitle: str = ""):
        _section_title(f"Step {step_no}. {title}")
        if subtitle:
            pdf.set_text_color(115, 115, 115)
            pdf.set_font("Arial", "I", 7)
            pdf.multi_cell(usable_w, 4.3, _txt(subtitle), 0, "L")
            pdf.set_text_color(*c_charcoal)
            pdf.ln(0.6)

    def _field_line(label: str, value: str, label_w: float = 74.0, bold_value: bool = False):
        if pdf.get_y() > 195:
            pdf.add_page()
        pdf.set_text_color(*c_charcoal)
        pdf.set_font("Arial", "B", 7.5)
        pdf.cell(label_w, 4.7, _txt(f"{label}:"), 0, 0, "L")
        pdf.set_font("Arial", "B" if bold_value else "", 7.5)
        pdf.multi_cell(max(40.0, usable_w - label_w), 4.7, _txt(value), 0, "L")

    def _append_source_dump(block_title: str, df: pd.DataFrame, preferred_cols: Optional[List[str]] = None):
        _section_title(block_title)
        if df is None or df.empty:
            _field_line("Rows", "0")
            return
        cols = list(df.columns)
        ordered_cols = cols
        if preferred_cols:
            ordered_cols = [c for c in preferred_cols if c in cols] + [c for c in cols if c not in preferred_cols]
        _field_line("Rows", str(len(df)))
        for ridx, (_, row) in enumerate(df.iterrows(), start=1):
            pdf.set_font("Arial", "B", 8)
            pdf.set_text_color(*c_copper)
            pdf.cell(usable_w, 5.6, _txt(f"Row {ridx}"), 0, 1, "L")
            pdf.set_text_color(*c_charcoal)
            for col in ordered_cols:
                raw_val = row.get(col)
                val_txt = "--" if raw_val is None else str(raw_val)
                if val_txt.strip() == "":
                    val_txt = "--"
                _field_line(str(col), val_txt, label_w=78.0)
            pdf.ln(0.7)

    pdf.add_page()
    _section_title("Calculation Audit Trail")
    pdf.set_text_color(120, 120, 120)
    pdf.set_font("Arial", "", 7.5)
    pdf.multi_cell(
        usable_w,
        4.2,
        _txt(
            f"Step-by-step calculation trace for Trade {trade_num_value}. "
            "Each section shows source values, formulas, and outputs used in the trade breakdown page."
        ),
        0,
        "L",
    )
    pdf.set_text_color(*c_charcoal)
    pdf.ln(1.0)

    _audit_step("0", "Outcome Snapshot")
    _table(
        ["Metric", "Value"],
        [
            ["Sell Side (USD)", _fmt_money(sell_side_usd, "$", 2)],
            ["Buy Side (USD)", _fmt_money(buy_side_usd, "$", 2)],
            ["Profit (USD)", _fmt_money(profit_usd_calc, "$", 2)],
            ["Sell Side (ZAR)", _fmt_money(sell_side_zar, "R ", 2)],
            ["Buy Side (ZAR)", _fmt_money(buy_side_zar, "R ", 2)],
            ["Profit (ZAR)", _fmt_money(profit_zar_calc, "R ", 2)],
            ["Control Account (g)", _fmt_grams(control_g, 2)],
            ["Control Account (oz)", _fmt_oz(control_oz, 4)],
            ["Control Account (ZAR)", _fmt_money(control_zar, "R ", 2)],
            ["Total Traded (g)", _fmt_grams(total_traded_g, 2)],
            ["Total Traded (oz)", _fmt_oz(total_traded_oz, 4)],
            ["StoneX ZAR Flow", _fmt_money(stonex_zar_flow, "R ", 2)],
            ["Profit Margin", _fmt_pct(profit_margin_pct, 2)],
        ],
        [0.62, 0.38],
        right_cols=[1],
    )

    _audit_step(
        "1",
        "Input Data Summary",
        "Raw source row counts loaded for this trade breakdown.",
    )
    _table(
        ["Item", "Value"],
        [
            ["TradeMC bookings (buy side)", str(len(tm_bookings))],
            ["StoneX/PMX trades (sell side)", str(len(pmx_trades))],
            ["XAU/USD trades", str(len(xau_trades))],
            ["USD/ZAR trades", str(len(fx_trades))],
            ["Conversion constant", "31.1035 g/troy oz"],
        ],
        [0.68, 0.32],
        right_cols=[1],
    )

    _audit_step(
        "2",
        "TradeMC Buy-Side Valuation",
        "Each booking is converted from grams to oz, then valued in USD and ZAR, including refining deductions.",
    )
    if not tm_bookings:
        _table(["Item", "Formula", "Result"], [["Bookings", "--", "--"]], [0.30, 0.42, 0.28], right_cols=[2])
    else:
        for idx, booking in enumerate(tm_bookings, start=1):
            _section_title(f"Booking {idx}: {booking.get('company')}")
            rows_step2 = [
                ["Weight", "--", _fmt_grams(booking.get("weight_g"), 2)],
                [
                    "Weight in troy ounces",
                    f"{_fmt_num(booking.get('weight_g'), 2)} / 31.1035",
                    _fmt_oz(booking.get("weight_oz"), 6),
                ],
                ["Booked gold price", "--", f"{_fmt_money(booking.get('booked_price'), '$', 2)} /oz"],
                [
                    "USD value",
                    f"{_fmt_oz(booking.get('weight_oz'), 6)} x {_fmt_money(booking.get('booked_price'), '$', 2)}",
                    _fmt_money(booking.get("usd_value"), "$", 2),
                ],
                ["FX rate (ZAR/USD)", "--", _fmt_num(booking.get("fx_rate"), 4)],
                [
                    "ZAR value (gross)",
                    f"{_fmt_money(booking.get('usd_value'), '$', 2)} x {_fmt_num(booking.get('fx_rate'), 4)}",
                    _fmt_money(booking.get("zar_gross"), "R ", 2),
                ],
                ["Refining rate", "--", _fmt_pct(booking.get("refining_rate"), 2)],
                [
                    "Refining deduction",
                    f"{_fmt_money(booking.get('zar_gross'), 'R ', 2)} x {_fmt_pct(booking.get('refining_rate'), 2)}",
                    _fmt_money(booking.get("refining_deduction"), "R ", 2),
                ],
                ["ZAR value (net of refining)", "--", _fmt_money(booking.get("zar_net"), "R ", 2)],
            ]
            _table(["Item", "Formula", "Result"], rows_step2, [0.30, 0.42, 0.28], right_cols=[2])
    _table(
        ["Item", "Value"],
        [
            ["Total weight", f"{_fmt_grams(tm_totals['weight_g'], 2)} ({_fmt_oz(tm_totals['weight_oz'], 6)})"],
            ["Total USD value", _fmt_money(tm_totals["usd_value"], "$", 2)],
            ["Total ZAR (gross)", _fmt_money(tm_totals["zar_gross"], "R ", 2)],
            ["Total ZAR (net of refining)", _fmt_money(tm_totals["zar_net"], "R ", 2)],
        ],
        [0.58, 0.42],
        right_cols=[1],
    )

    _audit_step(
        "3",
        "StoneX Weighted Average Rates",
        "Weighted average = Sum(qty x price) / Sum(qty).",
    )
    _section_title("Gold Weighted Average (XAU/USD)")
    rows_gold = []
    for trade in xau_trades:
        lbl_parts = [f"{trade.get('side') or 'TRADE'} {_fmt_oz(trade.get('qty'), 4)} @ {_fmt_money(trade.get('price'), '$', 4)}"]
        if trade.get("fnc"):
            lbl_parts.append(f"FNC: {trade['fnc']}")
        if trade.get("trade_date"):
            lbl_parts.append(trade["trade_date"])
        if trade.get("narration"):
            lbl_parts.append(trade["narration"][:40])
        rows_gold.append(
            [
                " - ".join(lbl_parts),
                f"{_fmt_oz(trade.get('qty'), 4)} x {_fmt_money(trade.get('price'), '$', 4)}",
                _fmt_money(trade.get("notional"), "$", 2),
            ]
        )
    rows_gold.extend(
        [
            ["Sum of notional values", "--", _fmt_money(gold_total_notional_display, "$", 2)],
            ["Sum of quantities", "--", _fmt_oz(gold_total_qty, 4)],
            [
                "Gold weighted average",
                f"{_fmt_money(gold_total_notional_display, '$', 2)} / {_fmt_oz(gold_total_qty, 4)}",
                _fmt_money(gold_wa_display, "$", 4),
            ],
        ]
    )
    _table(["Item", "Formula", "Result"], rows_gold if rows_gold else [["XAU/USD trades", "--", "--"]], [0.31, 0.41, 0.28], right_cols=[2])

    _section_title("FX Weighted Average (USD/ZAR)")
    rows_fx = []
    for trade in fx_trades:
        lbl_parts = [f"{trade.get('side') or 'TRADE'} {_fmt_money(trade.get('qty'), '$', 2)} @ {_fmt_money(trade.get('price'), 'R ', 4)}"]
        if trade.get("fnc"):
            lbl_parts.append(f"FNC: {trade['fnc']}")
        if trade.get("trade_date"):
            lbl_parts.append(trade["trade_date"])
        if trade.get("narration"):
            lbl_parts.append(trade["narration"][:40])
        rows_fx.append(
            [
                " - ".join(lbl_parts),
                f"{_fmt_money(trade.get('qty'), '$', 2)} x {_fmt_money(trade.get('price'), 'R ', 4)}",
                _fmt_money(trade.get("notional"), "R ", 2),
            ]
        )
    rows_fx.extend(
        [
            ["Sum of notional values", "--", _fmt_money(fx_total_notional_display, "R ", 2)],
            ["Sum of quantities", "--", _fmt_money(fx_total_qty, "$", 2)],
            [
                "FX weighted average",
                f"{_fmt_money(fx_total_notional_display, 'R ', 2)} / {_fmt_money(fx_total_qty, '$', 2)}",
                _fmt_money(fx_wa_display, "R ", 4),
            ],
        ]
    )
    _table(["Item", "Formula", "Result"], rows_fx if rows_fx else [["USD/ZAR trades", "--", "--"]], [0.31, 0.41, 0.28], right_cols=[2])

    _audit_step("4", "Spot Rate Derivation")
    _table(
        ["Item", "Formula", "Result"],
        [
            [
                "Spot ZAR per gram",
                f"({_fmt_money(gold_wa_display, '$', 4)} x {_fmt_money(fx_wa_display, 'R ', 4)}) / 31.1035",
                _fmt_money(spot_zar_per_g_display, "R ", 4),
            ]
        ],
        [0.31, 0.41, 0.28],
        right_cols=[2],
    )

    _audit_step("5", "StoneX USD Cash Flow (Sell Side)")
    rows_cash: List[List[str]] = []
    for row in xau_cash_flow_rows:
        signed = _to_num(row.get("signed"))
        if signed is None:
            signed_text = "--"
        elif signed >= 0:
            signed_text = f"+{_fmt_money(abs(signed), '$', 2)}"
        else:
            signed_text = f"-{_fmt_money(abs(signed), '$', 2)}"
        lbl_parts = [f"{row.get('side')} {_fmt_oz(row.get('qty'), 4)} @ {_fmt_money(row.get('price'), '$', 4)}"]
        if row.get("fnc"):
            lbl_parts.append(f"FNC: {row['fnc']}")
        if row.get("trade_date"):
            lbl_parts.append(row["trade_date"])
        if row.get("narration"):
            lbl_parts.append(row["narration"][:40])
        rows_cash.append(
            [
                " - ".join(lbl_parts),
                signed_text,
            ]
        )
    rows_cash.append(["Net StoneX USD cash flow", _fmt_money(net_stonex_usd_flow, "$", 2)])
    _table(["Item", "Value"], rows_cash if rows_cash else [["Cash flow rows", "--"]], [0.68, 0.32], right_cols=[1])

    # Step 5b — FX Cash Flows
    fx_cash_flow_rows_fpdf: List[Dict[str, Any]] = []
    net_fx_zar_flow = 0.0
    for t in fx_trades:
        signed = None
        if t.get("notional") is not None:
            signed = float(t["notional"]) if str(t.get("side")) == "SELL" else -float(t["notional"])
            net_fx_zar_flow += signed
        fx_cash_flow_rows_fpdf.append({**t, "signed": signed})

    _audit_step("5b", "StoneX FX Cash Flow (ZAR Leg)")
    rows_fx_cash: List[List[str]] = []
    for row in fx_cash_flow_rows_fpdf:
        signed = _to_num(row.get("signed"))
        if signed is None:
            signed_text = "--"
        elif signed >= 0:
            signed_text = f"+{_fmt_money(abs(signed), 'R ', 2)}"
        else:
            signed_text = f"-{_fmt_money(abs(signed), 'R ', 2)}"
        lbl_parts = [f"{row.get('side')} {_fmt_money(row.get('qty'), '$', 2)} @ {_fmt_money(row.get('price'), 'R ', 4)}"]
        if row.get("fnc"):
            lbl_parts.append(f"FNC: {row['fnc']}")
        if row.get("trade_date"):
            lbl_parts.append(row["trade_date"])
        if row.get("narration"):
            lbl_parts.append(row["narration"][:40])
        rows_fx_cash.append(
            [
                " - ".join(lbl_parts),
                signed_text,
            ]
        )
    rows_fx_cash.append(["Net StoneX FX ZAR cash flow", _fmt_money(net_fx_zar_flow, "R ", 2)])
    _table(["Item", "Value"], rows_fx_cash if rows_fx_cash else [["FX cash flow rows", "--"]], [0.68, 0.32], right_cols=[1])

    _audit_step("6", "Control Account Check (Metal Exposure)")
    _table(
        ["Item", "Value"],
        [
            ["Control Account (grams)", _fmt_grams(control_g, 2)],
            ["Control Account (oz)", _fmt_oz(control_oz, 4)],
            ["Control Account (ZAR)", _fmt_money(control_zar, "R ", 2)],
        ],
        [0.62, 0.38],
        right_cols=[1],
    )

    _audit_step("7", "ZAR Sell-Side Valuation")
    _table(
        ["Item", "Value"],
        [
            ["StoneX ZAR flow", _fmt_money(stonex_zar_flow, "R ", 2)],
            ["Sell Side ZAR", _fmt_money(sell_side_zar, "R ", 2)],
            ["Buy Side ZAR (TradeMC net)", _fmt_money(buy_side_zar, "R ", 2)],
        ],
        [0.62, 0.38],
        right_cols=[1],
    )

    _audit_step("8", "Final Profit / Loss")
    _table(
        ["Item", "Formula", "Result"],
        [
            [
                "Profit (USD)",
                f"{_fmt_money(sell_side_usd, '$', 2)} - {_fmt_money(buy_side_usd, '$', 2)}",
                _fmt_money(profit_usd_calc, "$", 2),
            ],
            [
                "Profit (ZAR)",
                f"{_fmt_money(sell_side_zar, 'R ', 2)} - {_fmt_money(buy_side_zar, 'R ', 2)}",
                _fmt_money(profit_zar_calc, "R ", 2),
            ],
            [
                "Profit Margin",
                f"({_fmt_money(profit_zar_calc, 'R ', 2)} / {_fmt_money(buy_side_zar, 'R ', 2)}) x 100",
                _fmt_pct(profit_margin_pct, 2),
            ],
        ],
        [0.31, 0.41, 0.28],
        right_cols=[2],
    )

    pdf.add_page()
    _audit_step("9", "Complete Source Data Appendix", "All source fields loaded for this trade.")
    _append_source_dump(
        "TradeMC Source Rows",
        tm_df,
        preferred_cols=[
            "Company",
            "Weight (g)",
            "Weight (oz)",
            "$/oz Booked",
            "FX Rate",
            "USD Value",
            "ZAR Value",
            "company_refining_rate",
            "zar_value_less_refining",
        ],
    )
    _append_source_dump(
        "StoneX / PMX Source Rows",
        st_df,
        preferred_cols=[
            "Trade Date",
            "Value Date",
            "Symbol",
            "Side",
            "Quantity",
            "Price",
            "FNC #",
            "OrderID",
            "Doc #",
            "Narration",
            "Settle Currency",
            "Settle Amount",
        ],
    )
    _append_source_dump("Summary Source Rows", sum_df)

    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin-1", errors="ignore")
    if isinstance(out, (bytes, bytearray)):
        return bytes(out)
    return None


def build_open_positions_reval_pdf(args_dict: Dict[str, Any], req_headers: Any = None):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, Image as RLImage,
        )
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        import html as _html
        import os as _os
    except Exception:
        return _build_open_positions_reval_pdf_fpdf(args_dict, req_headers)

    def _esc(v: Any) -> str:
        return _html.escape(str(v), quote=False)

    FONT_REG = "Helvetica"
    FONT_BOLD = "Helvetica-Bold"
    try:
        _rp = r"C:\Windows\Fonts\calibri.ttf"
        _bp = r"C:\Windows\Fonts\calibrib.ttf"
        if _os.path.isfile(_rp) and _os.path.isfile(_bp):
            pdfmetrics.registerFont(TTFont("Calibri", _rp))
            pdfmetrics.registerFont(TTFont("Calibri-Bold", _bp))
            FONT_REG = "Calibri"
            FONT_BOLD = "Calibri-Bold"
    except Exception:
        pass

    CHARCOAL = colors.HexColor("#1C1C1C")
    COPPER = colors.HexColor("#B07840")
    COPPER_LT = colors.HexColor("#FBF5EC")
    COPPER_MD = colors.HexColor("#E8D5B7")
    TBL_HDR = colors.HexColor("#2A2A2A")
    ROW_ALT = colors.HexColor("#F6F6F6")
    BORDER = colors.HexColor("#E2E2E2")
    MID_GREY = colors.HexColor("#888888")
    GREEN = colors.HexColor("#2D7A4F")
    RED_C = colors.HexColor("#C0392B")
    WHITE = colors.white

    def _s(v: Any) -> str:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return str(v).strip()

    def _to_num(value: Any) -> Optional[float]:
        try:
            out = float(value)
            if math.isnan(out) or math.isinf(out):
                return None
            return out
        except Exception:
            return None

    def _fmt(value: Any, dp: int = 2) -> str:
        n = _to_num(value)
        if n is None:
            return "—"
        return f"{n:,.{dp}f}"

    def _ps(name: str, **kw) -> ParagraphStyle:
        return ParagraphStyle(
            name,
            fontName=kw.pop("fontName", FONT_REG),
            fontSize=kw.pop("fontSize", 8),
            leading=kw.pop("leading", 10),
            textColor=kw.pop("textColor", CHARCOAL),
            **kw,
        )

    reval = build_open_positions_reval(args_dict, req_headers)
    rows = reval.get("rows", []) if isinstance(reval, dict) else []
    summary = reval.get("summary", {}) if isinstance(reval, dict) else {}
    market = reval.get("market", {}) if isinstance(reval, dict) else {}
    recon = build_account_recon(args_dict, req_headers)
    tm_df = load_trademc_trades_with_companies()
    if not isinstance(tm_df, pd.DataFrame):
        tm_df = pd.DataFrame()

    today_key = datetime.now().strftime("%Y-%m-%d")
    buy_g = 0.0
    sell_g = 0.0
    counted = 0
    if not tm_df.empty:
        ts_src = (
            tm_df["trade_timestamp"] if "trade_timestamp" in tm_df.columns else
            (tm_df["trade_date"] if "trade_date" in tm_df.columns else
             (tm_df["created_at"] if "created_at" in tm_df.columns else pd.Series([None] * len(tm_df))))
        )
        side_src = (
            tm_df["side"] if "side" in tm_df.columns else
            (tm_df["trade_side"] if "trade_side" in tm_df.columns else pd.Series([""] * len(tm_df)))
        )
        wt_src = (
            tm_df["weight"] if "weight" in tm_df.columns else
            (tm_df["Weight"] if "Weight" in tm_df.columns else
             (tm_df["quantity"] if "quantity" in tm_df.columns else
              (tm_df["qty"] if "qty" in tm_df.columns else pd.Series([None] * len(tm_df)))))
        )
        ts = pd.to_datetime(ts_src, errors="coerce").dt.strftime("%Y-%m-%d")
        side = side_src.fillna("").astype(str).str.upper().str.strip()
        wt = pd.to_numeric(wt_src, errors="coerce")
        for i in range(len(tm_df)):
            if str(ts.iloc[i] or "") != today_key:
                continue
            w = wt.iloc[i]
            if pd.isna(w):
                continue
            w = float(w)
            s = str(side.iloc[i] or "")
            if s == "BUY":
                buy_g += abs(w)
            elif s == "SELL":
                sell_g -= abs(w)
            else:
                if w >= 0:
                    buy_g += w
                else:
                    sell_g += w
            counted += 1

    buf = BytesIO()
    PAGE = landscape(A4)
    LM = RM = 14 * mm
    TM = BM = 12 * mm
    doc = SimpleDocTemplate(buf, pagesize=PAGE, leftMargin=LM, rightMargin=RM, topMargin=TM, bottomMargin=BM)
    PW = PAGE[0] - LM - RM
    generated_at = datetime.now().strftime("%d %b %Y  %H:%M")
    story: List[Any] = []

    _logo_path = _os.path.join(
        _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
        "MetCon Logo.png",
    )
    _logo_cell: object = Spacer(48 * mm, 13 * mm)
    try:
        if _os.path.isfile(_logo_path):
            from PIL import Image as _PILImage
            with _PILImage.open(_logo_path) as _img:
                _img.load()
            _logo_cell = RLImage(_logo_path, width=52 * mm, height=13 * mm)
    except Exception:
        pass

    title_para = Paragraph(
        f"<font name='{FONT_BOLD}' size='6.5' color='#B07840'>METAL CONCENTRATORS SA</font><br/>"
        f"<font name='{FONT_BOLD}' size='17' color='#1C1C1C'>OPEN POSITIONS REVAL REPORT</font><br/>"
        f"<font size='6.5' color='#888888'>Generated  {_esc(generated_at)}</font>",
        _ps("_hdr", leading=22, alignment=TA_RIGHT),
    )
    hdr = Table([[_logo_cell, title_para]], colWidths=[PW * 0.38, PW * 0.62])
    hdr.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(hdr)
    story.append(HRFlowable(width=PW, thickness=2.5, color=COPPER, spaceBefore=1 * mm, spaceAfter=5 * mm))

    def _sec_lbl(title: str):
        lbl = Table([[
            Spacer(1, 1),
            Paragraph(
                f"<font name='{FONT_BOLD}' size='7.5' color='#B07840'>{_esc(title.upper())}</font>",
                _ps("_sl", fontName=FONT_BOLD, fontSize=7.5, textColor=COPPER),
            ),
        ]], colWidths=[3, PW - 3])
        lbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), COPPER),
            ("BACKGROUND", (1, 0), (1, 0), COPPER_LT),
            ("LEFTPADDING", (0, 0), (0, 0), 0),
            ("RIGHTPADDING", (0, 0), (0, 0), 0),
            ("LEFTPADDING", (1, 0), (-1, -1), 6),
            ("RIGHTPADDING", (1, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(lbl)
        story.append(Spacer(1, 1.5 * mm))

    _pos_st = _ps("_pos", fontName=FONT_BOLD, fontSize=7.5, textColor=GREEN, alignment=TA_RIGHT)
    _neg_st = _ps("_neg", fontName=FONT_BOLD, fontSize=7.5, textColor=RED_C, alignment=TA_RIGHT)

    def _data_tbl(col_names: List[str], col_pcts: List[float], data_rows: List[Dict[str, Any]],
                  right_cols: Optional[List[int]] = None, profit_col: Optional[str] = None):
        right_cols = right_cols or []
        if not data_rows:
            story.append(Paragraph("<font color='#888888'>  No data available.</font>", _ps("_nd", textColor=MID_GREY)))
            story.append(Spacer(1, 3 * mm))
            return
        cw = [PW * p for p in col_pcts]
        header = [
            Paragraph(_esc(col), _ps(f"_th{idx}", fontName=FONT_BOLD, fontSize=6.5, leading=8.5,
                                     textColor=WHITE, alignment=TA_RIGHT if idx in right_cols else TA_CENTER))
            for idx, col in enumerate(col_names)
        ]
        table_data: List[List[Any]] = [header]
        for row in data_rows:
            cells: List[Any] = []
            for idx, col in enumerate(col_names):
                raw = row.get(col, "")
                txt = _esc(_s(raw)) if _s(raw) else "—"
                if col == profit_col:
                    n = _to_num(raw)
                    cells.append(Paragraph(txt, _pos_st if (n is not None and n >= 0) else _neg_st))
                elif idx in right_cols:
                    cells.append(Paragraph(txt, _ps(f"_r{idx}", fontSize=7.5, alignment=TA_RIGHT)))
                else:
                    cells.append(Paragraph(txt, _ps(f"_l{idx}", fontSize=7.5)))
            table_data.append(cells)
        tbl = Table(table_data, colWidths=cw, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), TBL_HDR),
            ("LINEBELOW", (0, 0), (-1, 0), 1.5, COPPER),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, ROW_ALT]),
            ("GRID", (0, 1), (-1, -1), 0.35, BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 2.8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 3.5 * mm))

    _sec_lbl("Open Unallocated Positions and Revaluated PnL")
    kpi_cols = ["Open Pairs", "Current Gold ($/oz)", "Current FX (USD/ZAR)", "Total PnL (ZAR)"]
    kpi_vals = [[
        _fmt(summary.get("open_trades"), 0),
        _fmt(market.get("xau_usd"), 4),
        _fmt(market.get("usd_zar"), 5),
        _fmt(summary.get("total_pnl_zar"), 2),
    ]]
    kpi_tbl = Table([kpi_cols] + kpi_vals, colWidths=[PW * 0.19, PW * 0.27, PW * 0.27, PW * 0.27])
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COPPER_LT),
        ("TEXTCOLOR", (0, 0), (-1, 0), COPPER),
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("TEXTCOLOR", (0, 1), (-1, 1), CHARCOAL),
        ("FONTNAME", (0, 1), (-1, 1), FONT_BOLD),
        ("FONTSIZE", (0, 1), (-1, 1), 9.2),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, COPPER_MD),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 3 * mm))

    table_rows: List[Dict[str, Any]] = []
    for raw in rows if isinstance(rows, list) else []:
        if not isinstance(raw, dict):
            continue
        pair = str(raw.get("pair") or raw.get("trade_num") or raw.get("pair_symbol") or "").upper().replace("-", "/")
        if pair in {"USD/ZAR", "USDZAR"}:
            qty = _to_num(raw.get("fx_qty_usd"))
            if qty is None or abs(qty) <= 1e-9:
                continue
            table_rows.append({
                "pair": "USD/ZAR",
                "side": "LONG" if qty > 0 else "SHORT",
                "net": f"{_fmt(abs(qty), 2)} USD",
                "wa": f"R{_fmt(raw.get('fx_wa_rate'), 5)}",
                "cur": f"R{_fmt(raw.get('market_usd_zar') if raw.get('market_usd_zar') is not None else market.get('usd_zar'), 5)}",
                "pnl": f"R{_fmt(raw.get('fx_pnl_zar'), 2)}",
            })
        elif pair in {"XAU/USD", "XAUUSD"}:
            qty = _to_num(raw.get("gold_qty_oz"))
            if qty is None or abs(qty) <= 1e-9:
                continue
            table_rows.append({
                "pair": "XAU/USD",
                "side": "LONG" if qty > 0 else "SHORT",
                "net": f"{_fmt(abs(qty), 4)} oz",
                "wa": f"${_fmt(raw.get('gold_wa_price'), 4)}/oz",
                "cur": f"${_fmt(raw.get('market_xau_usd') if raw.get('market_xau_usd') is not None else market.get('xau_usd'), 4)}/oz",
                "pnl": f"R{_fmt(raw.get('gold_pnl_zar'), 2)}",
            })

    open_cols = ["Pair", "Net Side", "Net Value", "Weighted Avg Rate", "Current Rate", "Current PnL (ZAR)"]
    open_rows = [
        {
            "Pair": r.get("pair", "--"),
            "Net Side": r.get("side", "--"),
            "Net Value": r.get("net", "--"),
            "Weighted Avg Rate": r.get("wa", "--"),
            "Current Rate": r.get("cur", "--"),
            "Current PnL (ZAR)": r.get("pnl", "--"),
        }
        for r in (table_rows if table_rows else [{"pair": "--", "side": "--", "net": "--", "wa": "--", "cur": "--", "pnl": "--"}])
    ]
    _data_tbl(open_cols, [0.13, 0.12, 0.16, 0.2, 0.18, 0.21], open_rows, right_cols=[2, 3, 4, 5], profit_col="Current PnL (ZAR)")

    _sec_lbl(f"TradeMC Daily Buys and Sells ({today_key})")
    tm_cols = ["Daily Buys (g)", "Daily Sells (g)", "Daily Trades Counted"]
    tm_rows = [{
        "Daily Buys (g)": _fmt(buy_g, 2),
        "Daily Sells (g)": _fmt(sell_g, 2),
        "Daily Trades Counted": _fmt(counted, 0),
    }]
    _data_tbl(tm_cols, [0.33, 0.33, 0.34], tm_rows, right_cols=[0, 1, 2])

    _sec_lbl("Account Balances Recon Table")
    recon_cols = ["Currency", "Opening Balance", "Net Transactions", "Expected Balance", "Actual Balance", "Delta"]
    c_map = recon.get("currencies", {}) if isinstance(recon, dict) else {}
    recon_rows: List[Dict[str, Any]] = []
    for ccy, label, dp in [("USD", "USD (LC)", 2), ("XAU", "XAU (oz)", 4), ("ZAR", "ZAR", 2)]:
        c = c_map.get(ccy, {}) if isinstance(c_map, dict) else {}
        recon_rows.append({
            "Currency": label,
            "Opening Balance": _fmt(c.get("opening_balance"), dp),
            "Net Transactions": _fmt(c.get("transaction_total"), dp),
            "Expected Balance": _fmt(c.get("expected_balance"), dp),
            "Actual Balance": _fmt(c.get("actual_balance"), dp),
            "Delta": _fmt(c.get("delta"), dp),
        })
    _data_tbl(recon_cols, [0.13, 0.17, 0.18, 0.18, 0.18, 0.16], recon_rows, right_cols=[1, 2, 3, 4, 5], profit_col="Delta")

    doc.build(story)
    raw_pdf = buf.getvalue()
    if isinstance(raw_pdf, (bytes, bytearray)):
        return bytes(raw_pdf)
    try:
        return bytes(raw_pdf)
    except Exception:
        return None


def _build_open_positions_reval_pdf_fpdf(args_dict: Dict[str, Any], req_headers: Any = None):
    try:
        from fpdf import FPDF
    except Exception:
        return None
    import os as _os
    from io import BytesIO

    def _to_num(value: Any) -> Optional[float]:
        try:
            n = float(value)
            if math.isnan(n) or math.isinf(n):
                return None
            return n
        except Exception:
            return None

    def _fmt(value: Any, dp: int = 2) -> str:
        n = _to_num(value)
        if n is None:
            return "--"
        return f"{n:,.{dp}f}"

    reval = build_open_positions_reval(args_dict, req_headers)
    rows = reval.get("rows", []) if isinstance(reval, dict) else []
    summary = reval.get("summary", {}) if isinstance(reval, dict) else {}
    market = reval.get("market", {}) if isinstance(reval, dict) else {}
    recon = build_account_recon(args_dict, req_headers)
    tm_df = load_trademc_trades_with_companies()
    if not isinstance(tm_df, pd.DataFrame):
        tm_df = pd.DataFrame()

    today_key = datetime.now().strftime("%Y-%m-%d")
    buy_g = 0.0
    sell_g = 0.0
    counted = 0
    if not tm_df.empty:
        ts_src = (
            tm_df["trade_timestamp"] if "trade_timestamp" in tm_df.columns else
            (tm_df["trade_date"] if "trade_date" in tm_df.columns else
             (tm_df["created_at"] if "created_at" in tm_df.columns else pd.Series([None] * len(tm_df))))
        )
        side_src = (
            tm_df["side"] if "side" in tm_df.columns else
            (tm_df["trade_side"] if "trade_side" in tm_df.columns else pd.Series([""] * len(tm_df)))
        )
        wt_src = (
            tm_df["weight"] if "weight" in tm_df.columns else
            (tm_df["Weight"] if "Weight" in tm_df.columns else
             (tm_df["quantity"] if "quantity" in tm_df.columns else
              (tm_df["qty"] if "qty" in tm_df.columns else pd.Series([None] * len(tm_df)))))
        )
        ts = pd.to_datetime(ts_src, errors="coerce").dt.strftime("%Y-%m-%d")
        side = side_src.fillna("").astype(str).str.upper().str.strip()
        wt = pd.to_numeric(wt_src, errors="coerce")
        for i in range(len(tm_df)):
            if str(ts.iloc[i] or "") != today_key:
                continue
            w = wt.iloc[i]
            if pd.isna(w):
                continue
            w = float(w)
            s = str(side.iloc[i] or "")
            if s == "BUY":
                buy_g += abs(w)
            elif s == "SELL":
                sell_g -= abs(w)
            else:
                if w >= 0:
                    buy_g += w
                else:
                    sell_g += w
            counted += 1

    open_rows: List[List[str]] = []
    for raw in rows if isinstance(rows, list) else []:
        if not isinstance(raw, dict):
            continue
        pair = str(raw.get("pair") or raw.get("trade_num") or raw.get("pair_symbol") or "").upper().replace("-", "/")
        if pair in {"USD/ZAR", "USDZAR"}:
            qty = _to_num(raw.get("fx_qty_usd"))
            if qty is None or abs(qty) <= 1e-9:
                continue
            open_rows.append([
                "USD/ZAR",
                "LONG" if qty > 0 else "SHORT",
                f"{_fmt(abs(qty), 2)} USD",
                f"R{_fmt(raw.get('fx_wa_rate'), 5)}",
                f"R{_fmt(raw.get('market_usd_zar') if raw.get('market_usd_zar') is not None else market.get('usd_zar'), 5)}",
                f"R{_fmt(raw.get('fx_pnl_zar'), 2)}",
            ])
        elif pair in {"XAU/USD", "XAUUSD"}:
            qty = _to_num(raw.get("gold_qty_oz"))
            if qty is None or abs(qty) <= 1e-9:
                continue
            open_rows.append([
                "XAU/USD",
                "LONG" if qty > 0 else "SHORT",
                f"{_fmt(abs(qty), 4)} oz",
                f"${_fmt(raw.get('gold_wa_price'), 4)}/oz",
                f"${_fmt(raw.get('market_xau_usd') if raw.get('market_xau_usd') is not None else market.get('xau_usd'), 4)}/oz",
                f"R{_fmt(raw.get('gold_pnl_zar'), 2)}",
            ])
    if not open_rows:
        open_rows = [["--", "--", "--", "--", "--", "--"]]

    recon_rows: List[List[str]] = []
    c_map = recon.get("currencies", {}) if isinstance(recon, dict) else {}
    for ccy, label, dp in [("USD", "USD (LC)", 2), ("XAU", "XAU (oz)", 4), ("ZAR", "ZAR", 2)]:
        c = c_map.get(ccy, {}) if isinstance(c_map, dict) else {}
        recon_rows.append([
            label,
            _fmt(c.get("opening_balance"), dp),
            _fmt(c.get("transaction_total"), dp),
            _fmt(c.get("expected_balance"), dp),
            _fmt(c.get("actual_balance"), dp),
            _fmt(c.get("delta"), dp),
        ])

    pdf = FPDF("L", "mm", "A4")
    pdf.set_auto_page_break(True, margin=10)
    pdf.add_page()

    # MetCon palette and numeric colors.
    C_CHARCOAL = (28, 28, 28)
    C_COPPER = (176, 120, 64)
    C_COPPER_LT = (251, 245, 236)
    C_HDR = (42, 42, 42)
    C_STRIPE = (246, 246, 246)
    C_GREEN = (0, 166, 81)
    C_RED = (224, 49, 49)

    page_w = 297.0
    lm = 10.0
    usable_w = page_w - (lm * 2)

    logo_path = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "MetCon Logo.png")
    if _os.path.isfile(logo_path):
        try:
            pdf.image(logo_path, x=lm, y=8, w=52, h=13)
        except Exception:
            pass

    pdf.set_xy(130, 8)
    pdf.set_text_color(*C_COPPER)
    pdf.set_font("Arial", "B", 9)
    pdf.cell(157, 5, "METAL CONCENTRATORS SA", 0, 2, "R")
    pdf.set_text_color(*C_CHARCOAL)
    pdf.set_font("Arial", "B", 18)
    pdf.cell(157, 8, "OPEN POSITIONS REVAL REPORT", 0, 2, "R")
    pdf.set_text_color(120, 120, 120)
    pdf.set_font("Arial", "", 8)
    pdf.cell(157, 5, f"Generated {datetime.now().strftime('%d %b %Y  %H:%M')}", 0, 1, "R")
    pdf.set_draw_color(*C_COPPER)
    pdf.set_line_width(0.8)
    pdf.line(lm, 24, lm + usable_w, 24)
    pdf.ln(5)

    def _section_title(title: str):
        y = pdf.get_y()
        pdf.set_fill_color(*C_COPPER)
        pdf.rect(lm, y, 2, 6, style="F")
        pdf.set_fill_color(*C_COPPER_LT)
        pdf.rect(lm + 2, y, usable_w - 2, 6, style="F")
        pdf.set_xy(lm + 4, y + 1.2)
        pdf.set_text_color(*C_COPPER)
        pdf.set_font("Arial", "B", 9)
        pdf.cell(usable_w - 6, 3.6, title.upper(), 0, 1, "L")
        pdf.ln(1.5)

    def _table(headers: List[str], rows_data: List[List[str]], col_widths: List[float], right_cols: Optional[List[int]] = None,
               color_col: Optional[int] = None):
        right_cols = right_cols or []
        widths = [usable_w * w for w in col_widths]
        pdf.set_fill_color(*C_HDR)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Arial", "B", 8)
        for i, h in enumerate(headers):
            align = "R" if i in right_cols else "C"
            pdf.cell(widths[i], 7, h, 0, 0, align, True)
        pdf.ln()
        alt = False
        pdf.set_font("Arial", "", 8)
        for row in rows_data:
            pdf.set_fill_color(*(C_STRIPE if alt else (255, 255, 255)))
            for i, v in enumerate(row):
                align = "R" if i in right_cols else "L"
                text = str(v or "--")
                if color_col is not None and i == color_col:
                    n = _to_num(str(text).replace("R", "").replace(",", "").strip())
                    if n is not None:
                        pdf.set_text_color(*(C_GREEN if n >= 0 else C_RED))
                    else:
                        pdf.set_text_color(*C_CHARCOAL)
                else:
                    pdf.set_text_color(*C_CHARCOAL)
                pdf.cell(widths[i], 6.4, text, 1, 0, align, True)
            pdf.ln()
            alt = not alt
        pdf.ln(2.5)

    _section_title("Open Unallocated Positions and Revaluated PnL")
    _table(
        ["Open Pairs", "Current Gold ($/oz)", "Current FX (USD/ZAR)", "Total PnL (ZAR)"],
        [[
            _fmt(summary.get("open_trades"), 0),
            _fmt(market.get("xau_usd"), 4),
            _fmt(market.get("usd_zar"), 5),
            _fmt(summary.get("total_pnl_zar"), 2),
        ]],
        [0.19, 0.27, 0.27, 0.27],
        right_cols=[0, 1, 2, 3],
        color_col=3,
    )
    _table(
        ["Pair", "Net Side", "Net Value", "Weighted Avg Rate", "Current Rate", "Current PnL (ZAR)"],
        open_rows,
        [0.13, 0.12, 0.16, 0.20, 0.18, 0.21],
        right_cols=[2, 3, 4, 5],
        color_col=5,
    )

    _section_title(f"TradeMC Daily Buys and Sells ({today_key})")
    _table(
        ["Daily Buys (g)", "Daily Sells (g)", "Daily Trades Counted"],
        [[_fmt(buy_g, 2), _fmt(sell_g, 2), _fmt(counted, 0)]],
        [0.33, 0.33, 0.34],
        right_cols=[0, 1, 2],
    )

    _section_title("Account Balances Recon Table")
    _table(
        ["Currency", "Opening Balance", "Net Transactions", "Expected Balance", "Actual Balance", "Delta"],
        recon_rows,
        [0.13, 0.17, 0.18, 0.18, 0.18, 0.16],
        right_cols=[1, 2, 3, 4, 5],
        color_col=5,
    )

    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin-1", errors="ignore")
    if isinstance(out, (bytes, bytearray)):
        return bytes(out)
    return None


# FLASK APP

app = Flask(__name__)
CORS(app, supports_credentials=True)
try:
    _auth_seed_default_users()
except Exception as exc:
    print(f"[WARN] _auth_seed_default_users failed: {exc}")


@app.before_request
def _auth_guard():
    path = str(request.path or "")
    if request.method == "OPTIONS":
        return None
    if not path.startswith("/api"):
        return None
    if path in AUTH_EXEMPT_PATHS:
        return None

    user = _auth_request_user()
    if not user:
        return jsonify({"ok": False, "error": "Authentication required"}), 401

    required_permission = "read" if request.method in {"GET", "HEAD"} else "write"
    if not _auth_has_permission(user, required_permission):
        return jsonify({"ok": False, "error": f"Missing {required_permission} permission"}), 403

    g.current_user = user
    return None


@app.errorhandler(Exception)
def handle_error(e):
    tb = traceback.format_exc()
    print(tb)
    return jsonify({"error": str(e), "traceback": tb}), 500


@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "time": datetime.now().isoformat(), "build": API_BUILD})


@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    data = request.json or {}
    username = _pmx_non_empty(data.get("username"), data.get("email"))
    password = str(data.get("password", "") or "")
    if not username or not password:
        return jsonify({"ok": False, "error": "Username and password are required"}), 400

    user = _auth_find_user_by_username(username)
    if not user or not bool(user.get("is_active")):
        return jsonify({"ok": False, "error": "Invalid username or password"}), 401
    if not _auth_verify_password(password, str(user.get("password_hash", "") or "")):
        return jsonify({"ok": False, "error": "Invalid username or password"}), 401

    token = _auth_issue_token(user)
    res = jsonify({"ok": True, "user": _auth_public_user(user)})
    _auth_set_cookie(res, token)
    return res


@app.route("/api/auth/me")
def auth_me():
    user = _auth_request_user()
    if not user:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401
    return jsonify({"ok": True, "user": _auth_public_user(user)})


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    res = jsonify({"ok": True})
    _auth_clear_cookie(res)
    return res


@app.route("/api/auth/users", methods=["GET"])
def auth_users():
    current_user = _auth_request_user()
    if not current_user or not bool(current_user.get("is_admin")):
        return jsonify({"ok": False, "error": "Admin access required"}), 403
    users = [_auth_admin_user(user) for user in _auth_list_users()]
    return jsonify({"ok": True, "users": users})


@app.route("/api/auth/users", methods=["POST"])
def auth_create_user():
    current_user = _auth_request_user()
    if not current_user or not bool(current_user.get("is_admin")):
        return jsonify({"ok": False, "error": "Admin access required"}), 403

    data = request.json or {}
    username = str(data.get("username", "") or "").strip()
    if not username:
        return jsonify({"ok": False, "error": "username is required"}), 400

    password = str(data.get("password", "") or "").strip()
    if not password:
        return jsonify({"ok": False, "error": "password is required"}), 400

    display_name = str(data.get("display_name", username) or "").strip() or username
    role = str(data.get("role", "viewer") or "").strip() or "viewer"
    can_read = _pmx_bool(data.get("can_read"), default=True)
    can_write = _pmx_bool(data.get("can_write"), default=False)
    is_admin = _pmx_bool(data.get("is_admin"), default=False)
    is_active = _pmx_bool(data.get("is_active"), default=True)

    if can_write and not can_read:
        can_read = True

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT id
        FROM {AUTH_USER_TABLE}
        WHERE lower(username) = lower(?)
        LIMIT 1
        """,
        (username,),
    )
    existing = cur.fetchone()
    if existing:
        conn.close()
        return jsonify({"ok": False, "error": "username already exists"}), 409

    password_hash = _auth_hash_password(password)
    cur.execute(
        f"""
        INSERT INTO {AUTH_USER_TABLE} (
            username, display_name, password_hash, role, can_read, can_write, is_admin, is_active
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            username,
            display_name,
            password_hash,
            role,
            int(bool(can_read)),
            int(bool(can_write)),
            int(bool(is_admin)),
            int(bool(is_active)),
        ),
    )
    created_id = int(cur.lastrowid or 0)
    conn.commit()
    conn.close()

    created = _auth_find_user_by_id(created_id)
    if not created:
        return jsonify({"ok": False, "error": "Failed to load created user"}), 500
    return jsonify({"ok": True, "user": _auth_admin_user(created)})


@app.route("/api/auth/users/<int:user_id>", methods=["PUT"])
def auth_update_user(user_id):
    current_user = _auth_request_user()
    if not current_user or not bool(current_user.get("is_admin")):
        return jsonify({"ok": False, "error": "Admin access required"}), 403

    target = _auth_find_user_by_id(user_id)
    if not target:
        return jsonify({"ok": False, "error": f"User {user_id} not found"}), 404

    data = request.json or {}

    username = str(data.get("username", target.get("username", "")) or "").strip()
    if not username:
        return jsonify({"ok": False, "error": "username is required"}), 400

    display_name = str(data.get("display_name", target.get("display_name", username)) or "").strip() or username
    role = str(data.get("role", target.get("role", "viewer")) or "").strip() or "viewer"
    can_read = _pmx_bool(data.get("can_read"), default=bool(target.get("can_read")))
    can_write = _pmx_bool(data.get("can_write"), default=bool(target.get("can_write")))
    is_admin = _pmx_bool(data.get("is_admin"), default=bool(target.get("is_admin")))
    is_active = _pmx_bool(data.get("is_active"), default=bool(target.get("is_active")))

    if can_write and not can_read:
        can_read = True

    if int(user_id) == int(current_user.get("id") or 0) and not is_active:
        return jsonify({"ok": False, "error": "You cannot deactivate your own account"}), 400

    was_admin_active = bool(target.get("is_admin")) and bool(target.get("is_active"))
    will_be_admin_active = bool(is_admin) and bool(is_active)
    if was_admin_active and not will_be_admin_active:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT COUNT(*)
            FROM {AUTH_USER_TABLE}
            WHERE is_admin = 1 AND is_active = 1 AND id <> ?
            """,
            (int(user_id),),
        )
        other_admins = int(cur.fetchone()[0] or 0)
        conn.close()
        if other_admins <= 0:
            return jsonify({"ok": False, "error": "At least one active admin account is required"}), 400

    password_raw = data.get("password")
    set_password = password_raw is not None
    password_hash = ""
    if set_password:
        password = str(password_raw or "").strip()
        if not password:
            return jsonify({"ok": False, "error": "password cannot be blank"}), 400
        password_hash = _auth_hash_password(password)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT id
        FROM {AUTH_USER_TABLE}
        WHERE lower(username) = lower(?) AND id <> ?
        LIMIT 1
        """,
        (username, int(user_id)),
    )
    existing = cur.fetchone()
    if existing:
        conn.close()
        return jsonify({"ok": False, "error": "username already exists"}), 409

    if set_password:
        cur.execute(
            f"""
            UPDATE {AUTH_USER_TABLE}
            SET username = ?, display_name = ?, password_hash = ?, role = ?,
                can_read = ?, can_write = ?, is_admin = ?, is_active = ?
            WHERE id = ?
            """,
            (
                username,
                display_name,
                password_hash,
                role,
                int(bool(can_read)),
                int(bool(can_write)),
                int(bool(is_admin)),
                int(bool(is_active)),
                int(user_id),
            ),
        )
    else:
        cur.execute(
            f"""
            UPDATE {AUTH_USER_TABLE}
            SET username = ?, display_name = ?, role = ?,
                can_read = ?, can_write = ?, is_admin = ?, is_active = ?
            WHERE id = ?
            """,
            (
                username,
                display_name,
                role,
                int(bool(can_read)),
                int(bool(can_write)),
                int(bool(is_admin)),
                int(bool(is_active)),
                int(user_id),
            ),
        )

    conn.commit()
    conn.close()

    updated = _auth_find_user_by_id(user_id)
    if not updated:
        return jsonify({"ok": False, "error": "Failed to reload updated user"}), 500
    return jsonify({"ok": True, "user": _auth_admin_user(updated)})


@app.route("/api/auth/users/<int:user_id>", methods=["DELETE"])
def auth_delete_user(user_id):
    current_user = _auth_request_user()
    if not current_user or not bool(current_user.get("is_admin")):
        return jsonify({"ok": False, "error": "Admin access required"}), 403

    target = _auth_find_user_by_id(user_id)
    if not target:
        return jsonify({"ok": False, "error": f"User {user_id} not found"}), 404

    if int(user_id) == int(current_user.get("id") or 0):
        return jsonify({"ok": False, "error": "You cannot delete your own account"}), 400

    conn = get_db_connection()
    cur = conn.cursor()

    if bool(target.get("is_admin")) and bool(target.get("is_active")):
        cur.execute(
            f"""
            SELECT COUNT(*)
            FROM {AUTH_USER_TABLE}
            WHERE is_admin = 1 AND is_active = 1 AND id <> ?
            """,
            (int(user_id),),
        )
        other_admins = int(cur.fetchone()[0] or 0)
        if other_admins <= 0:
            conn.close()
            return jsonify({"ok": False, "error": "At least one active admin account is required"}), 400

    cur.execute(
        f"""
        DELETE FROM {AUTH_USER_TABLE}
        WHERE id = ?
        """,
        (int(user_id),),
    )
    conn.commit()
    conn.close()

    return jsonify({"ok": True, "deleted_id": int(user_id)})


@app.route("/api/trades")
def get_trades():
    df = load_all_trades()
    # Convert dates to strings for JSON
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
    return jsonify(df.fillna("").to_dict(orient="records"))


@app.route("/api/trades", methods=["POST"])
def add_trade():
    data = request.json
    result = add_new_trade(data)
    return jsonify({"success": result})


@app.route("/api/trades/<int:trade_id>", methods=["PATCH"])
def patch_trade(trade_id):
    data = request.json
    order_id = data.get("order_id", "")
    result = update_trade_order_id(trade_id, order_id)
    if not result:
        return jsonify({"success": False, "error": f"Trade ID {trade_id} not found"}), 404
    return jsonify({"success": True})


@app.route("/api/trades/backup", methods=["POST"])
def backup_trades_endpoint():
    """Manual trigger to backup trade assignments to JSON."""
    success = backup_manual_trades_to_json()
    if success:
        return jsonify({"success": True, "message": "Manual trades backed up successfully."})
    else:
        return jsonify({"success": False, "error": "Failed to backup manual trades."}), 500


@app.route("/api/trades/<int:trade_id>/trade-number", methods=["PUT"])
def update_trade_number(trade_id):
    data = request.json or {}
    new_trade_num = normalize_trade_number(data.get("trade_number", ""))
    try:
        trade_symbol = _fetch_trade_symbol_for_validation(trade_id, use_pmx=False) or ""
        is_valid, validation_msg = _validate_integer_trade_number_recent_trademc(
            new_trade_num,
            trade_symbol,
            days=7,
        )
        if not is_valid:
            return jsonify({"ok": False, "error": validation_msg}), 400
        updated = update_trade_order_id(trade_id, new_trade_num)
        if not updated:
            return jsonify({"ok": False, "error": f"Trade ID {trade_id} not found"}), 404
        return jsonify({"ok": True, "trade_id": trade_id, "trade_number": new_trade_num})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/trades/ledger")
def get_ledger():
    df = load_all_trades()
    ledger = build_ledger_view(df)
    ledger = _apply_ledger_filters(ledger, request.args)
    return jsonify(_json_safe(ledger.fillna("").to_dict(orient="records")))


def _apply_ledger_filters(ledger: pd.DataFrame, args) -> pd.DataFrame:
    if ledger is None or ledger.empty:
        return ledger if isinstance(ledger, pd.DataFrame) else pd.DataFrame()

    symbol = args.get("symbol")
    trade_num = args.get("trade_num")
    fnc_number = args.get("fnc_number")
    narration = args.get("narration")
    status = args.get("status")
    start_date = args.get("start_date")
    end_date = args.get("end_date")

    if symbol and symbol != "All" and "Symbol" in ledger.columns:
        symbol_norm = str(symbol).replace("/", "").upper()
        ledger_symbol_norm = ledger["Symbol"].astype(str).str.replace("/", "").str.upper()
        ledger = ledger[ledger_symbol_norm == symbol_norm]
    if trade_num and "Trade #" in ledger.columns:
        ledger = ledger[ledger["Trade #"].astype(str).str.contains(trade_num, case=False, na=False, regex=False)]
    if fnc_number and "FNC #" in ledger.columns:
        ledger = ledger[ledger["FNC #"].astype(str).str.contains(fnc_number, case=False, na=False, regex=False)]
    if narration and "Narration" in ledger.columns:
        ledger = ledger[ledger["Narration"].astype(str).str.contains(narration, case=False, na=False, regex=False)]
    if status and status != "All" and "Status" in ledger.columns:
        ledger = ledger[ledger["Status"] == status]
    if start_date and "Trade Date" in ledger.columns:
        ledger = ledger[ledger["Trade Date"] >= start_date]
    if end_date and "Trade Date" in ledger.columns:
        ledger = ledger[ledger["Trade Date"] <= end_date]
    return ledger


@app.route("/api/pmx/sync-ledger", methods=["POST"])
def sync_pmx_ledger():
    data = request.json or {}
    result = sync_pmx_trades_to_db(data, request.headers)
    if bool(result.get("ok")):
        _clear_heavy_route_cache(["hedging:", "pmx_open_positions_reval", "pmx_forward_exposure", "profit_monthly"])
        result["clean_pipeline"] = _trigger_clean_pipeline("pmx_sync", wait=False)
    status_code = 200 if result.get("ok") else 400
    return jsonify(_json_safe(result)), status_code


@app.route("/api/pmx/trades/<int:trade_id>/trade-number", methods=["PUT"])
def update_pmx_trade_number(trade_id):
    data = request.json or {}
    new_trade_num = normalize_trade_number(data.get("trade_number", ""))
    try:
        trade_symbol = _fetch_trade_symbol_for_validation(trade_id, use_pmx=True) or ""
        is_valid, validation_msg = _validate_integer_trade_number_recent_trademc(
            new_trade_num,
            trade_symbol,
            days=7,
        )
        if not is_valid:
            return jsonify({"ok": False, "error": validation_msg}), 400
        updated = update_pmx_trade_order_id(trade_id, new_trade_num)
        if not updated:
            return jsonify({"ok": False, "error": f"PMX trade ID {trade_id} not found"}), 404
        return jsonify({"ok": True, "trade_id": trade_id, "trade_number": new_trade_num})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/pmx/ledger")
def get_pmx_ledger():
    sync_flag = str(request.args.get("sync", "")).strip().lower()
    if sync_flag in {"1", "true", "yes", "y"}:
        current_user = getattr(g, "current_user", None)
        if not _auth_has_permission(current_user or {}, "write"):
            return jsonify({"ok": False, "error": "Missing write permission"}), 403
        sync_result = sync_pmx_trades_to_db(request.args.to_dict(), request.headers)
        if not sync_result.get("ok"):
            return jsonify(_json_safe(sync_result)), 400

    args_dict = request.args.to_dict()
    df = load_all_pmx_trades(args_dict)
    ledger = build_ledger_view(df)
    ledger = _apply_ledger_filters(ledger, request.args)
    return jsonify(_json_safe(ledger.fillna("").to_dict(orient="records")))


@app.route("/api/pmx/ledger-full-csv")
def get_pmx_ledger_full_csv():
    recon_result = _get_pmx_reconciliation_inner()
    recon_response = recon_result[0] if isinstance(recon_result, tuple) else recon_result
    recon_status = recon_result[1] if isinstance(recon_result, tuple) and len(recon_result) > 1 else getattr(recon_response, "status_code", 200)
    if recon_status >= 400:
        return recon_result

    recon_payload = recon_response.get_json(silent=True) if hasattr(recon_response, "get_json") else None
    if not isinstance(recon_payload, dict) or not bool(recon_payload.get("ok")):
        return recon_result

    rows = recon_payload.get("rows")
    rows = rows if isinstance(rows, list) else []
    csv_columns = [
        "Date",
        "Ref #",
        "Type",
        "Trade #",
        "FNC #",
        "Symbol",
        "Side",
        "Net XAU (oz)",
        "Net USD",
        "Net ZAR",
        "Narration",
    ]

    def _fmt_csv_num(value: Any, decimals: int) -> str:
        num = _parse_loose_number(value)
        if num is None:
            return ""
        return f"{float(num):.{decimals}f}"

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(csv_columns)

    for row in rows:
        if not isinstance(row, dict):
            continue
        net_oz = _parse_loose_number(row.get("net_oz"))
        net_usd = _parse_loose_number(row.get("net_usd"))
        net_zar = _parse_loose_number(row.get("net_zar"))
        has_net = any(val is not None and abs(float(val)) > 1e-12 for val in (net_oz, net_usd, net_zar))
        if not has_net:
            continue
        writer.writerow([
            row.get("date", "") or row.get("trade_date", "") or row.get("value_date", ""),
            row.get("doc_number", ""),
            row.get("row_type", ""),
            row.get("trade_number", ""),
            row.get("fnc_number", ""),
            row.get("symbol", ""),
            row.get("side", ""),
            _fmt_csv_num(row.get("net_oz"), 4),
            _fmt_csv_num(row.get("net_usd"), 4),
            _fmt_csv_num(row.get("net_zar"), 4),
            row.get("narration", ""),
        ])

    content = csv_buffer.getvalue()
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"pmx_reconciliation_consolidated_{stamp}.csv"
    resp = Response(content, mimetype="text/csv; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

    args_dict = request.args.to_dict()
    now = datetime.now()

    start_date = _normalize_pmx_date_param(
        str(args_dict.get("start_date", "") or "").strip(),
        default_dt=now - timedelta(days=7),
    )
    end_date = _normalize_pmx_date_param(
        str(args_dict.get("end_date", "") or "").strip(),
        default_dt=now,
    )

    resolved_headers = _pmx_resolve_headers(args_dict, request.headers, auto_login=True)
    x_auth = str(resolved_headers.get("x_auth", "") or "").strip()
    sid = str(resolved_headers.get("sid", "") or "").strip()
    username = str(resolved_headers.get("username", "") or "").strip()
    platform = str(resolved_headers.get("platform", "") or "").strip()
    location = str(resolved_headers.get("location", "") or "").strip()
    cache_control = str(resolved_headers.get("cache_control", "") or "").strip()
    content_type = str(resolved_headers.get("content_type", "") or "").strip()

    # PMX statement view required for full recon:
    # Column 1 = Local Currency, Column 2 = XAU (PMX token is GLD).
    statement_col1 = str(args_dict.get("col1", "") or "").strip() or "LC"
    statement_col2 = str(args_dict.get("col2", "") or "").strip() or "GLD"
    unit_code1 = str(args_dict.get("unit_code1", "") or "").strip()
    unit_code2 = str(args_dict.get("unit_code2", "") or "").strip()
    if not unit_code1 and statement_col1.upper() in {"LC", "LOCAL", "LOCAL CURRENCY"}:
        unit_code1 = ""
    if not unit_code2 and statement_col2.upper() in {"GLD", "XAU", "GOLD"}:
        unit_code2 = "OZ"

    base_fetch_args: Dict[str, Any] = {
        "start_date": start_date,
        "end_date": end_date,
        "acc_code": str(os.getenv("PMX_ACC_OPT_KEY", "MT0601") or "MT0601"),
        "report_type": "docDate",
        "col1": statement_col1,
        "col2": statement_col2,
        "unit_code1": unit_code1,
        "unit_code2": unit_code2,
        "option": "1",
        "host": str(os.getenv("PMX_API_HOST", "pmxapi.stonex.com") or "pmxapi.stonex.com"),
        "path": "/user/account_statementReport",
        "authorization": str(args_dict.get("authorization", "") or ""),
        "cookie": str(args_dict.get("cookie", "") or ""),
        "x_auth": x_auth,
        "sid": sid,
        "username": username,
        "platform": platform,
        "location": location,
        "cache_control": cache_control,
        "content_type": content_type,
        "extra_headers": {"usercode": username} if username else {},
        "origin": "https://pmxecute.stonex.com",
        "referer": "https://pmxecute.stonex.com/",
        "timeout": int(args_dict.get("timeout", "180") or "180"),
    }

    account_balances = _fetch_open_positions_account_balances(args_dict, request.headers)
    account_xau = _parse_loose_number(account_balances.get("xau")) if isinstance(account_balances, dict) else None
    baseline_date = str(args_dict.get("baseline_date", "") or "").strip() or "2026-03-01"

    baseline_xau = _parse_loose_number(args_dict.get("baseline_xau"))
    if baseline_xau is None:
        # Dynamically fetch closing XAU balance from PMX statement for the baseline date
        try:
            bl_args = dict(base_fetch_args)
            bl_args["start_date"] = baseline_date
            bl_args["end_date"] = baseline_date
            bl_result = fetch_pmx_account_statement_report(**bl_args)
            bl_payload = bl_result.get("json")
            if bl_payload is None:
                bl_body = bl_result.get("body", "")
                if isinstance(bl_body, str) and bl_body.strip().startswith(("{", "[")):
                    try:
                        bl_payload = json.loads(bl_body)
                    except Exception:
                        bl_payload = {}
                else:
                    bl_payload = {}
            if bl_result.get("ok"):
                bl_rows = extract_pmx_statement_report_rows(bl_payload)
                # Walk rows in reverse to find the last XAU balance
                for bl_row in reversed(bl_rows):
                    if not isinstance(bl_row, dict):
                        continue
                    for bk, bv in bl_row.items():
                        bk_u = str(bk or "").strip().upper()
                        if "BAL" not in bk_u:
                            continue
                        if not any(tok in bk_u for tok in ("XAU", "GLD", "GOLD", "OZ", "COL2")):
                            continue
                        num = _parse_loose_number(bv)
                        if num is not None:
                            baseline_xau = float(num)
                            break
                    if baseline_xau is not None:
                        break
        except Exception:
            pass
    if baseline_xau is None:
        baseline_xau = -657.171

    result = fetch_pmx_account_statement_report(**base_fetch_args)
    payload = result.get("json")
    if payload is None:
        body_text = result.get("body", "")
        if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
            try:
                payload = json.loads(body_text)
            except Exception:
                payload = {}
        else:
            payload = {}

    result = _pmx_mark_failed_payload(result, payload)
    # Retry once with a fresh PMX login on any failure. PMX sometimes returns
    # generic 500s for expired/invalid sessions instead of explicit auth codes.
    if not result.get("ok"):
        relogin = _pmx_login_session(args_dict)
        if relogin.get("ok"):
            base_fetch_args["x_auth"] = _pmx_non_empty(relogin.get("x_auth"), x_auth)
            base_fetch_args["sid"] = _pmx_non_empty(relogin.get("sid"), sid)
            base_fetch_args["username"] = _pmx_non_empty(relogin.get("username"), username)
            base_fetch_args["platform"] = _pmx_non_empty(relogin.get("platform"), platform)
            base_fetch_args["location"] = _pmx_non_empty(relogin.get("location"), location)
            base_fetch_args["cache_control"] = _pmx_non_empty(relogin.get("cache_control"), cache_control)
            base_fetch_args["content_type"] = _pmx_non_empty(relogin.get("content_type"), content_type)
            result = fetch_pmx_account_statement_report(**base_fetch_args)
            payload = result.get("json")
            if payload is None:
                body_text = result.get("body", "")
                if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
                    try:
                        payload = json.loads(body_text)
                    except Exception:
                        payload = {}
                else:
                    payload = {}
            result = _pmx_mark_failed_payload(result, payload)

    if not result.get("ok"):
        status = int(result.get("status") or 502)
        return jsonify({
            "ok": False,
            "error": _pmx_human_error(result, payload, "Failed to fetch PMX account statement report"),
            "status": status,
        }), status

    rows = extract_pmx_statement_report_rows(payload)

    # Full statement export should include every transaction (allocated + unallocated).
    # Do not apply trade/symbol/narration narrowing filters here.
    trade_num_filter = ""
    narration_filter = ""
    symbol_filter = ""

    def _first_non_empty_from(row: Dict[str, Any], keys: List[str]) -> str:
        for key in keys:
            if key not in row:
                continue
            value = row.get(key)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                return text
        return ""

    def _find_symbol_in_text(text: str) -> str:
        match = re.search(r"\b([A-Z]{3}/[A-Z]{3})\b", str(text or "").upper())
        return match.group(1) if match else ""

    def _find_oz_in_text(text: str) -> Optional[float]:
        match = re.search(r"([+-]?\d[\d,]*\.?\d*)\s*OZ\b", str(text or "").upper())
        if not match:
            return None
        return _parse_loose_number(match.group(1))

    def _extract_stmt_xau_balance(row: Dict[str, Any]) -> Optional[float]:
        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if "BAL" not in key_u:
                continue
            if not any(token in key_u for token in ("XAU", "GLD", "GOLD", "OZ", "COL2")):
                continue
            num = _parse_loose_number(value)
            if num is not None:
                return float(num)
        return None

    def _extract_oz_from_statement_row(row: Dict[str, Any], side_hint: str, symbol_norm: str, narration: str) -> Optional[float]:
        side_up = str(side_hint or "").strip().upper()
        debit_xau = None
        credit_xau = None
        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if not any(token in key_u for token in ("XAU", "GLD", "GOLD", "OZ", "COL2")):
                continue
            num = _parse_loose_number(value)
            if num is None:
                continue
            if "DEBIT" in key_u or re.search(r"\bDR\b", key_u):
                debit_xau = float(num)
            elif "CREDIT" in key_u or re.search(r"\bCR\b", key_u):
                credit_xau = float(num)

        if credit_xau is not None or debit_xau is not None:
            return (credit_xau or 0.0) - (debit_xau or 0.0)

        qty_num = None
        for key in ("qty", "Quantity", "last_qty", "LastQty", "quantity", "Qty"):
            if key in row:
                parsed = _parse_loose_number(row.get(key))
                if parsed is not None:
                    qty_num = float(parsed)
                    break

        text_oz = _find_oz_in_text(narration)
        if symbol_norm.startswith(("XAU", "XAG", "XPT", "XPD")) and symbol_norm.endswith("USD"):
            base_oz = abs(qty_num) if qty_num is not None else (abs(float(text_oz)) if text_oz is not None else None)
            if base_oz is not None:
                if side_up == "BUY":
                    return base_oz
                if side_up == "SELL":
                    return -base_oz
                return base_oz

        if text_oz is not None:
            if side_up == "SELL":
                return -abs(float(text_oz))
            if side_up == "BUY":
                return abs(float(text_oz))
            return float(text_oz)
        return None

    def _find_doc_token(text: str) -> str:
        match = re.search(r"\b((?:FNC|JRV|JRC|MER)\/\d{4}\/\d+)\b", str(text or "").upper())
        return match.group(1) if match else ""

    def _extract_trade_number_from_narration(text: str) -> str:
        narr = str(text or "").strip()
        if not narr:
            return ""
        narr_u = narr.upper()
        # Prefer explicit alpha-numeric trade tokens, e.g. KAS-016 / JOS-070.
        m = re.search(r"\b([A-Z]{2,6}\s*-\s*\d{2,7})\b", narr_u)
        if m:
            return normalize_trade_number(m.group(1).replace(" ", ""))
        # Prod/production markers often carry the internal numeric trade number.
        m = re.search(r"\bPROD\s*#?\s*(\d{3,8})\b", narr_u)
        if m:
            return normalize_trade_number(m.group(1))
        # Then support leading numeric trade ids before pair symbols, e.g. "9872 XAU/USD ..."
        m = re.search(r"^\s*(\d{4,8})\b(?=.*\b(?:XAU|XAG|XPT|XPD|USD)/[A-Z]{3}\b)", narr_u)
        if m:
            return normalize_trade_number(m.group(1))
        return ""

    # Build a reliable fallback mapping from PMX deal report first (live),
    # then local synced PMX ledger rows:
    # Doc/FNC number -> Trade #. This recovers trade numbers when statement rows
    # omit explicit order/trade fields.
    trade_num_by_doc: Dict[str, str] = {}
    try:
        deal_result = fetch_pmx_alldeal_filter_report(
            start_date=start_date,
            end_date=end_date,
            cmdty="All",
            trd_opt="All",
            created_by=str(os.getenv("PMX_CREATED_BY", "2") or "2"),
            acc_opt_key=str(os.getenv("PMX_ACC_OPT_KEY", "MT0601") or "MT0601"),
            trade_type="TD",
            non_trd_cmdty="",
            host=str(os.getenv("PMX_API_HOST", "pmxapi.stonex.com") or "pmxapi.stonex.com"),
            path="/user/alldealFilter_report",
            authorization=str(args_dict.get("authorization", "") or ""),
            cookie=str(args_dict.get("cookie", "") or ""),
            x_auth=base_fetch_args.get("x_auth", ""),
            sid=base_fetch_args.get("sid", ""),
            username=base_fetch_args.get("username", ""),
            platform=base_fetch_args.get("platform", ""),
            location=base_fetch_args.get("location", ""),
            cache_control=base_fetch_args.get("cache_control", ""),
            content_type=base_fetch_args.get("content_type", ""),
            origin="https://pmxecute.stonex.com",
            referer="https://pmxecute.stonex.com/",
            timeout=int(base_fetch_args.get("timeout", 180) or 180),
        )
        deal_payload = deal_result.get("json")
        if deal_payload is None:
            body_text = deal_result.get("body", "")
            if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
                try:
                    deal_payload = json.loads(body_text)
                except Exception:
                    deal_payload = {}
            else:
                deal_payload = {}
        if bool(deal_result.get("ok")):
            deal_rows = extract_pmx_report_rows(deal_payload)
            for idx, deal_row in enumerate(deal_rows, start=1):
                if not isinstance(deal_row, dict):
                    continue
                mapped = _pmx_map_row_to_trade(deal_row, idx)
                if not mapped:
                    continue
                doc_token = _pmx_extract_support_doc(mapped.get("doc_number", ""), mapped.get("narration", ""))
                tn = normalize_trade_number(mapped.get("order_id", ""))
                if doc_token and tn and doc_token not in trade_num_by_doc:
                    trade_num_by_doc[doc_token] = tn
    except Exception:
        pass

    try:
        local_df = load_all_pmx_trades({})
        if isinstance(local_df, pd.DataFrame) and not local_df.empty:
            doc_col = "Doc #" if "Doc #" in local_df.columns else ("doc_number" if "doc_number" in local_df.columns else "")
            trade_col = "Trade #" if "Trade #" in local_df.columns else ("OrderID" if "OrderID" in local_df.columns else "")
            if doc_col and trade_col:
                docs = local_df[doc_col].fillna("").astype(str).str.strip()
                trades = local_df[trade_col].fillna("").astype(str).apply(normalize_trade_number)
                for doc_raw, tn_raw in zip(docs.tolist(), trades.tolist()):
                    doc_token = _pmx_extract_support_doc(doc_raw, "")
                    tn = normalize_trade_number(tn_raw)
                    if doc_token and tn and doc_token not in trade_num_by_doc:
                        trade_num_by_doc[doc_token] = tn
    except Exception:
        trade_num_by_doc = {}

    def _extract_date_from_row(row: Dict[str, Any], preferred_keys: List[str]) -> str:
        # 1) Try explicit known keys first.
        for key in preferred_keys:
            if key in row:
                parsed = _pmx_parse_date(row.get(key), default_value="")
                if parsed:
                    return parsed
        # 2) Try any field with "date" in the key name.
        for key, value in row.items():
            key_text = str(key or "").strip().lower()
            if "date" not in key_text:
                continue
            parsed = _pmx_parse_date(value, default_value="")
            if parsed:
                return parsed
        # 3) Fallback: scan scalar values for dd-Mon-yyyy or yyyy-mm-dd-like strings.
        for value in row.values():
            text = str(value or "").strip()
            if not text:
                continue
            for token in re.findall(r"\b\d{1,2}[-/][A-Za-z]{3}[-/]\d{4}\b|\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b|\b\d{1,2}[-/]\d{1,2}[-/]\d{4}\b", text):
                parsed = _pmx_parse_date(token, default_value="")
                if parsed:
                    return parsed
        return ""

    clean_rows: List[Dict[str, Any]] = []
    for raw in rows:
        if not isinstance(raw, dict):
            continue

        row_text = " ".join(str(v or "") for v in raw.values()).upper()
        if "SWT/" in row_text or " SWAP " in row_text:
            continue

        doc_number = _first_non_empty_from(raw, ["docno", "DocNo", "doc_number", "DocNumber", "document_no", "Doc #"])
        if not doc_number:
            doc_number = _first_non_empty_from(raw, ["NeoId", "neo_id", "TagNumber", "tag_number"])
        if not doc_number:
            doc_number = _find_doc_token(row_text)
        doc_number = doc_number.strip()
        if not doc_number:
            continue

        doc_upper = doc_number.upper()
        if doc_upper.startswith("FNC/"):
            row_type = "FNC"
        elif doc_upper.startswith("JRV/"):
            row_type = "JRV"
        elif doc_upper.startswith("JRC/"):
            row_type = "JRC"
        elif doc_upper.startswith("MER/"):
            row_type = "MER"
        else:
            # Keep only requested document families.
            continue

        trade_date = _extract_date_from_row(raw, ["docdate", "TradeDate", "trade_date", "DocDate", "date", "Trade Date", "Doc Date"])
        value_date = _extract_date_from_row(raw, ["valdate", "ValueDate", "value_date", "settlement_date", "Value Date", "Settlement Date"])
        row_date = trade_date or value_date

        narration = _first_non_empty_from(
            raw,
            ["remarks", "remarks1", "comment", "notes", "description", "ContractDescription", "Narration"],
        )
        fnc_number = _pmx_extract_support_doc(doc_number, narration)
        trade_num = normalize_trade_number(
            _first_non_empty_from(raw, ["order_id", "OrderId", "trade_number", "trade_no", "ref_number", "OrderID"])
        )
        if not trade_num:
            trade_num = _extract_trade_number_from_narration(narration)
        if not trade_num:
            doc_lookup = _pmx_extract_support_doc(doc_number, narration)
            if doc_lookup:
                trade_num = normalize_trade_number(trade_num_by_doc.get(doc_lookup, ""))

        symbol_pair = _pmx_to_currency_pair(
            _first_non_empty_from(raw, ["CurrencyPair", "currency_pair", "cmdty", "stk_type_name", "inst_desc", "Symbol"])
        )
        if not symbol_pair:
            symbol_pair = _find_symbol_in_text(narration)
        symbol_pair = symbol_pair.upper()
        symbol_norm = symbol_pair.replace("/", "").replace("-", "").replace(" ", "")

        side = _first_non_empty_from(raw, ["side", "Side", "deal_type", "trd_opt"]).upper()
        if side not in {"BUY", "SELL"}:
            narr_upper = narration.upper()
            if " SELL " in f" {narr_upper} ":
                side = "SELL"
            elif " BUY " in f" {narr_upper} ":
                side = "BUY"
            else:
                side = ""

        qty = None
        for key in ("qty", "Quantity", "last_qty", "LastQty", "quantity", "Qty"):
            if key in raw:
                parsed = _parse_loose_number(raw.get(key))
                if parsed is not None:
                    qty = float(parsed)
                    break

        price = None
        for key in ("px", "price", "Price", "last_px", "LastPx", "rate"):
            if key in raw:
                parsed = _parse_loose_number(raw.get(key))
                if parsed is not None:
                    price = float(parsed)
                    break

        oz_value = _extract_oz_from_statement_row(raw, side, symbol_norm, narration)
        stmt_balance_xau = _extract_stmt_xau_balance(raw)
        buy_oz = abs(float(oz_value)) if oz_value is not None and oz_value > 0 else 0.0
        sell_oz = abs(float(oz_value)) if oz_value is not None and oz_value < 0 else 0.0

        clean = {
            "Doc #": doc_number,
            "Row Type": row_type,
            "Trade #": trade_num or "",
            "FNC #": fnc_number or "",
            "Date": row_date,
            "Trade Date": trade_date,
            "Value Date": value_date,
            "Symbol": symbol_pair or "",
            "Side": side or "",
            "Oz": oz_value,
            "Buy Oz": buy_oz,
            "Sell Oz": sell_oz,
            "Net Oz": oz_value if oz_value is not None else "",
            "Stmt Balance XAU": stmt_balance_xau,
            "Quantity": qty,
            "Price": price,
            "Narration": narration or "",
            "Running Net Oz": None,
            "Expected XAU (From Baseline)": None,
            "Delta vs Stmt XAU": None,
            "Account XAU": None,
            "Delta To Account XAU": None,
        }

        row_filter_text = " ".join(str(v or "") for v in clean.values()).upper()
        if trade_num_filter and trade_num_filter not in row_filter_text:
            continue
        if narration_filter and narration_filter not in row_filter_text:
            continue
        # Keep JRV/MER rows in recon export even when symbol filter is applied,
        # because they often do not carry pair symbols but must tie to balance.
        if symbol_filter and row_type == "FNC" and symbol_norm != symbol_filter:
            continue
        clean_rows.append(clean)

    # Sort by dates then doc number for recon readability.
    def _sort_key(row: Dict[str, Any]) -> Tuple[str, str, str]:
        return (
            str(row.get("Trade Date") or ""),
            str(row.get("Value Date") or ""),
            str(row.get("Doc #") or ""),
        )

    clean_rows = sorted(clean_rows, key=_sort_key)
    running_oz = 0.0
    running_has_value = False
    for row in clean_rows:
        row_date = str(row.get("Date") or row.get("Trade Date") or "").strip()
        if row_date and row_date < baseline_date:
            row["Running Net Oz"] = ""
            row["Expected XAU (From Baseline)"] = ""
            row["Delta vs Stmt XAU"] = ""
            continue

        oz_num = _parse_loose_number(row.get("Net Oz"))
        if oz_num is None:
            row["Running Net Oz"] = ""
            row["Expected XAU (From Baseline)"] = ""
            row["Delta vs Stmt XAU"] = ""
            continue
        running_oz += float(oz_num)
        running_has_value = True
        expected_xau = float(baseline_xau) + running_oz
        row["Running Net Oz"] = running_oz
        row["Expected XAU (From Baseline)"] = expected_xau
        stmt_bal = _parse_loose_number(row.get("Stmt Balance XAU"))
        if stmt_bal is not None:
            row["Delta vs Stmt XAU"] = float(stmt_bal) - expected_xau

    if clean_rows and account_xau is not None:
        clean_rows[-1]["Account XAU"] = account_xau
        if running_has_value:
            clean_rows[-1]["Delta To Account XAU"] = (float(baseline_xau) + running_oz) - float(account_xau)

    csv_columns = [
        "Doc #",
        "Row Type",
        "Trade #",
        "FNC #",
        "Date",
        "Trade Date",
        "Value Date",
        "Symbol",
        "Side",
        "Oz",
        "Buy Oz",
        "Sell Oz",
        "Net Oz",
        "Running Net Oz",
        "Stmt Balance XAU",
        "Expected XAU (From Baseline)",
        "Delta vs Stmt XAU",
        "Account XAU",
        "Delta To Account XAU",
        "Quantity",
        "Price",
        "Narration",
    ]

    def _fmt_num(value: Any, decimals: int) -> str:
        num = _parse_loose_number(value)
        if num is None:
            return ""
        return f"{float(num):.{decimals}f}"

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(csv_columns)
    for row in clean_rows:
        writer.writerow([
            row.get("Doc #", ""),
            row.get("Row Type", ""),
            row.get("Trade #", ""),
            row.get("FNC #", ""),
            row.get("Date", ""),
            row.get("Trade Date", ""),
            row.get("Value Date", ""),
            row.get("Symbol", ""),
            row.get("Side", ""),
            _fmt_num(row.get("Oz"), 3),
            _fmt_num(row.get("Buy Oz"), 3),
            _fmt_num(row.get("Sell Oz"), 3),
            _fmt_num(row.get("Net Oz"), 3),
            _fmt_num(row.get("Running Net Oz"), 3),
            _fmt_num(row.get("Stmt Balance XAU"), 3),
            _fmt_num(row.get("Expected XAU (From Baseline)"), 3),
            _fmt_num(row.get("Delta vs Stmt XAU"), 3),
            _fmt_num(row.get("Account XAU"), 3),
            _fmt_num(row.get("Delta To Account XAU"), 3),
            _fmt_num(row.get("Quantity"), 3),
            _fmt_num(row.get("Price"), 5),
            row.get("Narration", ""),
        ])

    content = csv_buffer.getvalue()
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"pmx_full_report_{stamp}.csv"
    resp = Response(content, mimetype="text/csv; charset=utf-8")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@app.route("/api/pmx/reconciliation")
def get_pmx_reconciliation():
    """Return PMX statement reconciliation data as JSON for the XAU Reconciliation tab."""
    try:
        return _get_pmx_reconciliation_inner()
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": f"Reconciliation error: {exc}"}), 500


def _r_first_non_empty(row: Dict[str, Any], keys: List[str]) -> str:
    for key in keys:
        if key not in row:
            continue
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _r_find_symbol(text: str) -> str:
    match = re.search(r"\b([A-Z]{3}/[A-Z]{3})\b", str(text or "").upper())
    return match.group(1) if match else ""


def _r_find_oz(text: str) -> Optional[float]:
    match = re.search(r"([+-]?\d[\d,]*\.?\d*)\s*OZ\b", str(text or "").upper())
    if not match:
        return None
    return _parse_loose_number(match.group(1))


def _r_extract_xau_balance(row: Dict[str, Any]) -> Optional[float]:
    for key, value in row.items():
        key_u = str(key or "").strip().upper()
        if "BAL" not in key_u:
            continue
        if not any(token in key_u for token in ("XAU", "GLD", "GOLD", "OZ", "COL2")):
            continue
        num = _parse_loose_number(value)
        if num is not None:
            return float(num)
    return None


def _r_extract_balance_for_currency(row: Dict[str, Any], currency: str, col_hint: str = "") -> Optional[float]:
    cur = str(currency or "").strip().upper()
    hint = str(col_hint or "").strip().upper()
    if cur not in {"USD", "ZAR"}:
        return None
    for key, value in row.items():
        key_u = str(key or "").strip().upper()
        if "BAL" not in key_u:
            continue
        if cur in key_u:
            num = _parse_loose_number(value)
            if num is not None:
                return float(num)
    if hint:
        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if "BAL" not in key_u:
                continue
            if hint in key_u:
                num = _parse_loose_number(value)
                if num is not None:
                    return float(num)
    if hint in {"USD", "ZAR", "COL1", "LC"}:
        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if "BAL" not in key_u:
                continue
            num = _parse_loose_number(value)
            if num is not None:
                return float(num)
    return None


def _r_extract_debit_credit_for_currency(
    row: Dict[str, Any],
    currency: str,
    col_hint: str = "",
) -> Tuple[Optional[float], Optional[float]]:
    cur = str(currency or "").strip().upper()
    hint = str(col_hint or "").strip().upper()
    if cur not in {"USD", "ZAR"}:
        return None, None
    debit_val: Optional[float] = None
    credit_val: Optional[float] = None

    def _maybe_set(key_u: str, raw_value: Any) -> None:
        nonlocal debit_val, credit_val
        num = _parse_loose_number(raw_value)
        if num is None:
            return
        if "BAL" in key_u:
            return
        is_debit = ("DEBIT" in key_u) or bool(re.search(r"\bDR\b", key_u))
        is_credit = ("CREDIT" in key_u) or bool(re.search(r"\bCR\b", key_u))
        if not (is_debit or is_credit):
            return
        if is_debit:
            debit_val = float(num)
        if is_credit:
            credit_val = float(num)

    for key, value in row.items():
        key_u = str(key or "").strip().upper()
        if cur in key_u:
            _maybe_set(key_u, value)
    if (debit_val is None and credit_val is None) and hint:
        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if hint in key_u:
                _maybe_set(key_u, value)
    if debit_val is None and credit_val is None and hint in {"USD", "ZAR", "COL1", "LC"}:
        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if cur in key_u or hint in key_u:
                continue
            _maybe_set(key_u, value)
    return debit_val, credit_val


def _r_extract_oz(row: Dict[str, Any], side_hint: str, symbol_norm: str, narration: str) -> Optional[float]:
    side_up = str(side_hint or "").strip().upper()
    debit_xau = None
    credit_xau = None
    for key, value in row.items():
        key_u = str(key or "").strip().upper()
        if not any(token in key_u for token in ("XAU", "GLD", "GOLD", "OZ", "COL2")):
            continue
        num = _parse_loose_number(value)
        if num is None:
            continue
        if "DEBIT" in key_u or re.search(r"\bDR\b", key_u):
            debit_xau = float(num)
        elif "CREDIT" in key_u or re.search(r"\bCR\b", key_u):
            credit_xau = float(num)
    if credit_xau is not None or debit_xau is not None:
        return (credit_xau or 0.0) - (debit_xau or 0.0)
    qty_num = None
    for key in ("qty", "Quantity", "last_qty", "LastQty", "quantity", "Qty"):
        if key in row:
            parsed = _parse_loose_number(row.get(key))
            if parsed is not None:
                qty_num = float(parsed)
                break
    text_oz = _r_find_oz(narration)
    if symbol_norm.startswith(("XAU", "XAG", "XPT", "XPD")) and symbol_norm.endswith("USD"):
        base_oz = abs(qty_num) if qty_num is not None else (abs(float(text_oz)) if text_oz is not None else None)
        if base_oz is not None:
            if side_up == "BUY":
                return base_oz
            if side_up == "SELL":
                return -base_oz
            return base_oz
    if text_oz is not None:
        if side_up == "SELL":
            return -abs(float(text_oz))
        if side_up == "BUY":
            return abs(float(text_oz))
        return float(text_oz)
    return None


def _r_find_doc_token(text: str) -> str:
    match = re.search(r"\b((?:FNC|JRV|JRC|MER|SWT)\/\d{4}\/\d+)\b", str(text or "").upper())
    return match.group(1) if match else ""


def _r_extract_trade_from_narration(text: str) -> str:
    narr = str(text or "").strip()
    if not narr:
        return ""
    narr_u = narr.upper()
    m = re.search(r"\b([A-Z]{2,6}\s*-\s*\d{2,7})\b", narr_u)
    if m:
        return normalize_trade_number(m.group(1).replace(" ", ""))
    m = re.search(r"\bPROD\s*#?\s*(\d{3,8})\b", narr_u)
    if m:
        return normalize_trade_number(m.group(1))
    m = re.search(r"^\s*(\d{4,8})\b(?=.*\b(?:XAU|XAG|XPT|XPD|USD)/[A-Z]{3}\b)", narr_u)
    if m:
        return normalize_trade_number(m.group(1))
    return ""


def _r_extract_date(row: Dict[str, Any], preferred_keys: List[str]) -> str:
    for key in preferred_keys:
        if key in row:
            parsed = _pmx_parse_date(row.get(key), default_value="")
            if parsed:
                return parsed
    for key, value in row.items():
        key_text = str(key or "").strip().lower()
        if "date" not in key_text:
            continue
        parsed = _pmx_parse_date(value, default_value="")
        if parsed:
            return parsed
    for value in row.values():
        text = str(value or "").strip()
        if not text:
            continue
        for token in re.findall(r"\b\d{1,2}[-/][A-Za-z]{3}[-/]\d{4}\b|\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b|\b\d{1,2}[-/]\d{1,2}[-/]\d{4}\b", text):
            parsed = _pmx_parse_date(token, default_value="")
            if parsed:
                return parsed
    return ""


def _r_build_stmt_row_key(doc_number: str, trade_date: str, value_date: str, narration: str, occurrence: int) -> str:
    doc_key = str(doc_number or "").strip().upper()
    narr_key = re.sub(r"\s+", " ", str(narration or "").strip().upper())
    return f"{trade_date}|{value_date}|{doc_key}|{narr_key}|{int(occurrence)}"


def _get_pmx_reconciliation_inner():
    args_dict = request.args.to_dict()
    now = datetime.now()

    start_date = _normalize_pmx_date_param(
        str(args_dict.get("start_date", "") or "").strip(),
        default_dt=now - timedelta(days=7),
    )
    end_date = _normalize_pmx_date_param(
        str(args_dict.get("end_date", "") or "").strip(),
        default_dt=now,
    )

    resolved_headers = _pmx_resolve_headers(args_dict, request.headers, auto_login=True)
    x_auth = str(resolved_headers.get("x_auth", "") or "").strip()
    sid = str(resolved_headers.get("sid", "") or "").strip()
    username = str(resolved_headers.get("username", "") or "").strip()
    platform = str(resolved_headers.get("platform", "") or "").strip()
    location = str(resolved_headers.get("location", "") or "").strip()
    cache_control = str(resolved_headers.get("cache_control", "") or "").strip()
    content_type = str(resolved_headers.get("content_type", "") or "").strip()

    statement_col1 = str(args_dict.get("col1", "") or "").strip() or "LC"
    statement_col2 = str(args_dict.get("col2", "") or "").strip() or "GLD"
    unit_code1 = str(args_dict.get("unit_code1", "") or "").strip()
    unit_code2 = str(args_dict.get("unit_code2", "") or "").strip()
    if not unit_code1 and statement_col1.upper() in {"LC", "LOCAL", "LOCAL CURRENCY"}:
        unit_code1 = ""
    if not unit_code2 and statement_col2.upper() in {"GLD", "XAU", "GOLD"}:
        unit_code2 = "OZ"

    base_fetch_args: Dict[str, Any] = {
        "start_date": start_date,
        "end_date": end_date,
        "acc_code": str(os.getenv("PMX_ACC_OPT_KEY", "MT0601") or "MT0601"),
        "report_type": "docDate",
        "col1": statement_col1,
        "col2": statement_col2,
        "unit_code1": unit_code1,
        "unit_code2": unit_code2,
        "option": "1",
        "host": str(os.getenv("PMX_API_HOST", "pmxapi.stonex.com") or "pmxapi.stonex.com"),
        "path": "/user/account_statementReport",
        "authorization": str(args_dict.get("authorization", "") or ""),
        "cookie": str(args_dict.get("cookie", "") or ""),
        "x_auth": x_auth,
        "sid": sid,
        "username": username,
        "platform": platform,
        "location": location,
        "cache_control": cache_control,
        "content_type": content_type,
        "origin": "https://pmxecute.stonex.com",
        "referer": "https://pmxecute.stonex.com/",
        "timeout": int(args_dict.get("timeout", "180") or "180"),
    }

    account_balances = _fetch_open_positions_account_balances(args_dict, request.headers)
    account_xau = _parse_loose_number(account_balances.get("xau")) if isinstance(account_balances, dict) else None
    account_usd = _parse_loose_number(account_balances.get("usd")) if isinstance(account_balances, dict) else None
    account_zar = _parse_loose_number(account_balances.get("zar")) if isinstance(account_balances, dict) else None
    baseline_date = str(args_dict.get("baseline_date", "") or "").strip() or "2026-03-01"

    def _r_fetch_baseline_balances_for_date() -> Dict[str, Optional[float]]:
        out: Dict[str, Optional[float]] = {"xau": None, "usd": None, "zar": None}
        try:
            fetch_variants = [
                {"col1": base_fetch_args.get("col1", "LC"), "col2": base_fetch_args.get("col2", "GLD")},
                {"col1": "USD", "col2": "None"},
            ]
            for variant in fetch_variants:
                if out["xau"] is not None and out["usd"] is not None and out["zar"] is not None:
                    break
                bl_args = dict(base_fetch_args)
                bl_args["start_date"] = baseline_date
                bl_args["end_date"] = baseline_date
                bl_args["col1"] = str(variant.get("col1", bl_args.get("col1", "LC")) or "LC")
                bl_args["col2"] = str(variant.get("col2", bl_args.get("col2", "GLD")) or "None")
                bl_args["unit_code1"] = ""
                bl_args["unit_code2"] = ""
                bl_result = fetch_pmx_account_statement_report(**bl_args)
                bl_payload = bl_result.get("json")
                if bl_payload is None:
                    bl_body = bl_result.get("body", "")
                    if isinstance(bl_body, str) and bl_body.strip().startswith(("{", "[")):
                        try:
                            bl_payload = json.loads(bl_body)
                        except Exception:
                            bl_payload = {}
                    else:
                        bl_payload = {}
                if bool(bl_result.get("ok")):
                    bl_rows = extract_pmx_statement_report_rows(bl_payload)
                    for bl_row in reversed(bl_rows):
                        if not isinstance(bl_row, dict):
                            continue
                        for bk, bv in bl_row.items():
                            bk_u = str(bk or "").strip().upper()
                            if "BAL" not in bk_u:
                                continue
                            num = _parse_loose_number(bv)
                            if num is None:
                                continue
                            if out["xau"] is None and any(tok in bk_u for tok in ("XAU", "GLD", "GOLD", "OZ", "COL2")):
                                out["xau"] = float(num)
                            if out["usd"] is None and "USD" in bk_u:
                                out["usd"] = float(num)
                            if out["zar"] is None and "ZAR" in bk_u:
                                out["zar"] = float(num)
                        if out["xau"] is not None and out["usd"] is not None and out["zar"] is not None:
                            break
        except Exception:
            pass
        return out

    baseline_balances = _r_fetch_baseline_balances_for_date()

    baseline_xau = _parse_loose_number(args_dict.get("baseline_xau"))
    if baseline_xau is None:
        # Dynamically fetch closing XAU balance from PMX statement for the baseline date
        baseline_xau = baseline_balances.get("xau")
    if baseline_xau is None:
        baseline_xau = -657.171

    baseline_usd_input = _parse_loose_number(args_dict.get("baseline_usd"))
    baseline_usd = baseline_usd_input
    if baseline_usd is None:
        baseline_usd = baseline_balances.get("usd")
    if baseline_usd is None:
        baseline_usd = -1214928.733
    baseline_zar = _parse_loose_number(args_dict.get("baseline_zar"))
    if baseline_zar is None:
        baseline_zar = 249105842.750

    result = fetch_pmx_account_statement_report(**base_fetch_args)
    payload = result.get("json")
    if payload is None:
        body_text = result.get("body", "")
        if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
            try:
                payload = json.loads(body_text)
            except Exception:
                payload = {}
        else:
            payload = {}

    result = _pmx_mark_failed_payload(result, payload)
    if not result.get("ok") and _pmx_result_is_auth_failure(result, payload):
        relogin = _pmx_login_session(args_dict)
        if relogin.get("ok"):
            base_fetch_args["x_auth"] = _pmx_non_empty(relogin.get("x_auth"), x_auth)
            base_fetch_args["sid"] = _pmx_non_empty(relogin.get("sid"), sid)
            base_fetch_args["username"] = _pmx_non_empty(relogin.get("username"), username)
            base_fetch_args["platform"] = _pmx_non_empty(relogin.get("platform"), platform)
            base_fetch_args["location"] = _pmx_non_empty(relogin.get("location"), location)
            base_fetch_args["cache_control"] = _pmx_non_empty(relogin.get("cache_control"), cache_control)
            base_fetch_args["content_type"] = _pmx_non_empty(relogin.get("content_type"), content_type)
            result = fetch_pmx_account_statement_report(**base_fetch_args)
            payload = result.get("json")
            if payload is None:
                body_text = result.get("body", "")
                if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
                    try:
                        payload = json.loads(body_text)
                    except Exception:
                        payload = {}
                else:
                    payload = {}
            result = _pmx_mark_failed_payload(result, payload)

    if not result.get("ok"):
        status = int(result.get("status") or 502)
        return jsonify({
            "ok": False,
            "error": _pmx_human_error(result, payload, "Failed to fetch PMX account statement report"),
            "status": status,
        }), status

    rows = extract_pmx_statement_report_rows(payload)

    def _r_first_non_empty(row: Dict[str, Any], keys: List[str]) -> str:
        for key in keys:
            if key not in row:
                continue
            value = row.get(key)
            if value is None:
                continue
            text = str(value).strip()
            if text:
                return text
        return ""

    def _r_find_symbol(text: str) -> str:
        match = re.search(r"\b([A-Z]{3}/[A-Z]{3})\b", str(text or "").upper())
        return match.group(1) if match else ""

    def _r_find_oz(text: str) -> Optional[float]:
        match = re.search(r"([+-]?\d[\d,]*\.?\d*)\s*OZ\b", str(text or "").upper())
        if not match:
            return None
        return _parse_loose_number(match.group(1))

    def _r_extract_xau_balance(row: Dict[str, Any]) -> Optional[float]:
        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if "BAL" not in key_u:
                continue
            if not any(token in key_u for token in ("XAU", "GLD", "GOLD", "OZ", "COL2")):
                continue
            num = _parse_loose_number(value)
            if num is not None:
                return float(num)
        return None

    def _r_extract_balance_for_currency(row: Dict[str, Any], currency: str, col_hint: str = "") -> Optional[float]:
        cur = str(currency or "").strip().upper()
        hint = str(col_hint or "").strip().upper()
        if cur not in {"USD", "ZAR"}:
            return None
        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if "BAL" not in key_u:
                continue
            if cur in key_u:
                num = _parse_loose_number(value)
                if num is not None:
                    return float(num)
        if hint:
            for key, value in row.items():
                key_u = str(key or "").strip().upper()
                if "BAL" not in key_u:
                    continue
                if hint in key_u:
                    num = _parse_loose_number(value)
                    if num is not None:
                        return float(num)
        # In col1=USD/ZAR statement views, PMX can expose a generic "Balance"
        # column without currency tokens in the header.
        if hint in {"USD", "ZAR", "COL1", "LC"}:
            for key, value in row.items():
                key_u = str(key or "").strip().upper()
                if "BAL" not in key_u:
                    continue
                num = _parse_loose_number(value)
                if num is not None:
                    return float(num)
        return None

    def _r_extract_debit_credit_for_currency(
        row: Dict[str, Any],
        currency: str,
        col_hint: str = "",
    ) -> Tuple[Optional[float], Optional[float]]:
        cur = str(currency or "").strip().upper()
        hint = str(col_hint or "").strip().upper()
        if cur not in {"USD", "ZAR"}:
            return None, None
        debit_val: Optional[float] = None
        credit_val: Optional[float] = None

        def _maybe_set(key_u: str, raw_value: Any) -> None:
            nonlocal debit_val, credit_val
            num = _parse_loose_number(raw_value)
            if num is None:
                return
            if "BAL" in key_u:
                return
            is_debit = ("DEBIT" in key_u) or bool(re.search(r"\bDR\b", key_u))
            is_credit = ("CREDIT" in key_u) or bool(re.search(r"\bCR\b", key_u))
            if not (is_debit or is_credit):
                return
            if is_debit:
                debit_val = float(num)
            if is_credit:
                credit_val = float(num)

        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if cur in key_u:
                _maybe_set(key_u, value)
        if (debit_val is None and credit_val is None) and hint:
            for key, value in row.items():
                key_u = str(key or "").strip().upper()
                if hint in key_u:
                    _maybe_set(key_u, value)
        # In col1=USD/ZAR statement views, PMX can expose generic "Debit"/"Credit"
        # headers with no currency token. Use them as a fallback.
        if debit_val is None and credit_val is None and hint in {"USD", "ZAR", "COL1", "LC"}:
            for key, value in row.items():
                key_u = str(key or "").strip().upper()
                if cur in key_u or hint in key_u:
                    continue
                _maybe_set(key_u, value)
        return debit_val, credit_val

    def _r_extract_oz(row: Dict[str, Any], side_hint: str, symbol_norm: str, narration: str) -> Optional[float]:
        side_up = str(side_hint or "").strip().upper()
        debit_xau = None
        credit_xau = None
        for key, value in row.items():
            key_u = str(key or "").strip().upper()
            if not any(token in key_u for token in ("XAU", "GLD", "GOLD", "OZ", "COL2")):
                continue
            num = _parse_loose_number(value)
            if num is None:
                continue
            if "DEBIT" in key_u or re.search(r"\bDR\b", key_u):
                debit_xau = float(num)
            elif "CREDIT" in key_u or re.search(r"\bCR\b", key_u):
                credit_xau = float(num)
        if credit_xau is not None or debit_xau is not None:
            return (credit_xau or 0.0) - (debit_xau or 0.0)
        qty_num = None
        for key in ("qty", "Quantity", "last_qty", "LastQty", "quantity", "Qty"):
            if key in row:
                parsed = _parse_loose_number(row.get(key))
                if parsed is not None:
                    qty_num = float(parsed)
                    break
        text_oz = _r_find_oz(narration)
        if symbol_norm.startswith(("XAU", "XAG", "XPT", "XPD")) and symbol_norm.endswith("USD"):
            base_oz = abs(qty_num) if qty_num is not None else (abs(float(text_oz)) if text_oz is not None else None)
            if base_oz is not None:
                if side_up == "BUY":
                    return base_oz
                if side_up == "SELL":
                    return -base_oz
                return base_oz
        if text_oz is not None:
            if side_up == "SELL":
                return -abs(float(text_oz))
            if side_up == "BUY":
                return abs(float(text_oz))
            return float(text_oz)
        return None

    def _r_find_doc_token(text: str) -> str:
        match = re.search(r"\b((?:FNC|JRV|JRC|MER|SWT)\/\d{4}\/\d+)\b", str(text or "").upper())
        return match.group(1) if match else ""

    def _r_extract_trade_from_narration(text: str) -> str:
        narr = str(text or "").strip()
        if not narr:
            return ""
        narr_u = narr.upper()
        m = re.search(r"\b([A-Z]{2,6}\s*-\s*\d{2,7})\b", narr_u)
        if m:
            return normalize_trade_number(m.group(1).replace(" ", ""))
        m = re.search(r"\bPROD\s*#?\s*(\d{3,8})\b", narr_u)
        if m:
            return normalize_trade_number(m.group(1))
        m = re.search(r"^\s*(\d{4,8})\b(?=.*\b(?:XAU|XAG|XPT|XPD|USD)/[A-Z]{3}\b)", narr_u)
        if m:
            return normalize_trade_number(m.group(1))
        return ""

    def _r_extract_date(row: Dict[str, Any], preferred_keys: List[str]) -> str:
        for key in preferred_keys:
            if key in row:
                parsed = _pmx_parse_date(row.get(key), default_value="")
                if parsed:
                    return parsed
        for key, value in row.items():
            key_text = str(key or "").strip().lower()
            if "date" not in key_text:
                continue
            parsed = _pmx_parse_date(value, default_value="")
            if parsed:
                return parsed
        for value in row.values():
            text = str(value or "").strip()
            if not text:
                continue
            for token in re.findall(r"\b\d{1,2}[-/][A-Za-z]{3}[-/]\d{4}\b|\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b|\b\d{1,2}[-/]\d{1,2}[-/]\d{4}\b", text):
                parsed = _pmx_parse_date(token, default_value="")
                if parsed:
                    return parsed
        return ""

    def _r_build_stmt_row_key(doc_number: str, trade_date: str, value_date: str, narration: str, occurrence: int) -> str:
        doc_key = str(doc_number or "").strip().upper()
        narr_key = re.sub(r"\s+", " ", str(narration or "").strip().upper())
        return f"{trade_date}|{value_date}|{doc_key}|{narr_key}|{int(occurrence)}"

    def _r_fetch_statement_rows_for_recon(col1: str, col2: str, unit1: str = "", unit2: str = "") -> List[Dict[str, Any]]:
        """Fetch PMX statement rows for a specific col1/col2 view; best-effort only."""
        try:
            fetch_args = dict(base_fetch_args)
            fetch_args["col1"] = str(col1 or "").strip() or fetch_args.get("col1", "LC")
            fetch_args["col2"] = str(col2 or "").strip()
            fetch_args["unit_code1"] = str(unit1 or "").strip()
            fetch_args["unit_code2"] = str(unit2 or "").strip()

            fetch_result = fetch_pmx_account_statement_report(**fetch_args)
            fetch_payload = fetch_result.get("json")
            if fetch_payload is None:
                body_text = fetch_result.get("body", "")
                if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
                    try:
                        fetch_payload = json.loads(body_text)
                    except Exception:
                        fetch_payload = {}
                else:
                    fetch_payload = {}

            fetch_result = _pmx_mark_failed_payload(fetch_result, fetch_payload)
            if not fetch_result.get("ok") and _pmx_result_is_auth_failure(fetch_result, fetch_payload):
                relogin = _pmx_login_session(args_dict)
                if relogin.get("ok"):
                    fetch_args["x_auth"] = _pmx_non_empty(relogin.get("x_auth"), fetch_args.get("x_auth", ""))
                    fetch_args["sid"] = _pmx_non_empty(relogin.get("sid"), fetch_args.get("sid", ""))
                    fetch_args["username"] = _pmx_non_empty(relogin.get("username"), fetch_args.get("username", ""))
                    fetch_args["platform"] = _pmx_non_empty(relogin.get("platform"), fetch_args.get("platform", ""))
                    fetch_args["location"] = _pmx_non_empty(relogin.get("location"), fetch_args.get("location", ""))
                    fetch_args["cache_control"] = _pmx_non_empty(relogin.get("cache_control"), fetch_args.get("cache_control", ""))
                    fetch_args["content_type"] = _pmx_non_empty(relogin.get("content_type"), fetch_args.get("content_type", ""))
                    fetch_result = fetch_pmx_account_statement_report(**fetch_args)
                    fetch_payload = fetch_result.get("json")
                    if fetch_payload is None:
                        body_text = fetch_result.get("body", "")
                        if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
                            try:
                                fetch_payload = json.loads(body_text)
                            except Exception:
                                fetch_payload = {}
                        else:
                            fetch_payload = {}
                    fetch_result = _pmx_mark_failed_payload(fetch_result, fetch_payload)

            if not fetch_result.get("ok"):
                return []
            out_rows = extract_pmx_statement_report_rows(fetch_payload)
            return [r for r in out_rows if isinstance(r, dict)]
        except Exception:
            return []

    trade_num_by_doc: Dict[str, str] = {}
    price_by_doc: Dict[str, Dict[str, Any]] = {}
    try:
        deal_result = fetch_pmx_alldeal_filter_report(
            start_date=start_date, end_date=end_date,
            cmdty="All", trd_opt="All",
            created_by=str(os.getenv("PMX_CREATED_BY", "2") or "2"),
            acc_opt_key=str(os.getenv("PMX_ACC_OPT_KEY", "MT0601") or "MT0601"),
            trade_type="TD", non_trd_cmdty="",
            host=str(os.getenv("PMX_API_HOST", "pmxapi.stonex.com") or "pmxapi.stonex.com"),
            path="/user/alldealFilter_report",
            authorization=str(args_dict.get("authorization", "") or ""),
            cookie=str(args_dict.get("cookie", "") or ""),
            x_auth=base_fetch_args.get("x_auth", ""),
            sid=base_fetch_args.get("sid", ""),
            username=base_fetch_args.get("username", ""),
            platform=base_fetch_args.get("platform", ""),
            location=base_fetch_args.get("location", ""),
            cache_control=base_fetch_args.get("cache_control", ""),
            content_type=base_fetch_args.get("content_type", ""),
            origin="https://pmxecute.stonex.com",
            referer="https://pmxecute.stonex.com/",
            timeout=int(base_fetch_args.get("timeout", 180) or 180),
        )
        deal_payload = deal_result.get("json")
        if deal_payload is None:
            body_text = deal_result.get("body", "")
            if isinstance(body_text, str) and body_text.strip().startswith(("{", "[")):
                try:
                    deal_payload = json.loads(body_text)
                except Exception:
                    deal_payload = {}
            else:
                deal_payload = {}
        if bool(deal_result.get("ok")):
            deal_rows = extract_pmx_report_rows(deal_payload)
            for idx, deal_row in enumerate(deal_rows, start=1):
                if not isinstance(deal_row, dict):
                    continue
                mapped = _pmx_map_row_to_trade(deal_row, idx)
                if not mapped:
                    continue
                doc_token = _pmx_extract_support_doc(mapped.get("doc_number", ""), mapped.get("narration", ""))
                tn = normalize_trade_number(mapped.get("order_id", ""))
                if doc_token and tn and doc_token not in trade_num_by_doc:
                    trade_num_by_doc[doc_token] = tn
                # Also store price/qty/side/symbol so XAUUSD statement rows can compute net_usd.
                if doc_token and doc_token not in price_by_doc:
                    m_px = _parse_loose_number(mapped.get("price"))
                    m_qty = _parse_loose_number(mapped.get("quantity"))
                    m_side = str(mapped.get("side", "") or "").upper()
                    m_sym = str(mapped.get("symbol", "") or "").upper().replace("/", "").replace("-", "").replace(" ", "")
                    if m_px is not None and m_qty is not None:
                        price_by_doc[doc_token] = {"price": float(m_px), "qty": float(m_qty), "side": m_side, "symbol": m_sym}
    except Exception:
        pass

    try:
        local_df = load_all_pmx_trades({})
        if isinstance(local_df, pd.DataFrame) and not local_df.empty:
            doc_col = "Doc #" if "Doc #" in local_df.columns else ("doc_number" if "doc_number" in local_df.columns else "")
            trade_col = "Trade #" if "Trade #" in local_df.columns else ("OrderID" if "OrderID" in local_df.columns else "")
            if doc_col and trade_col:
                docs = local_df[doc_col].fillna("").astype(str).str.strip()
                trades = local_df[trade_col].fillna("").astype(str).apply(normalize_trade_number)
                for doc_raw, tn_raw in zip(docs.tolist(), trades.tolist()):
                    doc_token = _pmx_extract_support_doc(doc_raw, "")
                    tn = normalize_trade_number(tn_raw)
                    if doc_token and tn and doc_token not in trade_num_by_doc:
                        trade_num_by_doc[doc_token] = tn
    except Exception:
        trade_num_by_doc = {}

    def _r_collect_currency_rows(raw_rows: List[Dict[str, Any]], currency: str, source_col1: str) -> List[Dict[str, Any]]:
        cur = str(currency or "").strip().upper()
        out: List[Dict[str, Any]] = []
        occurrence_by_base_key: Dict[Tuple[str, str, str, str], int] = {}
        for row_index, raw in enumerate(raw_rows, start=1):
            if not isinstance(raw, dict):
                continue

            doc_number = _r_first_non_empty(raw, ["docno", "DocNo", "doc_number", "DocNumber", "document_no", "Doc #"])
            if not doc_number:
                doc_number = _r_first_non_empty(raw, ["NeoId", "neo_id", "TagNumber", "tag_number"])
            if not doc_number:
                row_text = " ".join(str(v or "") for v in raw.values()).upper()
                doc_number = _r_find_doc_token(row_text)
            doc_number = doc_number.strip()

            narration = _r_first_non_empty(raw, ["remarks", "remarks1", "comment", "notes", "description", "ContractDescription", "Narration"])
            trade_date = _r_extract_date(raw, ["docdate", "TradeDate", "trade_date", "DocDate", "date", "Trade Date", "Doc Date"])
            value_date = _r_extract_date(raw, ["valdate", "ValueDate", "value_date", "settlement_date", "Value Date", "Settlement Date"])

            symbol_pair = _pmx_to_currency_pair(
                _r_first_non_empty(raw, ["CurrencyPair", "currency_pair", "cmdty", "stk_type_name", "inst_desc", "Symbol"])
            )
            if not symbol_pair:
                symbol_pair = _r_find_symbol(narration)
            symbol_pair = symbol_pair.upper()
            symbol_norm = symbol_pair.replace("/", "").replace("-", "").replace(" ", "")

            side = _r_first_non_empty(raw, ["side", "Side", "deal_type", "trd_opt"]).upper()
            if side not in {"BUY", "SELL"}:
                narr_upper = narration.upper()
                if " SELL " in f" {narr_upper} ":
                    side = "SELL"
                elif " BUY " in f" {narr_upper} ":
                    side = "BUY"
                else:
                    side = ""

            qty = None
            for key in ("qty", "Quantity", "last_qty", "LastQty", "quantity", "Qty"):
                if key in raw:
                    parsed = _parse_loose_number(raw.get(key))
                    if parsed is not None:
                        qty = float(parsed)
                        break
            price = None
            for key in ("px", "price", "Price", "last_px", "LastPx", "rate"):
                if key in raw:
                    parsed = _parse_loose_number(raw.get(key))
                    if parsed is not None:
                        price = float(parsed)
                        break

            debit, credit = _r_extract_debit_credit_for_currency(raw, cur)
            if cur == "USD" and debit is None and credit is None:
                for _hint in ("COL1", str(source_col1 or "").strip().upper(), "USD"):
                    if not _hint:
                        continue
                    debit, credit = _r_extract_debit_credit_for_currency(raw, cur, col_hint=_hint)
                    if debit is not None or credit is not None:
                        break
            if cur == "ZAR" and debit is None and credit is None:
                for _hint in ("COL1", "LC", str(source_col1 or "").strip().upper(), "ZAR"):
                    if not _hint:
                        continue
                    debit, credit = _r_extract_debit_credit_for_currency(raw, cur, col_hint=_hint)
                    if debit is not None or credit is not None:
                        break

            stmt_balance = _r_extract_balance_for_currency(raw, cur)
            if stmt_balance is None and cur == "USD":
                for _hint in ("COL1", str(source_col1 or "").strip().upper(), "USD"):
                    if not _hint:
                        continue
                    stmt_balance = _r_extract_balance_for_currency(raw, cur, col_hint=_hint)
                    if stmt_balance is not None:
                        break
            if stmt_balance is None and cur == "ZAR":
                for _hint in ("COL1", "LC", str(source_col1 or "").strip().upper(), "ZAR"):
                    if not _hint:
                        continue
                    stmt_balance = _r_extract_balance_for_currency(raw, cur, col_hint=_hint)
                    if stmt_balance is not None:
                        break

            fnc_token = _pmx_extract_support_doc(doc_number, narration)
            deal_info = price_by_doc.get(fnc_token or "") if fnc_token else None
            deal_side = str((deal_info or {}).get("side", "") or "").upper() if deal_info else ""
            deal_qty = _parse_loose_number((deal_info or {}).get("qty")) if deal_info else None
            deal_price = _parse_loose_number((deal_info or {}).get("price")) if deal_info else None
            eff_side = side if side in {"BUY", "SELL"} else (deal_side if deal_side in {"BUY", "SELL"} else "")

            net_value: Optional[float] = None
            source_col1_u = str(source_col1 or "").strip().upper()
            allow_usd_inference = not (cur == "USD" and source_col1_u == "USD")
            if credit is not None or debit is not None:
                net_value = (credit or 0.0) - (debit or 0.0)
            elif allow_usd_inference and cur == "USD" and symbol_norm in RECON_USD_METAL_SYMBOLS and deal_price is not None:
                oz_value = _r_extract_oz(raw, side, symbol_norm, narration)
                if oz_value is not None:
                    net_value = -(float(oz_value) * float(deal_price))
            elif allow_usd_inference and cur == "USD" and symbol_norm == "USDZAR":
                if deal_qty is not None and deal_side in {"BUY", "SELL"}:
                    abs_qty = abs(float(deal_qty))
                    net_value = abs_qty if deal_side == "BUY" else -abs_qty
                elif qty is not None and eff_side in {"BUY", "SELL"}:
                    abs_qty = abs(float(qty))
                    net_value = abs_qty if eff_side == "BUY" else -abs_qty
            elif cur == "ZAR" and symbol_norm == "USDZAR":
                ref_price = float(deal_price or price or 0.0)
                if deal_qty is not None and ref_price and deal_side in {"BUY", "SELL"}:
                    abs_qty = abs(float(deal_qty))
                    net_value = -(abs_qty * ref_price) if deal_side == "BUY" else (abs_qty * ref_price)
                elif qty is not None and ref_price and eff_side in {"BUY", "SELL"}:
                    abs_qty = abs(float(qty))
                    net_value = -(abs_qty * ref_price) if eff_side == "BUY" else (abs_qty * ref_price)

            has_meaningful_content = any(
                (isinstance(value, str) and bool(value.strip())) or (not isinstance(value, str) and value is not None)
                for value in (doc_number, narration, debit, credit, net_value, stmt_balance)
            )
            if not has_meaningful_content:
                continue

            base_key = (
                str(trade_date or "").strip(),
                str(value_date or "").strip(),
                str(doc_number or "").strip().upper(),
                re.sub(r"\s+", " ", str(narration or "").strip().upper()),
            )
            occurrence = int(occurrence_by_base_key.get(base_key, 0)) + 1
            occurrence_by_base_key[base_key] = occurrence

            out.append({
                "row_key": _r_build_stmt_row_key(doc_number, trade_date, value_date, narration, occurrence),
                "source_index": row_index,
                "doc_number": doc_number,
                "date": trade_date,
                "trade_date": trade_date,
                "value_date": value_date,
                "trade_number": normalize_trade_number(
                    _r_first_non_empty(raw, ["order_id", "OrderId", "trade_number", "trade_no", "ref_number", "OrderID"])
                ),
                "fnc_number": _pmx_extract_support_doc(doc_number, narration),
                "symbol": symbol_pair or "",
                "side": side or "",
                "narration": narration or "",
                "debit": float(debit) if debit is not None else None,
                "credit": float(credit) if credit is not None else None,
                "net": float(net_value) if net_value is not None else None,
                "stmt_balance": float(stmt_balance) if stmt_balance is not None else None,
            })
        return out

    def _r_extract_statement_closing_balance(raw_rows: List[Dict[str, Any]], currency: str, source_col1: str) -> Optional[float]:
        cur = str(currency or "").strip().upper()
        for raw in reversed(raw_rows):
            if not isinstance(raw, dict):
                continue
            stmt_balance = _r_extract_balance_for_currency(raw, cur)
            if stmt_balance is None and cur == "USD":
                for _hint in ("COL1", str(source_col1 or "").strip().upper(), "USD"):
                    if not _hint:
                        continue
                    stmt_balance = _r_extract_balance_for_currency(raw, cur, col_hint=_hint)
                    if stmt_balance is not None:
                        break
            if stmt_balance is None and cur == "ZAR":
                for _hint in ("COL1", "LC", str(source_col1 or "").strip().upper(), "ZAR"):
                    if not _hint:
                        continue
                    stmt_balance = _r_extract_balance_for_currency(raw, cur, col_hint=_hint)
                    if stmt_balance is not None:
                        break
            if stmt_balance is not None:
                return float(stmt_balance)
        return None

    # Build USD rows from the USD statement view.
    # Do not blend LC-derived USD rows here: LC fallbacks can inject synthetic
    # notional movements that distort the USD reconciliation.
    usd_statement_rows = _r_fetch_statement_rows_for_recon(col1="USD", col2="None")
    zar_statement_rows = _r_fetch_statement_rows_for_recon(col1="ZAR", col2="None")
    usd_statement_detail_rows = _r_collect_currency_rows(usd_statement_rows, "USD", "USD")
    zar_statement_detail_rows = _r_collect_currency_rows(zar_statement_rows, "ZAR", "ZAR")
    usd_rows_by_key = {
        str(row.get("row_key") or "").strip(): row
        for row in usd_statement_detail_rows
        if isinstance(row, dict) and str(row.get("row_key") or "").strip()
    }
    zar_rows_by_key = {
        str(row.get("row_key") or "").strip(): row
        for row in zar_statement_detail_rows
        if isinstance(row, dict) and str(row.get("row_key") or "").strip()
    }
    usd_statement_closing_balance = _r_extract_statement_closing_balance(usd_statement_rows, "USD", "USD")

    clean_rows: List[Dict[str, Any]] = []
    occurrence_by_base_key: Dict[Tuple[str, str, str, str], int] = {}
    for row_index, raw in enumerate(rows, start=1):
        if not isinstance(raw, dict):
            continue
        doc_number = _r_first_non_empty(raw, ["docno", "DocNo", "doc_number", "DocNumber", "document_no", "Doc #"])
        if not doc_number:
            doc_number = _r_first_non_empty(raw, ["NeoId", "neo_id", "TagNumber", "tag_number"])
        if not doc_number:
            row_text = " ".join(str(v or "") for v in raw.values()).upper()
            doc_number = _r_find_doc_token(row_text)
        doc_number = doc_number.strip()
        doc_upper = doc_number.upper()
        narration = _r_first_non_empty(raw, ["remarks", "remarks1", "comment", "notes", "description", "ContractDescription", "Narration"])
        if not doc_number and str(narration or "").strip().upper().startswith("BALANCE "):
            row_type = "BAL"
        elif doc_upper.startswith("FNC/"):
            row_type = "FNC"
        elif doc_upper.startswith("JRV/"):
            row_type = "JRV"
        elif doc_upper.startswith("MER/"):
            row_type = "MER"
        elif doc_upper.startswith("SWT/"):
            row_type = "SWT"
        elif "/" in doc_upper:
            row_type = doc_upper.split("/")[0]
        else:
            row_type = "OTHER"

        trade_date = _r_extract_date(raw, ["docdate", "TradeDate", "trade_date", "DocDate", "date", "Trade Date", "Doc Date"])
        value_date = _r_extract_date(raw, ["valdate", "ValueDate", "value_date", "settlement_date", "Value Date", "Settlement Date"])
        row_date = trade_date or value_date

        fnc_number = _pmx_extract_support_doc(doc_number, narration)
        trade_num = normalize_trade_number(
            _r_first_non_empty(raw, ["order_id", "OrderId", "trade_number", "trade_no", "ref_number", "OrderID"])
        )
        if not trade_num:
            trade_num = _r_extract_trade_from_narration(narration)
        if not trade_num:
            doc_lookup = _pmx_extract_support_doc(doc_number, narration)
            if doc_lookup:
                trade_num = normalize_trade_number(trade_num_by_doc.get(doc_lookup, ""))

        symbol_pair = _pmx_to_currency_pair(
            _r_first_non_empty(raw, ["CurrencyPair", "currency_pair", "cmdty", "stk_type_name", "inst_desc", "Symbol"])
        )
        if not symbol_pair:
            symbol_pair = _r_find_symbol(narration)
        symbol_pair = symbol_pair.upper()
        symbol_norm = symbol_pair.replace("/", "").replace("-", "").replace(" ", "")

        side = _r_first_non_empty(raw, ["side", "Side", "deal_type", "trd_opt"]).upper()
        if side not in {"BUY", "SELL"}:
            narr_upper = narration.upper()
            if " SELL " in f" {narr_upper} ":
                side = "SELL"
            elif " BUY " in f" {narr_upper} ":
                side = "BUY"
            else:
                side = ""

        qty = None
        for key in ("qty", "Quantity", "last_qty", "LastQty", "quantity", "Qty"):
            if key in raw:
                parsed = _parse_loose_number(raw.get(key))
                if parsed is not None:
                    qty = float(parsed)
                    break
        price = None
        for key in ("px", "price", "Price", "last_px", "LastPx", "rate"):
            if key in raw:
                parsed = _parse_loose_number(raw.get(key))
                if parsed is not None:
                    price = float(parsed)
                    break

        oz_value = _r_extract_oz(raw, side, symbol_norm, narration)
        stmt_balance_xau = _r_extract_xau_balance(raw)
        stmt_balance_usd = _r_extract_balance_for_currency(raw, "USD")
        stmt_balance_zar = _r_extract_balance_for_currency(raw, "ZAR")

        # USD: check explicit USD-labelled columns only.
        debit_usd, credit_usd = _r_extract_debit_credit_for_currency(raw, "USD")

        # ZAR: check explicit ZAR columns, then try multiple local-currency column name
        # patterns (PMX may label the COL1 column as "COL1", "LC", or the col1 param value).
        debit_zar, credit_zar = _r_extract_debit_credit_for_currency(raw, "ZAR")
        col1_eligible = symbol_norm == "USDZAR" or not symbol_norm or row_type in RECON_JOURNAL_LIKE_ROW_TYPES
        if debit_zar is None and credit_zar is None and col1_eligible:
            for _zh in ("COL1", "LC", statement_col1.upper(), "ZAR"):
                if _zh:
                    debit_zar, credit_zar = _r_extract_debit_credit_for_currency(raw, "ZAR", col_hint=_zh)
                if debit_zar is not None or credit_zar is not None:
                    break
        if stmt_balance_zar is None and col1_eligible:
            for _zh in ("COL1", "LC", statement_col1.upper(), "ZAR"):
                if _zh:
                    stmt_balance_zar = _r_extract_balance_for_currency(raw, "ZAR", col_hint=_zh)
                if stmt_balance_zar is not None:
                    break

        # Resolve the deal-report entry for this statement row (used for both metals and FX).
        _fnc_token = _pmx_extract_support_doc(doc_number, narration)
        _deal_info = price_by_doc.get(_fnc_token or "") if _fnc_token else None
        _deal_side = str((_deal_info or {}).get("side", "") or "").upper() if _deal_info else ""
        _deal_qty = _parse_loose_number((_deal_info or {}).get("qty")) if _deal_info else None
        _deal_price = _parse_loose_number((_deal_info or {}).get("price")) if _deal_info else None
        # Prefer the statement's own side; fall back to the deal-report side.
        eff_side = side if side in {"BUY", "SELL"} else (_deal_side if _deal_side in {"BUY", "SELL"} else "")

        net_usd: Optional[float] = None

        # 1. Explicit USD debit/credit from statement columns (highest priority).
        if credit_usd is not None or debit_usd is not None:
            net_usd = (credit_usd or 0.0) - (debit_usd or 0.0)

        if net_usd is None and symbol_norm in RECON_USD_METAL_SYMBOLS and oz_value is not None and _deal_price is not None:
            net_usd = -(float(oz_value) * float(_deal_price))

        if net_usd is None and symbol_norm in RECON_USD_METAL_SYMBOLS and qty is not None and price is not None and eff_side in {"BUY", "SELL"}:
            abs_qty = abs(float(qty))
            net_usd = -(abs_qty * float(price)) if eff_side == "BUY" else (abs_qty * float(price))

        # 4. USDZAR: use deal-report qty+side when the statement row has no explicit side.
        if net_usd is None and symbol_norm == "USDZAR" and _deal_qty is not None and _deal_side in {"BUY", "SELL"}:
            abs_qty = abs(float(_deal_qty))
            net_usd = abs_qty if _deal_side == "BUY" else -abs_qty

        # 5. USDZAR: statement qty + effective side fallback.
        if net_usd is None and symbol_norm == "USDZAR" and qty is not None and eff_side in {"BUY", "SELL"}:
            abs_qty = abs(float(qty))
            net_usd = abs_qty if eff_side == "BUY" else -abs_qty

        net_zar: Optional[float] = None

        # 1. Explicit ZAR debit/credit from statement columns (incl. COL1/LC fallback).
        if credit_zar is not None or debit_zar is not None:
            net_zar = (credit_zar or 0.0) - (debit_zar or 0.0)

        # 2. USDZAR: use deal-report qty+price+side when statement columns are absent.
        if net_zar is None and symbol_norm == "USDZAR" and _deal_qty is not None and _deal_price is not None and _deal_side in {"BUY", "SELL"}:
            abs_qty = abs(float(_deal_qty))
            net_zar = -(abs_qty * float(_deal_price)) if _deal_side == "BUY" else (abs_qty * float(_deal_price))

        # 3. USDZAR: statement qty + effective side + price fallback.
        if net_zar is None and symbol_norm == "USDZAR" and qty is not None and price is not None and eff_side in {"BUY", "SELL"}:
            abs_qty = abs(float(qty))
            net_zar = -(abs_qty * float(price)) if eff_side == "BUY" else (abs_qty * float(price))

        _fx_rate = float(_deal_price or price or 0.0)
        if net_usd is None and net_zar is not None and symbol_norm == "USDZAR" and abs(_fx_rate) > 1e-12:
            net_usd = -float(net_zar) / _fx_rate
        if net_zar is None and net_usd is not None and symbol_norm == "USDZAR" and abs(_fx_rate) > 1e-12:
            net_zar = -float(net_usd) * _fx_rate

        buy_oz = abs(float(oz_value)) if oz_value is not None and oz_value > 0 else 0.0
        sell_oz = abs(float(oz_value)) if oz_value is not None and oz_value < 0 else 0.0

        has_meaningful_content = any(
            (isinstance(value, str) and bool(value.strip())) or (not isinstance(value, str) and value is not None)
            for value in (
                doc_number,
                narration,
                oz_value,
                stmt_balance_xau,
                debit_usd,
                credit_usd,
                net_usd,
                stmt_balance_usd,
                debit_zar,
                credit_zar,
                net_zar,
                stmt_balance_zar,
            )
        )
        if not has_meaningful_content:
            continue

        base_key = (
            str(trade_date or "").strip(),
            str(value_date or "").strip(),
            str(doc_number or "").strip().upper(),
            re.sub(r"\s+", " ", str(narration or "").strip().upper()),
        )
        occurrence = int(occurrence_by_base_key.get(base_key, 0)) + 1
        occurrence_by_base_key[base_key] = occurrence
        row_key = _r_build_stmt_row_key(doc_number, trade_date, value_date, narration, occurrence)

        clean = {
            "row_key": row_key,
            "source_index": row_index,
            "doc_number": doc_number,
            "row_type": row_type,
            "trade_number": trade_num or "",
            "fnc_number": fnc_number or "",
            "date": row_date,
            "trade_date": trade_date,
            "value_date": value_date,
            "symbol": symbol_pair or "",
            "side": side or "",
            "oz": oz_value,
            "buy_oz": buy_oz,
            "sell_oz": sell_oz,
            "net_oz": oz_value if oz_value is not None else None,
            "stmt_balance_xau": stmt_balance_xau,
            "debit_usd": debit_usd,
            "credit_usd": credit_usd,
            "net_usd": net_usd,
            "stmt_balance_usd": stmt_balance_usd,
            "debit_zar": debit_zar,
            "credit_zar": credit_zar,
            "net_zar": net_zar,
            "stmt_balance_zar": stmt_balance_zar,
            "quantity": qty,
            "price": price,
            "narration": narration or "",
            "running_net_oz": None,
            "expected_xau": None,
            "delta_vs_stmt": None,
            "account_xau": None,
            "delta_to_account": None,
            "running_net_usd": None,
            "expected_usd": None,
            "running_net_zar": None,
            "expected_zar": None,
        }
        clean_rows.append(clean)

    clean_row_keys = {str(row.get("row_key") or "").strip() for row in clean_rows if str(row.get("row_key") or "").strip()}
    for row in clean_rows:
        row_key = str(row.get("row_key") or "").strip()
        usd_info = usd_rows_by_key.get(row_key) or {}
        if usd_info:
            if row.get("debit_usd") is None and usd_info.get("debit") is not None:
                row["debit_usd"] = usd_info.get("debit")
            if row.get("credit_usd") is None and usd_info.get("credit") is not None:
                row["credit_usd"] = usd_info.get("credit")
            # USD statement view is the source of truth for cash-leg net USD.
            if usd_info.get("net") is not None:
                row["net_usd"] = usd_info.get("net")
            if row.get("stmt_balance_usd") is None and usd_info.get("stmt_balance") is not None:
                row["stmt_balance_usd"] = usd_info.get("stmt_balance")

        zar_info = zar_rows_by_key.get(row_key) or {}
        if zar_info:
            if zar_info.get("debit") is not None:
                row["debit_zar"] = zar_info.get("debit")
            if zar_info.get("credit") is not None:
                row["credit_zar"] = zar_info.get("credit")
            if zar_info.get("net") is not None:
                row["net_zar"] = zar_info.get("net")
            if zar_info.get("stmt_balance") is not None:
                row["stmt_balance_zar"] = zar_info.get("stmt_balance")

    for info in usd_statement_detail_rows:
        if not isinstance(info, dict):
            continue
        row_key = str(info.get("row_key") or "").strip()
        if not row_key or row_key in clean_row_keys:
            continue
        doc_number = str(info.get("doc_number") or "").strip()
        doc_upper = str(doc_number).upper()
        narration = str(info.get("narration") or "").strip()
        if not doc_number and narration.upper().startswith("BALANCE "):
            extra_row_type = "BAL"
        elif doc_upper.startswith("JRV/"):
            extra_row_type = "JRV"
        elif doc_upper.startswith("MER/"):
            extra_row_type = "MER"
        elif doc_upper.startswith("FNC/"):
            extra_row_type = "FNC"
        elif "/" in doc_upper:
            extra_row_type = doc_upper.split("/")[0]
        else:
            extra_row_type = "OTHER"
        clean_rows.append(
            {
                "row_key": row_key,
                "source_index": int(info.get("source_index") or 0),
                "doc_number": doc_number,
                "row_type": extra_row_type,
                "trade_number": info.get("trade_number") or "",
                "fnc_number": info.get("fnc_number") or "",
                "date": info.get("date") or "",
                "trade_date": info.get("trade_date") or "",
                "value_date": info.get("value_date") or "",
                "symbol": info.get("symbol") or "",
                "side": info.get("side") or "",
                "oz": None,
                "buy_oz": 0.0,
                "sell_oz": 0.0,
                "net_oz": None,
                "stmt_balance_xau": None,
                "debit_usd": info.get("debit"),
                "credit_usd": info.get("credit"),
                "net_usd": info.get("net"),
                "stmt_balance_usd": info.get("stmt_balance"),
                "debit_zar": None,
                "credit_zar": None,
                "net_zar": None,
                "stmt_balance_zar": None,
                "quantity": None,
                "price": None,
                "narration": info.get("narration") or "",
                "running_net_oz": None,
                "expected_xau": None,
                "delta_vs_stmt": None,
                "account_xau": None,
                "delta_to_account": None,
                "running_net_usd": None,
                "expected_usd": None,
                "running_net_zar": None,
                "expected_zar": None,
            }
        )
        clean_row_keys.add(row_key)

    for info in zar_statement_detail_rows:
        if not isinstance(info, dict):
            continue
        row_key = str(info.get("row_key") or "").strip()
        if not row_key or row_key in clean_row_keys:
            continue
        doc_number = str(info.get("doc_number") or "").strip()
        doc_upper = str(doc_number).upper()
        narration = str(info.get("narration") or "").strip()
        if not doc_number and narration.upper().startswith("BALANCE "):
            extra_row_type = "BAL"
        elif doc_upper.startswith("JRV/"):
            extra_row_type = "JRV"
        elif doc_upper.startswith("MER/"):
            extra_row_type = "MER"
        elif doc_upper.startswith("FNC/"):
            extra_row_type = "FNC"
        elif "/" in doc_upper:
            extra_row_type = doc_upper.split("/")[0]
        else:
            extra_row_type = "OTHER"
        clean_rows.append(
            {
                "row_key": row_key,
                "source_index": int(info.get("source_index") or 0),
                "doc_number": doc_number,
                "row_type": extra_row_type,
                "trade_number": info.get("trade_number") or "",
                "fnc_number": info.get("fnc_number") or "",
                "date": info.get("date") or "",
                "trade_date": info.get("trade_date") or "",
                "value_date": info.get("value_date") or "",
                "symbol": info.get("symbol") or "",
                "side": info.get("side") or "",
                "oz": None,
                "buy_oz": 0.0,
                "sell_oz": 0.0,
                "net_oz": None,
                "stmt_balance_xau": None,
                "debit_usd": None,
                "credit_usd": None,
                "net_usd": None,
                "stmt_balance_usd": None,
                "debit_zar": info.get("debit"),
                "credit_zar": info.get("credit"),
                "net_zar": info.get("net"),
                "stmt_balance_zar": info.get("stmt_balance"),
                "quantity": None,
                "price": None,
                "narration": info.get("narration") or "",
                "running_net_oz": None,
                "expected_xau": None,
                "delta_vs_stmt": None,
                "account_xau": None,
                "delta_to_account": None,
                "running_net_usd": None,
                "expected_usd": None,
                "running_net_zar": None,
                "expected_zar": None,
            }
        )
        clean_row_keys.add(row_key)

    # Sort by statement sequence so duplicate docs and swap legs stay aligned.
    def _r_sort_key(row: Dict[str, Any]) -> Tuple[str, str, int, str]:
        return (
            str(row.get("trade_date") or ""),
            str(row.get("value_date") or ""),
            int(row.get("source_index") or 0),
            str(row.get("row_key") or ""),
        )

    clean_rows = sorted(clean_rows, key=_r_sort_key)
    running_oz = 0.0
    running_has_value = False
    for row in clean_rows:
        row_date = str(row.get("date") or row.get("trade_date") or "").strip()
        if row_date and row_date < baseline_date:
            row["running_net_oz"] = None
            row["expected_xau"] = None
            row["delta_vs_stmt"] = None
            continue
        oz_num = _parse_loose_number(row.get("net_oz"))
        if oz_num is None:
            row["running_net_oz"] = None
            row["expected_xau"] = None
            row["delta_vs_stmt"] = None
            continue
        running_oz += float(oz_num)
        running_has_value = True
        expected_xau = float(baseline_xau) + running_oz
        row["running_net_oz"] = round(running_oz, 4)
        row["expected_xau"] = round(expected_xau, 4)
        stmt_bal = _parse_loose_number(row.get("stmt_balance_xau"))
        if stmt_bal is not None:
            row["delta_vs_stmt"] = round(float(stmt_bal) - expected_xau, 4)

    if clean_rows and account_xau is not None:
        clean_rows[-1]["account_xau"] = account_xau
        if running_has_value:
            clean_rows[-1]["delta_to_account"] = round((float(baseline_xau) + running_oz) - float(account_xau), 4)

    # USD reconciliation is statement-first.
    # Start from the baseline USD statement balance and derive movement from each
    # subsequent StoneX statement balance where available. Only fall back to row
    # `net_usd` when a statement balance is absent on that row.
    running_usd = 0.0
    running_usd_has_value = False
    usd_doc_contrib: Dict[str, float] = {}
    usd_row_trace: List[Dict[str, Any]] = []
    prev_stmt_balance_usd: float = float(baseline_usd)
    stmt_closing_usd: Optional[float] = None

    for idx, row in enumerate(clean_rows, start=1):
        row_date = str(row.get("date") or row.get("trade_date") or "").strip()
        doc_number = str(row.get("doc_number") or "").strip().upper()
        symbol = str(row.get("symbol") or "").strip().upper()
        row_type = str(row.get("row_type") or "").strip().upper()
        stmt_usd_row = _parse_loose_number(row.get("stmt_balance_usd"))
        raw_net_usd = _parse_loose_number(row.get("net_usd"))
        effective_net_usd: Optional[float] = None
        trace_reason = ""

        if row_date and row_date < baseline_date:
            row["running_net_usd"] = None
            row["expected_usd"] = None
            row["usd_recon_included"] = False
            row["usd_recon_reason"] = "before_baseline_date"
            row["usd_recon_effective_net"] = None
            row["usd_recon_running_after"] = None
            usd_row_trace.append({
                "row_index": idx,
                "doc_number": doc_number,
                "row_type": row_type,
                "symbol": symbol,
                "net_usd_raw": row.get("net_usd"),
                "effective_net_usd": None,
                "included": False,
                "reason": "before_baseline_date",
                "running_usd_after": None,
            })
            continue

        if stmt_usd_row is not None:
            effective_net_usd = float(stmt_usd_row) - float(prev_stmt_balance_usd)
            prev_stmt_balance_usd = float(stmt_usd_row)
            stmt_closing_usd = float(stmt_usd_row)
            row["net_usd"] = effective_net_usd
            trace_reason = "derived_from_stmt_balance"
        elif raw_net_usd is not None:
            effective_net_usd = float(raw_net_usd)
            trace_reason = "fallback_net_usd"

        if effective_net_usd is None:
            row["running_net_usd"] = None
            row["expected_usd"] = None
            row["usd_recon_included"] = False
            row["usd_recon_reason"] = "missing_usd_movement"
            row["usd_recon_effective_net"] = None
            row["usd_recon_running_after"] = None
            usd_row_trace.append({
                "row_index": idx,
                "doc_number": doc_number,
                "row_type": row_type,
                "symbol": symbol,
                "net_usd_raw": row.get("net_usd"),
                "effective_net_usd": None,
                "included": False,
                "reason": "missing_usd_movement",
                "running_usd_after": None,
            })
            continue

        running_usd += float(effective_net_usd)
        running_usd_has_value = True
        row["running_net_usd"] = round(running_usd, 4)
        row["expected_usd"] = round(float(baseline_usd) + running_usd, 4)
        row["usd_recon_included"] = True
        row["usd_recon_reason"] = trace_reason
        row["usd_recon_effective_net"] = round(float(effective_net_usd), 4)
        row["usd_recon_running_after"] = row["running_net_usd"]
        if doc_number:
            usd_doc_contrib[doc_number] = float(usd_doc_contrib.get(doc_number, 0.0)) + float(effective_net_usd)
        usd_row_trace.append({
            "row_index": idx,
            "doc_number": doc_number,
            "row_type": row_type,
            "symbol": symbol,
            "net_usd_raw": row.get("net_usd"),
            "effective_net_usd": round(float(effective_net_usd), 4),
            "included": True,
            "reason": trace_reason,
            "running_usd_after": row["running_net_usd"],
        })

    running_zar = 0.0
    running_zar_has_value = False
    for row in clean_rows:
        row_date = str(row.get("date") or row.get("trade_date") or "").strip()
        if row_date and row_date < baseline_date:
            row["running_net_zar"] = None
            row["expected_zar"] = None
            continue
        zar_num = _parse_loose_number(row.get("net_zar"))
        if zar_num is None:
            row["running_net_zar"] = None
            row["expected_zar"] = None
            continue
        running_zar += float(zar_num)
        running_zar_has_value = True
        row["running_net_zar"] = round(running_zar, 4)
        row["expected_zar"] = round(float(baseline_zar) + running_zar, 4)

    expected_closing = round(float(baseline_xau) + running_oz, 4) if running_has_value else None
    delta_to_account = None
    if expected_closing is not None and account_xau is not None:
        delta_to_account = round(expected_closing - float(account_xau), 4)

    if usd_statement_closing_balance is not None:
        stmt_closing_usd = float(usd_statement_closing_balance)
    expected_closing_usd = round(float(baseline_usd) + running_usd, 4)
    if stmt_closing_usd is not None:
        running_usd = float(stmt_closing_usd) - float(baseline_usd)
        running_usd_has_value = True
        expected_closing_usd = round(float(stmt_closing_usd), 4)
    effective_account_usd = stmt_closing_usd if stmt_closing_usd is not None else account_usd
    if effective_account_usd is not None:
        running_usd = float(effective_account_usd) - float(baseline_usd)
        running_usd_has_value = True
        expected_closing_usd = round(float(effective_account_usd), 4)
    delta_usd = None
    if effective_account_usd is not None:
        delta_usd = round(expected_closing_usd - float(effective_account_usd), 4)
    delta_usd_vs_stmt = None
    if stmt_closing_usd is not None:
        delta_usd_vs_stmt = round(expected_closing_usd - float(stmt_closing_usd), 4)
    account_vs_stmt_usd = None
    if account_usd is not None and stmt_closing_usd is not None:
        account_vs_stmt_usd = round(float(account_usd) - float(stmt_closing_usd), 4)

    expected_closing_zar = round(float(baseline_zar) + running_zar, 4) if running_zar_has_value else None
    delta_zar = None
    if expected_closing_zar is not None and account_zar is not None:
        delta_zar = round(expected_closing_zar - float(account_zar), 4)

    summary = {
        "baseline_date": baseline_date,
        "baseline_xau": float(baseline_xau),
        "total_net_oz": round(running_oz, 4) if running_has_value else 0.0,
        "expected_xau": expected_closing,
        "account_xau": round(float(account_xau), 4) if account_xau is not None else None,
        "delta_to_account": delta_to_account,
        "baseline_usd": float(baseline_usd),
        "total_net_usd": round(running_usd, 4) if running_usd_has_value else 0.0,
        "expected_usd": expected_closing_usd,
        "account_usd": round(float(effective_account_usd), 4) if effective_account_usd is not None else None,
        "account_usd_reported": round(float(account_usd), 4) if account_usd is not None else None,
        "account_usd_stmt_closing": round(float(stmt_closing_usd), 4) if stmt_closing_usd is not None else None,
        "account_usd_source": "statement" if stmt_closing_usd is not None else ("account_balance" if account_usd is not None else None),
        "delta_usd": delta_usd,
        "delta_usd_vs_stmt": delta_usd_vs_stmt,
        "account_vs_stmt_usd": account_vs_stmt_usd,
        "baseline_zar": float(baseline_zar),
        "total_net_zar": round(running_zar, 4) if running_zar_has_value else 0.0,
        "expected_zar": expected_closing_zar,
        "account_zar": round(float(account_zar), 4) if account_zar is not None else None,
        "delta_zar": delta_zar,
        "row_count": len(clean_rows),
        "start_date": start_date,
        "end_date": end_date,
    }
    usd_doc_breakdown = sorted(
        [
            {
                "doc_number": doc,
                "net_usd_counted": round(float(net), 4),
            }
            for doc, net in usd_doc_contrib.items()
        ],
        key=lambda rec: abs(float(rec.get("net_usd_counted") or 0.0)),
        reverse=True,
    )

    # Round numeric fields for readability
    for row in clean_rows:
        for fld in ("oz", "buy_oz", "sell_oz", "net_oz", "quantity", "debit_usd", "credit_usd", "net_usd", "debit_zar", "credit_zar", "net_zar"):
            val = row.get(fld)
            if val is not None and isinstance(val, (int, float)):
                row[fld] = round(float(val), 4)
        for fld in ("price",):
            val = row.get(fld)
            if val is not None and isinstance(val, (int, float)):
                row[fld] = round(float(val), 5)
        for fld in ("stmt_balance_xau", "stmt_balance_usd", "stmt_balance_zar"):
            val = row.get(fld)
            if val is not None and isinstance(val, (int, float)):
                row[fld] = round(float(val), 4)

    return jsonify(_json_safe({
        "ok": True,
        "rows": clean_rows,
        "summary": summary,
        "usd_recon_debug": {
            "baseline_usd": float(baseline_usd),
            "account_usd": round(float(account_usd), 4) if account_usd is not None else None,
            "counted_doc_count": len(usd_doc_contrib),
            "counted_row_count": sum(1 for trace in usd_row_trace if bool(trace.get("included"))),
            "excluded_row_count": sum(1 for trace in usd_row_trace if not bool(trace.get("included"))),
            "doc_breakdown": usd_doc_breakdown,
            "row_trace": usd_row_trace,
        },
        "account_balances": account_balances if isinstance(account_balances, dict) else {},
    }))


@app.route("/api/pmx/fnc-pdf", methods=["GET", "POST"])
def get_pmx_fnc_pdf():
    data = request.args.to_dict() if request.method == "GET" else (request.json or {})
    cell = str(data.get("cell") or data.get("fnc") or data.get("fnc_number") or "").strip()
    if not cell:
        return jsonify({"ok": False, "error": "Missing 'cell' (e.g. FNC/2026/048744)"}), 400

    doc_type = str(
        data.get("DocType")
        or data.get("doc_type")
        or (cell.split("/", 1)[0] if "/" in cell else "FNC")
        or "FNC"
    ).upper()
    result = _build_pmx_fnc_pdf_result(cell=cell, doc_type=doc_type, data=data, req_headers=request.headers)
    if not result.get("ok"):
        return jsonify(_json_safe(result)), 400

    pdf_bytes = result.get("bytes", b"")
    if not isinstance(pdf_bytes, (bytes, bytearray)):
        pdf_bytes = bytes(str(pdf_bytes), "utf-8")
    pdf_bytes = bytes(pdf_bytes)
    safe_name = str(result.get("filename") or ("Fixing_Invoice_" + str(cell).replace("/", "_").replace("\\", "_") + ".pdf"))
    content_disposition = f'attachment; filename="{safe_name}"'

    resp = Response(pdf_bytes, mimetype=result.get("content_type") or "application/pdf")
    resp.headers["Content-Disposition"] = content_disposition
    resp.headers["Content-Length"] = str(len(pdf_bytes))
    return resp


@app.route("/api/trades/open-positions")
def get_open_positions():
    df = load_all_trades()
    records, summary = build_open_positions(df)
    return jsonify({"positions": records, "summary": summary})


@app.route("/api/pmx/open-positions-reval")
def get_pmx_open_positions_reval():
    args_dict = request.args.to_dict()
    cache_key = _build_cache_key("pmx_open_positions_reval", args_dict)
    result = _get_cached_heavy_result(
        cache_key,
        lambda: build_open_positions_reval(args_dict, request.headers),
    )
    return jsonify(_json_safe(result))


@app.route("/api/pmx/open-positions-reval/pdf")
def get_pmx_open_positions_reval_pdf():
    args_dict = request.args.to_dict()
    try:
        pdf_bytes = build_open_positions_reval_pdf(args_dict, request.headers)
    except Exception as exc:
        return jsonify({"error": f"PDF generation failed: {exc}"}), 500
    if not pdf_bytes:
        return jsonify({"error": "PDF export requires the `reportlab` or `fpdf2` package."}), 400

    report_date = datetime.now().strftime("%Y-%m-%d")
    resp = Response(pdf_bytes, mimetype="application/pdf")
    resp.headers["Content-Disposition"] = f"attachment; filename=open_positions_reval_report_{report_date}.pdf"
    resp.headers["Content-Length"] = str(len(pdf_bytes))
    return resp


@app.route("/api/pmx/account-balances")
def get_pmx_account_balances():
    result = _fetch_open_positions_account_balances(request.args.to_dict(), request.headers)
    return jsonify(_json_safe(result))


@app.route("/api/pmx/account-recon")
def get_pmx_account_recon():
    args_dict = request.args.to_dict()
    cache_key = _build_cache_key("account_recon", args_dict)
    result = _get_cached_heavy_result(
        cache_key,
        lambda: build_account_recon(args_dict, request.headers),
    )
    return jsonify(_json_safe(result))


@app.route("/api/pmx/account-recon/opening-balance", methods=["POST"])
def set_pmx_account_recon_opening_balance():
    body = request.get_json(silent=True) or {}
    month = str(body.get("month", "") or "").strip()
    currency = str(body.get("currency", "") or "").strip().upper()
    if not month or not currency or currency not in ("XAU", "USD", "ZAR"):
        return jsonify({"ok": False, "error": "month and currency (XAU/USD/ZAR) are required"}), 400
    try:
        opening_balance = float(body.get("opening_balance", 0) or 0)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "opening_balance must be a number"}), 400
    try:
        conn = get_pmx_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO account_opening_balances (month, currency, opening_balance, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(month, currency) DO UPDATE SET opening_balance=excluded.opening_balance, updated_at=excluded.updated_at
            """,
            (month, currency, opening_balance, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        )
        conn.commit()
        conn.close()
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    # Invalidate cached recon results
    _clear_heavy_route_cache(["account_recon"])
    return jsonify({"ok": True})


@app.route("/api/pmx/account-recon/opening-balances")
def get_pmx_account_recon_opening_balances():
    try:
        conn = get_pmx_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, month, currency, opening_balance, updated_at FROM account_opening_balances ORDER BY month DESC")
        rows = [
            {"id": r[0], "month": r[1], "currency": r[2], "opening_balance": r[3], "updated_at": r[4]}
            for r in cur.fetchall()
        ]
        conn.close()
        return jsonify({"ok": True, "rows": rows})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc), "rows": []}), 500


@app.route("/api/pmx/forward-exposure")
def get_pmx_forward_exposure():
    args_dict = request.args.to_dict()
    cache_key = _build_cache_key("pmx_forward_exposure", args_dict)
    result = _get_cached_heavy_result(
        cache_key,
        lambda: build_forward_exposure(args_dict),
    )
    return jsonify(_json_safe(result))


@app.route("/api/trademc/trades")
def get_trademc_trades():
    kwargs = {}
    for key in ["status", "ref_filter", "company_id", "start_date", "end_date"]:
        val = request.args.get(key)
        if val:
            if key == "company_id":
                kwargs[key] = int(val)
            else:
                kwargs[key] = val

    df = load_trademc_trades_with_companies(**kwargs)
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
    return jsonify(df.fillna("").to_dict(orient="records"))


_trademc_sync_lock = threading.Lock()
_trademc_sync_status: Dict[str, Any] = {"running": False, "result": None}


def _run_trademc_sync_bg(include_weight: bool = False, incremental: bool = True, replace: bool = False):
    """Background worker: sync TradeMC trades using incremental/full mode."""
    global _trademc_sync_status
    try:
        company_result = sync_trademc_companies()
        trade_result = sync_trademc_trades(incremental=incremental, prune_missing=(replace and not incremental))
        if not bool(trade_result.get("success")):
            raise RuntimeError(str(trade_result.get("error", "TradeMC trade sync failed")))
        mode = (
            "incremental"
            if incremental
            else ("full_replace" if replace else "full")
        )
        result: Dict[str, Any] = {
            "companies": company_result,
            "trades": trade_result,
            "mode": mode,
        }
        if include_weight:
            result["weight_transactions"] = sync_trademc_weight_transactions(incremental=incremental)
        fiscal_cutoff = _pmx_filter_date(FISCAL_TRADES_START_DATE) if ENABLE_FISCAL_PURGE else ""
        result["fiscal_cutoff"] = fiscal_cutoff
        result["removed_pre_fiscal"] = {
            "trademc_trades": 0,
            "trademc_weight_transactions": 0,
            "fx_trades": 0,
        }
        if fiscal_cutoff:
            fiscal_purge = _purge_pre_fiscal_rows(cutoff_iso=fiscal_cutoff, purge_fx=True, purge_pmx=False)
            result["removed_pre_fiscal"] = {
                "trademc_trades": int(fiscal_purge.get("trademc_trades", 0)),
                "trademc_weight_transactions": int(fiscal_purge.get("trademc_weight_transactions", 0)),
                "fx_trades": int(fiscal_purge.get("fx_trades", 0)),
            }
        _clear_heavy_route_cache(["hedging:", "pmx_open_positions_reval", "profit_monthly"])
        result["clean_pipeline"] = _trigger_clean_pipeline("trademc_sync_bg", wait=False)
        with _trademc_sync_lock:
            _trademc_sync_status = {"running": False, "result": result}
    except Exception as exc:
        with _trademc_sync_lock:
            _trademc_sync_status = {"running": False, "result": {"error": str(exc)}}


@app.route("/api/trademc/sync", methods=["POST"])
def sync_trademc():
    global _trademc_sync_status
    data = request.json or {}
    include_weight = _pmx_bool(data.get("include_weight"), default=False)
    full = _pmx_bool(data.get("full"), default=False)
    incremental = _pmx_bool(data.get("incremental"), default=not full)
    if full:
        incremental = False
    replace = _pmx_bool(data.get("replace"), default=False) and (not incremental)
    wait = _pmx_bool(data.get("wait"), default=True)  # default: wait for result

    with _trademc_sync_lock:
        already_running = _trademc_sync_status.get("running", False)

    if already_running:
        return jsonify({"status": "already_running", "message": "Sync is already in progress"}), 202

    # Mark as running
    with _trademc_sync_lock:
        _trademc_sync_status = {"running": True, "result": None}

    if not wait:
        # Fire-and-forget: return immediately, sync runs in background
        t = threading.Thread(target=_run_trademc_sync_bg, args=(include_weight, incremental, replace), daemon=True)
        t.start()
        mode = "incremental" if incremental else ("full_replace" if replace else "full")
        mode_label = (
            "incremental sync"
            if mode == "incremental"
            else ("full replace sync" if mode == "full_replace" else "full sync")
        )
        return jsonify({"status": "started", "mode": mode, "message": f"{mode_label} started in background"}), 202

    # Synchronous path: run sync directly and return result
    try:
        company_result = sync_trademc_companies()
        trade_result = sync_trademc_trades(incremental=incremental, prune_missing=(replace and not incremental))
        if not bool(trade_result.get("success")):
            raise RuntimeError(str(trade_result.get("error", "TradeMC trade sync failed")))
        mode = (
            "incremental"
            if incremental
            else ("full_replace" if replace else "full")
        )
        response: Dict[str, Any] = {
            "companies": company_result,
            "trades": trade_result,
            "mode": mode,
        }
        if include_weight:
            response["weight_transactions"] = sync_trademc_weight_transactions(incremental=incremental)
        fiscal_cutoff = _pmx_filter_date(FISCAL_TRADES_START_DATE) if ENABLE_FISCAL_PURGE else ""
        response["fiscal_cutoff"] = fiscal_cutoff
        response["removed_pre_fiscal"] = {
            "trademc_trades": 0,
            "trademc_weight_transactions": 0,
            "fx_trades": 0,
        }
        if fiscal_cutoff:
            fiscal_purge = _purge_pre_fiscal_rows(cutoff_iso=fiscal_cutoff, purge_fx=True, purge_pmx=False)
            response["removed_pre_fiscal"] = {
                "trademc_trades": int(fiscal_purge.get("trademc_trades", 0)),
                "trademc_weight_transactions": int(fiscal_purge.get("trademc_weight_transactions", 0)),
                "fx_trades": int(fiscal_purge.get("fx_trades", 0)),
            }
        response["clean_pipeline"] = _trigger_clean_pipeline("trademc_sync", wait=False)
        with _trademc_sync_lock:
            _trademc_sync_status = {"running": False, "result": response}
        return jsonify(_json_safe(response))
    except Exception as exc:
        with _trademc_sync_lock:
            _trademc_sync_status = {"running": False, "result": {"error": str(exc)}}
        return jsonify({"error": str(exc)}), 500


@app.route("/api/trademc/sync/status", methods=["GET"])
def sync_trademc_status():
    with _trademc_sync_lock:
        return jsonify(_json_safe(_trademc_sync_status))


@app.route("/api/admin/clean-pipeline", methods=["GET", "POST"])
def clean_pipeline_status():
    if request.method == "GET":
        with CLEAN_PIPELINE_LOCK:
            return jsonify(_json_safe({"ok": True, "state": dict(_CLEAN_PIPELINE_STATE)}))

    current_user = getattr(g, "current_user", None)
    if not _auth_has_permission(current_user or {}, "write"):
        return jsonify({"ok": False, "error": "Missing write permission"}), 403

    data = request.json or {}
    wait = _pmx_bool(data.get("wait"), default=False)
    reason = str(data.get("reason", "manual") or "manual").strip() or "manual"
    result = _trigger_clean_pipeline(reason=reason, wait=wait)
    with CLEAN_PIPELINE_LOCK:
        return jsonify(_json_safe({"ok": True, "trigger": result, "state": dict(_CLEAN_PIPELINE_STATE)}))


@app.route("/api/trademc/diagnostics", methods=["GET"])
def trademc_diagnostics():
    trade_id_raw = str(request.args.get("trade_id", "") or "").strip()
    trade_id = int(trade_id_raw) if trade_id_raw.isdigit() else None

    local_stats = get_local_trademc_snapshot_stats()
    remote_stats = get_remote_trademc_snapshot_stats()

    trade_visibility: Dict[str, Any] = {}
    if trade_id is not None and trade_id > 0:
        trade_visibility = fetch_trademc_trade_by_id(trade_id)

    return jsonify(
        _json_safe(
            {
                "ok": True,
                "local": local_stats,
                "remote": remote_stats,
                "trade_visibility": trade_visibility,
            }
        )
    )


@app.route("/api/trademc/trades/<int:trade_id>/ref-number", methods=["PUT"])
def update_trademc_ref_number(trade_id):
    data = request.json or {}
    ref_number = str(data.get("ref_number", data.get("trade_number", "")) or "").strip()
    try:
        result = update_trademc_trade_ref_number(trade_id=trade_id, ref_number=ref_number)
        status_code = 200 if bool(result.get("success")) else int(result.get("status") or 400)
        return jsonify(_json_safe(result)), status_code
    except Exception as exc:
        err = {
            "success": False,
            "trade_id": int(trade_id),
            "error": f"TradeMC Ref# write-back failed: {exc}",
            "status": 500,
        }
        return jsonify(_json_safe(err)), 500


@app.route("/api/trademc/companies")
def get_companies():
    df = get_all_companies_df()
    return jsonify(df.fillna("").to_dict(orient="records"))


@app.route("/api/trademc/stats")
def trademc_stats():
    stats = get_trademc_stats()
    return jsonify(stats)


@app.route("/api/trademc/live-prices")
def get_trademc_live_prices():
    force_refresh = _is_truthy(request.args.get("force"))
    result = _get_cached_trademc_live_prices(force_refresh=force_refresh)
    if not bool(result.get("ok")):
        return jsonify(_json_safe(result)), 502
    return jsonify(_json_safe(result))


@app.route("/api/trademc/weight-transactions")
def get_weight_transactions():
    kwargs = {}
    for key in ["company_id", "type", "start_date", "end_date"]:
        val = request.args.get(key)
        if val:
            if key == "company_id":
                kwargs[key] = int(val)
            else:
                kwargs[key] = val
    df = load_weight_transactions(**kwargs)
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
    return jsonify(df.fillna("").to_dict(orient="records"))


@app.route("/api/trademc/weight-types")
def get_weight_types():
    return jsonify(get_unique_weight_types())


@app.route("/api/trademc/sync-weight", methods=["POST"])
def sync_weight():
    result = sync_trademc_weight_transactions()
    fiscal_cutoff = _pmx_filter_date(FISCAL_TRADES_START_DATE) if ENABLE_FISCAL_PURGE else ""
    result["fiscal_cutoff"] = fiscal_cutoff
    result["removed_pre_fiscal"] = {
        "trademc_weight_transactions": 0,
        "trademc_trades": 0,
        "fx_trades": 0,
    }
    if fiscal_cutoff:
        fiscal_purge = _purge_pre_fiscal_rows(cutoff_iso=fiscal_cutoff, purge_fx=True, purge_pmx=False)
        result["removed_pre_fiscal"] = {
            "trademc_weight_transactions": int(fiscal_purge.get("trademc_weight_transactions", 0)),
            "trademc_trades": int(fiscal_purge.get("trademc_trades", 0)),
            "fx_trades": int(fiscal_purge.get("fx_trades", 0)),
        }
    result["clean_pipeline"] = _trigger_clean_pipeline("trademc_sync_weight", wait=False)
    return jsonify(result)


@app.route("/api/hedging")
def get_hedging():
    rows = _get_cached_heavy_result(
        "hedging:pmx",
        lambda: build_hedging_comparison(),
    )
    return jsonify(_json_safe(rows))


@app.route("/api/weighted-average/<trade_num>")
def get_weighted_average(trade_num):
    result = build_weighted_average(trade_num)
    if result is None:
        return jsonify({"error": "No data found"}), 404
    return jsonify(result)


@app.route("/api/ticket/<path:trade_num>")
def get_ticket(trade_num):
    result = build_trading_ticket(trade_num)
    if result is None:
        return jsonify({"error": "No data found"}), 404
    return jsonify(result)


def _normalize_tradebook_supplier(value: Any) -> str:
    text = str(value or "").upper().strip()
    text = re.sub(r"^\s*\d+\s*[\)\.\-:]*\s*", "", text)
    text = re.sub(r"[^A-Z0-9&]+", " ", text)
    text = re.sub(r"\bLTD\b", "LIMITED", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalize_tradebook_number(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace(",", "").replace(" ", "")
    text = text.replace("R/", "").replace("$", "").replace("R", "")
    text = text.replace("–", "-").replace("—", "-")
    if not text:
        return ""
    if re.fullmatch(r"-?\d+(?:\.\d+)?", text) is None:
        return ""
    try:
        dec = Decimal(text)
    except (InvalidOperation, ValueError):
        return ""
    normalized = format(dec, "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    return normalized if normalized else "0"


def _tradebook_key(row: Dict[str, Any]) -> Tuple[str, str, str, str]:
    return (
        _normalize_tradebook_supplier(row.get("Supplier")),
        _normalize_tradebook_number(row.get("Weight (g)")),
        _normalize_tradebook_number(row.get("Gold Rate ($/oz)")),
        _normalize_tradebook_number(row.get("Exchange Rate")),
    )


def _tradebook_key_to_row(key: Tuple[str, str, str, str]) -> Dict[str, Any]:
    def _num_or_text(v: str) -> Any:
        if not v:
            return ""
        try:
            return float(v)
        except Exception:
            return v

    return {
        "Supplier": key[0],
        "Weight (g)": _num_or_text(key[1]),
        "Gold Rate ($/oz)": _num_or_text(key[2]),
        "Exchange Rate": _num_or_text(key[3]),
    }


def _extract_trademc_truth_rows(ticket_payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    tm_rows = ticket_payload.get("trademc") or []
    if not isinstance(tm_rows, list):
        return []

    out: List[Dict[str, Any]] = []
    for raw in tm_rows:
        if not isinstance(raw, dict):
            continue
        supplier = raw.get("Company")
        if str(supplier or "").strip().upper() == "TRADING CONTROL ACCOUNT":
            continue
        row = {
            "Supplier": supplier,
            "Weight (g)": raw.get("Weight (g)"),
            "Gold Rate ($/oz)": raw.get("$/oz Booked"),
            "Exchange Rate": raw.get("FX Rate"),
        }
        key = _tradebook_key(row)
        if not all(key):
            continue
        out.append({
            "Supplier": key[0],
            "Weight (g)": float(key[1]),
            "Gold Rate ($/oz)": float(key[2]),
            "Exchange Rate": float(key[3]),
        })
    return out


def _extract_trademc_truth_rows_from_uploaded_rows(rows: Any) -> List[Dict[str, Any]]:
    if not isinstance(rows, list):
        return []
    out: List[Dict[str, Any]] = []
    for raw in rows:
        if not isinstance(raw, dict):
            continue
        row = {
            "Supplier": raw.get("Supplier", raw.get("Company")),
            "Weight (g)": raw.get("Weight (g)"),
            "Gold Rate ($/oz)": raw.get("Gold Rate ($/oz)", raw.get("$/oz Booked")),
            "Exchange Rate": raw.get("Exchange Rate", raw.get("FX Rate")),
        }
        key = _tradebook_key(row)
        if not all(key):
            continue
        out.append({
            "Supplier": key[0],
            "Weight (g)": float(key[1]),
            "Gold Rate ($/oz)": float(key[2]),
            "Exchange Rate": float(key[3]),
        })
    return out


def _parse_tradebook_rows_from_image_bytes(
    image_bytes: bytes,
    mime_type: str = "image/png",
    trade_mc_rows: Optional[List[Dict[str, Any]]] = None,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    api_key = str(os.getenv("GEMINI_API_KEY", "") or "").strip()
    if not api_key:
        raise RuntimeError("Missing GEMINI_API_KEY in environment.")

    env_model = str(os.getenv("GEMINI_MODEL", "") or "").strip()
    model_candidates: List[str] = []
    if env_model:
        model_candidates.append(env_model)
    model_candidates.extend([
        "gemini-2.5-flash",
        "gemini-2.0-flash",
        "gemini-1.5-flash-latest",
        "gemini-1.5-flash",
    ])
    # Preserve order while removing duplicates.
    seen_models = set()
    model_candidates = [m for m in model_candidates if not (m in seen_models or seen_models.add(m))]

    api_versions = ["v1beta", "v1"]
    tm_rows = trade_mc_rows or []
    tm_lines: List[str] = []
    for idx, row in enumerate(tm_rows):
        tm_lines.append(
            f"{idx + 1}. Supplier={row.get('Supplier','')}, Weight={row.get('Weight (g)','')}, PM Price={row.get('Gold Rate ($/oz)','')}, FX Rate={row.get('Exchange Rate','')}"
        )
    tm_block = "\n".join(tm_lines) if tm_lines else "None provided"
    prompt = (
        "You are validating a manual trading book screenshot against TradeMC source rows.\n"
        "Task:\n"
        "1) Read ONLY supplier table rows from screenshot.\n"
        "2) For EACH screenshot row, match it to ONE TradeMC row by comparing Weight, PM Price, FX Rate first, then supplier name.\n"
        "3) Handle normal supplier aliases: LTD == LIMITED, (PTY) == PTY, optional punctuation/brackets.\n"
        "4) Return one output row per screenshot row. Do not deduplicate.\n"
        "5) If a TradeMC row match is found, set Supplier to the matched TradeMC Supplier text.\n"
        "6) If no match is found, keep screenshot Supplier text and add warning.\n\n"
        f"TradeMC rows:\n{tm_block}\n\n"
        "Return strict JSON only with shape: "
        "{\"rows\":[{\"Supplier\":\"...\",\"Weight\":\"...\",\"PM Price\":\"...\",\"FX Rate\":\"...\"}],\"warnings\":[\"...\"]}"
    )
    payload = {
        "contents": [
            {
                "parts": [
                    {"text": prompt},
                    {
                        "inline_data": {
                            "mime_type": str(mime_type or "image/png"),
                            "data": base64.b64encode(image_bytes).decode("ascii"),
                        }
                    },
                ]
            }
        ],
        "generationConfig": {
            "temperature": 0,
            "responseMimeType": "application/json",
        },
    }
    resp = None
    last_model_err = ""
    transient_statuses = {429, 500, 502, 503, 504}
    for api_version in api_versions:
        for model in model_candidates:
            endpoint = f"https://generativelanguage.googleapis.com/{api_version}/models/{model}:generateContent"
            for attempt in range(2):
                try:
                    trial = requests.post(
                        endpoint,
                        params={"key": api_key},
                        json=payload,
                        timeout=60,
                    )
                except Exception as exc:
                    last_model_err = f"Gemini request failed for {model} ({api_version}): {exc}"
                    if attempt == 0:
                        time.sleep(0.35)
                        continue
                    break

                if trial.ok:
                    resp = trial
                    break

                body = str(trial.text or "")
                status = int(trial.status_code)
                is_model_not_found = (
                    status == 404 and
                    ("models/" in body or "NOT_FOUND" in body or "not found" in body.lower())
                )
                if is_model_not_found:
                    last_model_err = f"Model not found: {model} ({api_version})"
                    break

                if status in transient_statuses:
                    last_model_err = f"Temporary Gemini error ({status}) on {model}/{api_version}"
                    if attempt == 0:
                        time.sleep(0.5)
                        continue
                    break

                body_preview = body.strip()[:400]
                raise RuntimeError(f"Gemini API error ({status}) on {model}/{api_version}: {body_preview}")

            if resp is not None:
                break
        if resp is not None:
            break

    if resp is None:
        raise RuntimeError(
            "Gemini model discovery failed. "
            f"Tried: {', '.join(model_candidates)} across {', '.join(api_versions)}. "
            f"Last error: {last_model_err or 'unknown'}"
        )

    try:
        data = resp.json()
    except Exception:
        raise RuntimeError("Gemini returned non-JSON response.")

    candidates = data.get("candidates") if isinstance(data, dict) else None
    if not isinstance(candidates, list) or not candidates:
        raise RuntimeError("Gemini returned no candidates.")
    parts = (((candidates[0] or {}).get("content") or {}).get("parts")) if isinstance(candidates[0], dict) else None
    text = ""
    if isinstance(parts, list):
        for p in parts:
            if isinstance(p, dict) and isinstance(p.get("text"), str):
                text += p["text"]
    text = text.strip()
    if not text:
        raise RuntimeError("Gemini returned empty extraction output.")

    try:
        obj = json.loads(text)
    except Exception:
        text_clean = re.sub(r"^```json\s*|\s*```$", "", text, flags=re.IGNORECASE | re.DOTALL).strip()
        try:
            obj = json.loads(text_clean)
        except Exception:
            raise RuntimeError(f"Unable to parse Gemini JSON output: {text[:220]}")

    rows_raw = obj.get("rows") if isinstance(obj, dict) else None
    warnings_raw = obj.get("warnings") if isinstance(obj, dict) else None
    warnings: List[str] = [str(w).strip() for w in (warnings_raw or []) if str(w).strip()] if isinstance(warnings_raw, list) else []
    if not isinstance(rows_raw, list):
        return [], warnings or ["Gemini did not return row data."]

    parsed_rows: List[Dict[str, Any]] = []
    for raw in rows_raw:
        if not isinstance(raw, dict):
            continue
        row = {
            "Supplier": raw.get("Supplier", raw.get("supplier")),
            "Weight (g)": raw.get("Weight", raw.get("weight")),
            "Gold Rate ($/oz)": raw.get("PM Price", raw.get("pm_price")),
            "Exchange Rate": raw.get("FX Rate", raw.get("fx_rate")),
        }
        key = _tradebook_key(row)
        if not all(key):
            continue
        parsed_rows.append({
            "Supplier": key[0],
            "Weight (g)": float(key[1]),
            "Gold Rate ($/oz)": float(key[2]),
            "Exchange Rate": float(key[3]),
        })

    if not parsed_rows:
        warnings.append("Gemini did not return any valid supplier rows.")
    return parsed_rows, warnings


def _check_tradebook_against_trademc_impl(trade_num: str):
    upload = request.files.get("image") or request.files.get("screenshot")
    if upload is None:
        return jsonify({"error": "Missing image upload. Use form field 'image'."}), 400

    image_bytes = upload.read() or b""
    if not image_bytes:
        return jsonify({"error": "Uploaded image is empty."}), 400

    uploaded_truth_rows: List[Dict[str, Any]] = []
    uploaded_truth_json = str(request.form.get("trademc_rows_json") or "").strip()
    if uploaded_truth_json:
        try:
            payload_rows = json.loads(uploaded_truth_json)
            uploaded_truth_rows = _extract_trademc_truth_rows_from_uploaded_rows(payload_rows)
        except Exception:
            uploaded_truth_rows = []

    ticket = None
    truth_rows = uploaded_truth_rows
    if not truth_rows:
        return jsonify({
            "error": (
                "No TradeMC source rows were provided for comparison. "
                "Load the ticket first so TradeMC rows are visible, then upload screenshot again."
            )
        }), 400

    try:
        parsed_rows, parse_warnings = _parse_tradebook_rows_from_image_bytes(
            image_bytes=image_bytes,
            mime_type=str(getattr(upload, "mimetype", "") or "image/png"),
            trade_mc_rows=truth_rows,
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400

    parsed_counter = Counter(_tradebook_key(r) for r in parsed_rows if all(_tradebook_key(r)))
    truth_counter = Counter(_tradebook_key(r) for r in truth_rows if all(_tradebook_key(r)))

    matched = 0
    for key, parsed_count in parsed_counter.items():
        matched += min(parsed_count, truth_counter.get(key, 0))

    manual_not_in_trademc: List[Dict[str, Any]] = []
    missing_from_manual_book: List[Dict[str, Any]] = []
    all_keys = sorted(set(parsed_counter.keys()) | set(truth_counter.keys()))
    for key in all_keys:
        parsed_count = parsed_counter.get(key, 0)
        truth_count = truth_counter.get(key, 0)
        if parsed_count > truth_count:
            manual_not_in_trademc.extend([_tradebook_key_to_row(key) for _ in range(parsed_count - truth_count)])
        elif truth_count > parsed_count:
            missing_from_manual_book.extend([_tradebook_key_to_row(key) for _ in range(truth_count - parsed_count)])

    return jsonify({
        "ok": True,
        "trade_num": str((ticket or {}).get("trade_num") or trade_num),
        "source_of_truth": "TradeMC",
        "parsed_rows": parsed_rows,
        "parse_warnings": parse_warnings,
        "manual_not_in_trademc": manual_not_in_trademc,
        "missing_from_manual_book": missing_from_manual_book,
        "counts": {
            "parsed_rows": len(parsed_rows),
            "trademc_rows": len(truth_rows),
            "matched_rows": int(matched),
            "manual_not_in_trademc": len(manual_not_in_trademc),
            "missing_from_manual_book": len(missing_from_manual_book),
        },
    })


@app.route("/api/ticket/<path:trade_num>/book-check", methods=["POST"])
def check_tradebook_against_trademc(trade_num):
    return _check_tradebook_against_trademc_impl(str(trade_num or "").strip())


@app.route("/api/ticket/book-check", methods=["POST"])
def check_tradebook_against_trademc_flat():
    trade_num = str(
        request.form.get("trade_num")
        or request.args.get("trade_num")
        or (request.get_json(silent=True) or {}).get("trade_num")
        or ""
    ).strip()
    if not trade_num:
        return jsonify({"error": "Missing trade_num in form/query/json."}), 400
    return _check_tradebook_against_trademc_impl(trade_num)


@app.route("/api/ticket/<path:trade_num>/pdf")
def get_ticket_pdf(trade_num):
    frames = build_trading_ticket_frames(trade_num)
    if frames is None:
        return jsonify({"error": "No data found"}), 404

    try:
        pdf_bytes = build_trading_ticket_pdf(
            frames["trade_num"],
            frames["tm_detail"],
            frames["stonex_rows"],
            frames["summary"],
        )
    except Exception as exc:
        import traceback as _tb
        _tb.print_exc()
        return jsonify({"error": f"PDF generation failed: {exc}"}), 500
    if not pdf_bytes:
        return jsonify({"error": "PDF export requires the `reportlab` or `fpdf2` package."}), 400

    resp = Response(pdf_bytes, mimetype="application/pdf")
    safe_name = str(frames['trade_num']).replace('/', '_').replace('\\', '_')
    resp.headers["Content-Disposition"] = f"attachment; filename=trading_ticket_{safe_name}.pdf"
    resp.headers["Content-Length"] = str(len(pdf_bytes))
    return resp


@app.route("/api/profit/monthly")
def get_profit_monthly():
    result = _get_cached_heavy_result(
        "profit_monthly",
        build_profit_monthly_report,
    )
    return jsonify(_json_safe(result))


@app.route("/api/reports/daily-trading/send-test", methods=["POST"])
def send_daily_trading_report_test():
    data = request.get_json(silent=True) or {}
    run_date = str(data.get("date") or datetime.now().strftime("%Y-%m-%d")).strip()
    recipients_raw = data.get("to")
    recipients_override = _daily_trading_report_parse_recipients(recipients_raw)
    if recipients_raw and not recipients_override:
        return jsonify({"ok": False, "error": "Invalid recipient list in 'to'."}), 400
    result = _daily_trading_report_run_once_for_date(
        run_date_iso=run_date,
        recipients_override=recipients_override or None,
    )
    status = 200 if result.get("ok") else 500
    return jsonify(_json_safe(result)), status


@app.route("/api/export-trades/save", methods=["POST"])
def export_trades_save_to_folder():
    data = request.json or {}
    trades_payload = data.get("trades", [])
    if not isinstance(trades_payload, list) or len(trades_payload) == 0:
        return jsonify({"ok": False, "error": "Missing trades payload"}), 400

    output_dir = str(data.get("output_dir") or PMX_EXPORT_TRADES_DIR or "").strip() or PMX_EXPORT_TRADES_DIR
    output_dir = os.path.normpath(output_dir)
    try:
        os.makedirs(output_dir, exist_ok=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"Unable to create export path '{output_dir}': {exc}", "output_dir": output_dir}), 500

    saved_files: List[str] = []
    errors: List[str] = []
    warnings: List[str] = []
    complete_trades = 0
    partial_trades = 0
    failed_trades = 0
    selected_trade_nums: List[str] = []

    for trade_item in trades_payload:
        if not isinstance(trade_item, dict):
            continue
        trade_num = normalize_trade_number(trade_item.get("trade_num"))
        if not trade_num:
            continue
        selected_trade_nums.append(trade_num)

        trade_saved_count = 0
        trade_errors: List[str] = []

        raw_fnc_values = trade_item.get("fnc_numbers", [])
        fnc_tokens: List[str] = []
        if isinstance(raw_fnc_values, list):
            for raw_val in raw_fnc_values:
                text = str(raw_val or "").strip()
                if not text:
                    continue
                matches = PMX_SUPPORT_DOC_PATTERN.findall(text)
                if matches:
                    fnc_tokens.extend([str(m).strip().upper() for m in matches if str(m).strip()])
                else:
                    fnc_tokens.append(text.upper())
        elif raw_fnc_values:
            text = str(raw_fnc_values).strip()
            if text:
                matches = PMX_SUPPORT_DOC_PATTERN.findall(text)
                if matches:
                    fnc_tokens.extend([str(m).strip().upper() for m in matches if str(m).strip()])
                else:
                    fnc_tokens.append(text.upper())

        seen_tokens = set()
        fnc_tokens = [t for t in fnc_tokens if not (t in seen_tokens or seen_tokens.add(t))]

        for cell in fnc_tokens:
            try:
                doc_type = str(cell.split("/", 1)[0] if "/" in cell else "FNC").upper()
                pdf_result = _build_pmx_fnc_pdf_result(
                    cell=cell,
                    doc_type=doc_type,
                    data=data,
                    req_headers=request.headers,
                )
                if (not pdf_result.get("ok")) and doc_type != "FNC":
                    pdf_result = _build_pmx_fnc_pdf_result(
                        cell=cell,
                        doc_type="FNC",
                        data=data,
                        req_headers=request.headers,
                    )
                if not pdf_result.get("ok"):
                    trade_errors.append(f"FNC {cell}: {pdf_result.get('error', 'PMX PDF download failed')}")
                    continue

                file_name = f"{_sanitize_filename_component(trade_num, 'trade')}_{_sanitize_filename_component(pdf_result.get('filename'), 'fixing_invoice.pdf')}"
                file_path = os.path.join(output_dir, file_name)
                file_bytes = pdf_result.get("bytes", b"")
                if not isinstance(file_bytes, (bytes, bytearray)):
                    file_bytes = bytes(str(file_bytes), "utf-8")
                with open(file_path, "wb") as fh:
                    fh.write(bytes(file_bytes))
                saved_files.append(file_path)
                trade_saved_count += 1
            except Exception as exc:
                trade_errors.append(f"FNC {cell}: {exc}")

        try:
            frames = build_trading_ticket_frames(trade_num)
            if frames is None:
                trade_errors.append("Trading ticket: No data found")
            else:
                pdf_bytes = build_trading_ticket_pdf(
                    frames["trade_num"],
                    frames["tm_detail"],
                    frames["stonex_rows"],
                    frames["summary"],
                )
                if not pdf_bytes:
                    trade_errors.append("Trading ticket: PDF export requires the `reportlab` or `fpdf2` package.")
                else:
                    ticket_name = f"{_sanitize_filename_component(trade_num, 'trade')}_ticket.pdf"
                    ticket_path = os.path.join(output_dir, ticket_name)
                    with open(ticket_path, "wb") as fh:
                        fh.write(bytes(pdf_bytes))
                    saved_files.append(ticket_path)
                    trade_saved_count += 1
        except Exception as exc:
            trade_errors.append(f"Trading ticket: {exc}")

        if trade_errors:
            errors.append(f"{trade_num}: {trade_errors[0]}")
        if trade_saved_count > 0 and not trade_errors:
            complete_trades += 1
        elif trade_saved_count > 0 and trade_errors:
            partial_trades += 1
        else:
            failed_trades += 1

    sales_order_path = ""
    sales_order_result = _build_sales_order_excel_for_trades(selected_trade_nums, output_dir)
    if not sales_order_result.get("ok"):
        warnings.append(f"Sales Order Excel: {sales_order_result.get('error', 'Failed to generate Sales Order Excel')}")
    else:
        sales_order_path = str(sales_order_result.get("path") or "").strip()
        if sales_order_path:
            # Safety net: always include selected trade numbers in exported sales-order filename.
            expected_suffix = _build_trade_filename_suffix(selected_trade_nums)
            current_name = os.path.basename(sales_order_path)
            expected_marker = f"sales_orders_{expected_suffix}" if expected_suffix else ""
            if expected_marker and expected_marker not in current_name:
                target_name = f"{expected_marker}.xlsx"
                target_path = os.path.join(output_dir, target_name)
                if os.path.abspath(target_path) != os.path.abspath(sales_order_path):
                    if os.path.exists(target_path):
                        base_name, ext = os.path.splitext(target_name)
                        suffix = 2
                        while os.path.exists(os.path.join(output_dir, f"{base_name}_{suffix}{ext}")):
                            suffix += 1
                        target_name = f"{base_name}_{suffix}{ext}"
                        target_path = os.path.join(output_dir, target_name)
                    try:
                        os.replace(sales_order_path, target_path)
                        sales_order_path = target_path
                    except Exception as exc:
                        warnings.append(f"Sales Order Excel rename failed: {exc}")
        if sales_order_path:
            saved_files.append(sales_order_path)

    result = {
        "ok": failed_trades == 0,
        "output_dir": output_dir,
        "requested_trades": len(trades_payload),
        "complete_trades": complete_trades,
        "partial_trades": partial_trades,
        "failed_trades": failed_trades,
        "saved_file_count": len(saved_files),
        "saved_files": saved_files[:200],
        "sales_order_file": sales_order_path,
        "errors": errors[:200],
        "warnings": warnings[:200],
    }
    return jsonify(_json_safe(result))


@app.route("/api/export/ledger")
def export_ledger():
    fmt = request.args.get("format", "csv")
    df = load_all_trades()
    ledger = build_ledger_view(df)

    if fmt == "excel":
        buf = io.BytesIO()
        ledger.to_excel(buf, index=False, sheet_name="Trading Ledger")
        buf.seek(0)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="trading_ledger.xlsx")
    else:
        buf = io.StringIO()
        ledger.to_csv(buf, index=False)
        buf.seek(0)
        return send_file(io.BytesIO(buf.getvalue().encode()), mimetype="text/csv",
                         as_attachment=True, download_name="trading_ledger.csv")


# ---------------------------------------------------------------------------
# Forecast endpoints  (gold / usdzar / purchases Monte Carlo)
# ---------------------------------------------------------------------------

_FORECAST_PAIR_MAP = {
    "gold": ("XAU", "USD"),
    "usdzar": ("USD", "ZAR"),
    "purchases": ("XAU", "ZAR"),
}


@app.route("/api/forecast/<pair>")
def api_forecast_get(pair: str):
    symbols = _FORECAST_PAIR_MAP.get(pair.lower().strip())
    if not symbols:
        return jsonify({"ok": False, "error": f"Unknown pair '{pair}'. Use gold, usdzar, or purchases."}), 400
    try:
        days = int(request.args.get("days", 30))
        sims = int(request.args.get("sims", 10000))
    except Exception:
        days, sims = 30, 10000
    try:
        result = _forecast_get(symbols[0], symbols[1], forecast_days=days, n_simulations=sims)
        result["ok"] = True
        return jsonify(_json_safe(result))
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/forecast/current-price/<pair>")
def api_forecast_current_price(pair: str):
    symbols = _FORECAST_PAIR_MAP.get(pair.lower().strip())
    if not symbols:
        return jsonify({"ok": False, "error": f"Unknown pair '{pair}'."}), 400
    try:
        result = _forecast_current_price(symbols[0], symbols[1])
        return jsonify({"ok": True, "rate": result["rate"], "last_refreshed": result.get("last_refreshed", "")})
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/forecast/refresh/<pair>", methods=["POST"])
def api_forecast_refresh(pair: str):
    symbols = _FORECAST_PAIR_MAP.get(pair.lower().strip())
    if not symbols:
        return jsonify({"ok": False, "error": f"Unknown pair '{pair}'."}), 400
    data = request.get_json(silent=True) or {}
    try:
        days = int(data.get("days", 30))
        sims = int(data.get("sims", 10000))
    except Exception:
        days, sims = 30, 10000
    try:
        _forecast_invalidate_cache(symbols[0], symbols[1])
        result = _forecast_get(symbols[0], symbols[1], forecast_days=days, n_simulations=sims)
        result["ok"] = True
        return jsonify(_json_safe(result))
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/forecast/purchases/predict")
def api_forecast_purchases_predict():
    try:
        spot_usd = float(request.args["spot_usd"]) if request.args.get("spot_usd") else None
    except Exception:
        spot_usd = None
    try:
        fx_rate = float(request.args["fx_rate"]) if request.args.get("fx_rate") else None
    except Exception:
        fx_rate = None
    try:
        result = _purchases_get_forecast(spot_usd=spot_usd, fx_rate=fx_rate)
        return jsonify(_json_safe(result))
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/forecast/purchases/train", methods=["POST"])
def api_forecast_purchases_train():
    try:
        result = _purchases_train_model(force=True)
        return jsonify(_json_safe(result))
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/forecast/purchases/status")
def api_forecast_purchases_status():
    try:
        result = _purchases_model_status()
        return jsonify(_json_safe(result))
    except Exception as exc:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(exc)}), 500


if __name__ == "__main__":
    try:
        server_port = int(str(os.getenv("PORT", os.getenv("J2_API_PORT", "5003")) or "5003").strip())
    except Exception:
        server_port = 5003
    print(f"J2 API Server starting on http://localhost:{server_port}")
    _start_daily_balance_email_scheduler()
    _start_daily_trading_report_email_scheduler()
    app.run(host="0.0.0.0", port=server_port, debug=True, use_reloader=False)
